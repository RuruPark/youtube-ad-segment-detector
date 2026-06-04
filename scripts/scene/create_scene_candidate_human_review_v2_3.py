#!/usr/bin/env python3
"""장면 전환 후보 검토용 v2.3 workbook을 만든다."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import shutil
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(".")
OLD_PROJECT_ROOT = Path("./_old_project_not_included")

INPUT_CANDIDATE_PATH = PROJECT_ROOT / "data/scene/scene_candidates_v2_3_merged_ffmpeg_fallback_mmss.csv"
INPUT_BOUNDARY_AUDIT_PATH = PROJECT_ROOT / "data/scene/scene_candidate_boundary_audit_v2_3_merged_ffmpeg_fallback.csv"
INPUT_AD_SEGMENT_PATH = PROJECT_ROOT / "data/segments/ad_interval_segments_v2_3.csv"
INPUT_LABEL_CONTEXT_PATH = PROJECT_ROOT / "data/segments/label_interval_context_segments_v2_3.csv"
INPUT_MANIFEST_PATH = PROJECT_ROOT / "data/video_metadata/video_manifest_v2_2.csv"
OPTIONAL_FFMPEG_REPORT_PATH = PROJECT_ROOT / "reports/ffmpeg_fallback_scene_change_failed4_v2_3_report.json"
OPTIONAL_FFMPEG_SUMMARY_PATH = PROJECT_ROOT / "reports/ffmpeg_fallback_scene_change_failed4_v2_3_summary.md"

REVIEW_DIR = PROJECT_ROOT / "data/review"
OUTPUT_XLSX_PATH = REVIEW_DIR / "scene_candidate_human_review_v2_3.xlsx"
OUTPUT_CANDIDATE_CSV_PATH = REVIEW_DIR / "scene_candidate_human_review_candidate_sheet_v2_3.csv"
OUTPUT_BOUNDARY_CSV_PATH = REVIEW_DIR / "scene_candidate_human_review_boundary_sheet_v2_3.csv"
REPORT_PATH = PROJECT_ROOT / "reports/scene_candidate_human_review_v2_3_report.json"
SUMMARY_PATH = PROJECT_ROOT / "reports/scene_candidate_human_review_v2_3_summary.md"
LOG_PATH = PROJECT_ROOT / "logs/scene_candidate_human_review_v2_3_run_log.txt"
README_PATH = PROJECT_ROOT / "README.md"
LATEST_DIR = PROJECT_ROOT / "outputs/latest_for_chatgpt"
LATEST_README_PATH = LATEST_DIR / "README_latest_files.md"

ESTIMATED_RUNTIME = "약 20분"
RUNTIME_ESTIMATION_REASON = (
    "scene candidate 707개 내외, 광고 boundary audit 21개 interval, "
    "Excel sheet/validation/report 생성 작업 기준"
)
SHEET_NAMES = [
    "scene_candidate_review",
    "ad_boundary_review",
    "video_summary_review",
    "value_options",
    "review_guide",
]
RUN_LOG: list[str] = []


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def log(message: str) -> None:
    line = f"[{now_iso()}] {message}"
    RUN_LOG.append(line)
    print(message, flush=True)


def readable_runtime(seconds: float) -> str:
    total = int(round(seconds))
    minutes, sec = divmod(total, 60)
    return f"{minutes}분 {sec}초"


def strip_bom_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(col).replace("\ufeff", "").strip() for col in df.columns]
    return df


def read_csv_optional(path: Path, missing_input_files: list[str], warnings: list[Any]) -> pd.DataFrame:
    if not path.exists():
        missing_input_files.append(str(path))
        warnings.append({"missing_input_file": str(path)})
        return pd.DataFrame()
    try:
        return strip_bom_columns(pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8-sig"))
    except Exception as exc:
        missing_input_files.append(str(path))
        warnings.append({"failed_to_read_input": str(path), "error": str(exc)})
        return pd.DataFrame()


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_float(value: Any) -> float | None:
    text = clean(value)
    if text == "":
        return None
    try:
        return float(text)
    except Exception:
        return None


def to_bool(value: Any) -> bool:
    text = clean(value).lower()
    return text in {"true", "1", "yes", "y", "t"}


def bool_text(value: bool) -> str:
    return "true" if value else "false"


def fmt_number(value: Any, digits: int = 3) -> Any:
    num = to_float(value)
    if num is None:
        return ""
    rounded = round(num, digits)
    if abs(rounded - round(rounded)) < 10 ** (-(digits + 1)):
        return int(round(rounded))
    return rounded


def mmss(seconds: Any) -> str:
    sec = to_float(seconds)
    if sec is None:
        return ""
    total = max(0, int(math.floor(sec)))
    return f"{total // 60:02d}분 {total % 60:02d}초"


def first_present(row: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value = clean(row.get(key))
        if value != "":
            return value
    return ""


def safe_row_dict(row: pd.Series) -> dict[str, Any]:
    return {str(k): clean(v) for k, v in row.to_dict().items()}


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in columns})


def old_project_snapshot() -> dict[str, Any]:
    if not OLD_PROJECT_ROOT.exists():
        return {"exists": False, "file_count": 0, "digest": ""}
    h = hashlib.sha256()
    file_count = 0
    for path in sorted(p for p in OLD_PROJECT_ROOT.rglob("*") if p.is_file()):
        try:
            stat = path.stat()
        except OSError:
            continue
        rel = path.relative_to(OLD_PROJECT_ROOT).as_posix()
        h.update(f"{rel}\t{stat.st_size}\t{stat.st_mtime_ns}\n".encode("utf-8", errors="replace"))
        file_count += 1
    return {"exists": True, "file_count": file_count, "digest": h.hexdigest()}


def build_manifest_lookup(manifest_df: pd.DataFrame) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    if manifest_df.empty:
        return lookup
    for _, row in manifest_df.iterrows():
        data = safe_row_dict(row)
        keys = [
            data.get("video_id", ""),
            data.get("label_mapping_video_id", ""),
            data.get("video_filename", ""),
            data.get("file_stem", ""),
        ]
        for key in keys:
            if clean(key):
                lookup[clean(key)] = data
    return lookup


def manifest_value(lookup: dict[str, dict[str, Any]], video_id: str, video_filename: str, keys: list[str]) -> str:
    row = lookup.get(clean(video_id)) or lookup.get(clean(video_filename)) or {}
    return first_present(row, keys)


def prepare_ad_intervals(ad_df: pd.DataFrame, boundary_df: pd.DataFrame) -> list[dict[str, Any]]:
    source_df = ad_df if not ad_df.empty else boundary_df
    if source_df.empty:
        return []
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for _, row in source_df.iterrows():
        data = safe_row_dict(row)
        ad_interval_id = clean(data.get("ad_interval_id")) or clean(data.get("segment_id"))
        if not ad_interval_id or ad_interval_id in seen:
            continue
        seen.add(ad_interval_id)
        start = first_present(data, ["ad_start_sec", "segment_start_sec"])
        end = first_present(data, ["ad_end_sec", "segment_end_sec"])
        rows.append(
            {
                "ad_interval_id": ad_interval_id,
                "video_id": clean(data.get("video_id")),
                "video_title": clean(data.get("video_title")),
                "video_filename": clean(data.get("video_filename")),
                "video_path": clean(data.get("video_path")),
                "ad_start_sec": start,
                "ad_end_sec": end,
                "ad_start_mmss": clean(data.get("ad_start_mmss")) or mmss(start),
                "ad_end_mmss": clean(data.get("ad_end_mmss")) or mmss(end),
                "video_duration_sec": clean(data.get("video_duration_sec")),
            }
        )
    return rows


def index_intervals_by_video(ad_intervals: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for interval in ad_intervals:
        by_video[clean(interval.get("video_id"))].append(interval)
    return by_video


def interval_metrics_for_candidate(
    video_intervals: list[dict[str, Any]], candidate_time: float | None
) -> dict[str, Any]:
    if candidate_time is None or not video_intervals:
        return {
            "nearest_ad_boundary_type": "",
            "nearest_ad_boundary_sec": "",
            "nearest_ad_boundary_mmss": "",
            "distance_to_nearest_ad_boundary_sec": "",
            "is_near_ad_start_3s": "false",
            "is_near_ad_start_5s": "false",
            "is_near_ad_end_3s": "false",
            "is_near_ad_end_5s": "false",
            "is_near_any_ad_boundary_5s": "false",
            "candidate_inside_ad_interval": False,
        }

    nearest: dict[str, Any] | None = None
    start_dists: list[float] = []
    end_dists: list[float] = []
    inside = False
    for interval in video_intervals:
        start = to_float(interval.get("ad_start_sec"))
        end = to_float(interval.get("ad_end_sec"))
        if start is not None:
            dist = abs(candidate_time - start)
            start_dists.append(dist)
            item = {"type": "ad_start", "sec": start, "dist": dist}
            if nearest is None or dist < nearest["dist"]:
                nearest = item
        if end is not None:
            dist = abs(candidate_time - end)
            end_dists.append(dist)
            item = {"type": "ad_end", "sec": end, "dist": dist}
            if nearest is None or dist < nearest["dist"]:
                nearest = item
        if start is not None and end is not None and start <= candidate_time <= end:
            inside = True

    min_start = min(start_dists) if start_dists else None
    min_end = min(end_dists) if end_dists else None
    near_start_3 = min_start is not None and min_start <= 3
    near_start_5 = min_start is not None and min_start <= 5
    near_end_3 = min_end is not None and min_end <= 3
    near_end_5 = min_end is not None and min_end <= 5
    nearest_sec = nearest["sec"] if nearest else ""
    nearest_dist = nearest["dist"] if nearest else ""
    return {
        "nearest_ad_boundary_type": nearest["type"] if nearest else "",
        "nearest_ad_boundary_sec": fmt_number(nearest_sec),
        "nearest_ad_boundary_mmss": mmss(nearest_sec),
        "distance_to_nearest_ad_boundary_sec": fmt_number(nearest_dist),
        "is_near_ad_start_3s": bool_text(near_start_3),
        "is_near_ad_start_5s": bool_text(near_start_5),
        "is_near_ad_end_3s": bool_text(near_end_3),
        "is_near_ad_end_5s": bool_text(near_end_5),
        "is_near_any_ad_boundary_5s": bool_text(near_start_5 or near_end_5),
        "candidate_inside_ad_interval": inside,
    }


def add_score_rank_columns(candidate_df: pd.DataFrame) -> pd.DataFrame:
    if candidate_df.empty:
        return candidate_df
    df = candidate_df.copy()
    if "scene_change_score" not in df.columns:
        df["score_rank_in_video"] = ""
        df["score_percentile_in_video"] = ""
        return df
    scores = pd.to_numeric(df["scene_change_score"], errors="coerce")
    df["_score_numeric"] = scores
    df["score_rank_in_video"] = (
        df.groupby("video_id")["_score_numeric"].rank(method="first", ascending=False).fillna("").astype(str)
    )
    percentile = df.groupby("video_id")["_score_numeric"].rank(pct=True, ascending=True) * 100
    df["score_percentile_in_video"] = percentile.round(2).fillna("").astype(str)
    df = df.drop(columns=["_score_numeric"])
    return df


def build_candidate_review(
    candidate_df: pd.DataFrame, ad_intervals_by_video: dict[str, list[dict[str, Any]]], manifest_lookup: dict[str, dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[str]]:
    requested_source_columns = [
        "video_id",
        "video_title",
        "video_filename",
        "video_path",
        "candidate_time_sec",
        "candidate_time_mmss",
        "candidate_time_mmss_floor",
        "candidate_time_mmss_round",
        "scene_change_score",
        "threshold",
        "score_rank_in_video",
        "score_percentile_in_video",
        "candidate_source",
        "method_used",
        "nearest_window_id",
        "nearest_ad_interval_id",
        "nearest_ad_boundary_type",
        "nearest_ad_boundary_sec",
        "nearest_ad_boundary_mmss",
        "distance_to_nearest_ad_boundary_sec",
        "is_near_ad_start_3s",
        "is_near_ad_start_5s",
        "is_near_ad_end_3s",
        "is_near_ad_end_5s",
        "is_near_any_ad_boundary_5s",
    ]
    if candidate_df.empty:
        return [], requested_source_columns

    df = strip_bom_columns(candidate_df)
    if "score_rank_in_video" not in df.columns or "score_percentile_in_video" not in df.columns:
        df = add_score_rank_columns(df)

    rows: list[dict[str, Any]] = []
    for _, series in df.iterrows():
        src = safe_row_dict(series)
        video_id = clean(src.get("video_id"))
        video_filename = clean(src.get("video_filename"))
        candidate_time = to_float(src.get("candidate_time_sec"))
        metrics = interval_metrics_for_candidate(ad_intervals_by_video.get(video_id, []), candidate_time)
        review_clip_start = max(0.0, (candidate_time or 0.0) - 5.0) if candidate_time is not None else None
        review_clip_end = (candidate_time + 5.0) if candidate_time is not None else None
        nearest_ad_interval_id = clean(src.get("nearest_ad_interval_id"))
        if metrics["is_near_any_ad_boundary_5s"] == "true":
            review_priority = "high"
        elif nearest_ad_interval_id or metrics["candidate_inside_ad_interval"]:
            review_priority = "medium"
        else:
            review_priority = "low"

        row = {
            "video_id": video_id,
            "video_title": clean(src.get("video_title"))
            or manifest_value(manifest_lookup, video_id, video_filename, ["video_title", "matched_label_video_title"]),
            "video_filename": video_filename,
            "video_path": clean(src.get("video_path")) or manifest_value(manifest_lookup, video_id, video_filename, ["video_path"]),
            "candidate_time_sec": fmt_number(src.get("candidate_time_sec")),
            "candidate_time_mmss": clean(src.get("candidate_time_mmss"))
            or clean(src.get("candidate_time_mmss_floor"))
            or mmss(src.get("candidate_time_sec")),
            "candidate_time_mmss_floor": clean(src.get("candidate_time_mmss_floor")) or mmss(src.get("candidate_time_sec")),
            "candidate_time_mmss_round": clean(src.get("candidate_time_mmss_round")) or mmss(round(candidate_time or 0)),
            "scene_change_score": fmt_number(src.get("scene_change_score"), digits=6),
            "threshold": clean(src.get("threshold")),
            "score_rank_in_video": fmt_number(src.get("score_rank_in_video"), digits=0),
            "score_percentile_in_video": fmt_number(src.get("score_percentile_in_video"), digits=2),
            "candidate_source": clean(src.get("candidate_source")),
            "method_used": clean(src.get("method_used")),
            "nearest_window_id": clean(src.get("nearest_window_id")),
            "nearest_ad_interval_id": nearest_ad_interval_id,
            "nearest_ad_boundary_type": metrics["nearest_ad_boundary_type"],
            "nearest_ad_boundary_sec": metrics["nearest_ad_boundary_sec"],
            "nearest_ad_boundary_mmss": metrics["nearest_ad_boundary_mmss"],
            "distance_to_nearest_ad_boundary_sec": metrics["distance_to_nearest_ad_boundary_sec"],
            "is_near_ad_start_3s": metrics["is_near_ad_start_3s"],
            "is_near_ad_start_5s": metrics["is_near_ad_start_5s"],
            "is_near_ad_end_3s": metrics["is_near_ad_end_3s"],
            "is_near_ad_end_5s": metrics["is_near_ad_end_5s"],
            "is_near_any_ad_boundary_5s": metrics["is_near_any_ad_boundary_5s"],
            "review_clip_start_sec": fmt_number(review_clip_start),
            "review_clip_end_sec": fmt_number(review_clip_end),
            "review_clip_start_mmss": mmss(review_clip_start),
            "review_clip_end_mmss": mmss(review_clip_end),
            "review_priority": review_priority,
            "review_status": "not_reviewed",
            "is_true_scene_change": "",
            "scene_change_strength": "",
            "scene_change_type": "",
            "is_ad_boundary_related": "",
            "false_positive_type": "",
            "keep_as_boundary_candidate": "",
            "review_note": "",
            "reviewer": "",
            "reviewed_at": "",
        }
        rows.append(row)

    priority_order = {"high": 0, "medium": 1, "low": 2}
    rows.sort(
        key=lambda r: (
            priority_order.get(clean(r.get("review_priority")), 99),
            clean(r.get("video_id")),
            to_float(r.get("candidate_time_sec")) if to_float(r.get("candidate_time_sec")) is not None else 10**12,
        )
    )
    return rows, requested_source_columns


def nearest_candidate_metrics(candidates: list[dict[str, Any]], video_id: str, target_sec: Any) -> tuple[str, str, str]:
    target = to_float(target_sec)
    if target is None:
        return "", "", ""
    best: tuple[float, float] | None = None
    for row in candidates:
        if clean(row.get("video_id")) != clean(video_id):
            continue
        cand = to_float(row.get("candidate_time_sec"))
        if cand is None:
            continue
        dist = abs(cand - target)
        if best is None or dist < best[0]:
            best = (dist, cand)
    if best is None:
        return "", "", ""
    return str(fmt_number(best[1])), mmss(best[1]), str(fmt_number(best[0]))


def count_candidates_in_range(candidates: list[dict[str, Any]], video_id: str, start: float, end: float) -> int:
    count = 0
    for row in candidates:
        if clean(row.get("video_id")) != clean(video_id):
            continue
        cand = to_float(row.get("candidate_time_sec"))
        if cand is not None and start <= cand <= end:
            count += 1
    return count


def has_candidate_near(candidates: list[dict[str, Any]], video_id: str, target: Any, radius: float) -> bool:
    sec = to_float(target)
    if sec is None:
        return False
    return count_candidates_in_range(candidates, video_id, sec - radius, sec + radius) > 0


def build_boundary_review(
    boundary_df: pd.DataFrame, ad_intervals: list[dict[str, Any]], candidates: list[dict[str, Any]], manifest_lookup: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    audit_by_id: dict[str, dict[str, Any]] = {}
    if not boundary_df.empty:
        for _, row in boundary_df.iterrows():
            data = safe_row_dict(row)
            if clean(data.get("ad_interval_id")):
                audit_by_id[clean(data.get("ad_interval_id"))] = data

    rows: list[dict[str, Any]] = []
    for interval in ad_intervals:
        ad_interval_id = clean(interval.get("ad_interval_id"))
        audit = audit_by_id.get(ad_interval_id, {})
        video_id = clean(audit.get("video_id")) or clean(interval.get("video_id"))
        video_filename = clean(audit.get("video_filename")) or clean(interval.get("video_filename"))
        start = first_present(audit, ["ad_start_sec"]) or clean(interval.get("ad_start_sec"))
        end = first_present(audit, ["ad_end_sec"]) or clean(interval.get("ad_end_sec"))
        start_f = to_float(start)
        end_f = to_float(end)
        duration = (end_f - start_f) if start_f is not None and end_f is not None else None

        nearest_start_sec, nearest_start_mmss, dist_start = nearest_candidate_metrics(candidates, video_id, start)
        nearest_end_sec, nearest_end_mmss, dist_end = nearest_candidate_metrics(candidates, video_id, end)
        start_hit_3s = has_candidate_near(candidates, video_id, start, 3)
        start_hit_5s = has_candidate_near(candidates, video_id, start, 5)
        end_hit_3s = has_candidate_near(candidates, video_id, end, 3)
        end_hit_5s = has_candidate_near(candidates, video_id, end, 5)
        pre10_count = count_candidates_in_range(candidates, video_id, max(0.0, (start_f or 0.0) - 10.0), (start_f or 0.0))
        post10_count = count_candidates_in_range(candidates, video_id, (end_f or 0.0), (end_f or 0.0) + 10.0)
        in_ad_count = (
            count_candidates_in_range(candidates, video_id, start_f, end_f)
            if start_f is not None and end_f is not None
            else 0
        )
        row = {
            "ad_interval_id": ad_interval_id,
            "video_id": video_id,
            "video_title": clean(audit.get("video_title"))
            or clean(interval.get("video_title"))
            or manifest_value(manifest_lookup, video_id, video_filename, ["video_title", "matched_label_video_title"]),
            "video_filename": video_filename,
            "ad_start_sec": fmt_number(start),
            "ad_start_mmss": clean(audit.get("ad_start_mmss")) or clean(interval.get("ad_start_mmss")) or mmss(start),
            "ad_end_sec": fmt_number(end),
            "ad_end_mmss": clean(audit.get("ad_end_mmss")) or clean(interval.get("ad_end_mmss")) or mmss(end),
            "ad_duration_sec": fmt_number(duration),
            "start_hit_3s": bool_text(to_bool(audit.get("start_hit_3s")) if clean(audit.get("start_hit_3s")) else start_hit_3s),
            "start_hit_5s": bool_text(to_bool(audit.get("start_hit_5s")) if clean(audit.get("start_hit_5s")) else start_hit_5s),
            "end_hit_3s": bool_text(to_bool(audit.get("end_hit_3s")) if clean(audit.get("end_hit_3s")) else end_hit_3s),
            "end_hit_5s": bool_text(to_bool(audit.get("end_hit_5s")) if clean(audit.get("end_hit_5s")) else end_hit_5s),
            "both_boundary_hit_5s": bool_text(
                (
                    to_bool(audit.get("both_boundary_hit_5s"))
                    if clean(audit.get("both_boundary_hit_5s"))
                    else start_hit_5s and end_hit_5s
                )
            ),
            "nearest_candidate_to_start_sec": fmt_number(audit.get("nearest_candidate_to_start_sec") or nearest_start_sec),
            "nearest_candidate_to_start_mmss": clean(audit.get("nearest_candidate_to_start_mmss")) or nearest_start_mmss,
            "distance_to_start_candidate_sec": fmt_number(audit.get("distance_to_start_candidate_sec") or dist_start),
            "nearest_candidate_to_end_sec": fmt_number(audit.get("nearest_candidate_to_end_sec") or nearest_end_sec),
            "nearest_candidate_to_end_mmss": clean(audit.get("nearest_candidate_to_end_mmss")) or nearest_end_mmss,
            "distance_to_end_candidate_sec": fmt_number(audit.get("distance_to_end_candidate_sec") or dist_end),
            "candidate_count_in_ad_interval": fmt_number(audit.get("candidate_count_in_ad_interval") or in_ad_count, digits=0),
            "candidate_count_in_pre10": fmt_number(audit.get("candidate_count_in_pre10") or pre10_count, digits=0),
            "candidate_count_in_post10": fmt_number(audit.get("candidate_count_in_post10") or post10_count, digits=0),
            "candidate_count_near_start_5s": fmt_number(
                audit.get("candidate_count_near_start_5s")
                or count_candidates_in_range(candidates, video_id, (start_f or 0.0) - 5.0, (start_f or 0.0) + 5.0),
                digits=0,
            ),
            "candidate_count_near_end_5s": fmt_number(
                audit.get("candidate_count_near_end_5s")
                or count_candidates_in_range(candidates, video_id, (end_f or 0.0) - 5.0, (end_f or 0.0) + 5.0),
                digits=0,
            ),
            "start_candidate_correct": "",
            "end_candidate_correct": "",
            "actual_start_transition_visible": "",
            "actual_end_transition_visible": "",
            "ad_start_boundary_quality": "",
            "ad_end_boundary_quality": "",
            "start_boundary_review_note": "",
            "end_boundary_review_note": "",
            "overall_boundary_review_status": "not_reviewed",
        }
        rows.append(row)
    rows.sort(key=lambda r: (clean(r.get("video_id")), to_float(r.get("ad_start_sec")) or 10**12))
    return rows


def distribution_text(values: list[str]) -> str:
    counter = Counter(v for v in values if clean(v))
    return "; ".join(f"{key}:{counter[key]}" for key in sorted(counter))


def build_video_summary(
    candidate_rows: list[dict[str, Any]], boundary_rows: list[dict[str, Any]], manifest_lookup: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    video_ids = sorted({clean(r.get("video_id")) for r in candidate_rows + boundary_rows if clean(r.get("video_id"))})
    candidates_by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    boundaries_by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in candidate_rows:
        candidates_by_video[clean(row.get("video_id"))].append(row)
    for row in boundary_rows:
        boundaries_by_video[clean(row.get("video_id"))].append(row)

    rows: list[dict[str, Any]] = []
    for video_id in video_ids:
        cands = candidates_by_video.get(video_id, [])
        bounds = boundaries_by_video.get(video_id, [])
        sample = (cands[0] if cands else bounds[0] if bounds else {})
        video_filename = clean(sample.get("video_filename"))
        duration = manifest_value(manifest_lookup, video_id, video_filename, ["duration_sec", "video_duration_sec"])
        duration_f = to_float(duration)
        candidate_count = len(cands)
        per_min = candidate_count / (duration_f / 60.0) if duration_f and duration_f > 0 else None
        start_near = sum(1 for row in cands if to_bool(row.get("is_near_ad_start_5s")))
        end_near = sum(1 for row in cands if to_bool(row.get("is_near_ad_end_5s")))
        ad_interval_count = len(bounds)
        boundary_count = ad_interval_count * 2
        hit_count = sum(1 for row in bounds if to_bool(row.get("start_hit_5s"))) + sum(
            1 for row in bounds if to_bool(row.get("end_hit_5s"))
        )
        hit_rate = hit_count / boundary_count if boundary_count else None
        scores = pd.to_numeric(pd.Series([row.get("scene_change_score") for row in cands]), errors="coerce").dropna()
        source_values = [clean(row.get("candidate_source")) for row in cands]
        method_values = [clean(row.get("method_used")) for row in cands]
        ffmpeg_count = sum(1 for value in source_values + method_values if "ffmpeg" in value.lower())
        if (ad_interval_count > 0 and hit_count == 0) or candidate_count > 100 or (per_min is not None and per_min > 8) or ffmpeg_count >= 30:
            priority = "high"
        elif ad_interval_count > 0 and hit_rate is not None and hit_rate >= 0.75 and (per_min is None or per_min <= 5):
            priority = "low"
        else:
            priority = "medium"
        rows.append(
            {
                "video_id": video_id,
                "video_title": clean(sample.get("video_title"))
                or manifest_value(manifest_lookup, video_id, video_filename, ["video_title", "matched_label_video_title"]),
                "video_filename": video_filename
                or manifest_value(manifest_lookup, video_id, video_filename, ["video_filename"]),
                "video_duration_sec": fmt_number(duration),
                "video_duration_mmss": mmss(duration),
                "candidate_count": candidate_count,
                "candidate_count_per_min": fmt_number(per_min, digits=3),
                "candidate_count_near_ad_start_5s": start_near,
                "candidate_count_near_ad_end_5s": end_near,
                "ad_interval_count": ad_interval_count,
                "boundary_count": boundary_count,
                "boundary_hit_5s_count": hit_count,
                "boundary_hit_5s_rate": fmt_number(hit_rate, digits=3),
                "candidate_source_distribution": distribution_text(source_values),
                "method_used_distribution": distribution_text(method_values),
                "max_scene_change_score": fmt_number(scores.max() if not scores.empty else ""),
                "median_scene_change_score": fmt_number(scores.median() if not scores.empty else ""),
                "p95_scene_change_score": fmt_number(scores.quantile(0.95) if not scores.empty else ""),
                "candidate_density_ok": "",
                "too_many_false_candidates": "",
                "too_few_candidates": "",
                "video_review_priority": priority,
                "video_review_note": "",
            }
        )
    return rows


def value_options_rows() -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    options = {
        "review_status": ["not_reviewed", "reviewed"],
        "yes_no_unclear": ["yes", "no", "unclear"],
        "scene_change_strength": ["strong", "medium", "weak", "unclear"],
        "scene_change_type": [
            "hard_cut",
            "fade",
            "broll",
            "camera_angle_change",
            "lighting_change",
            "text_overlay_only",
            "motion_only",
            "object_motion",
            "other",
            "unclear",
        ],
        "is_ad_boundary_related": ["ad_start", "ad_end", "inside_ad", "near_ad_but_not_boundary", "no", "unclear"],
        "false_positive_type": [
            "not_false_positive",
            "normal_cut",
            "camera_motion",
            "subtitle_change",
            "lighting_change",
            "object_motion",
            "duplicate_candidate",
            "score_noise",
            "unclear",
        ],
        "keep_as_boundary_candidate": ["yes", "no", "unclear"],
        "boundary_quality": ["clear", "weak", "none", "unclear"],
        "video_review_priority": ["high", "medium", "low"],
    }
    max_len = max(len(values) for values in options.values())
    rows = []
    for idx in range(max_len):
        rows.append({key: values[idx] if idx < len(values) else "" for key, values in options.items()})
    return rows, options


def review_guide_rows() -> list[dict[str, Any]]:
    return [
        {
            "section": "검토 목적",
            "guide": (
                "scene-change 후보가 실제 장면 전환인지 확인한다. 광고 시작/종료 boundary 후보로 사용할 수 있는지 "
                "판단한다. 이 검토는 최종 광고 탐지 성능 평가가 아니라 scene evidence 품질 검토이다."
            ),
        },
        {
            "section": "후보 검토 방법",
            "guide": (
                "`candidate_time_mmss` 기준으로 영상에서 해당 시점으로 이동한다. 후보 시점 앞 5초, 뒤 5초를 본다. "
                "실제 컷 전환인지, 카메라 움직임/자막 변화/밝기 변화인지 확인한다."
            ),
        },
        {
            "section": "is_true_scene_change 기준",
            "guide": (
                "yes: 장소, 구도, 배경, 화면 구성, 제품 B-roll 등 실제 장면 전환이 있음. "
                "no: 같은 장면에서 카메라 움직임, 자막 변화, 조명 변화, 물체 움직임만 있음. "
                "unclear: 장면 전환인지 애매함."
            ),
        },
        {
            "section": "is_ad_boundary_related 기준",
            "guide": (
                "ad_start: 일반 콘텐츠에서 광고 블록으로 넘어가는 전환. ad_end: 광고 블록에서 원래 콘텐츠로 복귀하는 전환. "
                "inside_ad: 광고 내부의 장면 전환. near_ad_but_not_boundary: 광고 근처이지만 시작/종료는 아님. "
                "no: 광고와 무관한 일반 전환."
            ),
        },
        {
            "section": "keep_as_boundary_candidate 기준",
            "guide": (
                "yes: 실제 전환이고 광고 boundary 후보로 사용할 만함. no: 오탐이거나 일반 컷 전환이어서 boundary 후보로 부적절함. "
                "unclear: 판단 보류."
            ),
        },
        {
            "section": "추천 검토 순서",
            "guide": (
                "1순위: review_priority = high. 2순위: ad_boundary_review sheet의 start/end hit 관련 row. "
                "3순위: ffmpeg_fallback_failed4 source 후보. 4순위: 각 영상별 score 상위 후보."
            ),
        },
    ]


def rows_to_df(rows: list[dict[str, Any]], columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame([{col: row.get(col, "") for col in columns} for row in rows], columns=columns)


def options_range_map(options: dict[str, list[str]]) -> dict[str, str]:
    keys = list(options.keys())
    ranges: dict[str, str] = {}
    for idx, key in enumerate(keys, start=1):
        col = get_column_letter(idx)
        ranges[key] = f"=value_options!${col}$2:${col}${len(options[key]) + 1}"
    return ranges


def apply_basic_sheet_format(ws, review_columns: set[str], wrap_columns: set[str]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    review_fill = PatternFill("solid", fgColor="F4B183")
    header_font = Font(color="FFFFFF", bold=True)
    review_font = Font(color="000000", bold=True)
    for cell in ws[1]:
        cell.fill = review_fill if clean(cell.value) in review_columns else header_fill
        cell.font = review_font if clean(cell.value) in review_columns else header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        header = clean(ws.cell(row=1, column=col_idx).value)
        max_len = len(header)
        for cell in list(column_cells)[1: min(ws.max_row, 80)]:
            max_len = max(max_len, len(clean(cell.value)))
            if header in wrap_columns:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        width = min(max(max_len + 2, 10), 55)
        if header.endswith("_note") or header in {"video_title", "video_filename", "video_path", "guide"}:
            width = min(max(width, 28), 65)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def col_index(ws, header: str) -> int | None:
    for idx, cell in enumerate(ws[1], start=1):
        if clean(cell.value) == header:
            return idx
    return None


def add_validation(ws, header: str, formula: str) -> bool:
    idx = col_index(ws, header)
    if idx is None or ws.max_row < 2:
        return False
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(idx)}2:{get_column_letter(idx)}{max(ws.max_row, 200)}")
    return True


def add_row_formula_fill(ws, header: str, value: str, fill_color: str, target_range: str | None = None) -> bool:
    idx = col_index(ws, header)
    if idx is None or ws.max_row < 2:
        return False
    col = get_column_letter(idx)
    first_row = 2
    formula = f'=${col}{first_row}="{value}"'
    fill = PatternFill("solid", fgColor=fill_color)
    ws.conditional_formatting.add(target_range or f"A2:{get_column_letter(ws.max_column)}{ws.max_row}", FormulaRule(formula=[formula], fill=fill))
    return True


def add_cell_formula_fill(ws, header: str, value: str, fill_color: str) -> bool:
    idx = col_index(ws, header)
    if idx is None or ws.max_row < 2:
        return False
    col = get_column_letter(idx)
    formula = f'=${col}2="{value}"'
    fill = PatternFill("solid", fgColor=fill_color)
    ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", FormulaRule(formula=[formula], fill=fill))
    return True


def apply_excel_formatting(path: Path, options: dict[str, list[str]]) -> tuple[bool, bool]:
    wb = load_workbook(path)
    option_ranges = options_range_map(options)

    review_columns = {
        "review_status",
        "is_true_scene_change",
        "scene_change_strength",
        "scene_change_type",
        "is_ad_boundary_related",
        "false_positive_type",
        "keep_as_boundary_candidate",
        "review_note",
        "reviewer",
        "reviewed_at",
        "start_candidate_correct",
        "end_candidate_correct",
        "actual_start_transition_visible",
        "actual_end_transition_visible",
        "ad_start_boundary_quality",
        "ad_end_boundary_quality",
        "start_boundary_review_note",
        "end_boundary_review_note",
        "overall_boundary_review_status",
        "candidate_density_ok",
        "too_many_false_candidates",
        "too_few_candidates",
        "video_review_priority",
        "video_review_note",
    }
    wrap_columns = {
        "video_title",
        "video_filename",
        "video_path",
        "candidate_source_distribution",
        "method_used_distribution",
        "review_note",
        "start_boundary_review_note",
        "end_boundary_review_note",
        "video_review_note",
        "guide",
    }
    for ws in wb.worksheets:
        apply_basic_sheet_format(ws, review_columns, wrap_columns)

    dropdowns_applied = True
    cf_applied = True
    ws = wb["scene_candidate_review"]
    dropdowns_applied &= add_validation(ws, "review_status", option_ranges["review_status"])
    dropdowns_applied &= add_validation(ws, "is_true_scene_change", option_ranges["yes_no_unclear"])
    dropdowns_applied &= add_validation(ws, "scene_change_strength", option_ranges["scene_change_strength"])
    dropdowns_applied &= add_validation(ws, "scene_change_type", option_ranges["scene_change_type"])
    dropdowns_applied &= add_validation(ws, "is_ad_boundary_related", option_ranges["is_ad_boundary_related"])
    dropdowns_applied &= add_validation(ws, "false_positive_type", option_ranges["false_positive_type"])
    dropdowns_applied &= add_validation(ws, "keep_as_boundary_candidate", option_ranges["keep_as_boundary_candidate"])
    cf_applied &= add_row_formula_fill(ws, "review_priority", "high", "FCE4D6")
    cf_applied &= add_cell_formula_fill(ws, "is_near_any_ad_boundary_5s", "true", "D9EAF7")
    cf_applied &= add_cell_formula_fill(ws, "review_status", "not_reviewed", "E7E6E6")
    cf_applied &= add_cell_formula_fill(ws, "keep_as_boundary_candidate", "yes", "E2F0D9")
    cf_applied &= add_cell_formula_fill(ws, "keep_as_boundary_candidate", "no", "FCE4D6")

    ws = wb["ad_boundary_review"]
    for header in [
        "start_candidate_correct",
        "end_candidate_correct",
        "actual_start_transition_visible",
        "actual_end_transition_visible",
    ]:
        dropdowns_applied &= add_validation(ws, header, option_ranges["yes_no_unclear"])
    dropdowns_applied &= add_validation(ws, "ad_start_boundary_quality", option_ranges["boundary_quality"])
    dropdowns_applied &= add_validation(ws, "ad_end_boundary_quality", option_ranges["boundary_quality"])
    dropdowns_applied &= add_validation(ws, "overall_boundary_review_status", option_ranges["review_status"])
    cf_applied &= add_cell_formula_fill(ws, "overall_boundary_review_status", "not_reviewed", "E7E6E6")

    ws = wb["video_summary_review"]
    for header in ["candidate_density_ok", "too_many_false_candidates", "too_few_candidates"]:
        dropdowns_applied &= add_validation(ws, header, option_ranges["yes_no_unclear"])
    dropdowns_applied &= add_validation(ws, "video_review_priority", option_ranges["video_review_priority"])
    cf_applied &= add_row_formula_fill(ws, "video_review_priority", "high", "FCE4D6")

    wb.save(path)
    return dropdowns_applied, cf_applied


def update_readme() -> None:
    section_title = "## Scene Candidate Human Review v2.3"
    section = f"""{section_title}

OpenCV/ffmpeg merged scene candidate를 사람이 검토하기 위한 Excel 파일을 생성했다.

검토 대상은 scene-change candidate의 실제 전환 여부와 광고 boundary 관련성이다.
이 검토 결과는 이후 rule-based detector의 boundary evidence 정제에 사용된다.

생성된 주요 파일:

- `data/review/scene_candidate_human_review_v2_3.xlsx`
- `data/review/scene_candidate_human_review_candidate_sheet_v2_3.csv`
- `data/review/scene_candidate_human_review_boundary_sheet_v2_3.csv`
- `reports/scene_candidate_human_review_v2_3_report.json`
- `reports/scene_candidate_human_review_v2_3_summary.md`
- `logs/scene_candidate_human_review_v2_3_run_log.txt`
"""
    text = README_PATH.read_text(encoding="utf-8") if README_PATH.exists() else "# youtube_ad_segment_detector\n"
    if section_title in text:
        before, _, rest = text.partition(section_title)
        marker = "\n## "
        if marker in rest:
            _, _, after = rest.partition(marker)
            text = before.rstrip() + "\n\n" + section.rstrip() + "\n\n## " + after.lstrip()
        else:
            text = before.rstrip() + "\n\n" + section.rstrip() + "\n"
    else:
        text = text.rstrip() + "\n\n" + section.rstrip() + "\n"
    README_PATH.write_text(text, encoding="utf-8")


def copy_latest() -> list[str]:
    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    files = [
        OUTPUT_XLSX_PATH,
        OUTPUT_CANDIDATE_CSV_PATH,
        OUTPUT_BOUNDARY_CSV_PATH,
        REPORT_PATH,
        SUMMARY_PATH,
        LOG_PATH,
    ]
    copied: list[str] = []
    for path in files:
        if path.exists():
            target = LATEST_DIR / path.name
            shutil.copy2(path, target)
            copied.append(str(target))
    latest_readme = """# latest_for_chatgpt files

이번 작업명: create_scene_candidate_human_review_xlsx_v2_3

OpenCV/ffmpeg merged scene candidate human review workbook, companion CSV, report, summary, and run log만 복사했다.
mp4, 원본 label xlsx, frame image, model, checkpoint, cache 파일은 복사하지 않는다.

복사 파일:

- `scene_candidate_human_review_v2_3.xlsx`
- `scene_candidate_human_review_candidate_sheet_v2_3.csv`
- `scene_candidate_human_review_boundary_sheet_v2_3.csv`
- `scene_candidate_human_review_v2_3_report.json`
- `scene_candidate_human_review_v2_3_summary.md`
- `scene_candidate_human_review_v2_3_run_log.txt`
"""
    LATEST_README_PATH.write_text(latest_readme, encoding="utf-8")
    copied.append(str(LATEST_README_PATH))
    return copied


def latest_safety_warnings() -> list[Any]:
    warnings: list[Any] = []
    allowed_xlsx = {"scene_candidate_human_review_v2_3.xlsx"}
    banned_parts = {"frame", "model", "checkpoint", "cache"}
    for path in LATEST_DIR.rglob("*"):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        name_lower = path.name.lower()
        if suffix == ".mp4":
            warnings.append({"disallowed_latest_mp4": str(path)})
        if suffix == ".xlsx" and path.name not in allowed_xlsx:
            warnings.append({"disallowed_latest_original_xlsx": str(path)})
        if any(part in name_lower for part in banned_parts):
            warnings.append({"latest_filename_contains_disallowed_part": str(path)})
    return warnings


def write_summary(report: dict[str, Any]) -> None:
    summary = f"""# Scene Candidate Human Review v2.3

## 생성 파일

- Review Excel: `{report.get("output_xlsx_path", "")}`
- Candidate companion CSV: `{report.get("output_candidate_review_csv_path", "")}`
- Boundary companion CSV: `{report.get("output_boundary_review_csv_path", "")}`
- QA report: `{REPORT_PATH}`
- Run log: `{LOG_PATH}`

## 요약

- 총 candidate 수: {report.get("candidate_row_count", 0)}
- high priority 후보 수: {report.get("high_priority_candidate_count", 0)}
- 광고 boundary review row 수: {report.get("boundary_review_row_count", 0)}
- video summary row 수: {report.get("video_summary_row_count", 0)}
- dropdown validation 적용: {report.get("dropdown_validation_applied")}
- conditional formatting 적용: {report.get("conditional_formatting_applied")}

## 먼저 검토할 위치

1. `scene_candidate_review` sheet의 `review_priority=high` row
2. `ad_boundary_review` sheet의 `start_hit_5s`, `end_hit_5s`, `nearest_candidate_*`, `distance_to_*` 컬럼
3. `scene_candidate_review` sheet의 `candidate_source`, `method_used`, `scene_change_score` 상위 후보

## 다음 작업

- 사람이 `is_true_scene_change`, `is_ad_boundary_related`, `keep_as_boundary_candidate`를 채운다.
- `ad_boundary_review`에서 start/end 후보가 실제 광고 boundary인지 검토한다.
- 검토 완료 후 rule-based detector의 boundary evidence 조건을 정제한다.
"""
    SUMMARY_PATH.write_text(summary, encoding="utf-8")


def write_log(report: dict[str, Any]) -> None:
    lines = [
        "작업명: create_scene_candidate_human_review_xlsx_v2_3",
        f"estimated_runtime: {report.get('estimated_runtime')}",
        f"runtime_estimation_reason: {report.get('runtime_estimation_reason')}",
        f"start_time: {report.get('start_time')}",
        f"end_time: {report.get('end_time')}",
        f"actual_runtime_seconds: {report.get('actual_runtime_seconds')}",
        f"actual_runtime_readable: {report.get('actual_runtime_readable')}",
        "",
        *RUN_LOG,
        "",
        "sub_agent_results:",
        json.dumps(report.get("sub_agent_results", {}), ensure_ascii=False, indent=2),
        "",
        "warnings:",
        json.dumps(report.get("warnings", []), ensure_ascii=False, indent=2),
        "",
        "errors:",
        json.dumps(report.get("errors", []), ensure_ascii=False, indent=2),
        "",
    ]
    LOG_PATH.write_text("\n".join(lines), encoding="utf-8")


def generate() -> int:
    start_time = now_iso()
    start_monotonic = time.monotonic()
    warnings: list[Any] = []
    errors: list[Any] = []
    missing_input_files: list[str] = []
    log(f"작업 시작: {start_time}")
    log(f"예상 작업 시간: {ESTIMATED_RUNTIME}")
    log(f"예상 근거: {RUNTIME_ESTIMATION_REASON}")

    before_old_snapshot = old_project_snapshot()
    log("기존 프로젝트 스냅샷 기록 완료")

    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

    candidate_df = read_csv_optional(INPUT_CANDIDATE_PATH, missing_input_files, warnings)
    boundary_df = read_csv_optional(INPUT_BOUNDARY_AUDIT_PATH, missing_input_files, warnings)
    ad_df = read_csv_optional(INPUT_AD_SEGMENT_PATH, missing_input_files, warnings)
    label_context_df = read_csv_optional(INPUT_LABEL_CONTEXT_PATH, missing_input_files, warnings)
    manifest_df = read_csv_optional(INPUT_MANIFEST_PATH, missing_input_files, warnings)
    if not OPTIONAL_FFMPEG_REPORT_PATH.exists():
        warnings.append({"missing_optional_file": str(OPTIONAL_FFMPEG_REPORT_PATH)})
    if not OPTIONAL_FFMPEG_SUMMARY_PATH.exists():
        warnings.append({"missing_optional_file": str(OPTIONAL_FFMPEG_SUMMARY_PATH)})

    critical_columns = {
        "candidate_csv": ["video_id", "video_title", "video_filename", "candidate_time_sec", "scene_change_score"],
        "boundary_audit_csv": ["ad_interval_id", "video_id", "ad_start_sec", "ad_end_sec"],
        "ad_segment_csv": ["ad_interval_id", "video_id", "ad_start_sec", "ad_end_sec"],
        "manifest_csv": ["video_id", "video_title", "video_filename", "duration_sec"],
    }
    missing_required_columns: dict[str, Any] = {}
    inputs_by_name = {
        "candidate_csv": candidate_df,
        "boundary_audit_csv": boundary_df,
        "ad_segment_csv": ad_df,
        "manifest_csv": manifest_df,
    }
    for name, columns in critical_columns.items():
        df = inputs_by_name[name]
        missing_required_columns[name] = [col for col in columns if col not in df.columns]

    manifest_lookup = build_manifest_lookup(manifest_df)
    ad_intervals = prepare_ad_intervals(ad_df, boundary_df)
    ad_intervals_by_video = index_intervals_by_video(ad_intervals)
    candidate_rows, requested_candidate_cols = build_candidate_review(candidate_df, ad_intervals_by_video, manifest_lookup)
    missing_requested_candidate_cols = [col for col in requested_candidate_cols if col not in candidate_df.columns]
    missing_required_columns["candidate_requested_source_columns_missing_or_computed"] = missing_requested_candidate_cols
    boundary_rows = build_boundary_review(boundary_df, ad_intervals, candidate_rows, manifest_lookup)
    video_summary_rows = build_video_summary(candidate_rows, boundary_rows, manifest_lookup)
    value_rows, options = value_options_rows()
    guide_rows = review_guide_rows()

    candidate_columns = [
        "video_id",
        "video_title",
        "video_filename",
        "video_path",
        "candidate_time_sec",
        "candidate_time_mmss",
        "candidate_time_mmss_floor",
        "candidate_time_mmss_round",
        "scene_change_score",
        "threshold",
        "score_rank_in_video",
        "score_percentile_in_video",
        "candidate_source",
        "method_used",
        "nearest_window_id",
        "nearest_ad_interval_id",
        "nearest_ad_boundary_type",
        "nearest_ad_boundary_sec",
        "nearest_ad_boundary_mmss",
        "distance_to_nearest_ad_boundary_sec",
        "is_near_ad_start_3s",
        "is_near_ad_start_5s",
        "is_near_ad_end_3s",
        "is_near_ad_end_5s",
        "is_near_any_ad_boundary_5s",
        "review_clip_start_sec",
        "review_clip_end_sec",
        "review_clip_start_mmss",
        "review_clip_end_mmss",
        "review_priority",
        "review_status",
        "is_true_scene_change",
        "scene_change_strength",
        "scene_change_type",
        "is_ad_boundary_related",
        "false_positive_type",
        "keep_as_boundary_candidate",
        "review_note",
        "reviewer",
        "reviewed_at",
    ]
    boundary_columns = [
        "ad_interval_id",
        "video_id",
        "video_title",
        "video_filename",
        "ad_start_sec",
        "ad_start_mmss",
        "ad_end_sec",
        "ad_end_mmss",
        "ad_duration_sec",
        "start_hit_3s",
        "start_hit_5s",
        "end_hit_3s",
        "end_hit_5s",
        "both_boundary_hit_5s",
        "nearest_candidate_to_start_sec",
        "nearest_candidate_to_start_mmss",
        "distance_to_start_candidate_sec",
        "nearest_candidate_to_end_sec",
        "nearest_candidate_to_end_mmss",
        "distance_to_end_candidate_sec",
        "candidate_count_in_ad_interval",
        "candidate_count_in_pre10",
        "candidate_count_in_post10",
        "candidate_count_near_start_5s",
        "candidate_count_near_end_5s",
        "start_candidate_correct",
        "end_candidate_correct",
        "actual_start_transition_visible",
        "actual_end_transition_visible",
        "ad_start_boundary_quality",
        "ad_end_boundary_quality",
        "start_boundary_review_note",
        "end_boundary_review_note",
        "overall_boundary_review_status",
    ]
    video_columns = [
        "video_id",
        "video_title",
        "video_filename",
        "video_duration_sec",
        "video_duration_mmss",
        "candidate_count",
        "candidate_count_per_min",
        "candidate_count_near_ad_start_5s",
        "candidate_count_near_ad_end_5s",
        "ad_interval_count",
        "boundary_count",
        "boundary_hit_5s_count",
        "boundary_hit_5s_rate",
        "candidate_source_distribution",
        "method_used_distribution",
        "max_scene_change_score",
        "median_scene_change_score",
        "p95_scene_change_score",
        "candidate_density_ok",
        "too_many_false_candidates",
        "too_few_candidates",
        "video_review_priority",
        "video_review_note",
    ]
    value_columns = list(options.keys())
    guide_columns = ["section", "guide"]

    log("CSV companion 파일 생성 중")
    write_csv(OUTPUT_CANDIDATE_CSV_PATH, candidate_rows, candidate_columns)
    write_csv(OUTPUT_BOUNDARY_CSV_PATH, boundary_rows, boundary_columns)

    log("Excel workbook 생성 중")
    with pd.ExcelWriter(OUTPUT_XLSX_PATH, engine="openpyxl") as writer:
        rows_to_df(candidate_rows, candidate_columns).to_excel(writer, sheet_name="scene_candidate_review", index=False)
        rows_to_df(boundary_rows, boundary_columns).to_excel(writer, sheet_name="ad_boundary_review", index=False)
        rows_to_df(video_summary_rows, video_columns).to_excel(writer, sheet_name="video_summary_review", index=False)
        rows_to_df(value_rows, value_columns).to_excel(writer, sheet_name="value_options", index=False)
        rows_to_df(guide_rows, guide_columns).to_excel(writer, sheet_name="review_guide", index=False)

    dropdown_applied, conditional_applied = apply_excel_formatting(OUTPUT_XLSX_PATH, options)
    log("Excel formatting, dropdown, conditional formatting 적용 완료")

    high_priority_count = sum(1 for row in candidate_rows if clean(row.get("review_priority")) == "high")
    near_boundary_count = sum(1 for row in candidate_rows if to_bool(row.get("is_near_any_ad_boundary_5s")))
    after_old_snapshot = old_project_snapshot()
    old_project_modified = before_old_snapshot != after_old_snapshot
    warnings.extend(latest_safety_warnings())
    if old_project_modified:
        warnings.append({"old_project_modified": True})

    end_time = now_iso()
    actual_runtime_seconds = round(time.monotonic() - start_monotonic, 3)
    report: dict[str, Any] = {
        "project_root": str(PROJECT_ROOT),
        "estimated_runtime": ESTIMATED_RUNTIME,
        "runtime_estimation_reason": RUNTIME_ESTIMATION_REASON,
        "start_time": start_time,
        "end_time": end_time,
        "actual_runtime_seconds": actual_runtime_seconds,
        "actual_runtime_readable": readable_runtime(actual_runtime_seconds),
        "input_candidate_path": str(INPUT_CANDIDATE_PATH),
        "input_boundary_audit_path": str(INPUT_BOUNDARY_AUDIT_PATH),
        "input_ad_segment_path": str(INPUT_AD_SEGMENT_PATH),
        "input_label_context_path": str(INPUT_LABEL_CONTEXT_PATH),
        "input_manifest_path": str(INPUT_MANIFEST_PATH),
        "candidate_row_count": len(candidate_df),
        "ad_interval_count": len(ad_intervals),
        "video_count": len(video_summary_rows),
        "output_xlsx_path": str(OUTPUT_XLSX_PATH),
        "output_candidate_review_csv_path": str(OUTPUT_CANDIDATE_CSV_PATH),
        "output_boundary_review_csv_path": str(OUTPUT_BOUNDARY_CSV_PATH),
        "sheet_names": SHEET_NAMES,
        "dropdown_validation_applied": dropdown_applied,
        "conditional_formatting_applied": conditional_applied,
        "missing_input_files": missing_input_files,
        "missing_required_columns": missing_required_columns,
        "candidate_review_row_count": len(candidate_rows),
        "boundary_review_row_count": len(boundary_rows),
        "video_summary_row_count": len(video_summary_rows),
        "high_priority_candidate_count": high_priority_count,
        "near_ad_boundary_candidate_count": near_boundary_count,
        "sub_agent_results": {
            "candidate_sheet_validation": {"status": "PENDING", "details": "Sub Agent 1 not recorded yet."},
            "boundary_video_summary_validation": {"status": "PENDING", "details": "Sub Agent 2 not recorded yet."},
            "output_safety_validation": {"status": "PENDING", "details": "Sub Agent 3 not recorded yet."},
        },
        "old_project_modified": old_project_modified,
        "old_project_snapshot_before": before_old_snapshot,
        "old_project_snapshot_after": after_old_snapshot,
        "latest_for_chatgpt_copied": [],
        "warnings": warnings,
        "errors": errors,
    }

    update_readme()
    write_summary(report)
    write_log(report)
    report["latest_for_chatgpt_copied"] = copy_latest()
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    write_summary(report)
    write_log(report)
    copy_latest()
    log(f"작업 종료: {end_time}")
    return 0 if not errors else 1


def record_sub_agent_results(results: list[list[str]]) -> int:
    if not REPORT_PATH.exists():
        raise FileNotFoundError(REPORT_PATH)
    report = json.loads(REPORT_PATH.read_text(encoding="utf-8"))
    start = report.get("start_time")
    for key, status, details in results:
        report.setdefault("sub_agent_results", {})[key] = {"status": status, "details": details}
        log(f"sub agent result recorded: {key}={status} - {details}")
    before_snapshot = report.get("old_project_snapshot_before") or old_project_snapshot()
    after_snapshot = old_project_snapshot()
    report["old_project_snapshot_after"] = after_snapshot
    report["old_project_modified"] = before_snapshot != after_snapshot
    end = now_iso()
    report["end_time"] = end
    if start:
        try:
            start_dt = datetime.fromisoformat(start)
            end_dt = datetime.fromisoformat(end)
            seconds = round((end_dt - start_dt).total_seconds(), 3)
            report["actual_runtime_seconds"] = seconds
            report["actual_runtime_readable"] = readable_runtime(seconds)
        except Exception:
            pass
    warnings = report.setdefault("warnings", [])
    warnings.extend(latest_safety_warnings())
    if report["old_project_modified"]:
        warnings.append({"old_project_modified_after_sub_agent_verification": True})
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    write_summary(report)
    write_log(report)
    copy_latest()
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--record-sub-agent",
        nargs=3,
        action="append",
        metavar=("KEY", "STATUS", "DETAILS"),
        help="Record a sub-agent PASS/WARN/FAIL result in the report/log.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.record_sub_agent:
        return record_sub_agent_results(args.record_sub_agent)
    return generate()


if __name__ == "__main__":
    raise SystemExit(main())
