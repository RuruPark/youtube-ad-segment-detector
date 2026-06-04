#!/usr/bin/env python3
"""v2.3 사람이 검토한 값을 옮기면서 OpenCV/FFmpeg v2.4 review workbook을 다시 만든다."""

from __future__ import annotations

import csv
import json
import math
import os
import shutil
import subprocess
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path(".")
OLD_PROJECT_ROOT = Path("./_old_project_not_included")
CHECK_ENV_SCRIPT = PROJECT_ROOT / "scripts/utils/check_cv_environment.py"

INPUT_V2_3_REVIEW_XLSX = PROJECT_ROOT / "data/review/scene_candidate_human_review_v2_3.xlsx"
LABELS_V2_4 = PROJECT_ROOT / "data/labels/clean_ad_labels_v0_scope_refreshed_v2_4.csv"
WINDOW_LABELS_V2_4 = PROJECT_ROOT / "data/windows/window_labels_5s_v2_4.csv"
AD_SEGMENTS_V2_4 = PROJECT_ROOT / "data/segments/ad_interval_segments_v2_4.csv"
COMBINED_SEGMENTS_V2_4 = PROJECT_ROOT / "data/segments/label_interval_context_segments_v2_4.csv"
CANDIDATES_V2_4 = PROJECT_ROOT / "data/scene/scene_candidates_v2_4_merged_ffmpeg_fallback_labelrefreshed.csv"
CANDIDATES_MMSS_V2_4 = PROJECT_ROOT / "data/scene/scene_candidates_v2_4_merged_ffmpeg_fallback_labelrefreshed_mmss.csv"
BOUNDARY_AUDIT_V2_4 = PROJECT_ROOT / "data/scene/scene_candidate_boundary_audit_v2_4_merged_ffmpeg_fallback.csv"
MANIFEST_PATH = PROJECT_ROOT / "data/video_metadata/video_manifest_v2_2.csv"

CANDIDATES_V2_3 = PROJECT_ROOT / "data/scene/scene_candidates_v2_3_merged_ffmpeg_fallback.csv"
CANDIDATES_MMSS_V2_3 = PROJECT_ROOT / "data/scene/scene_candidates_v2_3_merged_ffmpeg_fallback_mmss.csv"
BOUNDARY_AUDIT_V2_3 = PROJECT_ROOT / "data/scene/scene_candidate_boundary_audit_v2_3_merged_ffmpeg_fallback.csv"

REVIEW_DIR = PROJECT_ROOT / "data/review"
BACKUP_DIR = REVIEW_DIR / "backups"
OUTPUT_XLSX = REVIEW_DIR / "scene_candidate_human_review_v2_4.xlsx"
OUTPUT_CANDIDATE_CSV = REVIEW_DIR / "scene_candidate_human_review_candidate_sheet_v2_4.csv"
OUTPUT_BOUNDARY_CSV = REVIEW_DIR / "scene_candidate_human_review_boundary_sheet_v2_4.csv"
OUTPUT_VIDEO_SUMMARY_CSV = REVIEW_DIR / "scene_candidate_human_review_video_summary_v2_4.csv"
CANDIDATE_TRANSFER_AUDIT = REVIEW_DIR / "scene_candidate_review_transfer_audit_v2_3_to_v2_4.csv"
BOUNDARY_TRANSFER_AUDIT = REVIEW_DIR / "scene_boundary_review_transfer_audit_v2_3_to_v2_4.csv"
NEW_ROWS_CSV = REVIEW_DIR / "scene_candidate_new_rows_v2_4.csv"

REPORT_PATH = PROJECT_ROOT / "reports/rebuild_scene_candidate_review_v2_4_from_v2_3_xlsx_report.json"
SUMMARY_PATH = PROJECT_ROOT / "reports/rebuild_scene_candidate_review_v2_4_from_v2_3_xlsx_summary.md"
LOG_PATH = PROJECT_ROOT / "logs/rebuild_scene_candidate_review_v2_4_from_v2_3_xlsx_run_log.txt"
SCRIPT_PATH = PROJECT_ROOT / "scripts/scene/rebuild_scene_candidate_review_v2_4_from_v2_3_xlsx.py"
README_PATH = PROJECT_ROOT / "README.md"
LATEST_DIR = PROJECT_ROOT / "outputs/latest_for_chatgpt"
LATEST_README = LATEST_DIR / "README_latest_files.md"

ESTIMATED_RUNTIME = "약 15분"
RUNTIME_ESTIMATION_REASON = "v2_3 review xlsx 로드, v2_4 scene candidate/boundary audit 로드, 검토 컬럼 이관, Excel workbook 재생성, QA report 생성 범위 기준"
NEW_INTERVAL_ID = "A022"
NEW_VIDEO_ID = "12"
NEW_START_SEC = 1160.0
NEW_END_SEC = 1206.0
RUN_LOG: list[str] = []

CANDIDATE_REVIEW_COLS = [
    "review_status",
    "is_true_scene_change",
    "scene_change_strength",
    "scene_change_type",
    "is_ad_boundary_related",
    "false_positive_type",
    "keep_as_boundary_candidate",
    "review_note",
    "reviewer",
    "reviewed_at",
]
BOUNDARY_REVIEW_COLS = [
    "start_candidate_correct",
    "end_candidate_correct",
    "actual_start_transition_visible",
    "actual_end_transition_visible",
    "ad_start_boundary_quality",
    "ad_end_boundary_quality",
    "start_boundary_review_note",
    "end_boundary_review_note",
    "overall_boundary_review_status",
]
VIDEO_REVIEW_COLS = [
    "candidate_density_ok",
    "too_many_false_candidates",
    "too_few_candidates",
    "video_review_priority",
    "video_review_note",
]


def log(message: str) -> None:
    stamp = datetime.now().astimezone().isoformat(timespec="seconds")
    RUN_LOG.append(f"[{stamp}] {message}")
    print(message, flush=True)


def step(number: int, message: str) -> None:
    log(f"[STEP {number}/9] {message}")


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_float(value: Any) -> float | None:
    try:
        if value is None or pd.isna(value):
            return None
        text = str(value).strip()
        if not text:
            return None
        return float(text)
    except Exception:
        return None


def fmt(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        rounded = round(value, 6)
        return int(round(rounded)) if abs(rounded - round(rounded)) < 1e-9 else rounded
    return value


def bool_text(value: bool) -> str:
    return "true" if bool(value) else "false"


def is_true(value: Any) -> bool:
    return clean(value).lower() in {"true", "1", "yes", "y"}


def mmss_floor(seconds: Any) -> str:
    sec = to_float(seconds)
    if sec is None:
        return ""
    total = max(0, int(math.floor(sec)))
    return f"{total // 60:02d}분 {total % 60:02d}초"


def readable(seconds: float) -> str:
    total = int(round(seconds))
    return f"{total // 60}분 {total % 60}초"


def ensure_inside_project(path: Path) -> None:
    resolved = path.resolve()
    root = PROJECT_ROOT.resolve()
    old = OLD_PROJECT_ROOT.resolve()
    if resolved != root and not str(resolved).startswith(str(root) + os.sep):
        raise RuntimeError(f"Refusing to write outside project root: {resolved}")
    if resolved == old or str(resolved).startswith(str(old) + os.sep):
        raise RuntimeError(f"Refusing to write inside old project root: {resolved}")


def ensure_dirs() -> None:
    for path in [REVIEW_DIR, BACKUP_DIR, PROJECT_ROOT / "reports", PROJECT_ROOT / "logs", LATEST_DIR, SCRIPT_PATH.parent]:
        ensure_inside_project(path)
        path.mkdir(parents=True, exist_ok=True)


def read_csv(path: Path, required: bool, missing: list[str], warnings: list[Any], errors: list[Any]) -> pd.DataFrame:
    if not path.exists():
        missing.append(str(path))
        payload = {"missing_file": str(path)}
        if required:
            errors.append(payload)
        else:
            warnings.append(payload)
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str, keep_default_na=False)


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ensure_inside_project(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: fmt(row.get(col, "")) for col in columns})


def rows_to_df(rows: list[dict[str, Any]], columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame([{col: fmt(row.get(col, "")) for col in columns} for row in rows], columns=columns)


def verify_cv_environment(warnings: list[Any]) -> tuple[bool, str]:
    executable = sys.executable
    in_cv = "/envs/cv/" in executable or executable.endswith("/envs/cv/bin/python") or executable.endswith("/envs/cv/bin/python3.10")
    if CHECK_ENV_SCRIPT.exists():
        cmd = ["conda", "run", "-n", "cv", "python", str(CHECK_ENV_SCRIPT)]
        log("Command: " + " ".join(cmd))
        result = subprocess.run(cmd, cwd=PROJECT_ROOT, capture_output=True, text=True, check=False)
        log("cv check stdout: " + result.stdout.strip().replace("\n", " | "))
        if result.stderr.strip():
            log("cv check stderr: " + result.stderr.strip().replace("\n", " | "))
        if result.returncode != 0:
            warnings.append("cv_environment_check_failed")
            return False, executable
    else:
        cmd = ["conda", "run", "-n", "cv", "python", "-c", "import sys; print(sys.executable); import pandas as pd; import openpyxl; print('pandas', pd.__version__); print('openpyxl', openpyxl.__version__)"]
        log("Command: " + " ".join(cmd))
        result = subprocess.run(cmd, cwd=PROJECT_ROOT, capture_output=True, text=True, check=False)
        log("cv fallback stdout: " + result.stdout.strip().replace("\n", " | "))
        if result.stderr.strip():
            log("cv fallback stderr: " + result.stderr.strip().replace("\n", " | "))
        if result.returncode != 0:
            warnings.append("cv_environment_fallback_check_failed")
            return False, executable
    if not in_cv:
        warnings.append({"current_python_executable_not_in_cv": executable})
    return in_cv, executable


def old_project_snapshot() -> dict[str, Any]:
    if not OLD_PROJECT_ROOT.exists():
        return {"exists": False}
    stat = OLD_PROJECT_ROOT.stat()
    return {"exists": True, "mtime_ns": stat.st_mtime_ns, "size": stat.st_size}


def load_review_workbook(path: Path, warnings: list[Any], errors: list[Any]) -> dict[str, pd.DataFrame]:
    if not path.exists():
        errors.append({"missing_v2_3_review_xlsx": str(path)})
        return {}
    expected = ["scene_candidate_review", "ad_boundary_review", "video_summary_review", "value_options", "review_guide"]
    sheets: dict[str, pd.DataFrame] = {}
    try:
        xls = pd.ExcelFile(path)
        missing_sheets = [sheet for sheet in expected if sheet not in xls.sheet_names]
        if missing_sheets:
            warnings.append({"missing_v2_3_review_sheets": missing_sheets})
        for sheet in expected:
            if sheet in xls.sheet_names:
                sheets[sheet] = pd.read_excel(path, sheet_name=sheet, dtype=object, keep_default_na=False)
            else:
                sheets[sheet] = pd.DataFrame()
    except Exception as exc:
        errors.append({"v2_3_review_xlsx_read_failed": str(exc)})
    return sheets


def normalize_source(value: Any) -> str:
    text = clean(value).lower()
    if text in {"original_v2_3", "original_opencv_v2_3"}:
        return "original_opencv_v2_3"
    return text


def normalize_method(value: Any) -> str:
    return clean(value).lower()


def exact_key(row: dict[str, Any] | pd.Series) -> tuple[str, float | None, str, str]:
    return (
        clean(row.get("video_id")),
        None if to_float(row.get("candidate_time_sec")) is None else round(float(to_float(row.get("candidate_time_sec"))), 3),
        normalize_source(row.get("candidate_source")),
        normalize_method(row.get("method_used")),
    )


def similar_source(a: Any, b: Any) -> bool:
    aa = normalize_source(a)
    bb = normalize_source(b)
    return aa == bb or (aa and bb and (aa in bb or bb in aa))


def similar_method(a: Any, b: Any) -> bool:
    aa = normalize_method(a)
    bb = normalize_method(b)
    return aa == bb or (aa and bb and (aa in bb or bb in aa))


def build_v2_3_candidate_indexes(v2_3_df: pd.DataFrame) -> tuple[dict[tuple[str, float | None, str, str], list[tuple[int, dict[str, Any]]]], dict[str, list[tuple[int, dict[str, Any]]]]]:
    exact: dict[tuple[str, float | None, str, str], list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    by_video: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for idx, row in v2_3_df.iterrows():
        payload = row.to_dict()
        row_id = idx + 2
        exact[exact_key(payload)].append((row_id, payload))
        by_video[clean(payload.get("video_id"))].append((row_id, payload))
    return exact, by_video


def candidate_is_a022_related(row: dict[str, Any]) -> bool:
    if clean(row.get("video_id")) != NEW_VIDEO_ID:
        return False
    t = to_float(row.get("candidate_time_sec"))
    if clean(row.get("nearest_ad_interval_id")) == NEW_INTERVAL_ID:
        return True
    if t is None:
        return False
    return abs(t - NEW_START_SEC) <= 5 or abs(t - NEW_END_SEC) <= 5


def candidate_inside_any_ad(row: dict[str, Any], ad_by_video: dict[str, list[dict[str, Any]]]) -> bool:
    t = to_float(row.get("candidate_time_sec"))
    if t is None:
        return False
    for ad in ad_by_video.get(clean(row.get("video_id")), []):
        start = to_float(ad.get("ad_start_sec")) or to_float(ad.get("segment_start_sec"))
        end = to_float(ad.get("ad_end_sec")) or to_float(ad.get("segment_end_sec"))
        if start is not None and end is not None and start <= t <= end:
            return True
    return False


def review_priority(row: dict[str, Any], ad_by_video: dict[str, list[dict[str, Any]]]) -> str:
    if candidate_is_a022_related(row):
        return "very_high"
    if is_true(row.get("is_near_any_ad_boundary_5s")):
        return "high"
    if clean(row.get("nearest_ad_interval_id")) or candidate_inside_any_ad(row, ad_by_video):
        return "medium"
    return "low"


def transfer_candidate_reviews(v2_3_df: pd.DataFrame, v2_4_candidates: pd.DataFrame, ad_segments: pd.DataFrame, missing_required_columns: list[Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, int]]:
    missing_cols = [col for col in CANDIDATE_REVIEW_COLS if col not in v2_3_df.columns]
    if missing_cols:
        missing_required_columns.append({"candidate_review_missing_columns": missing_cols})
    exact_index, by_video = build_v2_3_candidate_indexes(v2_3_df)
    ad_by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for _, ad in ad_segments.iterrows():
        ad_by_video[clean(ad.get("video_id"))].append(ad.to_dict())

    rows: list[dict[str, Any]] = []
    audits: list[dict[str, Any]] = []
    new_rows: list[dict[str, Any]] = []
    counts = Counter()
    used_v2_3_ids: set[int] = set()

    df = v2_4_candidates.copy()
    if "score_rank_in_video" not in df.columns and "scene_change_score" in df.columns:
        df["score_rank_in_video"] = ""
        df["score_percentile_in_video"] = ""
        for video_id, idxs in df.groupby(df["video_id"].map(clean)).groups.items():
            sub = df.loc[list(idxs)].copy()
            scores = sub["scene_change_score"].map(lambda x: to_float(x) or 0.0)
            ranked_idx = scores.sort_values(ascending=False).index.tolist()
            for rank, idx in enumerate(ranked_idx, start=1):
                score = to_float(df.at[idx, "scene_change_score"]) or 0.0
                percentile = float((scores <= score).mean() * 100.0)
                df.at[idx, "score_rank_in_video"] = rank
                df.at[idx, "score_percentile_in_video"] = round(percentile, 3)

    for out_idx, (_, series) in enumerate(df.iterrows(), start=1):
        src = series.to_dict()
        t = to_float(src.get("candidate_time_sec"))
        if t is None:
            continue
        row = dict(src)
        row["candidate_time_sec"] = t
        row["candidate_time_mmss"] = clean(row.get("candidate_time_mmss")) or mmss_floor(t)
        row["candidate_time_mmss_floor"] = clean(row.get("candidate_time_mmss_floor")) or mmss_floor(t)
        row["candidate_time_mmss_round"] = clean(row.get("candidate_time_mmss_round")) or mmss_floor(round(t))
        row["review_clip_start_sec"] = max(0.0, t - 5.0)
        row["review_clip_end_sec"] = t + 5.0
        row["review_clip_start_mmss"] = mmss_floor(row["review_clip_start_sec"])
        row["review_clip_end_mmss"] = mmss_floor(row["review_clip_end_sec"])
        row["review_priority"] = review_priority(row, ad_by_video)
        row["is_new_or_changed_due_to_v2_4"] = bool_text(candidate_is_a022_related(row))

        transfer_status = "missing_v2_3_match"
        match_method = ""
        matched_row_id = ""
        matched_payload: dict[str, Any] | None = None
        time_diff = ""
        transferred_cols: list[str] = []
        warning = ""

        if candidate_is_a022_related(row):
            transfer_status = "new_v2_4_row"
            match_method = "A022_new_or_changed_due_to_v2_4"
            for col in CANDIDATE_REVIEW_COLS:
                row[col] = "not_reviewed" if col == "review_status" else ""
            new_rows.append({**row, "new_row_reason": "A022_related_candidate_or_boundary_window"})
        else:
            exact_matches = exact_index.get(exact_key(row), [])
            if len(exact_matches) == 1:
                matched_row_id, matched_payload = exact_matches[0]
                match_method = "exact"
                transfer_status = "transferred_exact"
                time_diff = 0.0
            elif len(exact_matches) > 1:
                transfer_status = "ambiguous_match_not_transferred"
                match_method = "exact_multiple"
                warning = f"multiple_exact_matches:{len(exact_matches)}"
            else:
                relaxed = []
                for row_id, candidate in by_video.get(clean(row.get("video_id")), []):
                    old_t = to_float(candidate.get("candidate_time_sec"))
                    if old_t is None:
                        continue
                    if abs(old_t - t) <= 0.25 and (similar_source(candidate.get("candidate_source"), row.get("candidate_source")) or similar_method(candidate.get("method_used"), row.get("method_used"))):
                        relaxed.append((row_id, candidate, abs(old_t - t)))
                if len(relaxed) == 1:
                    matched_row_id, matched_payload, time_diff = relaxed[0]
                    match_method = "relaxed_time_0.25"
                    transfer_status = "transferred_relaxed"
                elif len(relaxed) > 1:
                    transfer_status = "ambiguous_match_not_transferred"
                    match_method = "relaxed_multiple"
                    warning = f"multiple_relaxed_matches:{len(relaxed)}"
                else:
                    fallback = []
                    for row_id, candidate in by_video.get(clean(row.get("video_id")), []):
                        old_mmss = clean(candidate.get("candidate_time_mmss")) or clean(candidate.get("candidate_time_mmss_floor"))
                        new_mmss = clean(row.get("candidate_time_mmss")) or clean(row.get("candidate_time_mmss_floor"))
                        old_score = to_float(candidate.get("scene_change_score"))
                        new_score = to_float(row.get("scene_change_score"))
                        if old_mmss == new_mmss and old_score is not None and new_score is not None and abs(old_score - new_score) <= 1e-6:
                            fallback.append((row_id, candidate, abs((to_float(candidate.get("candidate_time_sec")) or t) - t)))
                    if len(fallback) == 1:
                        matched_row_id, matched_payload, time_diff = fallback[0]
                        match_method = "fallback_mmss_score"
                        transfer_status = "transferred_relaxed"
                    elif len(fallback) > 1:
                        transfer_status = "ambiguous_match_not_transferred"
                        match_method = "fallback_multiple"
                        warning = f"multiple_fallback_matches:{len(fallback)}"

            if matched_payload is not None:
                used_v2_3_ids.add(int(matched_row_id))
                for col in CANDIDATE_REVIEW_COLS:
                    if col in matched_payload:
                        row[col] = matched_payload.get(col, "")
                        transferred_cols.append(col)
                    else:
                        row[col] = "not_reviewed" if col == "review_status" else ""
                if not clean(row.get("review_status")):
                    row["review_status"] = "not_reviewed"
            elif transfer_status != "ambiguous_match_not_transferred":
                for col in CANDIDATE_REVIEW_COLS:
                    row[col] = "not_reviewed" if col == "review_status" else ""

        rows.append(row)
        counts[transfer_status] += 1
        audits.append(
            {
                "v2_4_row_id": out_idx + 1,
                "v2_4_video_id": clean(row.get("video_id")),
                "v2_4_candidate_time_sec": t,
                "v2_4_candidate_time_mmss": clean(row.get("candidate_time_mmss")),
                "matched_v2_3_row_id": matched_row_id,
                "match_method": match_method,
                "time_diff_sec": time_diff,
                "review_columns_transferred": ";".join(transferred_cols),
                "transfer_status": transfer_status,
                "transfer_warning": warning,
            }
        )

    counts["unmatched_v2_3_review_row_count"] = max(0, len(v2_3_df) - len(used_v2_3_ids))
    return rows, audits, new_rows, dict(counts)


def transfer_boundary_reviews(v2_3_boundary: pd.DataFrame, v2_4_audit: pd.DataFrame, ad_segments: pd.DataFrame, missing_required_columns: list[Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, int]]:
    missing_cols = [col for col in BOUNDARY_REVIEW_COLS if col not in v2_3_boundary.columns]
    if missing_cols:
        missing_required_columns.append({"boundary_review_missing_columns": missing_cols})
    old_by_id = {clean(row.get("ad_interval_id")): row.to_dict() for _, row in v2_3_boundary.iterrows()}
    seg_by_id = {clean(row.get("ad_interval_id")): row.to_dict() for _, row in ad_segments.iterrows()}
    rows: list[dict[str, Any]] = []
    audits: list[dict[str, Any]] = []
    counts = Counter()
    for _, series in v2_4_audit.iterrows():
        row = series.to_dict()
        ad_id = clean(row.get("ad_interval_id"))
        seg = seg_by_id.get(ad_id, {})
        start = to_float(row.get("ad_start_sec")) or to_float(seg.get("ad_start_sec")) or to_float(seg.get("segment_start_sec"))
        end = to_float(row.get("ad_end_sec")) or to_float(seg.get("ad_end_sec")) or to_float(seg.get("segment_end_sec"))
        row["ad_duration_sec"] = (end - start) if start is not None and end is not None else ""
        row["is_new_interval_v2_4"] = bool_text(ad_id == NEW_INTERVAL_ID)
        old = old_by_id.get(ad_id)
        transferred: list[str] = []
        if old and ad_id != NEW_INTERVAL_ID:
            for col in BOUNDARY_REVIEW_COLS:
                if col in old:
                    row[col] = old.get(col, "")
                    transferred.append(col)
                else:
                    row[col] = "not_reviewed" if col == "overall_boundary_review_status" else ""
            if not clean(row.get("overall_boundary_review_status")):
                row["overall_boundary_review_status"] = "not_reviewed"
            status = "transferred"
            counts["boundary_review_transferred_count"] += 1
        else:
            for col in BOUNDARY_REVIEW_COLS:
                row[col] = "not_reviewed" if col == "overall_boundary_review_status" else ""
            status = "new_interval_v2_4" if ad_id == NEW_INTERVAL_ID else "missing_v2_3_boundary_match"
            if ad_id == NEW_INTERVAL_ID:
                counts["boundary_review_new_interval_count"] += 1
        rows.append(row)
        audits.append(
            {
                "ad_interval_id": ad_id,
                "exists_in_v2_3": bool_text(bool(old)),
                "exists_in_v2_4": "true",
                "transfer_status": status,
                "transferred_columns": ";".join(transferred),
                "is_new_interval_v2_4": bool_text(ad_id == NEW_INTERVAL_ID),
                "transfer_warning": "" if old or ad_id == NEW_INTERVAL_ID else "missing_v2_3_boundary_review_row",
            }
        )
    return rows, audits, dict(counts)


def count_distribution(values: list[Any]) -> str:
    counter = Counter(clean(v) for v in values if clean(v))
    return ";".join(f"{k}:{v}" for k, v in sorted(counter.items()))


def make_video_summary(candidate_rows: list[dict[str, Any]], boundary_rows: list[dict[str, Any]], manifest: pd.DataFrame, ad_segments: pd.DataFrame, v2_3_video_summary: pd.DataFrame, missing_required_columns: list[Any]) -> list[dict[str, Any]]:
    missing_cols = [col for col in VIDEO_REVIEW_COLS if col not in v2_3_video_summary.columns]
    if missing_cols:
        missing_required_columns.append({"video_summary_review_missing_columns": missing_cols})
    old_by_video = {clean(row.get("video_id")): row.to_dict() for _, row in v2_3_video_summary.iterrows()}
    candidates_by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in candidate_rows:
        candidates_by_video[clean(row.get("video_id"))].append(row)
    boundaries_by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in boundary_rows:
        boundaries_by_video[clean(row.get("video_id"))].append(row)
    ad_count_by_video = Counter(clean(row.get("video_id")) for _, row in ad_segments.iterrows())
    rows: list[dict[str, Any]] = []
    for _, m in manifest.iterrows():
        vid = clean(m.get("label_mapping_video_id")) or clean(m.get("video_id"))
        title = clean(m.get("label_mapping_video_title")) or clean(m.get("matched_label_video_title")) or clean(m.get("video_title"))
        duration = to_float(m.get("duration_sec")) or 0.0
        cands = candidates_by_video.get(vid, [])
        bounds = boundaries_by_video.get(vid, [])
        scores = [to_float(row.get("scene_change_score")) for row in cands if to_float(row.get("scene_change_score")) is not None]
        boundary_hit_count = sum(int(is_true(row.get("start_hit_5s"))) + int(is_true(row.get("end_hit_5s"))) for row in bounds)
        boundary_count = len(bounds) * 2
        old = old_by_video.get(vid, {})
        row = {
            "video_id": vid,
            "video_title": title,
            "video_filename": clean(m.get("video_filename")),
            "video_duration_sec": duration,
            "video_duration_mmss": mmss_floor(duration),
            "candidate_count": len(cands),
            "candidate_count_per_min": len(cands) / (duration / 60.0) if duration else 0.0,
            "candidate_count_near_ad_start_5s": sum(1 for cand in cands if is_true(cand.get("is_near_ad_start_5s"))),
            "candidate_count_near_ad_end_5s": sum(1 for cand in cands if is_true(cand.get("is_near_ad_end_5s"))),
            "ad_interval_count": ad_count_by_video.get(vid, 0),
            "boundary_count": boundary_count,
            "boundary_hit_5s_count": boundary_hit_count,
            "boundary_hit_5s_rate": boundary_hit_count / boundary_count if boundary_count else 0.0,
            "candidate_source_distribution": count_distribution([cand.get("candidate_source") for cand in cands]),
            "method_used_distribution": count_distribution([cand.get("method_used") for cand in cands]),
            "max_scene_change_score": max(scores) if scores else "",
            "median_scene_change_score": pd.Series(scores).median() if scores else "",
            "p95_scene_change_score": pd.Series(scores).quantile(0.95) if scores else "",
            "has_new_interval_v2_4": bool_text(vid == NEW_VIDEO_ID),
        }
        for col in VIDEO_REVIEW_COLS:
            row[col] = old.get(col, "") if old else ""
        if not clean(row.get("video_review_priority")):
            row["video_review_priority"] = "high" if vid == NEW_VIDEO_ID else "medium"
        rows.append(row)
    rows.sort(key=lambda r: int(clean(r.get("video_id"))) if clean(r.get("video_id")).isdigit() else 999999)
    return rows


def default_value_options(v2_3_value_options: pd.DataFrame) -> pd.DataFrame:
    defaults = {
        "review_status": ["not_reviewed", "reviewed"],
        "yes_no_unclear": ["yes", "no", "unclear"],
        "scene_change_strength": ["strong", "medium", "weak", "unclear"],
        "scene_change_type": ["hard_cut", "fade", "broll", "camera_angle_change", "lighting_change", "text_overlay_only", "motion_only", "object_motion", "other", "unclear"],
        "is_ad_boundary_related": ["ad_start", "ad_end", "inside_ad", "near_ad_but_not_boundary", "no", "unclear"],
        "false_positive_type": ["not_false_positive", "normal_cut", "camera_motion", "subtitle_change", "lighting_change", "object_motion", "duplicate_candidate", "score_noise", "unclear"],
        "keep_as_boundary_candidate": ["yes", "no", "unclear"],
        "boundary_quality": ["clear", "weak", "none", "unclear"],
        "video_review_priority": ["high", "medium", "low"],
    }
    max_len = max(len(v) for v in defaults.values())
    rows = []
    for i in range(max_len):
        rows.append({col: values[i] if i < len(values) else "" for col, values in defaults.items()})
    df = pd.DataFrame(rows)
    if not v2_3_value_options.empty:
        for col in df.columns:
            if col in v2_3_value_options.columns:
                old_values = [clean(v) for v in v2_3_value_options[col].tolist() if clean(v)]
                merged = []
                for v in old_values + defaults[col]:
                    if v and v not in merged:
                        merged.append(v)
                for idx, value in enumerate(merged):
                    if idx >= len(df):
                        df.loc[idx] = {c: "" for c in df.columns}
                    df.at[idx, col] = value
    return df.fillna("")


def review_guide(v2_3_guide: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if not v2_3_guide.empty and {"section", "guide"}.issubset(v2_3_guide.columns):
        rows.extend(v2_3_guide[["section", "guide"]].to_dict("records"))
    rows.extend(
        [
            {"section": "v2.4 변경", "guide": "v2_4는 A022 광고 구간이 추가된 최신 라벨 기준이다."},
            {"section": "검토값 이관", "guide": "v2_3에서 입력한 review 값은 동일 candidate/ad_interval에 대해 이관했다. label-dependent nearest boundary 값은 v2_4 기준 값을 유지한다."},
            {"section": "A022", "guide": "A022 관련 row는 새로 검토해야 하므로 not_reviewed로 둔다."},
            {"section": "먼저 볼 항목", "guide": "1. ad_boundary_review에서 A022 row 2. scene_candidate_review에서 is_new_or_changed_due_to_v2_4=true 3. review_priority=very_high/high 4. 기존 not_reviewed row"},
        ]
    )
    return pd.DataFrame(rows)


def candidate_columns(rows: list[dict[str, Any]]) -> list[str]:
    preferred = [
        "video_id", "video_title", "video_filename", "video_path", "candidate_time_sec", "candidate_time_mmss", "candidate_time_mmss_floor", "candidate_time_mmss_round", "scene_change_score", "threshold", "score_rank_in_video", "score_percentile_in_video", "candidate_source", "method_used", "nearest_window_id", "nearest_ad_interval_id", "nearest_ad_boundary_type", "nearest_ad_boundary_sec", "nearest_ad_boundary_mmss", "distance_to_nearest_ad_boundary_sec", "is_near_ad_start_3s", "is_near_ad_start_5s", "is_near_ad_end_3s", "is_near_ad_end_5s", "is_near_any_ad_boundary_5s", "review_clip_start_sec", "review_clip_end_sec", "review_clip_start_mmss", "review_clip_end_mmss", "review_priority", "is_new_or_changed_due_to_v2_4", *CANDIDATE_REVIEW_COLS
    ]
    return [col for col in preferred if any(col in row for row in rows)]


def boundary_columns(rows: list[dict[str, Any]]) -> list[str]:
    preferred = [
        "ad_interval_id", "video_id", "video_title", "video_filename", "ad_start_sec", "ad_start_mmss", "ad_end_sec", "ad_end_mmss", "ad_duration_sec", "start_hit_3s", "start_hit_5s", "end_hit_3s", "end_hit_5s", "both_boundary_hit_5s", "nearest_candidate_to_start_sec", "nearest_candidate_to_start_mmss", "distance_to_start_candidate_sec", "nearest_candidate_to_end_sec", "nearest_candidate_to_end_mmss", "distance_to_end_candidate_sec", "candidate_count_in_ad_interval", "candidate_count_in_pre10", "candidate_count_in_post10", "candidate_count_near_start_5s", "candidate_count_near_end_5s", "is_new_interval_v2_4", *BOUNDARY_REVIEW_COLS
    ]
    return [col for col in preferred if any(col in row for row in rows)]


def add_validation(ws: Any, column_name: str, values: list[str]) -> bool:
    header = [cell.value for cell in ws[1]]
    if column_name not in header:
        return False
    col_idx = header.index(column_name) + 1
    col = get_column_letter(col_idx)
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{max(ws.max_row, 2)}")
    return True


def add_row_fill(ws: Any, column_name: str, expected_value: str, color: str) -> bool:
    header = [cell.value for cell in ws[1]]
    if column_name not in header or ws.max_row < 2:
        return False
    col = get_column_letter(header.index(column_name) + 1)
    fill = PatternFill("solid", fgColor=color)
    formula = f'${col}2="{expected_value}"'
    ws.conditional_formatting.add(f"A2:{get_column_letter(ws.max_column)}{ws.max_row}", FormulaRule(formula=[formula], fill=fill))
    return True


def style_workbook(path: Path, review_input_cols: set[str]) -> tuple[bool, bool]:
    dropdown_applied = False
    conditional_applied = False
    wb = load_workbook(path)
    dropdowns = {
        "review_status": ["not_reviewed", "reviewed"],
        "is_true_scene_change": ["yes", "no", "unclear"],
        "scene_change_strength": ["strong", "medium", "weak", "unclear"],
        "scene_change_type": ["hard_cut", "fade", "broll", "camera_angle_change", "lighting_change", "text_overlay_only", "motion_only", "object_motion", "other", "unclear"],
        "is_ad_boundary_related": ["ad_start", "ad_end", "inside_ad", "near_ad_but_not_boundary", "no", "unclear"],
        "false_positive_type": ["not_false_positive", "normal_cut", "camera_motion", "subtitle_change", "lighting_change", "object_motion", "duplicate_candidate", "score_noise", "unclear"],
        "keep_as_boundary_candidate": ["yes", "no", "unclear"],
        "start_candidate_correct": ["yes", "no", "unclear"],
        "end_candidate_correct": ["yes", "no", "unclear"],
        "actual_start_transition_visible": ["yes", "no", "unclear"],
        "actual_end_transition_visible": ["yes", "no", "unclear"],
        "ad_start_boundary_quality": ["clear", "weak", "none", "unclear"],
        "ad_end_boundary_quality": ["clear", "weak", "none", "unclear"],
        "candidate_density_ok": ["yes", "no", "unclear"],
        "too_many_false_candidates": ["yes", "no", "unclear"],
        "too_few_candidates": ["yes", "no", "unclear"],
        "video_review_priority": ["high", "medium", "low"],
    }
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        header = [cell.value for cell in ws[1]]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.value in review_input_cols:
                cell.fill = PatternFill("solid", fgColor="D9EAD3")
            else:
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 2000)):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 10
            for cell in list(column_cells)[:200]:
                max_len = max(max_len, min(len(str(cell.value or "")), 55))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
        for col, values in dropdowns.items():
            dropdown_applied = add_validation(ws, col, values) or dropdown_applied
        conditional_applied = add_row_fill(ws, "review_priority", "very_high", "F4B183") or conditional_applied
        conditional_applied = add_row_fill(ws, "review_priority", "high", "FCE4D6") or conditional_applied
        conditional_applied = add_row_fill(ws, "is_new_or_changed_due_to_v2_4", "true", "F4CCCC") or conditional_applied
        conditional_applied = add_row_fill(ws, "is_new_interval_v2_4", "true", "F4CCCC") or conditional_applied
        conditional_applied = add_row_fill(ws, "is_near_any_ad_boundary_5s", "true", "D9EAF7") or conditional_applied
        conditional_applied = add_row_fill(ws, "review_status", "not_reviewed", "E7E6E6") or conditional_applied
        conditional_applied = add_row_fill(ws, "keep_as_boundary_candidate", "yes", "D9EAD3") or conditional_applied
        conditional_applied = add_row_fill(ws, "keep_as_boundary_candidate", "no", "FCE4D6") or conditional_applied
    wb.save(path)
    return dropdown_applied, conditional_applied


def make_summary(report: dict[str, Any]) -> str:
    warning_text = "\n".join(f"- {w}" for w in report.get("warnings", [])) if report.get("warnings") else "- 주요 warning 없음"
    return f"""# rebuild_scene_candidate_review_v2_4_from_v2_3_xlsx summary

## Review Transfer

- v2_3 candidate review rows: {report.get('v2_3_candidate_review_row_count')}
- v2_4 candidate review rows: {report.get('v2_4_candidate_review_row_count')}
- transferred exact / relaxed: {report.get('candidate_review_transferred_exact_count')} / {report.get('candidate_review_transferred_relaxed_count')}
- new v2_4 rows: {report.get('candidate_review_new_v2_4_count')}
- ambiguous not transferred: {report.get('candidate_review_ambiguous_not_transferred_count')}

## Boundary Review

- v2_3 boundary rows: {report.get('v2_3_boundary_review_row_count')}
- v2_4 boundary rows: {report.get('v2_4_boundary_review_row_count')}
- boundary transferred: {report.get('boundary_review_transferred_count')}
- boundary new interval count: {report.get('boundary_review_new_interval_count')}
- A022 candidate rows: {report.get('A022_candidate_row_count')}
- A022 boundary row exists: {report.get('A022_boundary_row_exists')}

## Output

- backup created: {report.get('backup_created')}
- backup path: `{report.get('backup_path')}`
- output xlsx: `{report.get('output_v2_4_review_xlsx_path')}`
- transfer audit: `{CANDIDATE_TRANSFER_AUDIT}`
- boundary transfer audit: `{BOUNDARY_TRANSFER_AUDIT}`

## 먼저 볼 항목

1. `ad_boundary_review` sheet의 A022 row
2. `scene_candidate_review` sheet의 `is_new_or_changed_due_to_v2_4=true` row
3. `review_priority=very_high/high` row
4. 기존 `not_reviewed` row

## Warning

{warning_text}
"""


def update_readme() -> None:
    section = """## Scene Candidate Human Review v2.4 Rebuilt from v2.3

v2_3 review xlsx의 사람이 입력한 검토값을 v2_4 review xlsx로 이관했다. v2_4는 A022 광고 구간이 추가된 최신 라벨 기준이며, A022 관련 row는 새로 검토해야 한다.

이후 OpenCV/ffmpeg scene 후보 검토는 `data/review/scene_candidate_human_review_v2_4.xlsx` 기준으로 진행한다.
"""
    text = README_PATH.read_text(encoding="utf-8") if README_PATH.exists() else "# youtube_ad_segment_detector\n"
    marker = "## Scene Candidate Human Review v2.4 Rebuilt from v2.3"
    if marker in text:
        before, _, after = text.partition(marker)
        next_idx = after.find("\n## ")
        text = before.rstrip() + "\n\n" + section + (after[next_idx:] if next_idx >= 0 else "")
    else:
        text = text.rstrip() + "\n\n" + section
    README_PATH.write_text(text.rstrip() + "\n", encoding="utf-8")


def latest_copy(files: list[Path]) -> tuple[bool, list[str]]:
    ensure_inside_project(LATEST_DIR)
    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    for child in LATEST_DIR.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()
    forbidden_suffixes = {".mp4", ".jpg", ".jpeg", ".png", ".webp", ".pt", ".pth", ".ckpt", ".bin"}
    allowed_xlsx = {OUTPUT_XLSX.name}
    copied: list[str] = []
    for src in files:
        if not src.exists():
            continue
        if src.suffix.lower() in forbidden_suffixes:
            raise RuntimeError(f"Refusing to copy forbidden file to latest: {src}")
        if src.suffix.lower() in {".xlsx", ".xls"} and src.name not in allowed_xlsx:
            raise RuntimeError(f"Refusing to copy raw/non-output xlsx to latest: {src}")
        dst = LATEST_DIR / src.name
        shutil.copy2(src, dst)
        copied.append(str(dst))
    LATEST_README.write_text(
        "# latest_for_chatgpt files\n\n"
        "이번 작업명: rebuild_scene_candidate_review_v2_4_from_v2_3_xlsx\n\n"
        "v2_3 review 값 이관 후 재생성한 v2_4 OpenCV/ffmpeg review workbook, companion CSV, transfer audit, report/summary/log만 복사했다. mp4, 원본 raw xlsx, frame, model, checkpoint, cache 파일은 복사하지 않는다.\n",
        encoding="utf-8",
    )
    copied.append(str(LATEST_README))
    return True, copied


def write_report(report: dict[str, Any]) -> None:
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    SUMMARY_PATH.write_text(make_summary(report), encoding="utf-8")
    LOG_PATH.write_text("\n".join(RUN_LOG) + "\n", encoding="utf-8")


def main() -> None:
    ensure_dirs()
    start_dt = datetime.now().astimezone()
    start_time = start_dt.isoformat(timespec="seconds")
    started = time.time()
    warnings: list[Any] = []
    errors: list[Any] = []
    missing_input_files: list[str] = []
    missing_required_columns: list[Any] = []
    before_old = old_project_snapshot()
    before_v2_3_stat = INPUT_V2_3_REVIEW_XLSX.stat().st_mtime_ns if INPUT_V2_3_REVIEW_XLSX.exists() else None

    log(f"작업 시작 전 예상 작업 시간: {ESTIMATED_RUNTIME}")
    log(f"예상 근거: {RUNTIME_ESTIMATION_REASON}")
    log(f"작업 시작 시각: {start_time}")

    report: dict[str, Any] = {
        "project_root": str(PROJECT_ROOT),
        "estimated_runtime": ESTIMATED_RUNTIME,
        "runtime_estimation_reason": RUNTIME_ESTIMATION_REASON,
        "start_time": start_time,
        "end_time": "",
        "actual_runtime_seconds": 0,
        "actual_runtime_readable": "",
        "input_v2_3_review_xlsx_path": str(INPUT_V2_3_REVIEW_XLSX),
        "input_v2_4_candidate_path": str(CANDIDATES_V2_4),
        "input_v2_4_boundary_audit_path": str(BOUNDARY_AUDIT_V2_4),
        "output_v2_4_review_xlsx_path": str(OUTPUT_XLSX),
        "backup_created": False,
        "backup_path": "",
        "missing_input_files": missing_input_files,
        "missing_required_columns": missing_required_columns,
        "sub_agent_results": {},
        "old_project_modified": False,
        "warnings": warnings,
        "errors": errors,
    }

    try:
        step(1, "작업 시간 기록 및 cv 환경 확인")
        cv_ok, python_executable = verify_cv_environment(warnings)
        report["cv_environment_checked"] = cv_ok
        report["python_executable"] = python_executable
        if not cv_ok:
            errors.append("cv_environment_check_failed")
            raise RuntimeError("cv environment check failed")

        step(2, "입력 파일 로드")
        review_sheets = load_review_workbook(INPUT_V2_3_REVIEW_XLSX, warnings, errors)
        v2_4_candidates = read_csv(CANDIDATES_V2_4, True, missing_input_files, warnings, errors)
        v2_4_boundary = read_csv(BOUNDARY_AUDIT_V2_4, True, missing_input_files, warnings, errors)
        ad_segments = read_csv(AD_SEGMENTS_V2_4, True, missing_input_files, warnings, errors)
        combined_segments = read_csv(COMBINED_SEGMENTS_V2_4, False, missing_input_files, warnings, errors)
        manifest = read_csv(MANIFEST_PATH, False, missing_input_files, warnings, errors)
        _ = read_csv(LABELS_V2_4, False, missing_input_files, warnings, errors)
        _ = read_csv(WINDOW_LABELS_V2_4, False, missing_input_files, warnings, errors)
        _ = read_csv(CANDIDATES_MMSS_V2_4, False, missing_input_files, warnings, errors)
        _ = read_csv(CANDIDATES_V2_3, False, missing_input_files, warnings, errors)
        _ = read_csv(CANDIDATES_MMSS_V2_3, False, missing_input_files, warnings, errors)
        _ = read_csv(BOUNDARY_AUDIT_V2_3, False, missing_input_files, warnings, errors)
        if errors:
            warnings.append("critical inputs missing; output may be incomplete")
        if v2_4_candidates.empty or ad_segments.empty:
            raise RuntimeError("v2_4 candidates or ad segments are missing")

        step(3, "기존 v2_4 review xlsx backup 생성")
        backup_created = False
        backup_path = ""
        if OUTPUT_XLSX.exists():
            stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
            backup = BACKUP_DIR / f"scene_candidate_human_review_v2_4_backup_before_v2_3_transfer_{stamp}.xlsx"
            shutil.copy2(OUTPUT_XLSX, backup)
            backup_created = True
            backup_path = str(backup)
            log(f"Backup created: {backup}")
        report["backup_created"] = backup_created
        report["backup_path"] = backup_path

        step(4, "candidate review 값 이관")
        v2_3_candidate = review_sheets.get("scene_candidate_review", pd.DataFrame())
        candidate_rows, candidate_audit_rows, new_rows, candidate_counts = transfer_candidate_reviews(v2_3_candidate, v2_4_candidates, ad_segments, missing_required_columns)
        candidate_cols = candidate_columns(candidate_rows)
        candidate_audit_cols = ["v2_4_row_id", "v2_4_video_id", "v2_4_candidate_time_sec", "v2_4_candidate_time_mmss", "matched_v2_3_row_id", "match_method", "time_diff_sec", "review_columns_transferred", "transfer_status", "transfer_warning"]
        new_row_cols = candidate_cols + ["new_row_reason"]
        write_csv(OUTPUT_CANDIDATE_CSV, candidate_rows, candidate_cols)
        write_csv(CANDIDATE_TRANSFER_AUDIT, candidate_audit_rows, candidate_audit_cols)
        write_csv(NEW_ROWS_CSV, new_rows, new_row_cols)

        step(5, "boundary review 값 이관")
        v2_3_boundary = review_sheets.get("ad_boundary_review", pd.DataFrame())
        boundary_rows, boundary_audit_rows, boundary_counts = transfer_boundary_reviews(v2_3_boundary, v2_4_boundary, ad_segments, missing_required_columns)
        boundary_cols = boundary_columns(boundary_rows)
        boundary_audit_cols = ["ad_interval_id", "exists_in_v2_3", "exists_in_v2_4", "transfer_status", "transferred_columns", "is_new_interval_v2_4", "transfer_warning"]
        write_csv(OUTPUT_BOUNDARY_CSV, boundary_rows, boundary_cols)
        write_csv(BOUNDARY_TRANSFER_AUDIT, boundary_audit_rows, boundary_audit_cols)

        step(6, "video summary review 생성 및 이관")
        v2_3_video_summary = review_sheets.get("video_summary_review", pd.DataFrame())
        video_rows = make_video_summary(candidate_rows, boundary_rows, manifest, ad_segments, v2_3_video_summary, missing_required_columns)
        video_cols = [
            "video_id", "video_title", "video_filename", "video_duration_sec", "video_duration_mmss", "candidate_count", "candidate_count_per_min", "candidate_count_near_ad_start_5s", "candidate_count_near_ad_end_5s", "ad_interval_count", "boundary_count", "boundary_hit_5s_count", "boundary_hit_5s_rate", "candidate_source_distribution", "method_used_distribution", "max_scene_change_score", "median_scene_change_score", "p95_scene_change_score", "has_new_interval_v2_4", *VIDEO_REVIEW_COLS
        ]
        write_csv(OUTPUT_VIDEO_SUMMARY_CSV, video_rows, video_cols)

        step(7, "v2_4 Excel workbook 생성 및 서식 적용")
        value_options_df = default_value_options(review_sheets.get("value_options", pd.DataFrame()))
        guide_df = review_guide(review_sheets.get("review_guide", pd.DataFrame()))
        with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
            rows_to_df(candidate_rows, candidate_cols).to_excel(writer, sheet_name="scene_candidate_review", index=False)
            rows_to_df(boundary_rows, boundary_cols).to_excel(writer, sheet_name="ad_boundary_review", index=False)
            rows_to_df(video_rows, video_cols).to_excel(writer, sheet_name="video_summary_review", index=False)
            value_options_df.to_excel(writer, sheet_name="value_options", index=False)
            guide_df.to_excel(writer, sheet_name="review_guide", index=False)
        dropdown_applied, conditional_applied = style_workbook(OUTPUT_XLSX, set(CANDIDATE_REVIEW_COLS + BOUNDARY_REVIEW_COLS + VIDEO_REVIEW_COLS))

        step(8, "report/summary/log/README/latest_for_chatgpt 갱신")
        update_readme()
        a022_candidate_count = sum(1 for row in candidate_rows if candidate_is_a022_related(row))
        a022_boundary_exists = any(clean(row.get("ad_interval_id")) == NEW_INTERVAL_ID for row in boundary_rows)
        after_old = old_project_snapshot()
        old_modified = before_old != after_old
        after_v2_3_stat = INPUT_V2_3_REVIEW_XLSX.stat().st_mtime_ns if INPUT_V2_3_REVIEW_XLSX.exists() else None
        if before_v2_3_stat != after_v2_3_stat:
            errors.append("v2_3_review_xlsx_modified_unexpectedly")
        if old_modified:
            errors.append("old_project_modified_unexpectedly")
        elapsed = time.time() - started
        end_time = datetime.now().astimezone().isoformat(timespec="seconds")
        report.update(
            {
                "end_time": end_time,
                "actual_runtime_seconds": elapsed,
                "actual_runtime_readable": readable(elapsed),
                "v2_3_candidate_review_row_count": int(len(v2_3_candidate)),
                "v2_4_candidate_review_row_count": int(len(candidate_rows)),
                "candidate_review_transferred_exact_count": int(candidate_counts.get("transferred_exact", 0)),
                "candidate_review_transferred_relaxed_count": int(candidate_counts.get("transferred_relaxed", 0)),
                "candidate_review_new_v2_4_count": int(candidate_counts.get("new_v2_4_row", 0)),
                "candidate_review_ambiguous_not_transferred_count": int(candidate_counts.get("ambiguous_match_not_transferred", 0)),
                "candidate_review_missing_v2_3_match_count": int(candidate_counts.get("missing_v2_3_match", 0)),
                "candidate_review_unmatched_v2_3_review_row_count": int(candidate_counts.get("unmatched_v2_3_review_row_count", 0)),
                "v2_3_boundary_review_row_count": int(len(v2_3_boundary)),
                "v2_4_boundary_review_row_count": int(len(boundary_rows)),
                "boundary_review_transferred_count": int(boundary_counts.get("boundary_review_transferred_count", 0)),
                "boundary_review_new_interval_count": int(boundary_counts.get("boundary_review_new_interval_count", 0)),
                "A022_candidate_row_count": int(a022_candidate_count),
                "A022_boundary_row_exists": bool(a022_boundary_exists),
                "dropdown_validation_applied": bool(dropdown_applied),
                "conditional_formatting_applied": bool(conditional_applied),
                "missing_input_files": missing_input_files,
                "missing_required_columns": missing_required_columns,
                "old_project_modified": bool(old_modified),
                "generated_files": [str(p) for p in [OUTPUT_XLSX, OUTPUT_CANDIDATE_CSV, OUTPUT_BOUNDARY_CSV, OUTPUT_VIDEO_SUMMARY_CSV, CANDIDATE_TRANSFER_AUDIT, BOUNDARY_TRANSFER_AUDIT, NEW_ROWS_CSV, REPORT_PATH, SUMMARY_PATH, LOG_PATH, SCRIPT_PATH]],
                "latest_for_chatgpt_updated": False,
                "latest_for_chatgpt_files": [],
                "sub_agent_results": {
                    "sub_agent_1_review_transfer": "PENDING_EXTERNAL_REVIEW",
                    "sub_agent_2_v2_4_boundary": "PENDING_EXTERNAL_REVIEW",
                    "sub_agent_3_output_safety": "PENDING_EXTERNAL_REVIEW",
                },
                "warnings": warnings,
                "errors": errors,
            }
        )
        write_report(report)
        latest_files = [OUTPUT_XLSX, OUTPUT_CANDIDATE_CSV, OUTPUT_BOUNDARY_CSV, OUTPUT_VIDEO_SUMMARY_CSV, CANDIDATE_TRANSFER_AUDIT, BOUNDARY_TRANSFER_AUDIT, NEW_ROWS_CSV, REPORT_PATH, SUMMARY_PATH, LOG_PATH]
        latest_ok, latest_copied = latest_copy(latest_files)
        report["latest_for_chatgpt_updated"] = latest_ok
        report["latest_for_chatgpt_files"] = latest_copied
        write_report(report)
        for src in [REPORT_PATH, SUMMARY_PATH, LOG_PATH]:
            shutil.copy2(src, LATEST_DIR / src.name)
        log(f"작업 종료 시각: {end_time}")
        log(f"실제 작업 시간: {readable(elapsed)}")
        write_report(report)
        if latest_ok:
            shutil.copy2(LOG_PATH, LATEST_DIR / LOG_PATH.name)

        step(9, "완료")
        print(json.dumps({"candidate_rows_v2_3_v2_4": [len(v2_3_candidate), len(candidate_rows)], "candidate_transfer_exact": candidate_counts.get("transferred_exact", 0), "candidate_new_v2_4": candidate_counts.get("new_v2_4_row", 0), "boundary_rows_v2_3_v2_4": [len(v2_3_boundary), len(boundary_rows)], "A022_candidate_row_count": a022_candidate_count, "A022_boundary_row_exists": a022_boundary_exists, "backup_created": backup_created, "actual_runtime_readable": readable(elapsed), "errors": errors}, ensure_ascii=False, indent=2))
    except Exception as exc:
        errors.append({"fatal_error": str(exc)})
        elapsed = time.time() - started
        end_time = datetime.now().astimezone().isoformat(timespec="seconds")
        report.update({"end_time": end_time, "actual_runtime_seconds": elapsed, "actual_runtime_readable": readable(elapsed), "warnings": warnings, "errors": errors})
        write_report(report)
        log(f"작업 종료 시각: {end_time}")
        log(f"실제 작업 시간: {readable(elapsed)}")
        raise


if __name__ == "__main__":
    main()
