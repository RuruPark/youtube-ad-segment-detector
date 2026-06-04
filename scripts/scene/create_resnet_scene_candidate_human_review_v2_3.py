#!/usr/bin/env python3
"""ResNet 장면 전환 후보 검토용 v2.3 workbook을 만든다."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import shutil
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(".")
OLD_PROJECT_ROOT = Path("./_old_project_not_included")

INPUT_RESNET_CANDIDATE_RAW_PATH = PROJECT_ROOT / "data/scene/resnet_scene_candidates_v2_3.csv"
INPUT_RESNET_CANDIDATE_MMSS_PATH = PROJECT_ROOT / "data/scene/resnet_scene_candidates_v2_3_mmss.csv"
INPUT_RESNET_BOUNDARY_AUDIT_PATH = PROJECT_ROOT / "data/scene/resnet_scene_candidate_boundary_audit_v2_3.csv"
INPUT_RESNET_VIDEO_SUMMARY_PATH = PROJECT_ROOT / "data/scene/resnet_scene_video_summary_v2_3.csv"
INPUT_OPENCV_COMPARISON_PATH = PROJECT_ROOT / "data/scene/opencv_vs_resnet_scene_comparison_v2_3.csv"
INPUT_CANDIDATE_OVERLAP_PATH = PROJECT_ROOT / "data/scene/opencv_resnet_candidate_overlap_v2_3.csv"
INPUT_AD_SEGMENT_PATH = PROJECT_ROOT / "data/segments/ad_interval_segments_v2_3.csv"
INPUT_OPENCV_CANDIDATE_RAW_PATH = PROJECT_ROOT / "data/scene/scene_candidates_v2_3_merged_ffmpeg_fallback.csv"
INPUT_OPENCV_CANDIDATE_MMSS_PATH = PROJECT_ROOT / "data/scene/scene_candidates_v2_3_merged_ffmpeg_fallback_mmss.csv"
INPUT_OPENCV_BOUNDARY_AUDIT_PATH = PROJECT_ROOT / "data/scene/scene_candidate_boundary_audit_v2_3_merged_ffmpeg_fallback.csv"
OPTIONAL_RESNET_REPORT_PATH = PROJECT_ROOT / "reports/extract_resnet_embedding_scene_change_v2_3_report.json"
OPTIONAL_RESNET_SUMMARY_PATH = PROJECT_ROOT / "reports/extract_resnet_embedding_scene_change_v2_3_summary.md"

REVIEW_DIR = PROJECT_ROOT / "data/review"
OUTPUT_XLSX_PATH = REVIEW_DIR / "resnet_scene_candidate_human_review_v2_3.xlsx"
OUTPUT_CANDIDATE_CSV_PATH = REVIEW_DIR / "resnet_scene_candidate_human_review_candidate_sheet_v2_3.csv"
OUTPUT_BOUNDARY_CSV_PATH = REVIEW_DIR / "resnet_scene_candidate_human_review_boundary_sheet_v2_3.csv"
OUTPUT_RESNET_ONLY_CSV_PATH = REVIEW_DIR / "resnet_only_boundary_review_v2_3.csv"
OUTPUT_COMPARISON_CSV_PATH = REVIEW_DIR / "opencv_resnet_scene_comparison_review_v2_3.csv"
REPORT_PATH = PROJECT_ROOT / "reports/resnet_scene_candidate_human_review_v2_3_report.json"
SUMMARY_PATH = PROJECT_ROOT / "reports/resnet_scene_candidate_human_review_v2_3_summary.md"
LOG_PATH = PROJECT_ROOT / "logs/resnet_scene_candidate_human_review_v2_3_run_log.txt"
README_PATH = PROJECT_ROOT / "README.md"
LATEST_DIR = PROJECT_ROOT / "outputs/latest_for_chatgpt"
LATEST_README_PATH = LATEST_DIR / "README_latest_files.md"

ESTIMATED_RUNTIME = "약 25분"
RUNTIME_ESTIMATION_REASON = (
    "ResNet candidate 약 1093개, 광고 interval 21개, OpenCV/ResNet 비교 파일, "
    "Excel sheet/validation/report 생성 범위 기준"
)
SHEET_NAMES = [
    "resnet_candidate_review",
    "resnet_boundary_review",
    "resnet_only_boundary_review",
    "opencv_resnet_comparison",
    "video_summary_review",
    "value_options",
    "review_guide",
]
RUN_LOG: list[str] = []

CURRENT_LATEST_FILENAMES = {
    "resnet_scene_candidate_human_review_v2_3.xlsx",
    "resnet_scene_candidate_human_review_candidate_sheet_v2_3.csv",
    "resnet_scene_candidate_human_review_boundary_sheet_v2_3.csv",
    "resnet_only_boundary_review_v2_3.csv",
    "opencv_resnet_scene_comparison_review_v2_3.csv",
    "resnet_scene_candidate_human_review_v2_3_report.json",
    "resnet_scene_candidate_human_review_v2_3_summary.md",
    "resnet_scene_candidate_human_review_v2_3_run_log.txt",
    "README_latest_files.md",
}


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def log(message: str) -> None:
    line = f"[{now_iso()}] {message}"
    RUN_LOG.append(line)
    print(message, flush=True)


def readable_runtime(seconds: float) -> str:
    total = int(round(seconds))
    minutes, sec = divmod(total, 60)
    return f"{minutes}분 {sec}초"


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_float(value: Any) -> float | None:
    text = clean(value)
    if text == "":
        return None
    try:
        return float(text)
    except Exception:
        return None


def to_bool(value: Any) -> bool:
    return clean(value).lower() in {"true", "1", "yes", "y", "t"}


def bool_text(value: bool) -> str:
    return "true" if value else "false"


def fmt_number(value: Any, digits: int = 3) -> Any:
    num = to_float(value)
    if num is None:
        return ""
    rounded = round(num, digits)
    if abs(rounded - round(rounded)) < 10 ** (-(digits + 1)):
        return int(round(rounded))
    return rounded


def mmss(seconds: Any) -> str:
    sec = to_float(seconds)
    if sec is None:
        return ""
    total = max(0, int(math.floor(sec)))
    return f"{total // 60:02d}분 {total % 60:02d}초"


def strip_bom_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(col).replace("\ufeff", "").strip() for col in df.columns]
    return df


def read_csv_optional(path: Path, missing_input_files: list[str], warnings: list[Any]) -> pd.DataFrame:
    if not path.exists():
        missing_input_files.append(str(path))
        warnings.append({"missing_input_file": str(path)})
        return pd.DataFrame()
    try:
        return strip_bom_columns(pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8-sig"))
    except Exception as exc:
        missing_input_files.append(str(path))
        warnings.append({"failed_to_read_input": str(path), "error": str(exc)})
        return pd.DataFrame()


def load_json_optional(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def safe_row_dict(row: pd.Series) -> dict[str, Any]:
    return {str(k): clean(v) for k, v in row.to_dict().items()}


def first_present(row: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value = clean(row.get(key))
        if value:
            return value
    return ""


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in columns})


def old_project_snapshot() -> dict[str, Any]:
    if not OLD_PROJECT_ROOT.exists():
        return {"exists": False, "file_count": 0, "digest": ""}
    h = hashlib.sha256()
    file_count = 0
    for path in sorted(p for p in OLD_PROJECT_ROOT.rglob("*") if p.is_file()):
        try:
            stat = path.stat()
        except OSError:
            continue
        rel = path.relative_to(OLD_PROJECT_ROOT).as_posix()
        h.update(f"{rel}\t{stat.st_size}\t{stat.st_mtime_ns}\n".encode("utf-8", errors="replace"))
        file_count += 1
    return {"exists": True, "file_count": file_count, "digest": h.hexdigest()}


def build_ad_intervals(ad_df: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    if ad_df.empty:
        return rows
    for _, row in ad_df.iterrows():
        data = safe_row_dict(row)
        ad_interval_id = clean(data.get("ad_interval_id")) or clean(data.get("segment_id"))
        if not ad_interval_id or ad_interval_id in seen:
            continue
        seen.add(ad_interval_id)
        rows.append(
            {
                "ad_interval_id": ad_interval_id,
                "video_id": clean(data.get("video_id")),
                "video_title": clean(data.get("video_title")),
                "video_filename": clean(data.get("video_filename")),
                "video_path": clean(data.get("video_path")),
                "ad_start_sec": clean(data.get("ad_start_sec")) or clean(data.get("segment_start_sec")),
                "ad_end_sec": clean(data.get("ad_end_sec")) or clean(data.get("segment_end_sec")),
                "ad_start_mmss": mmss(clean(data.get("ad_start_sec")) or clean(data.get("segment_start_sec"))),
                "ad_end_mmss": mmss(clean(data.get("ad_end_sec")) or clean(data.get("segment_end_sec"))),
                "video_duration_sec": clean(data.get("video_duration_sec")),
            }
        )
    return rows


def index_by_id(rows: list[dict[str, Any]], key: str) -> dict[str, dict[str, Any]]:
    return {clean(row.get(key)): row for row in rows if clean(row.get(key))}


def intervals_by_video(ad_intervals: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for interval in ad_intervals:
        by_video[clean(interval.get("video_id"))].append(interval)
    return by_video


def candidate_inside_ad(video_intervals: list[dict[str, Any]], candidate_time: Any) -> bool:
    sec = to_float(candidate_time)
    if sec is None:
        return False
    for interval in video_intervals:
        start = to_float(interval.get("ad_start_sec"))
        end = to_float(interval.get("ad_end_sec"))
        if start is not None and end is not None and start <= sec <= end:
            return True
    return False


def candidate_times_by_video(df: pd.DataFrame) -> dict[str, list[float]]:
    by_video: dict[str, list[float]] = defaultdict(list)
    if df.empty:
        return by_video
    for _, row in df.iterrows():
        data = safe_row_dict(row)
        sec = to_float(data.get("candidate_time_sec"))
        if sec is not None:
            by_video[clean(data.get("video_id"))].append(sec)
    for values in by_video.values():
        values.sort()
    return by_video


def nearest_time(times: list[float], target: Any) -> tuple[float | None, float | None]:
    sec = to_float(target)
    if sec is None or not times:
        return None, None
    best = min(times, key=lambda value: abs(value - sec))
    return best, abs(best - sec)


def opencv_same_boundary_hit(
    opencv_boundary_by_id: dict[str, dict[str, Any]], ad_interval_id: str, boundary_type: str
) -> bool:
    row = opencv_boundary_by_id.get(clean(ad_interval_id), {})
    if boundary_type == "ad_start":
        return to_bool(row.get("start_hit_5s"))
    if boundary_type == "ad_end":
        return to_bool(row.get("end_hit_5s"))
    return False


def build_resnet_candidate_review(
    resnet_df: pd.DataFrame,
    opencv_times: dict[str, list[float]],
    opencv_boundary_by_id: dict[str, dict[str, Any]],
    ad_intervals_by_video: dict[str, list[dict[str, Any]]],
) -> tuple[list[dict[str, Any]], list[str]]:
    requested_source_columns = [
        "video_id",
        "video_title",
        "video_filename",
        "video_path",
        "candidate_time_sec",
        "candidate_time_mmss",
        "candidate_time_mmss_floor",
        "candidate_time_mmss_round",
        "scene_change_score",
        "threshold",
        "cosine_distance",
        "l2_distance",
        "score_rank_in_video",
        "score_percentile_in_video",
        "candidate_source",
        "method_used",
        "model_name",
        "feature_dim",
        "nearest_window_id",
        "nearest_ad_interval_id",
        "nearest_ad_boundary_type",
        "nearest_ad_boundary_sec",
        "nearest_ad_boundary_mmss",
        "distance_to_nearest_ad_boundary_sec",
        "is_near_ad_start_3s",
        "is_near_ad_start_5s",
        "is_near_ad_end_3s",
        "is_near_ad_end_5s",
        "is_near_any_ad_boundary_5s",
    ]
    rows: list[dict[str, Any]] = []
    if resnet_df.empty:
        return rows, requested_source_columns
    for _, series in resnet_df.iterrows():
        src = safe_row_dict(series)
        video_id = clean(src.get("video_id"))
        candidate_time = to_float(src.get("candidate_time_sec"))
        nearest_opencv_sec, nearest_opencv_dist = nearest_time(opencv_times.get(video_id, []), candidate_time)
        ad_interval_id = clean(src.get("nearest_ad_interval_id"))
        boundary_type = clean(src.get("nearest_ad_boundary_type"))
        near_boundary = to_bool(src.get("is_near_any_ad_boundary_5s"))
        opencv_same_hit = opencv_same_boundary_hit(opencv_boundary_by_id, ad_interval_id, boundary_type)
        inside_ad = candidate_inside_ad(ad_intervals_by_video.get(video_id, []), candidate_time)
        if near_boundary and not opencv_same_hit:
            review_priority = "very_high"
        elif near_boundary:
            review_priority = "high"
        elif ad_interval_id or inside_ad:
            review_priority = "medium"
        else:
            review_priority = "low"
        clip_start = max(0.0, (candidate_time or 0.0) - 5.0) if candidate_time is not None else None
        clip_end = candidate_time + 5.0 if candidate_time is not None else None
        row = {
            "video_id": video_id,
            "video_title": clean(src.get("video_title")),
            "video_filename": clean(src.get("video_filename")),
            "video_path": clean(src.get("video_path")),
            "candidate_time_sec": fmt_number(src.get("candidate_time_sec")),
            "candidate_time_mmss": clean(src.get("candidate_time_mmss")) or mmss(src.get("candidate_time_sec")),
            "candidate_time_mmss_floor": clean(src.get("candidate_time_mmss_floor")) or mmss(src.get("candidate_time_sec")),
            "candidate_time_mmss_round": clean(src.get("candidate_time_mmss_round")) or mmss(round(candidate_time or 0)),
            "scene_change_score": fmt_number(src.get("scene_change_score"), digits=6),
            "threshold": fmt_number(src.get("threshold"), digits=6),
            "cosine_distance": fmt_number(src.get("cosine_distance"), digits=6),
            "l2_distance": fmt_number(src.get("l2_distance"), digits=6),
            "score_rank_in_video": fmt_number(src.get("score_rank_in_video"), digits=0),
            "score_percentile_in_video": fmt_number(src.get("score_percentile_in_video"), digits=2),
            "candidate_source": clean(src.get("candidate_source")),
            "method_used": clean(src.get("method_used")),
            "model_name": clean(src.get("model_name")),
            "feature_dim": fmt_number(src.get("feature_dim"), digits=0),
            "nearest_window_id": clean(src.get("nearest_window_id")),
            "nearest_ad_interval_id": ad_interval_id,
            "nearest_ad_boundary_type": boundary_type,
            "nearest_ad_boundary_sec": fmt_number(src.get("nearest_ad_boundary_sec")),
            "nearest_ad_boundary_mmss": clean(src.get("nearest_ad_boundary_mmss")) or mmss(src.get("nearest_ad_boundary_sec")),
            "distance_to_nearest_ad_boundary_sec": fmt_number(src.get("distance_to_nearest_ad_boundary_sec")),
            "is_near_ad_start_3s": bool_text(to_bool(src.get("is_near_ad_start_3s"))),
            "is_near_ad_start_5s": bool_text(to_bool(src.get("is_near_ad_start_5s"))),
            "is_near_ad_end_3s": bool_text(to_bool(src.get("is_near_ad_end_3s"))),
            "is_near_ad_end_5s": bool_text(to_bool(src.get("is_near_ad_end_5s"))),
            "is_near_any_ad_boundary_5s": bool_text(near_boundary),
            "review_clip_start_sec": fmt_number(clip_start),
            "review_clip_end_sec": fmt_number(clip_end),
            "review_clip_start_mmss": mmss(clip_start),
            "review_clip_end_mmss": mmss(clip_end),
            "has_near_opencv_candidate_3s": bool_text(nearest_opencv_dist is not None and nearest_opencv_dist <= 3),
            "has_near_opencv_candidate_5s": bool_text(nearest_opencv_dist is not None and nearest_opencv_dist <= 5),
            "nearest_opencv_candidate_time_sec": fmt_number(nearest_opencv_sec),
            "nearest_opencv_candidate_mmss": mmss(nearest_opencv_sec),
            "distance_to_nearest_opencv_candidate_sec": fmt_number(nearest_opencv_dist),
            "review_priority": review_priority,
            "review_status": "not_reviewed",
            "is_true_scene_change": "",
            "scene_change_strength": "",
            "scene_change_type": "",
            "is_ad_boundary_related": "",
            "resnet_candidate_usefulness": "",
            "false_positive_type": "",
            "keep_as_boundary_candidate": "",
            "review_note": "",
            "reviewer": "",
            "reviewed_at": "",
        }
        rows.append(row)
    priority_order = {"very_high": 0, "high": 1, "medium": 2, "low": 3}
    rows.sort(
        key=lambda row: (
            priority_order.get(clean(row.get("review_priority")), 99),
            clean(row.get("video_id")),
            to_float(row.get("candidate_time_sec")) if to_float(row.get("candidate_time_sec")) is not None else 10**12,
        )
    )
    return rows, requested_source_columns


def audit_rows_by_id(df: pd.DataFrame) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    if df.empty:
        return rows
    for _, row in df.iterrows():
        data = safe_row_dict(row)
        if clean(data.get("ad_interval_id")):
            rows[clean(data.get("ad_interval_id"))] = data
    return rows


def build_boundary_review(
    ad_intervals: list[dict[str, Any]],
    resnet_audit_by_id: dict[str, dict[str, Any]],
    opencv_audit_by_id: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for interval in ad_intervals:
        ad_interval_id = clean(interval.get("ad_interval_id"))
        r = resnet_audit_by_id.get(ad_interval_id, {})
        o = opencv_audit_by_id.get(ad_interval_id, {})
        video_id = clean(r.get("video_id")) or clean(interval.get("video_id"))
        start = clean(r.get("ad_start_sec")) or clean(interval.get("ad_start_sec"))
        end = clean(r.get("ad_end_sec")) or clean(interval.get("ad_end_sec"))
        start_f = to_float(start)
        end_f = to_float(end)
        resnet_start = to_bool(r.get("start_hit_5s"))
        resnet_end = to_bool(r.get("end_hit_5s"))
        opencv_start = to_bool(o.get("start_hit_5s"))
        opencv_end = to_bool(o.get("end_hit_5s"))
        rows.append(
            {
                "ad_interval_id": ad_interval_id,
                "video_id": video_id,
                "video_title": clean(r.get("video_title")) or clean(interval.get("video_title")),
                "video_filename": clean(interval.get("video_filename")),
                "ad_start_sec": fmt_number(start),
                "ad_start_mmss": clean(r.get("ad_start_mmss")) or clean(interval.get("ad_start_mmss")) or mmss(start),
                "ad_end_sec": fmt_number(end),
                "ad_end_mmss": clean(r.get("ad_end_mmss")) or clean(interval.get("ad_end_mmss")) or mmss(end),
                "ad_duration_sec": fmt_number((end_f - start_f) if start_f is not None and end_f is not None else ""),
                "start_hit_3s": bool_text(to_bool(r.get("start_hit_3s"))),
                "start_hit_5s": bool_text(resnet_start),
                "end_hit_3s": bool_text(to_bool(r.get("end_hit_3s"))),
                "end_hit_5s": bool_text(resnet_end),
                "both_boundary_hit_5s": bool_text(to_bool(r.get("both_boundary_hit_5s"))),
                "nearest_candidate_to_start_sec": fmt_number(r.get("nearest_candidate_to_start_sec")),
                "nearest_candidate_to_start_mmss": clean(r.get("nearest_candidate_to_start_mmss")),
                "distance_to_start_candidate_sec": fmt_number(r.get("distance_to_start_candidate_sec")),
                "nearest_candidate_to_end_sec": fmt_number(r.get("nearest_candidate_to_end_sec")),
                "nearest_candidate_to_end_mmss": clean(r.get("nearest_candidate_to_end_mmss")),
                "distance_to_end_candidate_sec": fmt_number(r.get("distance_to_end_candidate_sec")),
                "candidate_count_in_ad_interval": fmt_number(r.get("candidate_count_in_ad_interval"), digits=0),
                "candidate_count_in_pre10": fmt_number(r.get("candidate_count_in_pre10"), digits=0),
                "candidate_count_in_post10": fmt_number(r.get("candidate_count_in_post10"), digits=0),
                "candidate_count_near_start_5s": fmt_number(r.get("candidate_count_near_start_5s"), digits=0),
                "candidate_count_near_end_5s": fmt_number(r.get("candidate_count_near_end_5s"), digits=0),
                "opencv_start_hit_5s": bool_text(opencv_start),
                "opencv_end_hit_5s": bool_text(opencv_end),
                "opencv_both_boundary_hit_5s": bool_text(to_bool(o.get("both_boundary_hit_5s"))),
                "resnet_additional_start_hit": bool_text(resnet_start and not opencv_start),
                "resnet_additional_end_hit": bool_text(resnet_end and not opencv_end),
                "opencv_only_start_hit": bool_text(opencv_start and not resnet_start),
                "opencv_only_end_hit": bool_text(opencv_end and not resnet_end),
                "start_candidate_correct": "",
                "end_candidate_correct": "",
                "actual_start_transition_visible": "",
                "actual_end_transition_visible": "",
                "ad_start_boundary_quality": "",
                "ad_end_boundary_quality": "",
                "resnet_boundary_better_than_opencv": "",
                "start_boundary_review_note": "",
                "end_boundary_review_note": "",
                "overall_boundary_review_status": "not_reviewed",
            }
        )
    rows.sort(key=lambda row: (clean(row.get("video_id")), to_float(row.get("ad_start_sec")) or 10**12))
    return rows


def best_resnet_candidate(
    candidate_rows: list[dict[str, Any]], video_id: str, ad_interval_id: str, boundary_type: str, boundary_sec: Any
) -> dict[str, Any]:
    target = to_float(boundary_sec)
    best: tuple[float, dict[str, Any]] | None = None
    for row in candidate_rows:
        if clean(row.get("video_id")) != clean(video_id):
            continue
        cand = to_float(row.get("candidate_time_sec"))
        if cand is None or target is None:
            continue
        dist = abs(cand - target)
        if dist > 5:
            continue
        same_boundary = (
            clean(row.get("nearest_ad_interval_id")) == clean(ad_interval_id)
            and clean(row.get("nearest_ad_boundary_type")) == clean(boundary_type)
        )
        rank_dist = dist if same_boundary else dist + 1000
        if best is None or rank_dist < best[0]:
            best = (rank_dist, row)
    return best[1] if best else {}


def build_resnet_only_boundary_review(
    boundary_rows: list[dict[str, Any]], candidate_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for boundary in boundary_rows:
        items = [
            (
                "ad_start",
                boundary.get("ad_start_sec"),
                boundary.get("ad_start_mmss"),
                to_bool(boundary.get("resnet_additional_start_hit")),
            ),
            (
                "ad_end",
                boundary.get("ad_end_sec"),
                boundary.get("ad_end_mmss"),
                to_bool(boundary.get("resnet_additional_end_hit")),
            ),
        ]
        for boundary_type, boundary_sec, boundary_mmss, include in items:
            if not include:
                continue
            candidate = best_resnet_candidate(
                candidate_rows,
                clean(boundary.get("video_id")),
                clean(boundary.get("ad_interval_id")),
                boundary_type,
                boundary_sec,
            )
            candidate_time = candidate.get("candidate_time_sec")
            clip_start = max(0.0, (to_float(candidate_time) or 0.0) - 5.0) if clean(candidate_time) else ""
            clip_end = (to_float(candidate_time) or 0.0) + 5.0 if clean(candidate_time) else ""
            rows.append(
                {
                    "ad_interval_id": clean(boundary.get("ad_interval_id")),
                    "video_id": clean(boundary.get("video_id")),
                    "video_title": clean(boundary.get("video_title")),
                    "video_filename": clean(boundary.get("video_filename")),
                    "boundary_type": boundary_type,
                    "ad_boundary_sec": fmt_number(boundary_sec),
                    "ad_boundary_mmss": boundary_mmss,
                    "resnet_candidate_time_sec": fmt_number(candidate_time),
                    "resnet_candidate_mmss": clean(candidate.get("candidate_time_mmss")) or mmss(candidate_time),
                    "distance_to_boundary_sec": fmt_number(
                        abs((to_float(candidate_time) or 0.0) - (to_float(boundary_sec) or 0.0)) if clean(candidate_time) else ""
                    ),
                    "resnet_scene_change_score": fmt_number(candidate.get("scene_change_score"), digits=6),
                    "resnet_score_percentile_in_video": fmt_number(candidate.get("score_percentile_in_video"), digits=2),
                    "nearest_opencv_candidate_time_sec": fmt_number(candidate.get("nearest_opencv_candidate_time_sec")),
                    "nearest_opencv_candidate_mmss": clean(candidate.get("nearest_opencv_candidate_mmss")),
                    "distance_to_nearest_opencv_candidate_sec": fmt_number(candidate.get("distance_to_nearest_opencv_candidate_sec")),
                    "opencv_hit_same_boundary_5s": "false",
                    "review_clip_start_mmss": mmss(clip_start),
                    "review_clip_end_mmss": mmss(clip_end),
                    "is_actual_scene_change": "",
                    "is_actual_ad_boundary": "",
                    "resnet_helped": "",
                    "should_add_to_combined_scene_evidence": "",
                    "failure_reason_for_opencv": "",
                    "review_note": "",
                    "reviewer": "",
                    "reviewed_at": "",
                }
            )
    rows.sort(key=lambda row: (clean(row.get("video_id")), to_float(row.get("ad_boundary_sec")) or 10**12))
    return rows


def unique_boundary_ids(boundary_rows: list[dict[str, Any]], prefix: str) -> list[str]:
    ids: list[str] = []
    for row in boundary_rows:
        ad_interval_id = clean(row.get("ad_interval_id"))
        if prefix == "resnet":
            if to_bool(row.get("resnet_additional_start_hit")):
                ids.append(f"{ad_interval_id}:start")
            if to_bool(row.get("resnet_additional_end_hit")):
                ids.append(f"{ad_interval_id}:end")
        if prefix == "opencv":
            if to_bool(row.get("opencv_only_start_hit")):
                ids.append(f"{ad_interval_id}:start")
            if to_bool(row.get("opencv_only_end_hit")):
                ids.append(f"{ad_interval_id}:end")
    return ids


def build_comparison_review(
    comparison_df: pd.DataFrame,
    overlap_df: pd.DataFrame,
    boundary_rows: list[dict[str, Any]],
    optional_report: dict[str, Any],
) -> list[dict[str, Any]]:
    overlap_3s = int(pd.to_numeric(overlap_df.get("overlap_candidate_count_3s", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not overlap_df.empty else 0
    overlap_5s = int(pd.to_numeric(overlap_df.get("overlap_candidate_count_5s", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not overlap_df.empty else 0
    resnet_ids = optional_report.get("unique_resnet_boundary_hit_ids") or unique_boundary_ids(boundary_rows, "resnet")
    opencv_ids = optional_report.get("unique_opencv_boundary_hit_ids") or unique_boundary_ids(boundary_rows, "opencv")
    rows: list[dict[str, Any]] = []
    if not comparison_df.empty:
        for _, row in comparison_df.iterrows():
            data = safe_row_dict(row)
            rows.append(
                {
                    "method": clean(data.get("method")),
                    "total_candidate_count": fmt_number(data.get("total_candidate_count"), digits=0),
                    "candidate_count_per_video_mean": fmt_number(data.get("candidate_count_per_video_mean"), digits=6),
                    "start_hit_3s_count": fmt_number(data.get("start_hit_3s_count"), digits=0),
                    "start_hit_5s_count": fmt_number(data.get("start_hit_5s_count"), digits=0),
                    "end_hit_3s_count": fmt_number(data.get("end_hit_3s_count"), digits=0),
                    "end_hit_5s_count": fmt_number(data.get("end_hit_5s_count"), digits=0),
                    "boundary_hit_5s_count": fmt_number(data.get("boundary_hit_5s_count"), digits=0),
                    "boundary_hit_5s_rate": fmt_number(data.get("boundary_hit_5s_rate"), digits=6),
                    "both_boundary_hit_5s_count": fmt_number(data.get("both_boundary_hit_5s_count"), digits=0),
                    "total_ad_intervals": fmt_number(data.get("total_ad_intervals"), digits=0),
                    "total_boundaries": fmt_number(data.get("total_boundaries"), digits=0),
                    "unique_resnet_boundary_hits": len(resnet_ids),
                    "unique_opencv_boundary_hits": len(opencv_ids),
                    "overlap_candidate_count_3s": overlap_3s,
                    "overlap_candidate_count_5s": overlap_5s,
                    "notes": clean(data.get("notes")),
                    "unique_resnet_boundary_hit_ids": "; ".join(resnet_ids),
                    "unique_opencv_boundary_hit_ids": "; ".join(opencv_ids),
                    "preferred_method_for_boundary": "",
                    "combination_strategy": "",
                    "comparison_review_note": "",
                }
            )
    return rows


def distribution_text(values: list[str]) -> str:
    counter = Counter(value for value in values if clean(value))
    return "; ".join(f"{key}:{counter[key]}" for key in sorted(counter))


def count_boundary_hits_by_video(boundary_rows: list[dict[str, Any]], key_start: str, key_end: str) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for row in boundary_rows:
        counts[clean(row.get("video_id"))] += int(to_bool(row.get(key_start))) + int(to_bool(row.get(key_end)))
    return counts


def build_video_summary_review(
    resnet_summary_df: pd.DataFrame,
    resnet_candidate_rows: list[dict[str, Any]],
    boundary_rows: list[dict[str, Any]],
    overlap_df: pd.DataFrame,
) -> list[dict[str, Any]]:
    cands_by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in resnet_candidate_rows:
        cands_by_video[clean(row.get("video_id"))].append(row)
    resnet_boundary_counts = count_boundary_hits_by_video(boundary_rows, "start_hit_5s", "end_hit_5s")
    opencv_boundary_counts = count_boundary_hits_by_video(boundary_rows, "opencv_start_hit_5s", "opencv_end_hit_5s")
    resnet_only_counts = count_boundary_hits_by_video(boundary_rows, "resnet_additional_start_hit", "resnet_additional_end_hit")
    opencv_only_counts = count_boundary_hits_by_video(boundary_rows, "opencv_only_start_hit", "opencv_only_end_hit")
    overlap_by_video: dict[str, dict[str, Any]] = {}
    if not overlap_df.empty:
        for _, row in overlap_df.iterrows():
            data = safe_row_dict(row)
            overlap_by_video[clean(data.get("video_id"))] = data

    summary_source: dict[str, dict[str, Any]] = {}
    if not resnet_summary_df.empty:
        for _, row in resnet_summary_df.iterrows():
            data = safe_row_dict(row)
            summary_source[clean(data.get("video_id"))] = data
    video_ids = sorted(set(summary_source) | set(cands_by_video) | set(overlap_by_video), key=lambda value: int(value) if value.isdigit() else value)
    rows: list[dict[str, Any]] = []
    for video_id in video_ids:
        source = summary_source.get(video_id, {})
        overlap = overlap_by_video.get(video_id, {})
        candidates = cands_by_video.get(video_id, [])
        scores = pd.to_numeric(pd.Series([row.get("scene_change_score") for row in candidates]), errors="coerce").dropna()
        duration = clean(source.get("duration_sec"))
        duration_f = to_float(duration)
        resnet_count = len(candidates) if candidates else int(to_float(source.get("candidate_count")) or 0)
        opencv_count = int(to_float(overlap.get("unique_opencv_candidate_count")) or 0)
        resnet_per_min = resnet_count / (duration_f / 60.0) if duration_f and duration_f > 0 else ""
        opencv_per_min = opencv_count / (duration_f / 60.0) if duration_f and duration_f > 0 else ""
        resnet_only = resnet_only_counts.get(video_id, 0)
        if resnet_only > 0 or resnet_count > 100 or (to_float(resnet_per_min) is not None and (to_float(resnet_per_min) or 0) > 8):
            priority = "high"
        elif resnet_boundary_counts.get(video_id, 0) >= opencv_boundary_counts.get(video_id, 0) and resnet_count <= 80:
            priority = "low"
        else:
            priority = "medium"
        rows.append(
            {
                "video_id": video_id,
                "video_title": clean(source.get("video_title")) or clean(overlap.get("video_title")),
                "video_filename": clean(source.get("video_filename")),
                "video_duration_sec": fmt_number(duration),
                "video_duration_mmss": mmss(duration),
                "resnet_candidate_count": resnet_count,
                "opencv_candidate_count": opencv_count,
                "resnet_candidate_count_per_min": fmt_number(resnet_per_min, digits=3),
                "opencv_candidate_count_per_min": fmt_number(opencv_per_min, digits=3),
                "resnet_boundary_hit_5s_count": resnet_boundary_counts.get(video_id, 0),
                "opencv_boundary_hit_5s_count": opencv_boundary_counts.get(video_id, 0),
                "resnet_only_boundary_hit_count": resnet_only,
                "opencv_only_boundary_hit_count": opencv_only_counts.get(video_id, 0),
                "candidate_overlap_5s_count": fmt_number(overlap.get("overlap_candidate_count_5s"), digits=0),
                "max_resnet_scene_change_score": fmt_number(scores.max() if not scores.empty else ""),
                "median_resnet_scene_change_score": fmt_number(scores.median() if not scores.empty else source.get("score_p50")),
                "p95_resnet_scene_change_score": fmt_number(scores.quantile(0.95) if not scores.empty else source.get("score_p95")),
                "resnet_candidate_density_ok": "",
                "resnet_too_many_false_candidates": "",
                "resnet_too_few_candidates": "",
                "resnet_video_review_priority": priority,
                "video_review_note": "",
            }
        )
    return rows


def value_options_rows() -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    options = {
        "review_status": ["not_reviewed", "reviewed"],
        "yes_no_unclear": ["yes", "no", "unclear"],
        "scene_change_strength": ["strong", "medium", "weak", "unclear"],
        "scene_change_type": [
            "hard_cut",
            "fade",
            "broll",
            "camera_angle_change",
            "semantic_change",
            "object_change",
            "lighting_change",
            "text_overlay_only",
            "motion_only",
            "object_motion",
            "other",
            "unclear",
        ],
        "is_ad_boundary_related": ["ad_start", "ad_end", "inside_ad", "near_ad_but_not_boundary", "no", "unclear"],
        "resnet_candidate_usefulness": ["useful", "redundant_with_opencv", "false_positive", "unclear"],
        "false_positive_type": [
            "not_false_positive",
            "normal_cut",
            "camera_motion",
            "subtitle_change",
            "lighting_change",
            "object_motion",
            "semantic_but_not_boundary",
            "duplicate_candidate",
            "score_noise",
            "unclear",
        ],
        "keep_as_boundary_candidate": ["yes", "no", "unclear"],
        "boundary_quality": ["clear", "weak", "none", "unclear"],
        "resnet_boundary_better_than_opencv": ["yes", "no", "similar", "unclear"],
        "resnet_helped": ["yes", "no", "unclear"],
        "should_add_to_combined_scene_evidence": ["yes", "no", "unclear"],
        "failure_reason_for_opencv": ["subtle_semantic_change", "low_pixel_change", "threshold_missed", "no_actual_boundary", "unclear"],
        "combination_strategy": [
            "opencv_only",
            "resnet_only",
            "opencv_or_resnet",
            "opencv_and_resnet",
            "opencv_primary_resnet_fallback",
            "unclear",
        ],
        "video_review_priority": ["high", "medium", "low"],
    }
    max_len = max(len(values) for values in options.values())
    rows = []
    for idx in range(max_len):
        rows.append({key: values[idx] if idx < len(values) else "" for key, values in options.items()})
    return rows, options


def review_guide_rows() -> list[dict[str, Any]]:
    return [
        {
            "section": "검토 목적",
            "guide": (
                "ResNet embedding scene-change 후보가 실제 장면 전환인지 확인한다. OpenCV/ffmpeg가 놓친 광고 boundary를 "
                "ResNet이 보완하는지 확인한다. 이 검토는 최종 광고 탐지 성능 평가가 아니라 scene evidence 품질 검토이다."
            ),
        },
        {
            "section": "먼저 볼 sheet",
            "guide": (
                "1순위: `resnet_only_boundary_review`. 2순위: `resnet_boundary_review`. "
                "3순위: `resnet_candidate_review` 중 review_priority = very_high/high. "
                "4순위: `opencv_resnet_comparison`."
            ),
        },
        {
            "section": "후보 검토 방법",
            "guide": (
                "candidate_time_mmss 기준으로 영상에서 해당 시점으로 이동한다. 후보 시점 앞 5초, 뒤 5초를 본다. "
                "실제 컷 전환인지, 의미적 장면 변화인지, 광고 경계와 관련 있는지 확인한다."
            ),
        },
        {
            "section": "ResNet 후보가 유효한 경우",
            "guide": (
                "OpenCV가 놓친 광고 시작/종료 전환을 ResNet이 잡음. 픽셀 변화는 작지만 의미적으로 장면이 바뀜. "
                "사람/일상 장면에서 제품/서비스 화면으로 전환됨. 광고 블록으로 넘어가는 경계가 보임."
            ),
        },
        {
            "section": "ResNet 후보가 부적절한 경우",
            "guide": (
                "의미적 변화는 있지만 광고 boundary와 무관함. 일반 컷 전환 또는 일반 B-roll임. "
                "같은 광고 내부 장면 변화일 뿐 시작/종료 경계는 아님. 카메라 움직임, 물체 움직임, 자막 변화에 반응함."
            ),
        },
        {
            "section": "최종 rule 반영 기준",
            "guide": (
                "ResNet-only 후보가 실제 광고 boundary라면 combined scene evidence에 포함 가능. ResNet 후보가 오탐이 많다면 "
                "OpenCV primary + ResNet fallback 또는 보수적 결합 사용. OpenCV와 ResNet이 모두 잡은 후보는 strong boundary evidence로 볼 수 있음."
            ),
        },
    ]


def rows_to_df(rows: list[dict[str, Any]], columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame([{col: row.get(col, "") for col in columns} for row in rows], columns=columns)


def option_ranges(options: dict[str, list[str]]) -> dict[str, str]:
    ranges: dict[str, str] = {}
    for idx, (key, values) in enumerate(options.items(), start=1):
        col = get_column_letter(idx)
        ranges[key] = f"=value_options!${col}$2:${col}${len(values) + 1}"
    return ranges


def apply_basic_sheet_format(ws, review_columns: set[str], wrap_columns: set[str]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    review_fill = PatternFill("solid", fgColor="F4B183")
    header_font = Font(color="FFFFFF", bold=True)
    review_font = Font(color="000000", bold=True)
    for cell in ws[1]:
        header = clean(cell.value)
        cell.fill = review_fill if header in review_columns else header_fill
        cell.font = review_font if header in review_columns else header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        header = clean(ws.cell(row=1, column=col_idx).value)
        max_len = len(header)
        for cell in list(column_cells)[1 : min(ws.max_row, 80)]:
            max_len = max(max_len, len(clean(cell.value)))
            if header in wrap_columns:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        width = min(max(max_len + 2, 10), 55)
        if header.endswith("_note") or header in {"video_title", "video_filename", "video_path", "guide", "notes"}:
            width = min(max(width, 28), 70)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def col_index(ws, header: str) -> int | None:
    for idx, cell in enumerate(ws[1], start=1):
        if clean(cell.value) == header:
            return idx
    return None


def add_validation(ws, header: str, formula: str) -> bool:
    idx = col_index(ws, header)
    if idx is None or ws.max_row < 2:
        return False
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(idx)}2:{get_column_letter(idx)}{max(ws.max_row, 200)}")
    return True


def add_row_formula_fill(ws, header: str, value: str, fill_color: str) -> bool:
    idx = col_index(ws, header)
    if idx is None or ws.max_row < 2:
        return False
    col = get_column_letter(idx)
    fill = PatternFill("solid", fgColor=fill_color)
    ws.conditional_formatting.add(
        f"A2:{get_column_letter(ws.max_column)}{ws.max_row}",
        FormulaRule(formula=[f'=${col}2="{value}"'], fill=fill),
    )
    return True


def add_cell_formula_fill(ws, header: str, value: str, fill_color: str) -> bool:
    idx = col_index(ws, header)
    if idx is None or ws.max_row < 2:
        return False
    col = get_column_letter(idx)
    fill = PatternFill("solid", fgColor=fill_color)
    ws.conditional_formatting.add(
        f"{col}2:{col}{ws.max_row}",
        FormulaRule(formula=[f'=${col}2="{value}"'], fill=fill),
    )
    return True


def apply_excel_formatting(path: Path, options: dict[str, list[str]]) -> tuple[bool, bool]:
    wb = load_workbook(path)
    ranges = option_ranges(options)
    review_columns = {
        "review_status",
        "is_true_scene_change",
        "scene_change_strength",
        "scene_change_type",
        "is_ad_boundary_related",
        "resnet_candidate_usefulness",
        "false_positive_type",
        "keep_as_boundary_candidate",
        "review_note",
        "reviewer",
        "reviewed_at",
        "start_candidate_correct",
        "end_candidate_correct",
        "actual_start_transition_visible",
        "actual_end_transition_visible",
        "ad_start_boundary_quality",
        "ad_end_boundary_quality",
        "resnet_boundary_better_than_opencv",
        "start_boundary_review_note",
        "end_boundary_review_note",
        "overall_boundary_review_status",
        "is_actual_scene_change",
        "is_actual_ad_boundary",
        "resnet_helped",
        "should_add_to_combined_scene_evidence",
        "failure_reason_for_opencv",
        "preferred_method_for_boundary",
        "combination_strategy",
        "comparison_review_note",
        "resnet_candidate_density_ok",
        "resnet_too_many_false_candidates",
        "resnet_too_few_candidates",
        "resnet_video_review_priority",
        "video_review_note",
    }
    wrap_columns = {
        "video_title",
        "video_filename",
        "video_path",
        "review_note",
        "start_boundary_review_note",
        "end_boundary_review_note",
        "comparison_review_note",
        "video_review_note",
        "guide",
        "notes",
        "unique_resnet_boundary_hit_ids",
        "unique_opencv_boundary_hit_ids",
    }
    for ws in wb.worksheets:
        apply_basic_sheet_format(ws, review_columns, wrap_columns)

    dropdowns_applied = True
    cf_applied = True

    ws = wb["resnet_candidate_review"]
    dropdowns_applied &= add_validation(ws, "review_status", ranges["review_status"])
    dropdowns_applied &= add_validation(ws, "is_true_scene_change", ranges["yes_no_unclear"])
    dropdowns_applied &= add_validation(ws, "scene_change_strength", ranges["scene_change_strength"])
    dropdowns_applied &= add_validation(ws, "scene_change_type", ranges["scene_change_type"])
    dropdowns_applied &= add_validation(ws, "is_ad_boundary_related", ranges["is_ad_boundary_related"])
    dropdowns_applied &= add_validation(ws, "resnet_candidate_usefulness", ranges["resnet_candidate_usefulness"])
    dropdowns_applied &= add_validation(ws, "false_positive_type", ranges["false_positive_type"])
    dropdowns_applied &= add_validation(ws, "keep_as_boundary_candidate", ranges["keep_as_boundary_candidate"])
    cf_applied &= add_row_formula_fill(ws, "review_priority", "very_high", "F8CBAD")
    cf_applied &= add_row_formula_fill(ws, "review_priority", "high", "FCE4D6")
    cf_applied &= add_cell_formula_fill(ws, "is_near_any_ad_boundary_5s", "true", "D9EAF7")
    cf_applied &= add_cell_formula_fill(ws, "review_status", "not_reviewed", "E7E6E6")
    cf_applied &= add_cell_formula_fill(ws, "keep_as_boundary_candidate", "yes", "E2F0D9")
    cf_applied &= add_cell_formula_fill(ws, "keep_as_boundary_candidate", "no", "FCE4D6")

    ws = wb["resnet_boundary_review"]
    for header in [
        "start_candidate_correct",
        "end_candidate_correct",
        "actual_start_transition_visible",
        "actual_end_transition_visible",
    ]:
        dropdowns_applied &= add_validation(ws, header, ranges["yes_no_unclear"])
    dropdowns_applied &= add_validation(ws, "ad_start_boundary_quality", ranges["boundary_quality"])
    dropdowns_applied &= add_validation(ws, "ad_end_boundary_quality", ranges["boundary_quality"])
    dropdowns_applied &= add_validation(ws, "resnet_boundary_better_than_opencv", ranges["resnet_boundary_better_than_opencv"])
    dropdowns_applied &= add_validation(ws, "overall_boundary_review_status", ranges["review_status"])
    cf_applied &= add_cell_formula_fill(ws, "overall_boundary_review_status", "not_reviewed", "E7E6E6")

    ws = wb["resnet_only_boundary_review"]
    for header in ["is_actual_scene_change", "is_actual_ad_boundary", "resnet_helped", "should_add_to_combined_scene_evidence"]:
        dropdowns_applied &= add_validation(ws, header, ranges["yes_no_unclear"])
    dropdowns_applied &= add_validation(ws, "failure_reason_for_opencv", ranges["failure_reason_for_opencv"])
    cf_applied &= add_row_formula_fill(ws, "opencv_hit_same_boundary_5s", "false", "F8CBAD")
    cf_applied &= add_cell_formula_fill(ws, "should_add_to_combined_scene_evidence", "yes", "E2F0D9")
    cf_applied &= add_cell_formula_fill(ws, "should_add_to_combined_scene_evidence", "no", "FCE4D6")

    ws = wb["opencv_resnet_comparison"]
    dropdowns_applied &= add_validation(ws, "preferred_method_for_boundary", ranges["combination_strategy"])
    dropdowns_applied &= add_validation(ws, "combination_strategy", ranges["combination_strategy"])

    ws = wb["video_summary_review"]
    for header in ["resnet_candidate_density_ok", "resnet_too_many_false_candidates", "resnet_too_few_candidates"]:
        dropdowns_applied &= add_validation(ws, header, ranges["yes_no_unclear"])
    dropdowns_applied &= add_validation(ws, "resnet_video_review_priority", ranges["video_review_priority"])
    cf_applied &= add_row_formula_fill(ws, "resnet_video_review_priority", "high", "FCE4D6")

    wb.save(path)
    return dropdowns_applied, cf_applied


def update_readme() -> None:
    section_title = "## ResNet Scene Candidate Human Review v2.3"
    section = f"""{section_title}

ResNet embedding scene-change 후보를 사람이 검토하기 위한 Excel 파일을 생성했다.

특히 OpenCV가 놓친 ResNet-only 광고 boundary 후보를 우선 검토한다.
검토 결과는 OpenCV/ResNet scene evidence 결합 rule 설계에 사용된다.

생성된 주요 파일:

- `data/review/resnet_scene_candidate_human_review_v2_3.xlsx`
- `data/review/resnet_scene_candidate_human_review_candidate_sheet_v2_3.csv`
- `data/review/resnet_scene_candidate_human_review_boundary_sheet_v2_3.csv`
- `data/review/resnet_only_boundary_review_v2_3.csv`
- `data/review/opencv_resnet_scene_comparison_review_v2_3.csv`
- `reports/resnet_scene_candidate_human_review_v2_3_report.json`
- `reports/resnet_scene_candidate_human_review_v2_3_summary.md`
- `logs/resnet_scene_candidate_human_review_v2_3_run_log.txt`
"""
    text = README_PATH.read_text(encoding="utf-8") if README_PATH.exists() else "# youtube_ad_segment_detector\n"
    if section_title in text:
        before, _, rest = text.partition(section_title)
        marker = "\n## "
        if marker in rest:
            _, _, after = rest.partition(marker)
            text = before.rstrip() + "\n\n" + section.rstrip() + "\n\n## " + after.lstrip()
        else:
            text = before.rstrip() + "\n\n" + section.rstrip() + "\n"
    else:
        text = text.rstrip() + "\n\n" + section.rstrip() + "\n"
    README_PATH.write_text(text, encoding="utf-8")


def latest_safety_warnings() -> list[Any]:
    warnings: list[Any] = []
    for path in LATEST_DIR.rglob("*"):
        if not path.is_file():
            continue
        if path.name not in CURRENT_LATEST_FILENAMES:
            continue
        name = path.name.lower()
        suffix = path.suffix.lower()
        if suffix == ".mp4":
            warnings.append({"disallowed_latest_mp4": str(path)})
        if suffix == ".xlsx" and path.name in {"new_ad_labeling.xlsx", "clean_ad_labels_v0_review.xlsx"}:
            warnings.append({"disallowed_latest_original_xlsx": str(path)})
        if any(part in name for part in ["frame", "model", "checkpoint", "cache"]):
            warnings.append({"latest_filename_contains_disallowed_part": str(path)})
    return warnings


def copy_latest() -> list[str]:
    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    files = [
        OUTPUT_XLSX_PATH,
        OUTPUT_CANDIDATE_CSV_PATH,
        OUTPUT_BOUNDARY_CSV_PATH,
        OUTPUT_RESNET_ONLY_CSV_PATH,
        OUTPUT_COMPARISON_CSV_PATH,
        REPORT_PATH,
        SUMMARY_PATH,
        LOG_PATH,
    ]
    copied: list[str] = []
    for path in files:
        if path.exists():
            target = LATEST_DIR / path.name
            shutil.copy2(path, target)
            copied.append(str(target))
    latest_readme = """# latest_for_chatgpt files

이번 작업명: create_resnet_scene_candidate_human_review_xlsx_v2_3

ResNet scene candidate human review workbook, companion CSV, report, summary, and run log만 복사했다.
mp4, 원본 label xlsx, frame image, model, checkpoint, cache 파일은 복사하지 않는다.

복사 파일:

- `resnet_scene_candidate_human_review_v2_3.xlsx`
- `resnet_scene_candidate_human_review_candidate_sheet_v2_3.csv`
- `resnet_scene_candidate_human_review_boundary_sheet_v2_3.csv`
- `resnet_only_boundary_review_v2_3.csv`
- `opencv_resnet_scene_comparison_review_v2_3.csv`
- `resnet_scene_candidate_human_review_v2_3_report.json`
- `resnet_scene_candidate_human_review_v2_3_summary.md`
- `resnet_scene_candidate_human_review_v2_3_run_log.txt`
"""
    LATEST_README_PATH.write_text(latest_readme, encoding="utf-8")
    copied.append(str(LATEST_README_PATH))
    return copied


def write_summary(report: dict[str, Any]) -> None:
    summary = f"""# ResNet Scene Candidate Human Review v2.3

## 생성 파일

- Review Excel: `{report.get("output_xlsx_path", "")}`
- Candidate companion CSV: `{report.get("output_candidate_review_csv_path", "")}`
- Boundary companion CSV: `{report.get("output_boundary_review_csv_path", "")}`
- ResNet-only boundary CSV: `{report.get("output_resnet_only_boundary_review_csv_path", "")}`
- OpenCV/ResNet comparison CSV: `{report.get("output_comparison_review_csv_path", "")}`
- QA report: `{REPORT_PATH}`
- Run log: `{LOG_PATH}`

## 요약

- 총 ResNet candidate 수: {report.get("candidate_row_count", 0)}
- ResNet-only boundary 검토 row 수: {report.get("resnet_only_boundary_review_row_count", 0)}
- 광고 boundary review row 수: {report.get("resnet_boundary_review_row_count", 0)}
- OpenCV/ResNet 비교 sheet 생성: {report.get("opencv_resnet_comparison_created", False)}
- video summary row 수: {report.get("video_summary_row_count", 0)}
- very high priority 후보 수: {report.get("very_high_priority_candidate_count", 0)}
- dropdown validation 적용: {report.get("dropdown_validation_applied")}
- conditional formatting 적용: {report.get("conditional_formatting_applied")}

## 먼저 검토할 위치

1. `resnet_only_boundary_review` sheet의 `is_actual_scene_change`, `is_actual_ad_boundary`, `resnet_helped`, `should_add_to_combined_scene_evidence`
2. `resnet_boundary_review` sheet의 `resnet_additional_*`, `opencv_only_*`, `resnet_boundary_better_than_opencv`
3. `resnet_candidate_review` sheet의 `review_priority=very_high/high` row
4. `opencv_resnet_comparison` sheet의 `preferred_method_for_boundary`, `combination_strategy`

## 다음 작업

- ResNet-only 8개 boundary 후보를 먼저 사람이 확인한다.
- 유효한 후보는 combined scene evidence에 포함할지 기록한다.
- 검토 결과를 바탕으로 OpenCV primary + ResNet fallback 또는 결합 rule을 설계한다.
"""
    SUMMARY_PATH.write_text(summary, encoding="utf-8")


def write_log(report: dict[str, Any]) -> None:
    lines = [
        "작업명: create_resnet_scene_candidate_human_review_xlsx_v2_3",
        f"estimated_runtime: {report.get('estimated_runtime')}",
        f"runtime_estimation_reason: {report.get('runtime_estimation_reason')}",
        f"start_time: {report.get('start_time')}",
        f"end_time: {report.get('end_time')}",
        f"actual_runtime_seconds: {report.get('actual_runtime_seconds')}",
        f"actual_runtime_readable: {report.get('actual_runtime_readable')}",
        "",
        *RUN_LOG,
        "",
        "sub_agent_results:",
        json.dumps(report.get("sub_agent_results", {}), ensure_ascii=False, indent=2),
        "",
        "warnings:",
        json.dumps(report.get("warnings", []), ensure_ascii=False, indent=2),
        "",
        "errors:",
        json.dumps(report.get("errors", []), ensure_ascii=False, indent=2),
        "",
    ]
    LOG_PATH.write_text("\n".join(lines), encoding="utf-8")


def missing_columns(df: pd.DataFrame, required: list[str]) -> list[str]:
    return [col for col in required if col not in df.columns]


def generate() -> int:
    start_time = now_iso()
    start_monotonic = time.monotonic()
    warnings: list[Any] = []
    errors: list[Any] = []
    missing_input_files: list[str] = []
    log(f"작업 시작: {start_time}")
    log(f"예상 작업 시간: {ESTIMATED_RUNTIME}")
    log(f"예상 근거: {RUNTIME_ESTIMATION_REASON}")

    before_old_snapshot = old_project_snapshot()
    log("기존 프로젝트 스냅샷 기록 완료")

    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

    resnet_raw_df = read_csv_optional(INPUT_RESNET_CANDIDATE_RAW_PATH, missing_input_files, warnings)
    resnet_df = read_csv_optional(INPUT_RESNET_CANDIDATE_MMSS_PATH, missing_input_files, warnings)
    if resnet_df.empty and not resnet_raw_df.empty:
        resnet_df = resnet_raw_df
        warnings.append({"using_resnet_raw_candidate_csv_because_mmss_empty": str(INPUT_RESNET_CANDIDATE_RAW_PATH)})
    resnet_boundary_df = read_csv_optional(INPUT_RESNET_BOUNDARY_AUDIT_PATH, missing_input_files, warnings)
    resnet_video_summary_df = read_csv_optional(INPUT_RESNET_VIDEO_SUMMARY_PATH, missing_input_files, warnings)
    comparison_df = read_csv_optional(INPUT_OPENCV_COMPARISON_PATH, missing_input_files, warnings)
    overlap_df = read_csv_optional(INPUT_CANDIDATE_OVERLAP_PATH, missing_input_files, warnings)
    ad_df = read_csv_optional(INPUT_AD_SEGMENT_PATH, missing_input_files, warnings)
    _opencv_raw_df = read_csv_optional(INPUT_OPENCV_CANDIDATE_RAW_PATH, missing_input_files, warnings)
    opencv_df = read_csv_optional(INPUT_OPENCV_CANDIDATE_MMSS_PATH, missing_input_files, warnings)
    if opencv_df.empty and not _opencv_raw_df.empty:
        opencv_df = _opencv_raw_df
    opencv_boundary_df = read_csv_optional(INPUT_OPENCV_BOUNDARY_AUDIT_PATH, missing_input_files, warnings)
    optional_report = load_json_optional(OPTIONAL_RESNET_REPORT_PATH)
    if not OPTIONAL_RESNET_REPORT_PATH.exists():
        warnings.append({"missing_optional_file": str(OPTIONAL_RESNET_REPORT_PATH)})
    if not OPTIONAL_RESNET_SUMMARY_PATH.exists():
        warnings.append({"missing_optional_file": str(OPTIONAL_RESNET_SUMMARY_PATH)})

    missing_required_columns = {
        "resnet_candidate_csv": missing_columns(
            resnet_df,
            ["video_id", "candidate_time_sec", "candidate_time_mmss", "scene_change_score", "nearest_ad_interval_id"],
        ),
        "resnet_boundary_audit_csv": missing_columns(
            resnet_boundary_df,
            ["ad_interval_id", "video_id", "ad_start_sec", "ad_end_sec", "start_hit_5s", "end_hit_5s"],
        ),
        "opencv_comparison_csv": missing_columns(comparison_df, ["method", "total_candidate_count", "boundary_hit_5s_count"]),
        "candidate_overlap_csv": missing_columns(overlap_df, ["video_id", "overlap_candidate_count_5s"]),
        "ad_segment_csv": missing_columns(ad_df, ["ad_interval_id", "video_id", "ad_start_sec", "ad_end_sec"]),
        "opencv_candidate_csv": missing_columns(opencv_df, ["video_id", "candidate_time_sec"]),
        "opencv_boundary_audit_csv": missing_columns(opencv_boundary_df, ["ad_interval_id", "start_hit_5s", "end_hit_5s"]),
    }

    ad_intervals = build_ad_intervals(ad_df)
    ad_by_id = index_by_id(ad_intervals, "ad_interval_id")
    resnet_audit_by_id = audit_rows_by_id(resnet_boundary_df)
    opencv_audit_by_id = audit_rows_by_id(opencv_boundary_df)
    opencv_times = candidate_times_by_video(opencv_df)
    resnet_candidate_rows, requested_candidate_cols = build_resnet_candidate_review(
        resnet_df,
        opencv_times,
        opencv_audit_by_id,
        intervals_by_video(ad_intervals),
    )
    missing_required_columns["resnet_requested_source_columns_missing_or_computed"] = [
        col for col in requested_candidate_cols if col not in resnet_df.columns
    ]
    boundary_rows = build_boundary_review(ad_intervals, resnet_audit_by_id, opencv_audit_by_id)
    resnet_only_rows = build_resnet_only_boundary_review(boundary_rows, resnet_candidate_rows)
    comparison_rows = build_comparison_review(comparison_df, overlap_df, boundary_rows, optional_report)
    video_summary_rows = build_video_summary_review(
        resnet_video_summary_df,
        resnet_candidate_rows,
        boundary_rows,
        overlap_df,
    )
    value_rows, options = value_options_rows()
    guide_rows = review_guide_rows()

    candidate_columns = [
        "video_id",
        "video_title",
        "video_filename",
        "video_path",
        "candidate_time_sec",
        "candidate_time_mmss",
        "candidate_time_mmss_floor",
        "candidate_time_mmss_round",
        "scene_change_score",
        "threshold",
        "cosine_distance",
        "l2_distance",
        "score_rank_in_video",
        "score_percentile_in_video",
        "candidate_source",
        "method_used",
        "model_name",
        "feature_dim",
        "nearest_window_id",
        "nearest_ad_interval_id",
        "nearest_ad_boundary_type",
        "nearest_ad_boundary_sec",
        "nearest_ad_boundary_mmss",
        "distance_to_nearest_ad_boundary_sec",
        "is_near_ad_start_3s",
        "is_near_ad_start_5s",
        "is_near_ad_end_3s",
        "is_near_ad_end_5s",
        "is_near_any_ad_boundary_5s",
        "review_clip_start_sec",
        "review_clip_end_sec",
        "review_clip_start_mmss",
        "review_clip_end_mmss",
        "has_near_opencv_candidate_3s",
        "has_near_opencv_candidate_5s",
        "nearest_opencv_candidate_time_sec",
        "nearest_opencv_candidate_mmss",
        "distance_to_nearest_opencv_candidate_sec",
        "review_priority",
        "review_status",
        "is_true_scene_change",
        "scene_change_strength",
        "scene_change_type",
        "is_ad_boundary_related",
        "resnet_candidate_usefulness",
        "false_positive_type",
        "keep_as_boundary_candidate",
        "review_note",
        "reviewer",
        "reviewed_at",
    ]
    boundary_columns = [
        "ad_interval_id",
        "video_id",
        "video_title",
        "video_filename",
        "ad_start_sec",
        "ad_start_mmss",
        "ad_end_sec",
        "ad_end_mmss",
        "ad_duration_sec",
        "start_hit_3s",
        "start_hit_5s",
        "end_hit_3s",
        "end_hit_5s",
        "both_boundary_hit_5s",
        "nearest_candidate_to_start_sec",
        "nearest_candidate_to_start_mmss",
        "distance_to_start_candidate_sec",
        "nearest_candidate_to_end_sec",
        "nearest_candidate_to_end_mmss",
        "distance_to_end_candidate_sec",
        "candidate_count_in_ad_interval",
        "candidate_count_in_pre10",
        "candidate_count_in_post10",
        "candidate_count_near_start_5s",
        "candidate_count_near_end_5s",
        "opencv_start_hit_5s",
        "opencv_end_hit_5s",
        "opencv_both_boundary_hit_5s",
        "resnet_additional_start_hit",
        "resnet_additional_end_hit",
        "opencv_only_start_hit",
        "opencv_only_end_hit",
        "start_candidate_correct",
        "end_candidate_correct",
        "actual_start_transition_visible",
        "actual_end_transition_visible",
        "ad_start_boundary_quality",
        "ad_end_boundary_quality",
        "resnet_boundary_better_than_opencv",
        "start_boundary_review_note",
        "end_boundary_review_note",
        "overall_boundary_review_status",
    ]
    resnet_only_columns = [
        "ad_interval_id",
        "video_id",
        "video_title",
        "video_filename",
        "boundary_type",
        "ad_boundary_sec",
        "ad_boundary_mmss",
        "resnet_candidate_time_sec",
        "resnet_candidate_mmss",
        "distance_to_boundary_sec",
        "resnet_scene_change_score",
        "resnet_score_percentile_in_video",
        "nearest_opencv_candidate_time_sec",
        "nearest_opencv_candidate_mmss",
        "distance_to_nearest_opencv_candidate_sec",
        "opencv_hit_same_boundary_5s",
        "review_clip_start_mmss",
        "review_clip_end_mmss",
        "is_actual_scene_change",
        "is_actual_ad_boundary",
        "resnet_helped",
        "should_add_to_combined_scene_evidence",
        "failure_reason_for_opencv",
        "review_note",
        "reviewer",
        "reviewed_at",
    ]
    comparison_columns = [
        "method",
        "total_candidate_count",
        "candidate_count_per_video_mean",
        "start_hit_3s_count",
        "start_hit_5s_count",
        "end_hit_3s_count",
        "end_hit_5s_count",
        "boundary_hit_5s_count",
        "boundary_hit_5s_rate",
        "both_boundary_hit_5s_count",
        "total_ad_intervals",
        "total_boundaries",
        "unique_resnet_boundary_hits",
        "unique_opencv_boundary_hits",
        "overlap_candidate_count_3s",
        "overlap_candidate_count_5s",
        "notes",
        "unique_resnet_boundary_hit_ids",
        "unique_opencv_boundary_hit_ids",
        "preferred_method_for_boundary",
        "combination_strategy",
        "comparison_review_note",
    ]
    video_columns = [
        "video_id",
        "video_title",
        "video_filename",
        "video_duration_sec",
        "video_duration_mmss",
        "resnet_candidate_count",
        "opencv_candidate_count",
        "resnet_candidate_count_per_min",
        "opencv_candidate_count_per_min",
        "resnet_boundary_hit_5s_count",
        "opencv_boundary_hit_5s_count",
        "resnet_only_boundary_hit_count",
        "opencv_only_boundary_hit_count",
        "candidate_overlap_5s_count",
        "max_resnet_scene_change_score",
        "median_resnet_scene_change_score",
        "p95_resnet_scene_change_score",
        "resnet_candidate_density_ok",
        "resnet_too_many_false_candidates",
        "resnet_too_few_candidates",
        "resnet_video_review_priority",
        "video_review_note",
    ]
    value_columns = list(options.keys())
    guide_columns = ["section", "guide"]

    log("CSV companion 파일 생성 중")
    write_csv(OUTPUT_CANDIDATE_CSV_PATH, resnet_candidate_rows, candidate_columns)
    write_csv(OUTPUT_BOUNDARY_CSV_PATH, boundary_rows, boundary_columns)
    write_csv(OUTPUT_RESNET_ONLY_CSV_PATH, resnet_only_rows, resnet_only_columns)
    write_csv(OUTPUT_COMPARISON_CSV_PATH, comparison_rows, comparison_columns)

    log("Excel workbook 생성 중")
    with pd.ExcelWriter(OUTPUT_XLSX_PATH, engine="openpyxl") as writer:
        rows_to_df(resnet_candidate_rows, candidate_columns).to_excel(writer, sheet_name="resnet_candidate_review", index=False)
        rows_to_df(boundary_rows, boundary_columns).to_excel(writer, sheet_name="resnet_boundary_review", index=False)
        rows_to_df(resnet_only_rows, resnet_only_columns).to_excel(writer, sheet_name="resnet_only_boundary_review", index=False)
        rows_to_df(comparison_rows, comparison_columns).to_excel(writer, sheet_name="opencv_resnet_comparison", index=False)
        rows_to_df(video_summary_rows, video_columns).to_excel(writer, sheet_name="video_summary_review", index=False)
        rows_to_df(value_rows, value_columns).to_excel(writer, sheet_name="value_options", index=False)
        rows_to_df(guide_rows, guide_columns).to_excel(writer, sheet_name="review_guide", index=False)
    dropdown_applied, conditional_applied = apply_excel_formatting(OUTPUT_XLSX_PATH, options)
    log("Excel formatting, dropdown, conditional formatting 적용 완료")

    very_high_count = sum(1 for row in resnet_candidate_rows if clean(row.get("review_priority")) == "very_high")
    high_count = sum(1 for row in resnet_candidate_rows if clean(row.get("review_priority")) == "high")
    after_old_snapshot = old_project_snapshot()
    old_project_modified = before_old_snapshot != after_old_snapshot
    warnings.extend(latest_safety_warnings())
    if old_project_modified:
        warnings.append({"old_project_modified": True})

    end_time = now_iso()
    actual_runtime_seconds = round(time.monotonic() - start_monotonic, 3)
    report: dict[str, Any] = {
        "project_root": str(PROJECT_ROOT),
        "estimated_runtime": ESTIMATED_RUNTIME,
        "runtime_estimation_reason": RUNTIME_ESTIMATION_REASON,
        "start_time": start_time,
        "end_time": end_time,
        "actual_runtime_seconds": actual_runtime_seconds,
        "actual_runtime_readable": readable_runtime(actual_runtime_seconds),
        "input_resnet_candidate_path": str(INPUT_RESNET_CANDIDATE_MMSS_PATH),
        "input_resnet_candidate_raw_path": str(INPUT_RESNET_CANDIDATE_RAW_PATH),
        "input_resnet_boundary_audit_path": str(INPUT_RESNET_BOUNDARY_AUDIT_PATH),
        "input_opencv_comparison_path": str(INPUT_OPENCV_COMPARISON_PATH),
        "input_candidate_overlap_path": str(INPUT_CANDIDATE_OVERLAP_PATH),
        "input_ad_segment_path": str(INPUT_AD_SEGMENT_PATH),
        "input_opencv_candidate_path": str(INPUT_OPENCV_CANDIDATE_MMSS_PATH),
        "input_opencv_boundary_audit_path": str(INPUT_OPENCV_BOUNDARY_AUDIT_PATH),
        "candidate_row_count": len(resnet_df),
        "ad_interval_count": len(ad_intervals),
        "video_count": len(video_summary_rows),
        "output_xlsx_path": str(OUTPUT_XLSX_PATH),
        "output_candidate_review_csv_path": str(OUTPUT_CANDIDATE_CSV_PATH),
        "output_boundary_review_csv_path": str(OUTPUT_BOUNDARY_CSV_PATH),
        "output_resnet_only_boundary_review_csv_path": str(OUTPUT_RESNET_ONLY_CSV_PATH),
        "output_comparison_review_csv_path": str(OUTPUT_COMPARISON_CSV_PATH),
        "sheet_names": SHEET_NAMES,
        "dropdown_validation_applied": dropdown_applied,
        "conditional_formatting_applied": conditional_applied,
        "missing_input_files": missing_input_files,
        "missing_required_columns": missing_required_columns,
        "resnet_candidate_review_row_count": len(resnet_candidate_rows),
        "resnet_boundary_review_row_count": len(boundary_rows),
        "resnet_only_boundary_review_row_count": len(resnet_only_rows),
        "opencv_resnet_comparison_row_count": len(comparison_rows),
        "opencv_resnet_comparison_created": len(comparison_rows) > 0,
        "video_summary_row_count": len(video_summary_rows),
        "very_high_priority_candidate_count": very_high_count,
        "high_priority_candidate_count": high_count,
        "resnet_only_boundary_count": len(resnet_only_rows),
        "unique_resnet_boundary_hit_ids": [f"{row['ad_interval_id']}:{'start' if row['boundary_type'] == 'ad_start' else 'end'}" for row in resnet_only_rows],
        "sub_agent_results": {
            "resnet_candidate_sheet_validation": {"status": "PENDING", "details": "Sub Agent 1 not recorded yet."},
            "boundary_comparison_sheet_validation": {"status": "PENDING", "details": "Sub Agent 2 not recorded yet."},
            "output_safety_validation": {"status": "PENDING", "details": "Sub Agent 3 not recorded yet."},
        },
        "old_project_modified": old_project_modified,
        "old_project_snapshot_before": before_old_snapshot,
        "old_project_snapshot_after": after_old_snapshot,
        "latest_for_chatgpt_copied": [],
        "warnings": warnings,
        "errors": errors,
    }

    update_readme()
    write_summary(report)
    write_log(report)
    report["latest_for_chatgpt_copied"] = copy_latest()
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    write_summary(report)
    write_log(report)
    copy_latest()
    log(f"작업 종료: {end_time}")
    return 0 if not errors else 1


def record_sub_agent_results(results: list[list[str]]) -> int:
    if not REPORT_PATH.exists():
        raise FileNotFoundError(REPORT_PATH)
    report = json.loads(REPORT_PATH.read_text(encoding="utf-8"))
    start = report.get("start_time")
    for key, status, details in results:
        report.setdefault("sub_agent_results", {})[key] = {"status": status, "details": details}
        log(f"sub agent result recorded: {key}={status} - {details}")
    before_snapshot = report.get("old_project_snapshot_before") or old_project_snapshot()
    after_snapshot = old_project_snapshot()
    report["old_project_snapshot_after"] = after_snapshot
    report["old_project_modified"] = before_snapshot != after_snapshot
    end = now_iso()
    report["end_time"] = end
    if start:
        try:
            start_dt = datetime.fromisoformat(start)
            end_dt = datetime.fromisoformat(end)
            seconds = round((end_dt - start_dt).total_seconds(), 3)
            report["actual_runtime_seconds"] = seconds
            report["actual_runtime_readable"] = readable_runtime(seconds)
        except Exception:
            pass
    warnings = report.setdefault("warnings", [])
    warnings.extend(latest_safety_warnings())
    if report["old_project_modified"]:
        warnings.append({"old_project_modified_after_sub_agent_verification": True})
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    write_summary(report)
    write_log(report)
    copy_latest()
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--record-sub-agent",
        nargs=3,
        action="append",
        metavar=("KEY", "STATUS", "DETAILS"),
        help="Record a sub-agent PASS/WARN/FAIL result in the report/log.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.record_sub_agent:
        return record_sub_agent_results(args.record_sub_agent)
    return generate()


if __name__ == "__main__":
    raise SystemExit(main())
