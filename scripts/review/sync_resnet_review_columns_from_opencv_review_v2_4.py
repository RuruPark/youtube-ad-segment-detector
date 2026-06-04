#!/usr/bin/env python3
from __future__ import annotations

import csv, hashlib, json, os, shutil, subprocess, sys, time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path('.')
OLD_PROJECT_ROOT = Path('./_old_project_not_included')
CHECK_ENV_SCRIPT = PROJECT_ROOT / 'scripts/utils/check_cv_environment.py'
TASK_NAME = 'sync_resnet_review_columns_from_opencv_review_v2_4'
ESTIMATED_RUNTIME = '약 15분'
RUNTIME_ESTIMATION_REASON = 'OpenCV review workbook에서 사용자 review 컬럼을 추출하고, ResNet v2_4 workbook의 관련 sheet에 누락 컬럼/dropdown/formatting을 반영하는 작업 기준'

SOURCE_WORKBOOK = PROJECT_ROOT / 'data/review/scene_candidate_human_review_v2_3.xlsx'
TARGET_WORKBOOK = PROJECT_ROOT / 'data/review/resnet_scene_candidate_human_review_v2_4.xlsx'
BACKUP_DIR = PROJECT_ROOT / 'data/review/backups'
CANDIDATE_CSV = PROJECT_ROOT / 'data/review/resnet_scene_candidate_human_review_candidate_sheet_v2_4.csv'
BOUNDARY_CSV = PROJECT_ROOT / 'data/review/resnet_scene_candidate_human_review_boundary_sheet_v2_4.csv'
RESNET_ONLY_CSV = PROJECT_ROOT / 'data/review/resnet_only_boundary_review_v2_4.csv'
COMPARISON_CSV = PROJECT_ROOT / 'data/review/opencv_resnet_scene_comparison_review_v2_4.csv'
VIDEO_SUMMARY_CSV = PROJECT_ROOT / 'data/review/resnet_scene_candidate_human_review_video_summary_v2_4.csv'
SYNC_AUDIT_CSV = PROJECT_ROOT / 'data/review/resnet_review_columns_sync_audit_v2_4.csv'
SOURCE_SCHEMA_CSV = PROJECT_ROOT / 'data/review/resnet_review_columns_source_schema_v2_3.csv'
TARGET_SCHEMA_CSV = PROJECT_ROOT / 'data/review/resnet_review_columns_target_schema_after_patch_v2_4.csv'
REPORT_PATH = PROJECT_ROOT / 'reports/sync_resnet_review_columns_from_opencv_review_v2_4_report.json'
SUMMARY_PATH = PROJECT_ROOT / 'reports/sync_resnet_review_columns_from_opencv_review_v2_4_summary.md'
LOG_PATH = PROJECT_ROOT / 'logs/sync_resnet_review_columns_from_opencv_review_v2_4_run_log.txt'
README_PATH = PROJECT_ROOT / 'README.md'
LATEST_DIR = PROJECT_ROOT / 'outputs/latest_for_chatgpt'
LATEST_README = LATEST_DIR / 'README_latest_files.md'

SOURCE_SHEETS = ['scene_candidate_review','ad_boundary_review','video_summary_review','value_options','review_guide']
TARGET_SHEETS = ['resnet_candidate_review','resnet_boundary_review','resnet_only_boundary_review','opencv_resnet_comparison','video_summary_review','value_options','review_guide']
EXPECTED_SOURCE_REVIEW = {
    'scene_candidate_review': ['review_status','is_true_scene_change','scene_change_strength','scene_change_type','is_ad_boundary_related','false_positive_type','keep_as_boundary_candidate','review_note','reviewer','reviewed_at'],
    'ad_boundary_review': ['start_candidate_correct','end_candidate_correct','ad_start_boundary_quality','ad_end_boundary_quality','start_boundary_review_note','end_boundary_review_note','overall_boundary_review_status'],
    'video_summary_review': ['candidate_density_ok','too_many_false_candidates','too_few_candidates','video_review_priority','video_review_note'],
}
DEPRECATED_COLUMNS = ['actual_start_transition_visible','actual_end_transition_visible']
RESNET_SPECIFIC = ['resnet_candidate_usefulness','resnet_helped','should_add_to_combined_scene_evidence','failure_reason_for_opencv','resnet_boundary_better_than_opencv','preferred_method_for_boundary','combination_strategy','comparison_review_note','resnet_candidate_density_ok','resnet_too_many_false_candidates','resnet_too_few_candidates','resnet_video_review_priority']
TARGET_MAPPING = {
    'resnet_candidate_review': 'scene_candidate_review',
    'resnet_only_boundary_review': 'scene_candidate_review',
    'resnet_boundary_review': 'ad_boundary_review',
    'video_summary_review': 'video_summary_review',
    'opencv_resnet_comparison': '__comparison__',
}
COMPARISON_REQUIRED = ['preferred_method_for_boundary','combination_strategy','comparison_review_note']
RESNET_ONLY_SPECIFIC_REQUIRED = ['resnet_helped','should_add_to_combined_scene_evidence','failure_reason_for_opencv']
COMPANION_BY_SHEET = {'resnet_candidate_review': CANDIDATE_CSV, 'resnet_boundary_review': BOUNDARY_CSV, 'resnet_only_boundary_review': RESNET_ONLY_CSV, 'opencv_resnet_comparison': COMPARISON_CSV, 'video_summary_review': VIDEO_SUMMARY_CSV}
REQUIRED_OPTIONS = {
    'review_status': ['not_reviewed','reviewed'],
    'yes_no_unclear': ['yes','no','unclear'],
    'scene_change_strength': ['strong','medium','weak','unclear'],
    'scene_change_type': ['hard_cut','fade','fade_or_dissolve','broll','broll_transition','camera_angle_change','semantic_change','object_change','lighting_change','text_overlay_only','motion_only','object_motion','no_visible_transition','other','unclear'],
    'is_ad_boundary_related': ['ad_start','ad_end','inside_ad','near_ad_but_not_boundary','no','unclear'],
    'false_positive_type': ['not_false_positive','normal_cut','camera_motion','subtitle_change','lighting_change','object_motion','semantic_but_not_boundary','broll_not_ad_boundary','duplicate_candidate','score_noise','unclear'],
    'keep_as_boundary_candidate': ['yes','no','unclear'],
    'boundary_quality': ['clear','weak','none','unclear'],
    'video_review_priority': ['high','medium','low'],
    'resnet_helped': ['yes','no','unclear'],
    'should_add_to_combined_scene_evidence': ['yes','no','unclear'],
    'failure_reason_for_opencv': ['subtle_semantic_change','low_pixel_change','threshold_missed','no_actual_boundary','unclear'],
    'resnet_boundary_better_than_opencv': ['yes','no','similar','unclear'],
    'resnet_candidate_usefulness': ['useful','redundant_with_opencv','false_positive','unclear'],
    'combination_strategy': ['opencv_only','resnet_only','opencv_or_resnet','opencv_and_resnet','opencv_primary_resnet_fallback','unclear'],
}
DROPDOWN = {
    'review_status':'review_status','is_true_scene_change':'yes_no_unclear','scene_change_strength':'scene_change_strength','scene_change_type':'scene_change_type','is_ad_boundary_related':'is_ad_boundary_related','false_positive_type':'false_positive_type','keep_as_boundary_candidate':'keep_as_boundary_candidate',
    'start_candidate_correct':'yes_no_unclear','end_candidate_correct':'yes_no_unclear','ad_start_boundary_quality':'boundary_quality','ad_end_boundary_quality':'boundary_quality','overall_boundary_review_status':'review_status',
    'candidate_density_ok':'yes_no_unclear','too_many_false_candidates':'yes_no_unclear','too_few_candidates':'yes_no_unclear','video_review_priority':'video_review_priority',
    'resnet_candidate_density_ok':'yes_no_unclear','resnet_too_many_false_candidates':'yes_no_unclear','resnet_too_few_candidates':'yes_no_unclear','resnet_video_review_priority':'video_review_priority',
    'resnet_candidate_usefulness':'resnet_candidate_usefulness','resnet_helped':'resnet_helped','should_add_to_combined_scene_evidence':'should_add_to_combined_scene_evidence','failure_reason_for_opencv':'failure_reason_for_opencv','resnet_boundary_better_than_opencv':'resnet_boundary_better_than_opencv','preferred_method_for_boundary':'combination_strategy','combination_strategy':'combination_strategy'
}
RUN_LOG=[]

def now_iso(): return datetime.now().astimezone().isoformat(timespec='seconds')
def log(msg):
    RUN_LOG.append(f'[{now_iso()}] {msg}')
    print(msg, flush=True)
def clean(v):
    if v is None: return ''
    try:
        if pd.isna(v): return ''
    except Exception: pass
    if isinstance(v,float) and v.is_integer(): return str(int(v))
    return str(v).strip()
def readable(sec):
    total=int(round(sec)); m,s=divmod(total,60); return f'{m}분 {s}초'
def headers(ws): return [clean(c.value) for c in ws[1]] if ws.max_row>=1 else []
def hidx(ws,name):
    for i,h in enumerate(headers(ws),1):
        if h==name: return i
    return None
def row_count(ws): return max(ws.max_row-1,0)
def file_hash(path):
    if not path.exists(): return ''
    h=hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda:f.read(1024*1024), b''): h.update(chunk)
    return h.hexdigest()
def old_snapshot():
    if not OLD_PROJECT_ROOT.exists(): return {'exists':False,'file_count':0,'digest':''}
    h=hashlib.sha256(); n=0
    for p in sorted(x for x in OLD_PROJECT_ROOT.rglob('*') if x.is_file()):
        try: st=p.stat()
        except OSError: continue
        h.update(f'{p.relative_to(OLD_PROJECT_ROOT).as_posix()}\t{st.st_size}\t{st.st_mtime_ns}\n'.encode('utf-8','replace')); n+=1
    return {'exists':True,'file_count':n,'digest':h.hexdigest()}
def ensure_dirs():
    for p in [BACKUP_DIR, REPORT_PATH.parent, LOG_PATH.parent, LATEST_DIR, PROJECT_ROOT/'scripts/review']:
        p.mkdir(parents=True, exist_ok=True)
def verify_cv(errors):
    exe=sys.executable
    cmd=['conda','run','-n','cv','python',str(CHECK_ENV_SCRIPT)] if CHECK_ENV_SCRIPT.exists() else ['conda','run','-n','cv','python','-c',"import sys; print(sys.executable); import pandas as pd; import openpyxl; print('pandas', pd.__version__); print('openpyxl', openpyxl.__version__)"]
    r=subprocess.run(cmd,cwd=PROJECT_ROOT,capture_output=True,text=True,check=False)
    log('cv environment command: '+' '.join(cmd)); log('cv environment stdout: '+r.stdout.strip().replace('\n',' | '))
    if r.stderr.strip(): log('cv environment stderr: '+r.stderr.strip().replace('\n',' | '))
    if r.returncode!=0: errors.append({'cv_environment_check_failed':r.returncode,'stderr':r.stderr.strip()}); return False, exe
    if '/envs/cv/' not in exe: errors.append({'current_python_executable_not_in_cv':exe}); return False, exe
    return True, exe
def ws_values(ws):
    cols=headers(ws); out=[]
    for row in ws.iter_rows(min_row=2, values_only=True):
        out.append({cols[i]: (row[i] if i<len(row) else '') for i in range(len(cols))})
    return out
def compare_values(before, after, before_cols):
    if len(before)!=len(after): return False
    for i,b in enumerate(before):
        for c in before_cols:
            if clean(b.get(c))!=clean(after[i].get(c)): return False
    return True
def write_dict_csv(path, rows, cols):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w',encoding='utf-8-sig',newline='') as f:
        w=csv.DictWriter(f,fieldnames=cols,extrasaction='ignore'); w.writeheader()
        for r in rows: w.writerow({c:r.get(c,'') for c in cols})
def source_schema(source_wb, warnings):
    rows=[]; review_by={}; deprecated_by={}
    for sheet in SOURCE_SHEETS:
        if sheet not in source_wb.sheetnames:
            warnings.append({'missing_source_sheet':sheet}); review_by[sheet]=[]; deprecated_by[sheet]=[]; continue
        hs=headers(source_wb[sheet]); expected=EXPECTED_SOURCE_REVIEW.get(sheet,[])
        review=[c for c in expected if c in hs]; deprecated=[c for c in DEPRECATED_COLUMNS if c in hs]
        review_by[sheet]=review; deprecated_by[sheet]=deprecated
        missing=[c for c in expected if c not in hs]
        if missing: warnings.append({'source_missing_expected_review_columns':{sheet:missing}})
        for idx,c in enumerate(hs,1):
            dep=c in DEPRECATED_COLUMNS; exp=c in expected; detected=exp and not dep
            rows.append({'source_sheet':sheet,'column_name':c,'column_index':idx,'detected_as_review_column':str(detected).lower(),'expected_review_column':str(exp).lower(),'deprecated_column':str(dep).lower(),'note':'deprecated_not_synced' if dep else ('source_review_column' if detected else '')})
    return rows, review_by, deprecated_by
def read_options(ws):
    out={}
    for ci,h in enumerate(headers(ws),1):
        if not h: continue
        vals=[]
        for ri in range(2, ws.max_row+1):
            v=clean(ws.cell(ri,ci).value)
            if v and v not in vals: vals.append(v)
        out[h]=vals
    return out
def merge_options(source_wb, target_wb):
    merged={}
    for wb in [target_wb, source_wb]:
        if 'value_options' not in wb.sheetnames: continue
        for k,vals in read_options(wb['value_options']).items():
            merged.setdefault(k,[])
            for v in vals:
                if v not in merged[k]: merged[k].append(v)
    for k,vals in REQUIRED_OPTIONS.items():
        merged.setdefault(k,[])
        for v in vals:
            if v not in merged[k]: merged[k].append(v)
    if 'value_options' in target_wb.sheetnames: target_wb.remove(target_wb['value_options'])
    ws=target_wb.create_sheet('value_options')
    for ci,k in enumerate(merged,1):
        ws.cell(1,ci).value=k
        for ri,v in enumerate(merged[k],2): ws.cell(ri,ci).value=v
    return merged
def add_col(ws,c):
    if c in headers(ws): return False
    ci=ws.max_column+1; ws.cell(1,ci).value=c
    if c in {'review_status','overall_boundary_review_status'}:
        for ri in range(2, ws.max_row+1): ws.cell(ri,ci).value='not_reviewed'
    return True
def patch_sheet(ws, source_cols, extra_cols):
    added=[]
    for c in source_cols+extra_cols:
        if c in DEPRECATED_COLUMNS: continue
        if add_col(ws,c): added.append(c)
    return added
def append_guide(wb):
    if 'review_guide' not in wb.sheetnames:
        ws=wb.create_sheet('review_guide'); ws.cell(1,1).value='section'; ws.cell(1,2).value='guide'
    ws=wb['review_guide']; marker='ResNet Review Columns Synced from OpenCV Review v2.4'
    existing='\n'.join(clean(cell.value) for row in ws.iter_rows() for cell in row)
    if marker in existing: return
    items=[(marker,'ResNet review workbook은 OpenCV/ffmpeg review workbook의 공통 사용자 review 컬럼 구조를 따른다.'),('resnet_only_boundary_review','resnet_only_boundary_review에서도 scene-change 유형(scene_change_type)과 false positive 유형(false_positive_type)을 기록할 수 있다.'),('deprecated','actual_start_transition_visible, actual_end_transition_visible은 사용하지 않는다.'),('핵심 검토 순서','1. resnet_only_boundary_review 2. resnet_boundary_review 3. resnet_candidate_review의 very_high/high row 4. A022 관련 row'),('최종 rule 판단','keep_as_boundary_candidate와 should_add_to_combined_scene_evidence를 함께 본다.')]
    start=ws.max_row+1
    for off,(s,g) in enumerate(items): ws.cell(start+off,1).value=s; ws.cell(start+off,2).value=g
def opt_formula(options,key): return '"'+','.join(options.get(key,[]))+'"'
def add_validation(ws,c,key,options):
    idx=hidx(ws,c)
    if idx is None or ws.max_row<2: return False
    dv=DataValidation(type='list', formula1=opt_formula(options,key), allow_blank=True); ws.add_data_validation(dv)
    col=get_column_letter(idx); dv.add(f'{col}2:{col}{max(ws.max_row,500)}'); return True
def add_row_fill(ws,c,val,color):
    idx=hidx(ws,c)
    if idx is None or ws.max_row<2: return False
    col=get_column_letter(idx); fill=PatternFill('solid',fgColor=color)
    ws.conditional_formatting.add(f'A2:{get_column_letter(ws.max_column)}{ws.max_row}', FormulaRule(formula=[f'=${col}2="{val}"'], fill=fill)); return True
def add_cell_fill(ws,c,val,color):
    idx=hidx(ws,c)
    if idx is None or ws.max_row<2: return False
    col=get_column_letter(idx); fill=PatternFill('solid',fgColor=color)
    ws.conditional_formatting.add(f'{col}2:{col}{ws.max_row}', FormulaRule(formula=[f'=${col}2="{val}"'], fill=fill)); return True
def format_validate(wb, source_derived, options):
    applied=defaultdict(list); source_fill=PatternFill('solid',fgColor='F4B183'); resnet_fill=PatternFill('solid',fgColor='B4C6E7'); analysis_fill=PatternFill('solid',fgColor='1F4E78'); white=Font(color='FFFFFF',bold=True); black=Font(color='000000',bold=True)
    wrap={'video_title','video_filename','video_path','review_note','reviewer','reviewed_at','start_boundary_review_note','end_boundary_review_note','comparison_review_note','video_review_note','guide','notes'}
    for ws in wb.worksheets:
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
        if ws.title in TARGET_MAPPING or ws.title in {'value_options','review_guide'}:
            try: ws.data_validations.dataValidation=[]
            except Exception: pass
        for cell in ws[1]:
            name=clean(cell.value)
            if name in source_derived: cell.fill=source_fill; cell.font=black
            elif name in RESNET_SPECIFIC or name in COMPARISON_REQUIRED: cell.fill=resnet_fill; cell.font=black
            else: cell.fill=analysis_fill; cell.font=white
            cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        for ci,cells in enumerate(ws.columns,1):
            name=clean(ws.cell(1,ci).value); max_len=len(name)
            for cell in list(cells)[1:min(ws.max_row,120)]:
                max_len=max(max_len,min(len(clean(cell.value)),60))
                if name in wrap: cell.alignment=Alignment(wrap_text=True,vertical='top')
            width=min(max(max_len+2,10),55)
            if name in wrap: width=min(max(width,24),70)
            ws.column_dimensions[get_column_letter(ci)].width=width
        for c,k in DROPDOWN.items():
            if c in headers(ws) and add_validation(ws,c,k,options): applied[ws.title].append(c)
        add_row_fill(ws,'review_status','not_reviewed','E7E6E6'); add_row_fill(ws,'overall_boundary_review_status','not_reviewed','E7E6E6')
        add_cell_fill(ws,'keep_as_boundary_candidate','yes','E2F0D9'); add_cell_fill(ws,'keep_as_boundary_candidate','no','FCE4D6')
        add_cell_fill(ws,'should_add_to_combined_scene_evidence','yes','E2F0D9'); add_cell_fill(ws,'should_add_to_combined_scene_evidence','no','FCE4D6')
        add_row_fill(ws,'is_new_interval_v2_4','true','F4CCCC'); add_row_fill(ws,'ad_interval_id','A022','F4CCCC'); add_row_fill(ws,'review_priority','very_high','F8CBAD'); add_row_fill(ws,'review_priority','high','FCE4D6')
    return True, True, dict(applied)
def write_sheet_csv(ws,path):
    with path.open('w',encoding='utf-8-sig',newline='') as f:
        rows=list(ws.iter_rows(values_only=True)); w=csv.writer(f)
        if not rows: return
        cols=[clean(v) for v in rows[0]]; w.writerow(cols)
        for row in rows[1:]: w.writerow(['' if v is None else v for v in row[:len(cols)]])
def target_schema(wb, source_derived, dropdown):
    rows=[]; review=set(source_derived)|set(RESNET_SPECIFIC)|set(COMPARISON_REQUIRED)
    for sheet in TARGET_SHEETS:
        if sheet not in wb.sheetnames: continue
        for idx,c in enumerate(headers(wb[sheet]),1):
            rows.append({'sheet_name':sheet,'column_name':c,'column_index':idx,'is_review_column':str(c in review).lower(),'source_derived':str(c in source_derived).lower(),'resnet_specific':str(c in RESNET_SPECIFIC or c in COMPARISON_REQUIRED).lower(),'deprecated':str(c in DEPRECATED_COLUMNS).lower(),'dropdown_applied':str(c in dropdown.get(sheet,[])).lower(),'note':'deprecated_existing_column' if c in DEPRECATED_COLUMNS else ''})
    return rows
def latest_forbidden():
    bad=[]; forbidden_suffix={'.mp4','.mov','.avi','.mkv','.jpg','.jpeg','.png','.webp','.pt','.pth','.ckpt','.bin'}; raw={'new_ad_labeling.xlsx','clean_ad_labels_v0_review.xlsx'}
    if not LATEST_DIR.exists(): return bad
    for p in LATEST_DIR.rglob('*'):
        if not p.is_file(): continue
        name=p.name.lower()
        if p.suffix.lower() in forbidden_suffix or p.name in raw or any(x in name for x in ['frame','model','checkpoint','cache']): bad.append(str(p))
    return sorted(set(bad))
def copy_latest(files):
    LATEST_DIR.mkdir(parents=True,exist_ok=True); copied=[]
    for src in files:
        if not src.exists(): continue
        if src.suffix.lower() in {'.mp4','.mov','.avi','.mkv','.jpg','.jpeg','.png','.webp','.pt','.pth','.ckpt','.bin'} or src.name in {SOURCE_WORKBOOK.name,'new_ad_labeling.xlsx','clean_ad_labels_v0_review.xlsx'}: raise RuntimeError(f'Refusing latest copy: {src}')
        dst=LATEST_DIR/src.name; shutil.copy2(src,dst); copied.append(str(dst))
    LATEST_README.write_text('# latest_for_chatgpt files\n\n이번 작업명: sync_resnet_review_columns_from_opencv_review_v2_4\n\nResNet v2.4 review workbook, companion CSV, audit, report/summary/log만 복사했다. mp4, raw xlsx, frame, model, checkpoint, cache 파일은 복사하지 않는다.\n',encoding='utf-8')
    copied.append(str(LATEST_README)); bad=latest_forbidden(); return len(bad)==0,copied,bad
def update_readme():
    title='## ResNet Review Columns Synced from OpenCV Review v2.4'
    section=f"""{title}\n\nOpenCV/ffmpeg review workbook의 사용자 review 컬럼 구조를 기준으로 ResNet review workbook을 보강했다.\n\n특히 `resnet_only_boundary_review` sheet에 scene-change 유형, false positive 유형, keep 여부 등 공통 review 컬럼을 추가했으며, 기존 ResNet-specific 컬럼도 보존했다.\n\n기존 검토값은 덮어쓰지 않았고, 이후 ResNet 후보 검토는 `data/review/resnet_scene_candidate_human_review_v2_4.xlsx` 기준으로 진행한다.\n"""
    text=README_PATH.read_text(encoding='utf-8') if README_PATH.exists() else '# youtube_ad_segment_detector\n'
    if title in text:
        before,_,after=text.partition(title); idx=after.find('\n## '); text=before.rstrip()+'\n\n'+section.rstrip()+('\n'+after[idx+1:].lstrip() if idx>=0 else '')
    else: text=text.rstrip()+'\n\n'+section.rstrip()
    README_PATH.write_text(text.rstrip()+'\n',encoding='utf-8')
def make_summary(report):
    added='\n'.join(f"- `{s}`: {', '.join(cols) if cols else '추가 없음'}" for s,cols in report.get('added_columns_by_target_sheet',{}).items())
    return f"""# sync_resnet_review_columns_from_opencv_review_v2_4 summary\n\n- Source workbook: `{report.get('source_workbook_path')}`\n- Target workbook: `{report.get('target_workbook_path')}`\n- Backup: `{report.get('backup_path')}`\n\n## Sheet별 추가 컬럼\n\n{added}\n\n## resnet_only_boundary_review\n\n- `scene_change_type` 존재: {report.get('resnet_only_scene_change_type_exists')}\n- `false_positive_type` 존재: {report.get('resnet_only_false_positive_type_exists')}\n- 기존 review 값 보존: {report.get('existing_review_values_preserved')}\n\n## 사용자가 입력해야 할 핵심 컬럼\n\n- `is_true_scene_change`\n- `scene_change_strength`\n- `scene_change_type`\n- `is_ad_boundary_related`\n- `false_positive_type`\n- `keep_as_boundary_candidate`\n- `resnet_helped`\n- `should_add_to_combined_scene_evidence`\n- `failure_reason_for_opencv`\n- `review_note`\n\n## Sub Agent Results\n\n```json\n{json.dumps(report.get('sub_agent_results',{}),ensure_ascii=False,indent=2)}\n```\n"""
def write_report(report):
    REPORT_PATH.write_text(json.dumps(report,ensure_ascii=False,indent=2)+'\n',encoding='utf-8'); SUMMARY_PATH.write_text(make_summary(report),encoding='utf-8'); LOG_PATH.write_text('\n'.join(RUN_LOG)+'\n',encoding='utf-8')
def sub_agents(report, source_hash_before, source_hash_after, old_before, old_after, latest_bad, value_options_ok, preserved):
    target_missing={r['target_sheet']:r['missing_after_patch'] for r in report.get('sync_audit_rows',[]) if r.get('missing_after_patch')}
    rows_ok=all(v[0]==v[1] for v in report.get('row_counts_before_after_by_sheet',{}).values())
    source_cols=report.get('source_review_columns_by_sheet',{})
    return {
      'sub_agent_1_source_column': {'status':'PASS' if source_cols else 'FAIL','checks':{'source_sheets_read':report.get('source_sheets_read'),'source_review_columns_by_sheet':source_cols,'source_deprecated_columns':report.get('deprecated_source_columns',{})},'warnings':report.get('deprecated_source_columns',{})},
      'sub_agent_2_target_patch': {'status':'PASS' if not target_missing and rows_ok and preserved and report.get('resnet_only_scene_change_type_exists') and report.get('resnet_only_false_positive_type_exists') else 'FAIL','checks':{'missing_after_patch':target_missing,'row_counts_preserved':rows_ok,'existing_review_values_preserved':preserved,'resnet_only_scene_change_type_exists':report.get('resnet_only_scene_change_type_exists'),'resnet_only_false_positive_type_exists':report.get('resnet_only_false_positive_type_exists')},'warnings':[]},
      'sub_agent_3_validation_formatting': {'status':'PASS' if report.get('dropdown_validation_applied') and report.get('conditional_formatting_applied') and value_options_ok else 'FAIL','checks':{'dropdown_validation_applied':report.get('dropdown_validation_applied'),'conditional_formatting_applied':report.get('conditional_formatting_applied'),'value_options_contains_required_options':value_options_ok},'warnings':[]},
      'sub_agent_4_safety_output': {'status':'PASS' if report.get('backup_created') and source_hash_before==source_hash_after and not latest_bad and old_before==old_after else 'FAIL','checks':{'backup_created':report.get('backup_created'),'source_workbook_unchanged':source_hash_before==source_hash_after,'latest_for_chatgpt_forbidden_files_absent':not latest_bad,'old_project_unmodified':old_before==old_after},'warnings':latest_bad}
    }
def main():
    ensure_dirs(); start=time.monotonic(); start_time=now_iso(); warnings=[]; errors=[]
    log(f'작업 시작 전 예상 작업 시간: {ESTIMATED_RUNTIME}'); log(f'예상 근거: {RUNTIME_ESTIMATION_REASON}'); log(f'작업 시작 시각: {start_time}')
    old_before=old_snapshot(); source_hash_before=file_hash(SOURCE_WORKBOOK)
    report={'project_root':str(PROJECT_ROOT),'estimated_runtime':ESTIMATED_RUNTIME,'runtime_estimation_reason':RUNTIME_ESTIMATION_REASON,'start_time':start_time,'end_time':'','actual_runtime_seconds':0,'actual_runtime_readable':'','source_workbook_path':str(SOURCE_WORKBOOK),'target_workbook_path':str(TARGET_WORKBOOK),'backup_created':False,'backup_path':'','source_sheets_read':[],'target_sheets_patched':[],'source_review_columns_by_sheet':{},'added_columns_by_target_sheet':{},'deprecated_existing_columns':{},'row_counts_before_after_by_sheet':{},'dropdown_validation_applied':False,'conditional_formatting_applied':False,'companion_csv_updated':False,'audit_paths':[str(SYNC_AUDIT_CSV),str(SOURCE_SCHEMA_CSV),str(TARGET_SCHEMA_CSV)],'sub_agent_results':{},'old_project_modified':False,'warnings':warnings,'errors':errors}
    try:
        log('[STEP 1/9] cv 환경 확인'); ok,exe=verify_cv(errors); report['cv_environment_checked']=ok; report['python_executable']=exe
        if not ok: raise RuntimeError('cv environment check failed')
        if not SOURCE_WORKBOOK.exists(): errors.append({'missing_source_workbook':str(SOURCE_WORKBOOK)}); raise RuntimeError('source workbook missing')
        if not TARGET_WORKBOOK.exists(): errors.append({'missing_target_workbook':str(TARGET_WORKBOOK)}); raise RuntimeError('target workbook missing')
        log('[STEP 2/9] source workbook schema 추출'); source_wb=load_workbook(SOURCE_WORKBOOK); source_rows,source_review,source_dep=source_schema(source_wb,warnings)
        report['source_sheets_read']=[s for s in SOURCE_SHEETS if s in source_wb.sheetnames]; report['source_review_columns_by_sheet']=source_review; report['deprecated_source_columns']={k:v for k,v in source_dep.items() if v}
        write_dict_csv(SOURCE_SCHEMA_CSV,source_rows,['source_sheet','column_name','column_index','detected_as_review_column','expected_review_column','deprecated_column','note'])
        log('[STEP 3/9] target workbook backup 생성'); stamp=datetime.now().astimezone().strftime('%Y%m%d_%H%M%S'); backup=BACKUP_DIR/f'resnet_scene_candidate_human_review_v2_4_backup_before_sync_from_opencv_review_{stamp}.xlsx'; shutil.copy2(TARGET_WORKBOOK,backup); report['backup_created']=True; report['backup_path']=str(backup); log(f'Backup created: {backup}')
        log('[STEP 4/9] target sheet별 누락 review 컬럼 추가'); wb=load_workbook(TARGET_WORKBOOK); missing=[s for s in TARGET_SHEETS if s not in wb.sheetnames]
        if missing: errors.append({'missing_target_sheets':missing}); raise RuntimeError('missing target sheets')
        before_headers={}; before_rows={}; added_by={}; dep_existing={}; audit=[]; source_derived=set()
        for target,source in TARGET_MAPPING.items():
            ws=wb[target]; before_headers[target]=headers(ws)[:]; before_rows[target]=ws_values(ws); rb=row_count(ws); cb=ws.max_column
            if source=='__comparison__': source_cols=[]; extras=COMPARISON_REQUIRED
            else:
                source_cols=source_review.get(source,[]); extras=[]; source_derived.update(source_cols)
                if target=='resnet_only_boundary_review': extras=RESNET_ONLY_SPECIFIC_REQUIRED
            added=patch_sheet(ws,source_cols,extras); added_by[target]=added; dep=[c for c in DEPRECATED_COLUMNS if c in headers(ws)]; dep_existing[target]=dep
            missing_after=[c for c in (source_cols if source!='__comparison__' else COMPARISON_REQUIRED) if c not in headers(ws)]
            if target=='resnet_only_boundary_review': missing_after += [c for c in RESNET_ONLY_SPECIFIC_REQUIRED if c not in headers(ws)]
            ra=row_count(ws); ca=ws.max_column; report['row_counts_before_after_by_sheet'][target]=[rb,ra]
            audit.append({'target_sheet':target,'source_sheet':source,'row_count_before':rb,'row_count_after':ra,'column_count_before':cb,'column_count_after':ca,'source_review_columns':';'.join(source_cols),'target_existing_review_columns_before':';'.join([c for c in before_headers[target] if c in set(source_cols)|set(RESNET_SPECIFIC)|set(COMPARISON_REQUIRED)]),'added_columns':';'.join(added),'preserved_existing_columns':';'.join(before_headers[target]),'deprecated_existing_columns':';'.join(dep),'missing_after_patch':';'.join(missing_after),'dropdown_applied_columns':'','formatting_applied':'','patch_status':'PASS' if rb==ra and not missing_after else 'FAIL','patch_warning':'deprecated_existing_column:'+ ';'.join(dep) if dep else ''})
        report['target_sheets_patched']=list(TARGET_MAPPING); report['added_columns_by_target_sheet']=added_by; report['deprecated_existing_columns']=dep_existing
        log('[STEP 5/9] value_options 병합 및 review_guide 갱신'); options=merge_options(source_wb,wb); append_guide(wb)
        log('[STEP 6/9] dropdown validation 및 formatting 적용'); dd_ok,cf_ok,dd=format_validate(wb,source_derived,options); report['dropdown_validation_applied']=dd_ok; report['conditional_formatting_applied']=cf_ok
        for r in audit: r['dropdown_applied_columns']=';'.join(dd.get(r['target_sheet'],[])); r['formatting_applied']='true'
        log('[STEP 7/9] workbook 저장 및 companion CSV 갱신'); wb.save(TARGET_WORKBOOK); wb=load_workbook(TARGET_WORKBOOK)
        for sheet,path in COMPANION_BY_SHEET.items(): write_sheet_csv(wb[sheet],path)
        report['companion_csv_updated']=True
        preserved=True
        for sheet,rows in before_rows.items():
            if not compare_values(rows,ws_values(wb[sheet]),before_headers[sheet]): preserved=False; warnings.append({'existing_values_changed':sheet})
        report['existing_review_values_preserved']=preserved; report['resnet_only_scene_change_type_exists']='scene_change_type' in headers(wb['resnet_only_boundary_review']); report['resnet_only_false_positive_type_exists']='false_positive_type' in headers(wb['resnet_only_boundary_review'])
        log('[STEP 8/9] audit/report/README/latest_for_chatgpt 갱신'); schema=target_schema(wb,source_derived,dd)
        write_dict_csv(SYNC_AUDIT_CSV,audit,['target_sheet','source_sheet','row_count_before','row_count_after','column_count_before','column_count_after','source_review_columns','target_existing_review_columns_before','added_columns','preserved_existing_columns','deprecated_existing_columns','missing_after_patch','dropdown_applied_columns','formatting_applied','patch_status','patch_warning'])
        write_dict_csv(TARGET_SCHEMA_CSV,schema,['sheet_name','column_name','column_index','is_review_column','source_derived','resnet_specific','deprecated','dropdown_applied','note'])
        update_readme(); value_options_ok=all(k in options and all(v in options[k] for v in vals) for k,vals in REQUIRED_OPTIONS.items()); old_after=old_snapshot(); source_hash_after=file_hash(SOURCE_WORKBOOK); old_modified=old_before!=old_after; report['old_project_modified']=old_modified
        if old_modified: errors.append('old_project_modified_unexpectedly')
        if source_hash_before!=source_hash_after: errors.append('source_workbook_modified_unexpectedly')
        elapsed=time.monotonic()-start; end_time=now_iso(); report.update({'end_time':end_time,'actual_runtime_seconds':round(elapsed,3),'actual_runtime_readable':readable(elapsed),'sync_audit_rows':audit,'warnings':warnings,'errors':errors}); write_report(report)
        files=[TARGET_WORKBOOK,CANDIDATE_CSV,BOUNDARY_CSV,RESNET_ONLY_CSV,COMPARISON_CSV,VIDEO_SUMMARY_CSV,SYNC_AUDIT_CSV,SOURCE_SCHEMA_CSV,TARGET_SCHEMA_CSV,REPORT_PATH,SUMMARY_PATH,LOG_PATH]
        latest_ok,copied,bad=copy_latest(files); report['latest_for_chatgpt_updated']=latest_ok; report['latest_for_chatgpt_files']=copied; report['latest_for_chatgpt_forbidden_files']=bad; report['sub_agent_results']=sub_agents(report,source_hash_before,source_hash_after,old_before,old_after,bad,value_options_ok,preserved); write_report(report)
        for p in [REPORT_PATH,SUMMARY_PATH,LOG_PATH]: shutil.copy2(p,LATEST_DIR/p.name)
        log(f'작업 종료 시각: {end_time}'); log(f'실제 작업 시간: {readable(elapsed)}'); write_report(report); shutil.copy2(LOG_PATH,LATEST_DIR/LOG_PATH.name); log('[STEP 9/9] 완료')
        print(json.dumps({'estimated_runtime':ESTIMATED_RUNTIME,'actual_runtime_readable':report['actual_runtime_readable'],'backup_created':report['backup_created'],'added_columns_by_target_sheet':added_by,'resnet_only_scene_change_type_exists':report['resnet_only_scene_change_type_exists'],'resnet_only_false_positive_type_exists':report['resnet_only_false_positive_type_exists'],'errors':errors},ensure_ascii=False,indent=2))
        return 0 if not errors else 1
    except Exception as exc:
        errors.append({'fatal_error':str(exc)}); elapsed=time.monotonic()-start; end_time=now_iso(); report.update({'end_time':end_time,'actual_runtime_seconds':round(elapsed,3),'actual_runtime_readable':readable(elapsed),'warnings':warnings,'errors':errors}); log(f'작업 종료 시각: {end_time}'); log(f'실제 작업 시간: {readable(elapsed)}'); write_report(report); raise
if __name__=='__main__':
    raise SystemExit(main())
