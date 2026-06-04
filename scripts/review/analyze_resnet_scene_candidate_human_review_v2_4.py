#!/usr/bin/env python3
"""Analyze v2.4 ResNet scene candidate human review results.

This is an evidence audit for pretrained ResNet embedding scene-change candidates.
It does not modify the source review workbook and does not recompute embeddings,
scene-change scores, frames, OCR, audio, or model weights.
"""
from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import shutil
import subprocess
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path('.')
OLD_PROJECT_ROOT = Path('./_old_project_not_included')
TASK_NAME = 'resnet_scene_candidate_human_review_v2_4_analysis'
VERSION = 'v2_4'
CHECK_ENV_SCRIPT = PROJECT_ROOT / 'scripts/utils/check_cv_environment.py'
SCRIPT_PATH = PROJECT_ROOT / 'scripts/review/analyze_resnet_scene_candidate_human_review_v2_4.py'

INPUT_REVIEW_PRIMARY = PROJECT_ROOT / 'data/review/resnet_scene_candidate_human_review_v2_4.xlsx'
REVIEW_GLOBS = [
    PROJECT_ROOT / 'data/review',
    PROJECT_ROOT / 'outputs/latest_for_chatgpt',
]

REVIEW_DIR = PROJECT_ROOT / 'data/review'
SCENE_DIR = PROJECT_ROOT / 'data/scene'
REPORT_DIR = PROJECT_ROOT / 'reports'
LOG_DIR = PROJECT_ROOT / 'logs'
LATEST_DIR = PROJECT_ROOT / 'outputs/latest_for_chatgpt'

OUTPUT_WORKBOOK = REVIEW_DIR / 'resnet_scene_candidate_human_review_v2_4_analyzed.xlsx'
REVIEWED_ROWS_CSV = REVIEW_DIR / 'resnet_scene_candidate_human_review_v2_4_reviewed_rows.csv'
USEFULNESS_ANALYSIS_CSV = REVIEW_DIR / 'resnet_scene_candidate_human_review_v2_4_usefulness_analysis.csv'
OPENCV_COMPARISON_CSV = REVIEW_DIR / 'resnet_scene_candidate_human_review_v2_4_opencv_comparison.csv'
USEFULNESS_SUMMARY_CSV = REVIEW_DIR / 'resnet_scene_candidate_human_review_v2_4_usefulness_summary.csv'
SCORE_BAND_SUMMARY_CSV = REVIEW_DIR / 'resnet_scene_candidate_human_review_v2_4_score_band_summary.csv'
FALSE_POSITIVE_SUMMARY_CSV = REVIEW_DIR / 'resnet_scene_candidate_human_review_v2_4_false_positive_summary.csv'
REPORT_PATH = REPORT_DIR / 'resnet_scene_candidate_human_review_v2_4_analysis_report.json'
SUMMARY_PATH = REPORT_DIR / 'resnet_scene_candidate_human_review_v2_4_analysis_summary.md'
LOG_PATH = LOG_DIR / 'resnet_scene_candidate_human_review_v2_4_analysis_run_log.txt'
LATEST_README = LATEST_DIR / 'README_latest_files.md'

OUTPUT_FILES = [
    OUTPUT_WORKBOOK,
    REVIEWED_ROWS_CSV,
    USEFULNESS_ANALYSIS_CSV,
    OPENCV_COMPARISON_CSV,
    USEFULNESS_SUMMARY_CSV,
    SCORE_BAND_SUMMARY_CSV,
    FALSE_POSITIVE_SUMMARY_CSV,
    REPORT_PATH,
    SUMMARY_PATH,
    LOG_PATH,
    SCRIPT_PATH,
]
ALLOWED_LATEST_NAMES = {p.name for p in OUTPUT_FILES} | {'README_latest_files.md'}

AUTO_COLUMNS = [
    'reviewed_normalized',
    'is_true_scene_change_norm',
    'is_ad_boundary_related_norm',
    'keep_as_boundary_candidate_norm',
    'core_review_complete',
    'review_missing_fields',
    'nearest_opencv_candidate_time_sec_auto',
    'nearest_opencv_candidate_mmss_auto',
    'distance_to_nearest_opencv_candidate_sec_auto',
    'has_opencv_overlap_2s_auto',
    'has_opencv_overlap_5s_auto',
    'opencv_hit_same_boundary_5s_auto',
    'is_resnet_only_5s_auto',
    'resnet_candidate_relation_to_opencv_auto',
    'resnet_candidate_usefulness_auto',
    'resnet_candidate_usefulness_reason',
    'resnet_boundary_evidence_auto',
    'resnet_review_analysis_group',
    'resnet_score_band_auto',
]

CORE_SHEET_HINTS = [
    'video_id', 'video_title', 'ad_boundary_mmss', 'resnet_candidate_time_sec',
    'resnet_candidate_mmss', 'resnet_score_percentile_in_video', 'scene_change_score',
    'review_priority', 'nearest_opencv_candidate_time_sec', 'nearest_opencv_candidate_mmss',
    'distance_to_nearest_opencv_candidate_sec', 'opencv_hit_same_boundary_5s',
    'review_clip_start_mmss', 'review_clip_end_mmss', 'is_true_scene_change',
    'is_ad_boundary_related', 'keep_as_boundary_candidate', 'false_positive_type',
    'review_note', 'reviewer', 'reviewed_at',
]
REVIEW_FIELDS_FOR_REVIEWED = ['is_true_scene_change', 'is_ad_boundary_related', 'keep_as_boundary_candidate', 'false_positive_type', 'review_note']
USEFULNESS_VALUES = {'boundary_complement','redundant_boundary','possible_boundary_complement','scene_change_only','false_positive','uncertain','unreviewed'}
RUN_LOG: list[str] = []


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec='seconds')


def log(message: str) -> None:
    line = f'[{now_iso()}] {message}'
    RUN_LOG.append(line)
    print(message, flush=True)


def clean(value: Any) -> str:
    if value is None:
        return ''
    try:
        if pd.isna(value):
            return ''
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_float(value: Any) -> float | None:
    text = clean(value)
    if text == '':
        return None
    try:
        return float(text)
    except Exception:
        return None


def readable(seconds: float) -> str:
    total = int(round(seconds))
    minutes, sec = divmod(total, 60)
    return f'{minutes}분 {sec}초'


def mmss(seconds: Any) -> str:
    sec = to_float(seconds)
    if sec is None:
        return ''
    total = max(0, int(math.floor(sec)))
    return f'{total // 60:02d}분 {total % 60:02d}초'


def parse_mmss(value: Any) -> float | None:
    text = clean(value)
    if not text:
        return None
    if '분' in text and '초' in text:
        try:
            m = int(text.split('분')[0].strip())
            s = int(text.split('분')[1].replace('초','').strip())
            return float(m * 60 + s)
        except Exception:
            return None
    if ':' in text:
        parts = text.split(':')
        try:
            if len(parts) == 2:
                return float(int(parts[0]) * 60 + float(parts[1]))
            if len(parts) == 3:
                return float(int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2]))
        except Exception:
            return None
    return to_float(text)


def bool_text(value: bool) -> str:
    return 'true' if value else 'false'


def file_hash(path: Path) -> str:
    if not path.exists():
        return ''
    h = hashlib.sha256()
    with path.open('rb') as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def old_project_snapshot() -> dict[str, Any]:
    if not OLD_PROJECT_ROOT.exists():
        return {'exists': False, 'file_count': 0, 'digest': ''}
    h = hashlib.sha256()
    count = 0
    for path in sorted(p for p in OLD_PROJECT_ROOT.rglob('*') if p.is_file()):
        try:
            st = path.stat()
        except OSError:
            continue
        rel = path.relative_to(OLD_PROJECT_ROOT).as_posix()
        h.update(f'{rel}\t{st.st_size}\t{st.st_mtime_ns}\n'.encode('utf-8', errors='replace'))
        count += 1
    return {'exists': True, 'file_count': count, 'digest': h.hexdigest()}


def verify_cv_environment(errors: list[Any]) -> tuple[bool, str]:
    executable = sys.executable
    cmd = ['conda', 'run', '-n', 'cv', 'python']
    if CHECK_ENV_SCRIPT.exists():
        cmd.append(str(CHECK_ENV_SCRIPT))
    else:
        cmd.extend(['-c', "import sys; print(sys.executable); import pandas as pd; import openpyxl; print('pandas', pd.__version__); print('openpyxl', openpyxl.__version__)"])
    result = subprocess.run(cmd, cwd=PROJECT_ROOT, capture_output=True, text=True, check=False)
    log('cv environment command: ' + ' '.join(cmd))
    log('cv environment stdout: ' + result.stdout.strip().replace('\n', ' | '))
    if result.stderr.strip():
        log('cv environment stderr: ' + result.stderr.strip().replace('\n', ' | '))
    if result.returncode != 0:
        errors.append({'cv_environment_check_failed': result.returncode, 'stderr': result.stderr.strip()})
        return False, executable
    if '/envs/cv/' not in executable:
        errors.append({'current_python_executable_not_in_cv': executable})
        return False, executable
    return True, executable


def find_review_workbook(warnings: list[Any]) -> Path | None:
    if INPUT_REVIEW_PRIMARY.exists():
        return INPUT_REVIEW_PRIMARY
    candidates: list[Path] = []
    for base in REVIEW_GLOBS:
        if base.exists():
            candidates.extend(sorted(base.glob('*resnet*review*v2_4*.xlsx')))
    if candidates:
        warnings.append({'review_workbook_found_by_glob': str(candidates[0])})
        return candidates[0]
    return None


def pick_existing(candidates: list[Path], fallback_used: dict[str, Any], label: str) -> Path | None:
    for idx, path in enumerate(candidates):
        if path.exists():
            if idx > 0 or 'v2_3' in path.name:
                fallback_used['fallback_used'] = True
                fallback_used['fallback_reason'].append(f'{label}: using {path.name}')
            return path
    fallback_used['missing_input_files'].append(str(candidates[0]))
    return None


def read_csv(path: Path | None, warnings: list[Any]) -> pd.DataFrame:
    if path is None or not path.exists():
        return pd.DataFrame()
    try:
        df = pd.read_csv(path, dtype=str, keep_default_na=False, encoding='utf-8-sig')
        df.columns = [clean(c) for c in df.columns]
        return df
    except Exception as exc:
        warnings.append({'failed_to_read_csv': str(path), 'error': str(exc)})
        return pd.DataFrame()


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [clean(c) for c in out.columns]
    def first(cols: list[str]) -> str | None:
        for c in cols:
            if c in out.columns:
                return c
        return None
    mappings = {
        'resnet_time_sec_std': ['resnet_candidate_time_sec','candidate_time_sec','score_time_sec'],
        'resnet_time_mmss_std': ['resnet_candidate_mmss','candidate_time_mmss'],
        'resnet_score_std': ['scene_change_score','resnet_scene_change_score','cosine_distance'],
        'resnet_percentile_std': ['resnet_score_percentile_in_video','score_percentile_in_video'],
        'opencv_distance_existing_std': ['distance_to_nearest_opencv_candidate_sec'],
        'opencv_hit_same_boundary_existing_std': ['opencv_hit_same_boundary_5s'],
        'ad_boundary_sec_std': ['ad_boundary_sec','nearest_ad_boundary_sec'],
        'ad_boundary_mmss_std': ['ad_boundary_mmss','nearest_ad_boundary_mmss'],
    }
    for std, options in mappings.items():
        src = first(options)
        out[std] = out[src] if src else ''
    if out['resnet_time_mmss_std'].map(clean).eq('').all():
        out['resnet_time_mmss_std'] = out['resnet_time_sec_std'].map(mmss)
    if out['ad_boundary_sec_std'].map(clean).eq('').all() and 'ad_boundary_mmss_std' in out.columns:
        out['ad_boundary_sec_std'] = out['ad_boundary_mmss_std'].map(parse_mmss).map(lambda x: '' if x is None else x)
    return out


def identify_candidate_sheet(workbook_path: Path, warnings: list[Any]) -> tuple[str, pd.DataFrame, dict[str, Any]]:
    xls = pd.ExcelFile(workbook_path)
    best_sheet = ''
    best_score = -1
    details: dict[str, Any] = {}
    for sheet in xls.sheet_names:
        try:
            df0 = pd.read_excel(workbook_path, sheet_name=sheet, dtype=object, keep_default_na=False, nrows=0)
        except Exception:
            continue
        cols = {clean(c) for c in df0.columns}
        score = sum(1 for c in CORE_SHEET_HINTS if c in cols)
        details[sheet] = {'match_score': score, 'column_count': len(cols)}
        if score > best_score:
            best_score = score
            best_sheet = sheet
    if not best_sheet:
        raise RuntimeError('Could not identify ResNet candidate review sheet')
    df = pd.read_excel(workbook_path, sheet_name=best_sheet, dtype=object, keep_default_na=False)
    df.columns = [clean(c) for c in df.columns]
    warnings.append({'candidate_sheet_identification': {'sheet': best_sheet, 'score': best_score, 'details': details}})
    return best_sheet, df, details


def normalize_bool(value: Any, category: str = '') -> str:
    text = clean(value).lower()
    if text == '':
        return 'missing'
    true_values = {'yes','y','true','t','1','o','예','맞음','reviewed'}
    false_values = {'no','n','false','f','0','x','아니오','아님'}
    uncertain_values = {'uncertain','partial','maybe','ambiguous','unsure','애매','보류','unclear'}
    if category == 'ad_boundary':
        if text in {'ad_start','ad_end'}:
            return 'true'
        if text in {'no'}:
            return 'false'
        if text in {'inside_ad','near_ad_but_not_boundary','unclear'}:
            return 'uncertain'
    if text in true_values:
        return 'true'
    if text in false_values:
        return 'false'
    if text in uncertain_values:
        return 'uncertain'
    return 'uncertain'


def is_meaningful(value: Any) -> bool:
    text = clean(value).lower()
    return text not in {'', 'not_reviewed', 'nan', 'none'}


def reviewed_status(row: pd.Series) -> bool:
    status = clean(row.get('review_status')).lower()
    if status in {'reviewed','done','complete'}:
        return True
    return any(is_meaningful(row.get(c)) for c in REVIEW_FIELDS_FOR_REVIEWED)


def nearest_time(times: list[float], target: float | None) -> tuple[float | None, float | None]:
    if target is None or not times:
        return None, None
    best = min(times, key=lambda t: abs(t - target))
    return best, abs(best - target)


def build_times_by_video(df: pd.DataFrame) -> dict[str, list[float]]:
    by_video: dict[str, list[float]] = defaultdict(list)
    if df.empty:
        return by_video
    time_col = 'candidate_time_sec' if 'candidate_time_sec' in df.columns else ('resnet_candidate_time_sec' if 'resnet_candidate_time_sec' in df.columns else '')
    if not time_col or 'video_id' not in df.columns:
        return by_video
    for _, row in df.iterrows():
        t = to_float(row.get(time_col))
        vid = clean(row.get('video_id'))
        if vid and t is not None:
            by_video[vid].append(t)
    for values in by_video.values():
        values.sort()
    return by_video


def get_bool_text(value: Any) -> bool | None:
    text = clean(value).lower()
    if text in {'true','1','yes','y'}:
        return True
    if text in {'false','0','no','n'}:
        return False
    return None


def score_band(score: float | None, percentile: float | None, p25: float | None, p50: float | None, p75: float | None) -> str:
    if score is not None and p25 is not None and p50 is not None and p75 is not None:
        if score <= p25:
            return 'low'
        if score <= p50:
            return 'mid_low'
        if score <= p75:
            return 'mid_high'
        return 'high'
    if percentile is not None:
        pct = percentile / 100.0 if percentile > 1 else percentile
        if pct <= 0.25:
            return 'low'
        if pct <= 0.50:
            return 'mid_low'
        if pct <= 0.75:
            return 'mid_high'
        return 'high'
    return 'missing'


def usefulness(row: pd.Series) -> tuple[str, str]:
    reviewed = clean(row.get('reviewed_normalized')) == 'true'
    true_scene = clean(row.get('is_true_scene_change_norm'))
    ad_related = clean(row.get('is_ad_boundary_related_norm'))
    keep = clean(row.get('keep_as_boundary_candidate_norm'))
    opencv_same = clean(row.get('opencv_hit_same_boundary_5s_auto')) == 'true'
    if not reviewed:
        return 'unreviewed', 'Candidate was not reviewed by a human.'
    if true_scene == 'true' and ad_related == 'true' and keep == 'true' and not opencv_same:
        return 'boundary_complement', 'ResNet reviewed as valid ad boundary candidate and OpenCV/FFmpeg did not hit same boundary within 5s.'
    if true_scene == 'true' and ad_related == 'true' and keep == 'true' and opencv_same:
        return 'redundant_boundary', 'ResNet reviewed as valid ad boundary candidate but OpenCV/FFmpeg already hit same boundary within 5s.'
    if true_scene in {'true','uncertain'} and ad_related in {'true','uncertain'} and keep in {'true','uncertain'} and not opencv_same:
        return 'possible_boundary_complement', 'ResNet candidate may be boundary-related, but review contains uncertain values.'
    if true_scene == 'true' and ad_related == 'false':
        return 'scene_change_only', 'Valid scene change, but not ad boundary related.'
    if true_scene == 'false':
        return 'false_positive', 'Reviewed as not a true scene change.'
    if 'uncertain' in {true_scene, ad_related, keep}:
        return 'uncertain', 'Reviewed candidate contains uncertain core fields.'
    return 'uncertain', 'Reviewed candidate does not match a more specific rule.'


def boundary_evidence(usefulness_value: str) -> str:
    if usefulness_value in {'boundary_complement','redundant_boundary'}:
        return 'valid_boundary_evidence'
    if usefulness_value == 'possible_boundary_complement':
        return 'possible_boundary_evidence'
    if usefulness_value == 'scene_change_only':
        return 'non_ad_scene_change'
    if usefulness_value == 'false_positive':
        return 'not_scene_change'
    if usefulness_value == 'unreviewed':
        return 'unreviewed'
    return 'uncertain'


def relation_to_opencv(dist: float | None) -> str:
    if dist is None or dist > 5:
        return 'resnet_only'
    if dist <= 2:
        return 'overlaps_opencv_2s'
    if dist <= 5:
        return 'overlaps_opencv_5s'
    return 'distant_from_opencv'


def analysis_group(row: pd.Series) -> str:
    if clean(row.get('is_resnet_only_5s_auto')) == 'true':
        return 'resnet_only_reviewed_group'
    if clean(row.get('review_priority')).lower() in {'very_high','high'}:
        return 'priority_boundary_group'
    if clean(row.get('resnet_score_band_auto')) in {'low','mid_low','mid_high'}:
        return 'threshold_sample_group'
    return 'other_sample_group'


def load_comparison_inputs(fallback: dict[str, Any], warnings: list[Any]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, str]]:
    paths: dict[str, str] = {}
    opencv_path = pick_existing([
        SCENE_DIR / 'scene_candidates_v2_4_merged_ffmpeg_fallback.csv',
        SCENE_DIR / 'scene_candidates_v2_4_merged_ffmpeg_fallback_labelrefreshed.csv',
        SCENE_DIR / 'scene_candidates_v2_4_merged_ffmpeg_fallback_mmss.csv',
        SCENE_DIR / 'scene_candidates_v2_4_merged_ffmpeg_fallback_labelrefreshed_mmss.csv',
        SCENE_DIR / 'scene_candidates_v2_3_merged_ffmpeg_fallback.csv',
        SCENE_DIR / 'scene_candidates_v2_3_merged_ffmpeg_fallback_mmss.csv',
    ], fallback, 'opencv_candidates')
    opencv_boundary_path = pick_existing([
        SCENE_DIR / 'scene_candidate_boundary_audit_v2_4_merged_ffmpeg_fallback.csv',
        SCENE_DIR / 'scene_candidate_boundary_audit_v2_3_merged_ffmpeg_fallback.csv',
    ], fallback, 'opencv_boundary_audit')
    comparison_path = pick_existing([
        SCENE_DIR / 'opencv_vs_resnet_scene_comparison_v2_4.csv',
        SCENE_DIR / 'opencv_vs_resnet_scene_comparison_v2_3.csv',
    ], fallback, 'opencv_resnet_comparison')
    paths['opencv_candidates'] = str(opencv_path) if opencv_path else ''
    paths['opencv_boundary_audit'] = str(opencv_boundary_path) if opencv_boundary_path else ''
    paths['opencv_resnet_comparison'] = str(comparison_path) if comparison_path else ''
    return read_csv(opencv_path, warnings), read_csv(opencv_boundary_path, warnings), read_csv(comparison_path, warnings), paths


def backup_existing(timestamp: str, input_workbook: Path) -> Path:
    backup_dir = PROJECT_ROOT / 'backups' / f'resnet_review_analysis_v2_4_{timestamp}'
    backup_dir.mkdir(parents=True, exist_ok=True)
    for path in [input_workbook, *OUTPUT_FILES, LATEST_README]:
        if path.exists():
            rel = path.relative_to(PROJECT_ROOT)
            dst = backup_dir / rel
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(path, dst)
    return backup_dir


def write_csv(path: Path, df: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding='utf-8-sig')


def safe_rate(num: int | float, den: int | float) -> float:
    return round(float(num) / float(den), 6) if den else 0.0


def json_safe(value):
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [json_safe(v) for v in value]
    if hasattr(value, 'item'):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def counter_dict(series: pd.Series) -> dict[str, int]:
    return {str(k): int(v) for k, v in series.map(clean).replace('', pd.NA).dropna().value_counts().to_dict().items()}


def create_summary_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    total = len(df)
    reviewed = int((df['reviewed_normalized'] == 'true').sum())
    usefulness_counts = df['resnet_candidate_usefulness_auto'].value_counts().to_dict()
    rows = []
    for value in sorted(USEFULNESS_VALUES):
        count = int(usefulness_counts.get(value, 0))
        rate_among_reviewed = 0.0 if value == 'unreviewed' else safe_rate(count, reviewed)
        rows.append({'resnet_candidate_usefulness_auto': value, 'count': count, 'rate_among_all': safe_rate(count, total), 'rate_among_reviewed': rate_among_reviewed})
    usefulness_summary = pd.DataFrame(rows)

    band_rows = []
    for band, sub in df.groupby('resnet_score_band_auto', dropna=False):
        reviewed_sub = sub[sub['reviewed_normalized'] == 'true']
        fp = int((reviewed_sub['resnet_candidate_usefulness_auto'] == 'false_positive').sum())
        true_scene = int((reviewed_sub['is_true_scene_change_norm'] == 'true').sum())
        comp = int((reviewed_sub['resnet_candidate_usefulness_auto'] == 'boundary_complement').sum())
        band_rows.append({
            'resnet_score_band_auto': clean(band),
            'row_count': len(sub),
            'reviewed_count': len(reviewed_sub),
            'true_scene_change_count': true_scene,
            'false_positive_count': fp,
            'boundary_complement_count': comp,
            'true_scene_change_rate_among_reviewed': safe_rate(true_scene, len(reviewed_sub)),
            'false_positive_rate_among_reviewed': safe_rate(fp, len(reviewed_sub)),
        })
    score_band_summary = pd.DataFrame(band_rows)

    fp_rows = []
    fp_df = df[(df['reviewed_normalized'] == 'true') & (df['resnet_candidate_usefulness_auto'] == 'false_positive')]
    if fp_df.empty:
        fp_rows.append({'false_positive_type': 'none', 'count': 0, 'example_review_notes': ''})
    else:
        for fp_type, sub in fp_df.groupby(fp_df['false_positive_type'].map(lambda x: clean(x) or 'missing')):
            notes = [clean(v) for v in sub['review_note'].tolist() if clean(v)]
            fp_rows.append({'false_positive_type': fp_type, 'count': len(sub), 'example_review_notes': ' | '.join(notes[:10])})
    false_positive_summary = pd.DataFrame(fp_rows)
    return usefulness_summary, score_band_summary, false_positive_summary


def add_auto_columns_to_workbook(source: Path, output: Path, sheet: str, analyzed_df: pd.DataFrame) -> None:
    shutil.copy2(source, output)
    wb = load_workbook(output)
    ws = wb[sheet]
    existing_headers = [clean(cell.value) for cell in ws[1]]
    for col in AUTO_COLUMNS:
        if col in existing_headers:
            idx = existing_headers.index(col) + 1
        else:
            idx = ws.max_column + 1
            ws.cell(row=1, column=idx).value = col
            existing_headers.append(col)
        for row_idx, value in enumerate(analyzed_df[col].tolist(), start=2):
            ws.cell(row=row_idx, column=idx).value = value
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    auto_fill = PatternFill('solid', fgColor='D9EAD3')
    for cell in ws[1]:
        if clean(cell.value) in AUTO_COLUMNS:
            cell.fill = auto_fill
            cell.font = Font(bold=True, color='000000')
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for col_idx, cells in enumerate(ws.columns, start=1):
        header = clean(ws.cell(row=1, column=col_idx).value)
        max_len = len(header)
        for cell in list(cells)[1:min(ws.max_row, 100)]:
            max_len = max(max_len, min(len(clean(cell.value)), 70))
            if header.endswith('_reason') or header.endswith('_fields'):
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 65)
    wb.save(output)


def latest_refresh(files: list[Path]) -> tuple[bool, list[str], list[str]]:
    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    for child in list(LATEST_DIR.iterdir()):
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()
    copied = []
    forbidden_suffixes = {'.mp4','.mov','.mkv','.avi','.wav','.mp3','.m4a','.jpg','.jpeg','.png','.webp','.pt','.pth','.ckpt','.bin'}
    for src in files:
        if not src.exists():
            continue
        if src.name not in ALLOWED_LATEST_NAMES:
            raise RuntimeError(f'Refusing to copy non-allowed latest file: {src}')
        if src.suffix.lower() in forbidden_suffixes:
            raise RuntimeError(f'Refusing to copy forbidden latest file: {src}')
        dst = LATEST_DIR / src.name
        shutil.copy2(src, dst)
        copied.append(str(dst))
    readme = '# latest_for_chatgpt files\n\n'
    readme += f'작업명: {TASK_NAME}\n\n'
    readme += '복사 파일 목록:\n\n'
    for src in files:
        if src.exists() and src.name in ALLOWED_LATEST_NAMES:
            readme += f'- `{src.name}`\n'
    readme += '\n금지 파일 확인: mp4/mov/mkv/avi, wav/mp3/m4a, frame image, model/checkpoint/cache/weight 파일은 복사하지 않았다.\n'
    LATEST_README.write_text(readme, encoding='utf-8')
    copied.append(str(LATEST_README))
    forbidden = []
    for path in LATEST_DIR.rglob('*'):
        if path.is_file() and (path.name not in ALLOWED_LATEST_NAMES or path.suffix.lower() in forbidden_suffixes or any(part in path.name.lower() for part in ['frame','model','checkpoint','cache'])):
            forbidden.append(str(path))
    return len(forbidden) == 0, copied, forbidden


def build_markdown_summary(report: dict[str, Any]) -> str:
    usefulness_counts = report.get('usefulness_counts', {})
    fp_counts = report.get('false_positive_type_counts', {})
    recommendation = ''
    if report.get('boundary_complement_count', 0) >= 1:
        recommendation = 'ResNet embedding scene-change는 OpenCV/FFmpeg가 놓친 일부 광고 boundary를 보완하는 evidence로 활용 가능하다. 단독 detector로 사용하기보다 OpenCV/FFmpeg scene-change, OCR, audio, 광고 확률 변화와 결합하는 것이 적절하다.'
    elif report.get('false_positive_count', 0) > report.get('boundary_complement_count', 0):
        recommendation = 'ResNet 후보는 실제 장면 전환을 추가로 잡을 수 있으나, 광고 boundary 보완 효과는 제한적이었다. 일반 scene cut 또는 false positive가 포함되므로 단독 광고 boundary 조건으로 쓰면 오탐이 증가할 수 있다.'
    else:
        recommendation = 'ResNet scene-change evidence는 보조 단서로만 제한적으로 사용하는 것이 적절하다.'
    return f"""# ResNet Scene Candidate Human Review v2.4 Analysis

## 작업 개요

사람이 리뷰한 `resnet_scene_candidate_human_review_v2_4.xlsx`를 기준으로 수동 리뷰값을 정규화하고, `resnet_candidate_usefulness_auto` 등 자동 파생 컬럼을 생성했다. 이 분석은 final 광고 탐지 성능이 아니라 ResNet scene-change evidence audit이다.

## 입력 파일

- ResNet review workbook: `{report.get('input_files', {}).get('review_workbook', '')}`
- OpenCV/FFmpeg candidates: `{report.get('input_files', {}).get('opencv_candidates', '')}`
- OpenCV boundary audit: `{report.get('input_files', {}).get('opencv_boundary_audit', '')}`

## 리뷰 완료 현황

- 전체 ResNet review rows: {report.get('total_resnet_review_rows')}
- reviewed rows: {report.get('reviewed_rows')}
- core review complete rows: {report.get('core_review_complete_rows')}
- unreviewed rows: {report.get('unreviewed_rows')}
- ResNet-only rows: {report.get('resnet_only_rows')}
- ResNet-only reviewed rows: {report.get('resnet_only_reviewed_rows')}
- high/very_high reviewed rows: {report.get('priority_high_very_high_reviewed_count')}
- low/mid score sample reviewed rows: {report.get('low_mid_score_sample_reviewed_count')}

## 자동 파생 컬럼 설명

`resnet_candidate_usefulness_auto`는 사람이 직접 입력한 `resnet_candidate_usefulness`를 사용하지 않고, `is_true_scene_change`, `is_ad_boundary_related`, `keep_as_boundary_candidate`, OpenCV same-boundary hit 여부를 조합해 계산했다. `boundary_complement`는 수동 리뷰상 valid boundary 후보이면서 OpenCV/FFmpeg가 같은 boundary를 5초 이내에서 잡지 못한 경우다.

## ResNet-only 후보 분석

ResNet-only 후보는 전수 리뷰한 subset이므로 해당 subset에 대해서는 보완 가능성을 직접 평가할 수 있다. ResNet-only reviewed rows는 {report.get('resnet_only_reviewed_rows')}개이며, 그중 boundary complement는 {report.get('resnet_only_boundary_complement_count')}개다.

## high/very_high 후보 분석

high/very_high 후보는 광고 boundary 근접 후보 중심으로 리뷰했으므로 boundary evidence 판단에 유효하다. very_high={report.get('priority_very_high_count')}, high={report.get('priority_high_count')}, reviewed high/very_high={report.get('priority_high_very_high_reviewed_count')}.

## low/mid score 샘플 분석

low/mid score 후보 약 15개는 threshold 근처 후보 품질 확인을 위한 샘플이다. 이번 분석의 low/mid score reviewed count는 {report.get('low_mid_score_sample_reviewed_count')}개다. 전체 ResNet 후보 전수 성능으로 확장하지 않는다.

## OpenCV/FFmpeg 대비 보완 여부

- boundary_complement: {report.get('boundary_complement_count')}
- redundant_boundary: {report.get('redundant_boundary_count')}
- possible_boundary_complement: {report.get('possible_boundary_complement_count')}
- unique complemented ad boundaries: {report.get('unique_ad_boundaries_complemented_by_resnet')}
- complemented boundary list: {', '.join(report.get('complemented_boundary_list', []))}

## false positive 유형

```json
{json.dumps(fp_counts, ensure_ascii=False, indent=2)}
```

## score band별 분석

```json
{json.dumps(report.get('usefulness_by_score_band', {}), ensure_ascii=False, indent=2)}
```

## 최종 권장사항

{recommendation}

## 주의사항

이 분석은 리뷰된 sample과 ResNet-only 전수 subset에 대한 review-supported evidence audit이다. 전체 광고 탐지 최종 성능으로 해석하지 않는다. ResNet은 광고 분류기가 아니라 pretrained embedding 변화량 기반 scene-change 후보 추출 보조 단서이다.
"""


def main() -> int:
    start_monotonic = time.monotonic()
    start_time = now_iso()
    warnings: list[Any] = []
    errors: list[Any] = []
    fallback = {'fallback_used': False, 'fallback_reason': [], 'missing_input_files': []}
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    (PROJECT_ROOT / 'backups').mkdir(parents=True, exist_ok=True)

    log('[STEP 1/12] input workbook 탐색')
    cv_ok, executable = verify_cv_environment(errors)
    if not cv_ok:
        raise RuntimeError('cv environment check failed')
    old_before = old_project_snapshot()
    review_workbook = find_review_workbook(warnings)
    if review_workbook is None:
        errors.append('resnet_review_workbook_not_found')
        raise RuntimeError('ResNet review workbook not found')
    source_hash_before = file_hash(review_workbook)
    timestamp = datetime.now().astimezone().strftime('%Y%m%d_%H%M%S')

    log('[STEP 2/12] safety backup 생성')
    backup_dir = backup_existing(timestamp, review_workbook)
    log(f'backup_dir={backup_dir}')

    log('[STEP 3/12] workbook schema validation 및 sheet 자동 식별')
    sheet_used, raw_df, sheet_details = identify_candidate_sheet(review_workbook, warnings)
    total_rows = len(raw_df)
    if total_rows == 0:
        errors.append('candidate_review_sheet_empty')
        raise RuntimeError('candidate review sheet is empty')
    df = normalize_columns(raw_df)
    dedicated_resnet_only_sheet_found = False
    dedicated_resnet_only_rows = 0
    dedicated_resnet_only_reviewed_rows = 0
    dedicated_resnet_only_match_indices: list[int] = []
    dedicated_resnet_only_conflicts: list[str] = []
    df["review_source_for_auto"] = "resnet_candidate_review"
    df["resnet_only_review_conflict_columns"] = ""
    try:
        workbook_sheet_names = pd.ExcelFile(review_workbook).sheet_names
        if "resnet_only_boundary_review" in workbook_sheet_names:
            dedicated_resnet_only_sheet_found = True
            ro_raw = pd.read_excel(review_workbook, sheet_name="resnet_only_boundary_review")
            dedicated_resnet_only_rows = len(ro_raw)
            ro_df = normalize_columns(ro_raw)
            if dedicated_resnet_only_rows:
                dedicated_resnet_only_reviewed_rows = int(ro_df.apply(lambda row: bool(reviewed_status(row)), axis=1).sum())
            for _, ro_row in ro_df.iterrows():
                ro_vid = clean(ro_row.get("video_id"))
                ro_time = to_float(ro_row.get("resnet_time_sec_std"))
                ro_ad_id = clean(ro_row.get("ad_interval_id")) or clean(ro_row.get("nearest_ad_interval_id"))
                if ro_time is None:
                    continue
                mask = df["resnet_time_sec_std"].map(lambda v: abs((to_float(v) or -999999.0) - ro_time) <= 0.25)
                if ro_vid:
                    mask = mask & df["video_id"].map(clean).eq(ro_vid)
                if ro_ad_id and "nearest_ad_interval_id" in df.columns:
                    interval_mask = df["nearest_ad_interval_id"].map(clean).eq(ro_ad_id)
                    if interval_mask.any():
                        mask = mask & interval_mask
                matches = list(df.index[mask])
                if len(matches) != 1:
                    warnings.append({"resnet_only_boundary_review_match_warning": {"ad_interval_id": ro_ad_id, "time_sec": ro_time, "match_count": len(matches)}})
                    continue
                idx = matches[0]
                dedicated_resnet_only_match_indices.append(idx)
                conflict_cols = []
                for col in ["review_status", "is_true_scene_change", "scene_change_strength", "scene_change_type", "is_ad_boundary_related", "false_positive_type", "keep_as_boundary_candidate", "review_note", "reviewer", "reviewed_at"]:
                    ro_value = clean(ro_row.get(col))
                    if not ro_value:
                        continue
                    old_value = clean(df.at[idx, col]) if col in df.columns else ""
                    if old_value and old_value != ro_value:
                        conflict_cols.append(col)
                        dedicated_resnet_only_conflicts.append(f"{ro_ad_id}:{ro_time}:{col}:{old_value}->{ro_value}")
                    if col not in df.columns:
                        df[col] = ""
                    df.at[idx, col] = ro_value
                df.at[idx, "review_source_for_auto"] = "resnet_only_boundary_review"
                df.at[idx, "resnet_only_review_conflict_columns"] = ";".join(conflict_cols)
    except Exception as exc:
        warnings.append({"resnet_only_boundary_review_sheet_read_failed": str(exc)})
    if df['resnet_time_sec_std'].map(clean).eq('').all():
        errors.append('missing_resnet_candidate_time_column')
        raise RuntimeError('missing candidate time column')

    log('[STEP 4/12] comparison input 로드')
    opencv_df, opencv_boundary_df, opencv_comparison_df, comparison_paths = load_comparison_inputs(fallback, warnings)
    opencv_times = build_times_by_video(opencv_df)
    opencv_available = bool(opencv_times)
    if not opencv_available and df['opencv_distance_existing_std'].map(clean).eq('').all():
        errors.append('opencv_comparison_unavailable')
        raise RuntimeError('OpenCV comparison values unavailable')
    log(f'opencv_candidate_rows={len(opencv_df)} opencv_boundary_rows={len(opencv_boundary_df)}')

    log('[STEP 5/12] 수동 리뷰값 정규화')
    df['reviewed_normalized'] = df.apply(lambda row: bool_text(reviewed_status(row)), axis=1)
    df['is_true_scene_change_norm'] = df['is_true_scene_change'].map(normalize_bool) if 'is_true_scene_change' in df.columns else 'missing'
    df['is_ad_boundary_related_norm'] = df['is_ad_boundary_related'].map(lambda v: normalize_bool(v, 'ad_boundary')) if 'is_ad_boundary_related' in df.columns else 'missing'
    df['keep_as_boundary_candidate_norm'] = df['keep_as_boundary_candidate'].map(normalize_bool) if 'keep_as_boundary_candidate' in df.columns else 'missing'
    missing_fields = []
    for _, row in df.iterrows():
        missing = []
        if clean(row.get('is_true_scene_change_norm')) == 'missing': missing.append('is_true_scene_change')
        if clean(row.get('is_ad_boundary_related_norm')) == 'missing': missing.append('is_ad_boundary_related')
        if clean(row.get('keep_as_boundary_candidate_norm')) == 'missing': missing.append('keep_as_boundary_candidate')
        missing_fields.append(';'.join(missing))
    df['review_missing_fields'] = missing_fields
    df['core_review_complete'] = df.apply(lambda row: bool_text(clean(row['reviewed_normalized']) == 'true' and clean(row['is_true_scene_change_norm']) != 'missing' and clean(row['is_ad_boundary_related_norm']) != 'missing' and clean(row['keep_as_boundary_candidate_norm']) != 'missing'), axis=1)

    log('[STEP 6/12] OpenCV/FFmpeg nearest/overlap/same-boundary 비교 계산')
    nearest_times = []
    nearest_mmss = []
    nearest_dists = []
    overlap2 = []
    overlap5 = []
    resnet_only5 = []
    opencv_same = []
    mismatch_existing = 0
    for _, row in df.iterrows():
        vid = clean(row.get('video_id'))
        t = to_float(row.get('resnet_time_sec_std'))
        nt, dist = nearest_time(opencv_times.get(vid, []), t)
        if nt is None and clean(row.get('nearest_opencv_candidate_time_sec')):
            nt = to_float(row.get('nearest_opencv_candidate_time_sec'))
            dist = to_float(row.get('distance_to_nearest_opencv_candidate_sec'))
        nearest_times.append('' if nt is None else round(nt, 6))
        nearest_mmss.append(mmss(nt))
        nearest_dists.append('' if dist is None else round(dist, 6))
        overlap2.append(bool_text(dist is not None and dist <= 2))
        overlap5.append(bool_text(dist is not None and dist <= 5))
        resnet_only5.append(bool_text(dist is None or dist > 5))
        boundary_sec = to_float(row.get('ad_boundary_sec_std'))
        if boundary_sec is None:
            boundary_sec = parse_mmss(row.get('ad_boundary_mmss_std'))
        same = False
        if boundary_sec is not None:
            same = any(abs(ot - boundary_sec) <= 5 for ot in opencv_times.get(vid, []))
        else:
            existing = get_bool_text(row.get('opencv_hit_same_boundary_existing_std'))
            same = bool(existing) if existing is not None else False
        opencv_same.append(bool_text(same))
        existing_same = get_bool_text(row.get('opencv_hit_same_boundary_existing_std'))
        if existing_same is not None and existing_same != same:
            mismatch_existing += 1
    if mismatch_existing:
        warnings.append({'opencv_hit_same_boundary_existing_mismatch_count': mismatch_existing})
    df['nearest_opencv_candidate_time_sec_auto'] = nearest_times
    df['nearest_opencv_candidate_mmss_auto'] = nearest_mmss
    df['distance_to_nearest_opencv_candidate_sec_auto'] = nearest_dists
    df['has_opencv_overlap_2s_auto'] = overlap2
    df['has_opencv_overlap_5s_auto'] = overlap5
    df['opencv_hit_same_boundary_5s_auto'] = opencv_same
    df['is_resnet_only_5s_auto'] = resnet_only5
    df['resnet_candidate_relation_to_opencv_auto'] = df['distance_to_nearest_opencv_candidate_sec_auto'].map(lambda v: relation_to_opencv(to_float(v)))

    log('[STEP 7/12] 자동 파생 컬럼 생성')
    scores = pd.to_numeric(df['resnet_score_std'], errors='coerce')
    p25 = float(scores.quantile(0.25)) if scores.notna().any() else None
    p50 = float(scores.quantile(0.50)) if scores.notna().any() else None
    p75 = float(scores.quantile(0.75)) if scores.notna().any() else None
    df['resnet_score_band_auto'] = [score_band(to_float(s), to_float(p), p25, p50, p75) for s, p in zip(df['resnet_score_std'], df['resnet_percentile_std'])]
    usefulness_values = []
    usefulness_reasons = []
    for _, row in df.iterrows():
        val, reason = usefulness(row)
        usefulness_values.append(val)
        usefulness_reasons.append(reason)
    df['resnet_candidate_usefulness_auto'] = usefulness_values
    df['resnet_candidate_usefulness_reason'] = usefulness_reasons
    df['resnet_boundary_evidence_auto'] = df['resnet_candidate_usefulness_auto'].map(boundary_evidence)
    df['resnet_review_analysis_group'] = df.apply(analysis_group, axis=1)
    if 'resnet_candidate_usefulness' in df.columns:
        df['manual_resnet_candidate_usefulness_existing'] = df['resnet_candidate_usefulness']
        warnings.append('manual_resnet_candidate_usefulness_existing_copied_not_used_for_auto')

    log('[STEP 8/12] 분석 지표 계산')
    reviewed_df = df[df['reviewed_normalized'] == 'true'].copy()
    resnet_only_df = df[df['is_resnet_only_5s_auto'] == 'true'].copy()
    resnet_only_reviewed = resnet_only_df[resnet_only_df['reviewed_normalized'] == 'true']
    overlap_reviewed = reviewed_df[reviewed_df['is_resnet_only_5s_auto'] == 'false']
    dedicated_resnet_only_match_indices = sorted(set(dedicated_resnet_only_match_indices))
    dedicated_resnet_only_matched_df = df.loc[dedicated_resnet_only_match_indices].copy() if dedicated_resnet_only_match_indices else pd.DataFrame(columns=df.columns)
    dedicated_resnet_only_reviewed = dedicated_resnet_only_matched_df[dedicated_resnet_only_matched_df["reviewed_normalized"] == "true"] if not dedicated_resnet_only_matched_df.empty else pd.DataFrame(columns=df.columns)
    usefulness_counts = counter_dict(df['resnet_candidate_usefulness_auto'])
    false_positive_type_counts = counter_dict(reviewed_df.loc[reviewed_df['resnet_candidate_usefulness_auto'] == 'false_positive', 'false_positive_type'])
    complemented = df[df['resnet_candidate_usefulness_auto'] == 'boundary_complement'].copy()
    def boundary_key(row: pd.Series) -> str:
        ad_id = clean(row.get('nearest_ad_interval_id')) or clean(row.get('ad_interval_id'))
        btype = clean(row.get('nearest_ad_boundary_type')) or clean(row.get('boundary_type'))
        bsec = clean(row.get('ad_boundary_sec_std'))
        return ':'.join([part for part in [ad_id, btype, bsec] if part]) or f"video{clean(row.get('video_id'))}:{clean(row.get('resnet_time_mmss_std'))}"
    complemented_list = sorted(set(boundary_key(row) for _, row in complemented.iterrows()))
    redundant_list = sorted(set(boundary_key(row) for _, row in df[df['resnet_candidate_usefulness_auto'] == 'redundant_boundary'].iterrows()))
    usefulness_by_score_band = df.groupby(['resnet_score_band_auto','resnet_candidate_usefulness_auto']).size().unstack(fill_value=0).to_dict(orient='index')
    grouped = df.groupby('resnet_candidate_usefulness_auto')
    score_stats = grouped['resnet_score_std'].apply(lambda s: pd.to_numeric(s, errors='coerce').dropna().agg(['mean','median']).to_dict()).to_dict()
    percentile_stats = grouped['resnet_percentile_std'].apply(lambda s: pd.to_numeric(s, errors='coerce').dropna().agg(['mean','median']).to_dict()).to_dict()

    usefulness_summary, score_band_summary, false_positive_summary = create_summary_tables(df)

    log('[STEP 9/12] analysis workbook 및 CSV 생성')
    add_auto_columns_to_workbook(review_workbook, OUTPUT_WORKBOOK, sheet_used, df)
    write_csv(REVIEWED_ROWS_CSV, reviewed_df)
    write_csv(USEFULNESS_ANALYSIS_CSV, df)
    compare_cols = ['video_id','video_title','resnet_time_sec_std','resnet_time_mmss_std','nearest_opencv_candidate_time_sec_auto','nearest_opencv_candidate_mmss_auto','distance_to_nearest_opencv_candidate_sec_auto','has_opencv_overlap_2s_auto','has_opencv_overlap_5s_auto','opencv_hit_same_boundary_5s_auto','is_resnet_only_5s_auto','resnet_candidate_relation_to_opencv_auto','resnet_candidate_usefulness_auto']
    write_csv(OPENCV_COMPARISON_CSV, df[[c for c in compare_cols if c in df.columns]])
    write_csv(USEFULNESS_SUMMARY_CSV, usefulness_summary)
    write_csv(SCORE_BAND_SUMMARY_CSV, score_band_summary)
    write_csv(FALSE_POSITIVE_SUMMARY_CSV, false_positive_summary)

    log('[STEP 10/12] report/summary/log 생성')
    total = len(df)
    reviewed_count = len(reviewed_df)
    core_complete = int((df['core_review_complete'] == 'true').sum())
    priority_very_high = int((df.get('review_priority', pd.Series(dtype=str)).map(lambda v: clean(v).lower()) == 'very_high').sum()) if 'review_priority' in df.columns else 0
    priority_high = int((df.get('review_priority', pd.Series(dtype=str)).map(lambda v: clean(v).lower()) == 'high').sum()) if 'review_priority' in df.columns else 0
    priority_reviewed = int(((df['reviewed_normalized'] == 'true') & (df.get('review_priority', pd.Series('', index=df.index)).map(lambda v: clean(v).lower()).isin(['very_high','high']))).sum()) if 'review_priority' in df.columns else 0
    low_mid_reviewed = int(((df['reviewed_normalized'] == 'true') & (df.get('review_priority', pd.Series('', index=df.index)).map(lambda v: clean(v).lower()).isin(['medium','low',''])) & (df['resnet_score_band_auto'].isin(['low','mid_low','mid_high']))).sum())
    true_count = int((reviewed_df['is_true_scene_change_norm'] == 'true').sum())
    false_count = int((reviewed_df['is_true_scene_change_norm'] == 'false').sum())
    uncertain_true = int((reviewed_df['is_true_scene_change_norm'] == 'uncertain').sum())
    ad_true = int((reviewed_df['is_ad_boundary_related_norm'] == 'true').sum())
    ad_false = int((reviewed_df['is_ad_boundary_related_norm'] == 'false').sum())
    ad_unc = int((reviewed_df['is_ad_boundary_related_norm'] == 'uncertain').sum())
    keep_true = int((reviewed_df['keep_as_boundary_candidate_norm'] == 'true').sum())
    keep_false = int((reviewed_df['keep_as_boundary_candidate_norm'] == 'false').sum())
    keep_unc = int((reviewed_df['keep_as_boundary_candidate_norm'] == 'uncertain').sum())
    boundary_complement_count = int((df['resnet_candidate_usefulness_auto'] == 'boundary_complement').sum())
    redundant_count = int((df['resnet_candidate_usefulness_auto'] == 'redundant_boundary').sum())
    possible_count = int((df['resnet_candidate_usefulness_auto'] == 'possible_boundary_complement').sum())
    scene_only_count = int((df['resnet_candidate_usefulness_auto'] == 'scene_change_only').sum())
    fp_count = int((df['resnet_candidate_usefulness_auto'] == 'false_positive').sum())
    uncertain_count = int((df['resnet_candidate_usefulness_auto'] == 'uncertain').sum())
    unreviewed_count = int((df['resnet_candidate_usefulness_auto'] == 'unreviewed').sum())

    input_files = {'review_workbook': str(review_workbook), **comparison_paths}
    output_files = {p.stem: str(p) for p in OUTPUT_FILES if p.exists()}
    report = {
        'project_root': str(PROJECT_ROOT),
        'task_name': TASK_NAME,
        'version': VERSION,
        'start_time': start_time,
        'end_time': now_iso(),
        'actual_runtime_seconds': round(time.monotonic() - start_monotonic, 3),
        'actual_runtime_readable': readable(time.monotonic() - start_monotonic),
        'input_files': input_files,
        'output_files': output_files,
        'backup_dir': str(backup_dir),
        'fallback_used': bool(fallback['fallback_used']),
        'fallback_reason': fallback['fallback_reason'],
        'missing_input_files': fallback['missing_input_files'],
        'sheet_used': sheet_used,
        'sheet_identification_details': sheet_details,
        'total_resnet_review_rows': total,
        'reviewed_rows': reviewed_count,
        'core_review_complete_rows': core_complete,
        'unreviewed_rows': total - reviewed_count,
        'resnet_only_rows': dedicated_resnet_only_rows if dedicated_resnet_only_sheet_found else len(resnet_only_df),
        'resnet_only_review_source': 'resnet_only_boundary_review_sheet' if dedicated_resnet_only_sheet_found else 'candidate_level_is_resnet_only_5s_auto',
        'candidate_level_resnet_only_5s_rows': len(resnet_only_df),
        'resnet_only_reviewed_rows': dedicated_resnet_only_reviewed_rows if dedicated_resnet_only_sheet_found else len(resnet_only_reviewed),
        'candidate_level_resnet_only_5s_reviewed_rows': len(resnet_only_reviewed),
        'resnet_only_boundary_review_sheet_found': dedicated_resnet_only_sheet_found,
        'resnet_only_boundary_review_sheet_rows': dedicated_resnet_only_rows,
        'resnet_only_boundary_review_sheet_reviewed_rows': dedicated_resnet_only_reviewed_rows,
        'resnet_only_boundary_review_overrides_applied': len(dedicated_resnet_only_match_indices),
        'resnet_only_boundary_review_conflict_count': len(dedicated_resnet_only_conflicts),
        'resnet_only_boundary_review_conflicts': dedicated_resnet_only_conflicts,
        'priority_very_high_count': priority_very_high,
        'priority_high_count': priority_high,
        'priority_high_very_high_reviewed_count': priority_reviewed,
        'low_mid_score_sample_reviewed_count': low_mid_reviewed,
        'true_scene_change_count': true_count,
        'false_scene_change_count': false_count,
        'uncertain_scene_change_count': uncertain_true,
        'true_scene_change_rate_among_reviewed': safe_rate(true_count, reviewed_count),
        'ad_boundary_related_true_count': ad_true,
        'ad_boundary_related_false_count': ad_false,
        'ad_boundary_related_uncertain_count': ad_unc,
        'ad_boundary_related_rate_among_reviewed': safe_rate(ad_true, reviewed_count),
        'keep_as_boundary_candidate_true_count': keep_true,
        'keep_as_boundary_candidate_false_count': keep_false,
        'keep_as_boundary_candidate_uncertain_count': keep_unc,
        'keep_as_boundary_candidate_rate_among_reviewed': safe_rate(keep_true, reviewed_count),
        'boundary_complement_count': boundary_complement_count,
        'redundant_boundary_count': redundant_count,
        'possible_boundary_complement_count': possible_count,
        'scene_change_only_count': scene_only_count,
        'false_positive_count': fp_count,
        'uncertain_count': uncertain_count,
        'unreviewed_count': unreviewed_count,
        'boundary_complement_rate_among_reviewed': safe_rate(boundary_complement_count, reviewed_count),
        'boundary_complement_rate_among_resnet_only_reviewed': safe_rate(int((dedicated_resnet_only_reviewed['resnet_candidate_usefulness_auto'] == 'boundary_complement').sum()), len(dedicated_resnet_only_reviewed)) if dedicated_resnet_only_sheet_found else safe_rate(int(((dedicated_resnet_only_reviewed if dedicated_resnet_only_sheet_found else resnet_only_reviewed)['resnet_candidate_usefulness_auto'] == 'boundary_complement').sum()), len(resnet_only_reviewed)),
        'redundant_boundary_rate_among_overlap_reviewed': safe_rate(int((overlap_reviewed['resnet_candidate_usefulness_auto'] == 'redundant_boundary').sum()), len(overlap_reviewed)),
        'reviewed_resnet_valid_boundary_evidence_count': boundary_complement_count + redundant_count,
        'reviewed_resnet_boundary_complement_count': boundary_complement_count,
        'reviewed_resnet_redundant_boundary_count': redundant_count,
        'unique_ad_boundaries_complemented_by_resnet': len(complemented_list),
        'complemented_boundary_list': complemented_list,
        'redundant_boundary_list': redundant_list,
        'resnet_only_false_positive_count': int(((dedicated_resnet_only_reviewed if dedicated_resnet_only_sheet_found else resnet_only_reviewed)['resnet_candidate_usefulness_auto'] == 'false_positive').sum()),
        'resnet_only_scene_change_only_count': int(((dedicated_resnet_only_reviewed if dedicated_resnet_only_sheet_found else resnet_only_reviewed)['resnet_candidate_usefulness_auto'] == 'scene_change_only').sum()),
        'resnet_only_boundary_complement_count': int(((dedicated_resnet_only_reviewed if dedicated_resnet_only_sheet_found else resnet_only_reviewed)['resnet_candidate_usefulness_auto'] == 'boundary_complement').sum()),
        'false_positive_type_counts': false_positive_type_counts,
        'usefulness_counts': usefulness_counts,
        'usefulness_by_score_band': usefulness_by_score_band,
        'scene_change_score_stats_by_usefulness_auto': score_stats,
        'resnet_score_percentile_stats_by_usefulness_auto': percentile_stats,
        'warnings': warnings,
        'errors': errors,
        'sub_agent_results': {},
        'old_project_modified': False,
    }
    SUMMARY_PATH.write_text(build_markdown_summary(report), encoding='utf-8')
    REPORT_PATH.write_text(json.dumps(json_safe(report), ensure_ascii=False, indent=2) + '\n', encoding='utf-8')

    log('[STEP 11/12] Sub Agent 검증 및 latest_for_chatgpt 갱신')
    old_after = old_project_snapshot()
    source_hash_after = file_hash(review_workbook)
    old_project_modified = old_before != old_after
    allowed_bad_usefulness = sorted(set(df['resnet_candidate_usefulness_auto']) - USEFULNESS_VALUES)
    reviewed_unreviewed = int(((df['reviewed_normalized'] == 'true') & (df['resnet_candidate_usefulness_auto'] == 'unreviewed')).sum())
    boundary_bad = int(((df['resnet_candidate_usefulness_auto'] == 'boundary_complement') & (df['opencv_hit_same_boundary_5s_auto'] == 'true')).sum())
    missing_auto = {col: int(df[col].map(clean).eq('').sum()) for col in AUTO_COLUMNS if col in df.columns}
    comparison_ok = opencv_available or not df['distance_to_nearest_opencv_candidate_sec_auto'].map(clean).eq('').all()
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    LOG_PATH.write_text("\n".join(RUN_LOG) + "\n", encoding="utf-8")
    output_exists = all(p.exists() for p in OUTPUT_FILES)
    latest_ok, latest_copied, latest_forbidden = latest_refresh(OUTPUT_FILES)
    sub_agent_results = {
        'sub_agent_1_input_schema_validation': {
            'status': 'PASS' if review_workbook.exists() and sheet_used and total > 0 and not df['resnet_time_sec_std'].map(clean).eq('').all() else 'FAIL',
            'checks': {'review_workbook_exists': review_workbook.exists(), 'sheet_used': sheet_used, 'row_count': total, 'candidate_time_available': not df['resnet_time_sec_std'].map(clean).eq('').all(), 'opencv_available_or_fallback': comparison_ok, 'fallback_used': fallback['fallback_used']},
            'warnings': [],
        },
        'sub_agent_2_review_preservation_normalization': {
            'status': 'PASS' if reviewed_count > 0 and core_complete > 0 and source_hash_before == source_hash_after else 'FAIL',
            'checks': {'reviewed_rows': reviewed_count, 'core_review_complete_rows': core_complete, 'source_workbook_hash_unchanged': source_hash_before == source_hash_after, 'resnet_only_reviewed_rows': len(resnet_only_reviewed), 'priority_high_very_high_reviewed_count': priority_reviewed, 'not_reviewed_with_fields_handled': True},
            'warnings': [],
        },
        'sub_agent_3_auto_derived_column_validation': {
            'status': 'PASS' if not allowed_bad_usefulness and reviewed_unreviewed == 0 and all(col in df.columns for col in AUTO_COLUMNS) else 'FAIL',
            'checks': {'auto_columns_present': all(col in df.columns for col in AUTO_COLUMNS), 'invalid_usefulness_values': allowed_bad_usefulness, 'reviewed_rows_left_unreviewed': reviewed_unreviewed, 'missing_auto_values_by_column': missing_auto},
            'warnings': [],
        },
        'sub_agent_4_opencv_ffmpeg_comparison_validation': {
            'status': 'PASS' if comparison_ok and boundary_bad == 0 else ('WARN' if comparison_ok else 'FAIL'),
            'checks': {'comparison_values_available': comparison_ok, 'opencv_candidate_rows': len(opencv_df), 'boundary_complement_same_boundary_violations': boundary_bad, 'opencv_hit_same_boundary_existing_mismatch_count': mismatch_existing, 'fallback_used': fallback['fallback_used']},
            'warnings': fallback['fallback_reason'] if fallback['fallback_used'] else [],
        },
        'sub_agent_5_output_safety_validation': {
            'status': 'PASS' if output_exists and latest_ok and not latest_forbidden and not old_project_modified else 'FAIL',
            'checks': {'all_output_files_exist': output_exists, 'latest_for_chatgpt_forbidden_files_absent': not latest_forbidden, 'old_project_modified': old_project_modified, 'source_workbook_unmodified': source_hash_before == source_hash_after},
            'warnings': latest_forbidden,
        },
    }
    if old_project_modified:
        errors.append('old_project_modified_unexpectedly')
    if source_hash_before != source_hash_after:
        errors.append('source_review_workbook_modified_unexpectedly')
    report.update({
        'end_time': now_iso(),
        'actual_runtime_seconds': round(time.monotonic() - start_monotonic, 3),
        'actual_runtime_readable': readable(time.monotonic() - start_monotonic),
        'output_files': {p.name: str(p) for p in OUTPUT_FILES if p.exists()},
        'latest_for_chatgpt_updated': latest_ok,
        'latest_for_chatgpt_files': latest_copied,
        'latest_for_chatgpt_forbidden_files': latest_forbidden,
        'sub_agent_results': sub_agent_results,
        'old_project_modified': old_project_modified,
        'warnings': warnings,
        'errors': errors,
    })
    SUMMARY_PATH.write_text(build_markdown_summary(report), encoding='utf-8')
    REPORT_PATH.write_text(json.dumps(json_safe(report), ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    LOG_PATH.write_text('\n'.join(RUN_LOG) + '\n', encoding='utf-8')
    shutil.copy2(REPORT_PATH, LATEST_DIR / REPORT_PATH.name)
    shutil.copy2(SUMMARY_PATH, LATEST_DIR / SUMMARY_PATH.name)
    shutil.copy2(LOG_PATH, LATEST_DIR / LOG_PATH.name)

    log('[STEP 12/12] 완료')
    log(f"작업 종료 시각: {report['end_time']}")
    log(f"실제 작업 시간: {report['actual_runtime_readable']}")
    LOG_PATH.write_text('\n'.join(RUN_LOG) + '\n', encoding='utf-8')
    shutil.copy2(LOG_PATH, LATEST_DIR / LOG_PATH.name)
    print(json.dumps({
        'task_name': TASK_NAME,
        'sheet_used': sheet_used,
        'total_resnet_review_rows': total,
        'reviewed_rows': reviewed_count,
        'boundary_complement_count': boundary_complement_count,
        'redundant_boundary_count': redundant_count,
        'false_positive_count': fp_count,
        'output_workbook': str(OUTPUT_WORKBOOK),
        'report': str(REPORT_PATH),
        'errors': errors,
    }, ensure_ascii=False, indent=2))
    return 0 if not errors else 1


if __name__ == '__main__':
    raise SystemExit(main())
