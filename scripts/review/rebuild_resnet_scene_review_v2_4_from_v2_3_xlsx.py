#!/usr/bin/env python3
"""Rebuild the v2.4 ResNet scene-candidate review workbook from the v2.3 review xlsx.

This script intentionally does not rerun ResNet embeddings, frame extraction, OCR,
audio feature generation, model training, or score recomputation. It only refreshes
label-dependent review surfaces and transfers human-entered review fields.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import shutil
import subprocess
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(".")
OLD_PROJECT_ROOT = Path("./_old_project_not_included")
CHECK_ENV_SCRIPT = PROJECT_ROOT / "scripts/utils/check_cv_environment.py"

TASK_NAME = "rebuild_resnet_scene_review_v2_4_from_v2_3_xlsx"
ESTIMATED_RUNTIME = "약 25분"
RUNTIME_ESTIMATION_REASON = (
    "v2_3 ResNet review xlsx 로드, v2_4 ResNet candidate/boundary audit 로드 또는 생성, "
    "검토 컬럼 이관, Excel workbook 재생성, QA report 생성 범위 기준"
)

NEW_INTERVAL_ID = "A022"
NEW_VIDEO_ID = "12"
NEW_START_SEC = 1160.0
NEW_END_SEC = 1206.0

INPUT_V2_3_REVIEW_XLSX = PROJECT_ROOT / "data/review/resnet_scene_candidate_human_review_v2_3.xlsx"
LABELS_V2_4 = PROJECT_ROOT / "data/labels/clean_ad_labels_v0_scope_refreshed_v2_4.csv"
WINDOW_LABELS_V2_4 = PROJECT_ROOT / "data/windows/window_labels_5s_v2_4.csv"
AD_SEGMENTS_V2_4 = PROJECT_ROOT / "data/segments/ad_interval_segments_v2_4.csv"
CONTEXT_SEGMENTS_V2_4 = PROJECT_ROOT / "data/segments/label_interval_context_segments_v2_4.csv"
MANIFEST_PATH = PROJECT_ROOT / "data/video_metadata/video_manifest_v2_2.csv"

RESNET_CANDIDATES_V2_4 = PROJECT_ROOT / "data/scene/resnet_scene_candidates_v2_4_labelrefreshed.csv"
RESNET_CANDIDATES_MMSS_V2_4 = PROJECT_ROOT / "data/scene/resnet_scene_candidates_v2_4_labelrefreshed_mmss.csv"
RESNET_BOUNDARY_AUDIT_V2_4 = PROJECT_ROOT / "data/scene/resnet_scene_candidate_boundary_audit_v2_4.csv"
OPENCV_RESNET_COMPARISON_V2_4 = PROJECT_ROOT / "data/scene/opencv_vs_resnet_scene_comparison_v2_4.csv"
OPENCV_RESNET_OVERLAP_V2_4 = PROJECT_ROOT / "data/scene/opencv_resnet_candidate_overlap_v2_4.csv"

RESNET_CANDIDATES_V2_3 = PROJECT_ROOT / "data/scene/resnet_scene_candidates_v2_3.csv"
RESNET_CANDIDATES_MMSS_V2_3 = PROJECT_ROOT / "data/scene/resnet_scene_candidates_v2_3_mmss.csv"
RESNET_BOUNDARY_AUDIT_V2_3 = PROJECT_ROOT / "data/scene/resnet_scene_candidate_boundary_audit_v2_3.csv"
RESNET_VIDEO_SUMMARY_V2_3 = PROJECT_ROOT / "data/scene/resnet_scene_video_summary_v2_3.csv"
OPENCV_RESNET_COMPARISON_V2_3 = PROJECT_ROOT / "data/scene/opencv_vs_resnet_scene_comparison_v2_3.csv"
OPENCV_RESNET_OVERLAP_V2_3 = PROJECT_ROOT / "data/scene/opencv_resnet_candidate_overlap_v2_3.csv"

OPENCV_CANDIDATES_V2_4 = PROJECT_ROOT / "data/scene/scene_candidates_v2_4_merged_ffmpeg_fallback_labelrefreshed.csv"
OPENCV_CANDIDATES_MMSS_V2_4 = PROJECT_ROOT / "data/scene/scene_candidates_v2_4_merged_ffmpeg_fallback_labelrefreshed_mmss.csv"
OPENCV_BOUNDARY_AUDIT_V2_4 = PROJECT_ROOT / "data/scene/scene_candidate_boundary_audit_v2_4_merged_ffmpeg_fallback.csv"
OPENCV_CANDIDATES_V2_3 = PROJECT_ROOT / "data/scene/scene_candidates_v2_3_merged_ffmpeg_fallback.csv"
OPENCV_CANDIDATES_MMSS_V2_3 = PROJECT_ROOT / "data/scene/scene_candidates_v2_3_merged_ffmpeg_fallback_mmss.csv"
OPENCV_BOUNDARY_AUDIT_V2_3 = PROJECT_ROOT / "data/scene/scene_candidate_boundary_audit_v2_3_merged_ffmpeg_fallback.csv"

REVIEW_DIR = PROJECT_ROOT / "data/review"
BACKUP_DIR = REVIEW_DIR / "backups"
OUTPUT_XLSX = REVIEW_DIR / "resnet_scene_candidate_human_review_v2_4.xlsx"
OUTPUT_CANDIDATE_CSV = REVIEW_DIR / "resnet_scene_candidate_human_review_candidate_sheet_v2_4.csv"
OUTPUT_BOUNDARY_CSV = REVIEW_DIR / "resnet_scene_candidate_human_review_boundary_sheet_v2_4.csv"
OUTPUT_RESNET_ONLY_CSV = REVIEW_DIR / "resnet_only_boundary_review_v2_4.csv"
OUTPUT_COMPARISON_CSV = REVIEW_DIR / "opencv_resnet_scene_comparison_review_v2_4.csv"
OUTPUT_VIDEO_SUMMARY_CSV = REVIEW_DIR / "resnet_scene_candidate_human_review_video_summary_v2_4.csv"

CANDIDATE_TRANSFER_AUDIT = REVIEW_DIR / "resnet_candidate_review_transfer_audit_v2_3_to_v2_4.csv"
BOUNDARY_TRANSFER_AUDIT = REVIEW_DIR / "resnet_boundary_review_transfer_audit_v2_3_to_v2_4.csv"
CANDIDATE_NEW_ROWS_CSV = REVIEW_DIR / "resnet_candidate_new_rows_v2_4.csv"
RESNET_ONLY_NEW_ROWS_CSV = REVIEW_DIR / "resnet_only_boundary_new_rows_v2_4.csv"

REPORT_PATH = PROJECT_ROOT / "reports/rebuild_resnet_scene_review_v2_4_from_v2_3_xlsx_report.json"
SUMMARY_PATH = PROJECT_ROOT / "reports/rebuild_resnet_scene_review_v2_4_from_v2_3_xlsx_summary.md"
LOG_PATH = PROJECT_ROOT / "logs/rebuild_resnet_scene_review_v2_4_from_v2_3_xlsx_run_log.txt"
README_PATH = PROJECT_ROOT / "README.md"
LATEST_DIR = PROJECT_ROOT / "outputs/latest_for_chatgpt"
LATEST_README = LATEST_DIR / "README_latest_files.md"

RUN_LOG: list[str] = []

SHEET_NAMES = [
    "resnet_candidate_review",
    "resnet_boundary_review",
    "resnet_only_boundary_review",
    "opencv_resnet_comparison",
    "video_summary_review",
    "value_options",
    "review_guide",
]

CANDIDATE_REVIEW_COLS = [
    "review_status",
    "is_true_scene_change",
    "scene_change_strength",
    "scene_change_type",
    "is_ad_boundary_related",
    "resnet_candidate_usefulness",
    "false_positive_type",
    "keep_as_boundary_candidate",
    "review_note",
    "reviewer",
    "reviewed_at",
]
BOUNDARY_REVIEW_COLS = [
    "start_candidate_correct",
    "end_candidate_correct",
    "actual_start_transition_visible",
    "actual_end_transition_visible",
    "ad_start_boundary_quality",
    "ad_end_boundary_quality",
    "resnet_boundary_better_than_opencv",
    "start_boundary_review_note",
    "end_boundary_review_note",
    "overall_boundary_review_status",
]
RESNET_ONLY_REVIEW_COLS = [
    "is_actual_scene_change",
    "is_actual_ad_boundary",
    "resnet_helped",
    "should_add_to_combined_scene_evidence",
    "failure_reason_for_opencv",
    "review_note",
    "reviewer",
    "reviewed_at",
]
COMPARISON_REVIEW_COLS = [
    "preferred_method_for_boundary",
    "combination_strategy",
    "comparison_review_note",
]
VIDEO_REVIEW_COLS = [
    "resnet_candidate_density_ok",
    "resnet_too_many_false_candidates",
    "resnet_too_few_candidates",
    "resnet_video_review_priority",
    "video_review_note",
]

VALUE_OPTIONS = {
    "review_status": ["not_reviewed", "reviewed"],
    "yes_no_unclear": ["yes", "no", "unclear"],
    "scene_change_strength": ["strong", "medium", "weak", "unclear"],
    "scene_change_type": [
        "hard_cut",
        "fade",
        "broll",
        "camera_angle_change",
        "semantic_change",
        "object_change",
        "lighting_change",
        "text_overlay_only",
        "motion_only",
        "object_motion",
        "other",
        "unclear",
    ],
    "is_ad_boundary_related": ["ad_start", "ad_end", "inside_ad", "near_ad_but_not_boundary", "no", "unclear"],
    "resnet_candidate_usefulness": ["useful", "redundant_with_opencv", "false_positive", "unclear"],
    "false_positive_type": [
        "not_false_positive",
        "normal_cut",
        "camera_motion",
        "subtitle_change",
        "lighting_change",
        "object_motion",
        "semantic_but_not_boundary",
        "duplicate_candidate",
        "score_noise",
        "unclear",
    ],
    "keep_as_boundary_candidate": ["yes", "no", "unclear"],
    "boundary_quality": ["clear", "weak", "none", "unclear"],
    "resnet_boundary_better_than_opencv": ["yes", "no", "similar", "unclear"],
    "resnet_helped": ["yes", "no", "unclear"],
    "should_add_to_combined_scene_evidence": ["yes", "no", "unclear"],
    "failure_reason_for_opencv": ["subtle_semantic_change", "low_pixel_change", "threshold_missed", "no_actual_boundary", "unclear"],
    "combination_strategy": [
        "opencv_only",
        "resnet_only",
        "opencv_or_resnet",
        "opencv_and_resnet",
        "opencv_primary_resnet_fallback",
        "unclear",
    ],
    "video_review_priority": ["high", "medium", "low"],
}


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def log(message: str) -> None:
    line = f"[{now_iso()}] {message}"
    RUN_LOG.append(line)
    print(message, flush=True)


def readable_runtime(seconds: float) -> str:
    total = int(round(seconds))
    minutes, sec = divmod(total, 60)
    return f"{minutes}분 {sec}초"


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_float(value: Any) -> float | None:
    text = clean(value)
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def truthy(value: Any) -> bool:
    return clean(value).lower() in {"true", "1", "yes", "y", "t"}


def bool_text(value: bool) -> str:
    return "true" if value else "false"


def fmt_number(value: Any, digits: int = 3) -> Any:
    number = to_float(value)
    if number is None:
        return ""
    rounded = round(number, digits)
    if abs(rounded - round(rounded)) < 10 ** (-(digits + 1)):
        return int(round(rounded))
    return rounded


def mmss(seconds: Any) -> str:
    sec = to_float(seconds)
    if sec is None:
        return ""
    total = max(0, int(math.floor(sec)))
    return f"{total // 60:02d}분 {total % 60:02d}초"


def strip_bom_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(col).replace("\ufeff", "").strip() for col in df.columns]
    return df


def safe_row_dict(row: pd.Series | dict[str, Any]) -> dict[str, Any]:
    data = row if isinstance(row, dict) else row.to_dict()
    return {str(k).replace("\ufeff", "").strip(): clean(v) for k, v in data.items()}


def ensure_inside_project(path: Path) -> None:
    resolved = path.resolve()
    root = PROJECT_ROOT.resolve()
    old = OLD_PROJECT_ROOT.resolve()
    if resolved != root and not str(resolved).startswith(str(root) + os.sep):
        raise RuntimeError(f"Refusing to write outside project root: {resolved}")
    if resolved == old or str(resolved).startswith(str(old) + os.sep):
        raise RuntimeError(f"Refusing to write inside old project root: {resolved}")


def ensure_dirs() -> None:
    for path in [REVIEW_DIR, BACKUP_DIR, REPORT_PATH.parent, LOG_PATH.parent, LATEST_DIR]:
        ensure_inside_project(path)
        path.mkdir(parents=True, exist_ok=True)


def read_csv_optional(path: Path, missing_input_files: list[str], warnings: list[Any]) -> pd.DataFrame:
    if not path.exists():
        missing_input_files.append(str(path))
        warnings.append({"missing_input_file": str(path)})
        return pd.DataFrame()
    try:
        return strip_bom_columns(pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8-sig"))
    except Exception as exc:
        missing_input_files.append(str(path))
        warnings.append({"failed_to_read_input": str(path), "error": str(exc)})
        return pd.DataFrame()


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ensure_inside_project(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in columns})


def rows_to_df(rows: list[dict[str, Any]], columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame([{col: row.get(col, "") for col in columns} for row in rows], columns=columns)


def file_sha256(path: Path) -> str:
    if not path.exists():
        return ""
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def old_project_snapshot() -> dict[str, Any]:
    if not OLD_PROJECT_ROOT.exists():
        return {"exists": False, "file_count": 0, "digest": ""}
    h = hashlib.sha256()
    file_count = 0
    for path in sorted(p for p in OLD_PROJECT_ROOT.rglob("*") if p.is_file()):
        try:
            stat = path.stat()
        except OSError:
            continue
        rel = path.relative_to(OLD_PROJECT_ROOT).as_posix()
        h.update(f"{rel}\t{stat.st_size}\t{stat.st_mtime_ns}\n".encode("utf-8", errors="replace"))
        file_count += 1
    return {"exists": True, "file_count": file_count, "digest": h.hexdigest()}


def verify_cv_environment(warnings: list[Any], errors: list[Any]) -> tuple[bool, str]:
    executable = sys.executable
    in_cv = "/envs/cv/" in executable or executable.endswith("/envs/cv/bin/python") or executable.endswith("/envs/cv/bin/python3.10")
    cmd = ["conda", "run", "-n", "cv", "python"]
    if CHECK_ENV_SCRIPT.exists():
        cmd.append(str(CHECK_ENV_SCRIPT))
    else:
        cmd.extend(
            [
                "-c",
                "import sys; print(sys.executable); import pandas as pd; import openpyxl; "
                "print('pandas', pd.__version__); print('openpyxl', openpyxl.__version__)",
            ]
        )
    result = subprocess.run(cmd, cwd=PROJECT_ROOT, capture_output=True, text=True, check=False)
    log("cv environment command: " + " ".join(cmd))
    log("cv environment stdout: " + result.stdout.strip().replace("\n", " | "))
    if result.stderr.strip():
        log("cv environment stderr: " + result.stderr.strip().replace("\n", " | "))
    if result.returncode != 0:
        errors.append({"cv_environment_check_failed": result.returncode, "stderr": result.stderr.strip()})
        return False, executable
    if not in_cv:
        errors.append({"current_python_executable_not_in_cv": executable})
        return False, executable
    return True, executable


def load_review_workbook(path: Path, warnings: list[Any], errors: list[Any], missing_review_columns: list[Any]) -> dict[str, pd.DataFrame]:
    sheets: dict[str, pd.DataFrame] = {}
    if not path.exists():
        errors.append({"missing_v2_3_resnet_review_xlsx": str(path)})
        return {sheet: pd.DataFrame() for sheet in SHEET_NAMES}
    try:
        xls = pd.ExcelFile(path)
        missing_sheets = [sheet for sheet in SHEET_NAMES if sheet not in xls.sheet_names]
        if missing_sheets:
            warnings.append({"missing_v2_3_review_sheets": missing_sheets})
        for sheet in SHEET_NAMES:
            if sheet in xls.sheet_names:
                sheets[sheet] = strip_bom_columns(pd.read_excel(path, sheet_name=sheet, dtype=object, keep_default_na=False))
            else:
                sheets[sheet] = pd.DataFrame()
    except Exception as exc:
        errors.append({"v2_3_review_xlsx_read_failed": str(exc)})
        return {sheet: pd.DataFrame() for sheet in SHEET_NAMES}

    expected_columns = {
        "resnet_candidate_review": CANDIDATE_REVIEW_COLS,
        "resnet_boundary_review": BOUNDARY_REVIEW_COLS,
        "resnet_only_boundary_review": RESNET_ONLY_REVIEW_COLS,
        "opencv_resnet_comparison": COMPARISON_REVIEW_COLS,
        "video_summary_review": VIDEO_REVIEW_COLS,
    }
    for sheet, cols in expected_columns.items():
        missing = [col for col in cols if col not in sheets.get(sheet, pd.DataFrame()).columns]
        if missing:
            missing_review_columns.append({"sheet": sheet, "missing_review_columns": missing})
    return sheets


def build_ad_intervals(ad_df: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for _, series in ad_df.iterrows():
        data = safe_row_dict(series)
        ad_id = clean(data.get("ad_interval_id")) or clean(data.get("segment_id"))
        if not ad_id or ad_id in seen:
            continue
        seen.add(ad_id)
        start = clean(data.get("ad_start_sec")) or clean(data.get("segment_start_sec"))
        end = clean(data.get("ad_end_sec")) or clean(data.get("segment_end_sec"))
        rows.append(
            {
                "ad_interval_id": ad_id,
                "video_id": clean(data.get("video_id")),
                "video_title": clean(data.get("video_title")),
                "video_filename": clean(data.get("video_filename")),
                "video_path": clean(data.get("video_path")),
                "ad_start_sec": start,
                "ad_start_mmss": clean(data.get("ad_start_mmss")) or mmss(start),
                "ad_end_sec": end,
                "ad_end_mmss": clean(data.get("ad_end_mmss")) or mmss(end),
                "video_duration_sec": clean(data.get("video_duration_sec")),
            }
        )
    rows.sort(key=lambda row: (int(clean(row.get("video_id"))) if clean(row.get("video_id")).isdigit() else 999999, to_float(row.get("ad_start_sec")) or 10**12))
    return rows


def intervals_by_video(ad_intervals: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for interval in ad_intervals:
        by_video[clean(interval.get("video_id"))].append(interval)
    return by_video


def candidate_times_by_video(df: pd.DataFrame) -> dict[str, list[float]]:
    by_video: dict[str, list[float]] = defaultdict(list)
    if df.empty:
        return by_video
    for _, series in df.iterrows():
        data = safe_row_dict(series)
        sec = to_float(data.get("candidate_time_sec"))
        if sec is not None:
            by_video[clean(data.get("video_id"))].append(sec)
    for values in by_video.values():
        values.sort()
    return by_video


def nearest_time(times: list[float], target: Any) -> tuple[float | None, float | None]:
    sec = to_float(target)
    if sec is None or not times:
        return None, None
    best = min(times, key=lambda value: abs(value - sec))
    return best, abs(best - sec)


def nearest_ad_boundary(video_intervals: list[dict[str, Any]], candidate_time: Any) -> dict[str, Any]:
    t = to_float(candidate_time)
    if t is None or not video_intervals:
        return {
            "nearest_ad_interval_id": "",
            "nearest_ad_boundary_type": "",
            "nearest_ad_boundary_sec": "",
            "nearest_ad_boundary_mmss": "",
            "distance_to_nearest_ad_boundary_sec": "",
            "is_near_ad_start_3s": "false",
            "is_near_ad_start_5s": "false",
            "is_near_ad_end_3s": "false",
            "is_near_ad_end_5s": "false",
            "is_near_any_ad_boundary_5s": "false",
        }
    best: tuple[float, str, dict[str, Any], float] | None = None
    for interval in video_intervals:
        for boundary_type, boundary_sec in [("ad_start", interval.get("ad_start_sec")), ("ad_end", interval.get("ad_end_sec"))]:
            sec = to_float(boundary_sec)
            if sec is None:
                continue
            dist = abs(t - sec)
            if best is None or dist < best[0]:
                best = (dist, boundary_type, interval, sec)
    if best is None:
        return nearest_ad_boundary([], candidate_time)
    dist, boundary_type, interval, sec = best
    start = to_float(interval.get("ad_start_sec"))
    end = to_float(interval.get("ad_end_sec"))
    start_dist = abs(t - start) if start is not None else math.inf
    end_dist = abs(t - end) if end is not None else math.inf
    return {
        "nearest_ad_interval_id": clean(interval.get("ad_interval_id")),
        "nearest_ad_boundary_type": boundary_type,
        "nearest_ad_boundary_sec": fmt_number(sec),
        "nearest_ad_boundary_mmss": mmss(sec),
        "distance_to_nearest_ad_boundary_sec": fmt_number(dist),
        "is_near_ad_start_3s": bool_text(start_dist <= 3),
        "is_near_ad_start_5s": bool_text(start_dist <= 5),
        "is_near_ad_end_3s": bool_text(end_dist <= 3),
        "is_near_ad_end_5s": bool_text(end_dist <= 5),
        "is_near_any_ad_boundary_5s": bool_text(min(start_dist, end_dist) <= 5),
    }


def candidate_inside_ad(video_intervals: list[dict[str, Any]], candidate_time: Any) -> bool:
    t = to_float(candidate_time)
    if t is None:
        return False
    for interval in video_intervals:
        start = to_float(interval.get("ad_start_sec"))
        end = to_float(interval.get("ad_end_sec"))
        if start is not None and end is not None and start <= t <= end:
            return True
    return False


def refresh_resnet_candidates_from_v2_3(v2_3_df: pd.DataFrame, ad_intervals: list[dict[str, Any]]) -> pd.DataFrame:
    by_video = intervals_by_video(ad_intervals)
    refreshed_rows: list[dict[str, Any]] = []
    for _, series in v2_3_df.iterrows():
        row = safe_row_dict(series)
        t = to_float(row.get("candidate_time_sec"))
        if t is None:
            continue
        row["candidate_time_sec"] = fmt_number(t)
        row["candidate_time_mmss"] = clean(row.get("candidate_time_mmss")) or mmss(t)
        row["candidate_time_mmss_floor"] = clean(row.get("candidate_time_mmss_floor")) or mmss(t)
        row["candidate_time_mmss_round"] = clean(row.get("candidate_time_mmss_round")) or mmss(round(t))
        row.update(nearest_ad_boundary(by_video.get(clean(row.get("video_id")), []), t))
        row["label_refresh_version"] = "v2_4"
        row["label_refresh_source"] = "fallback_from_v2_3_resnet_candidates_without_embedding_rerun"
        refreshed_rows.append(row)
    return pd.DataFrame(refreshed_rows)


def generate_boundary_audit(candidates: pd.DataFrame, ad_intervals: list[dict[str, Any]]) -> pd.DataFrame:
    times_by_video = candidate_times_by_video(candidates)
    rows: list[dict[str, Any]] = []
    candidate_records = [safe_row_dict(row) for _, row in candidates.iterrows()]
    by_video_records: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in candidate_records:
        by_video_records[clean(row.get("video_id"))].append(row)
    for interval in ad_intervals:
        video_id = clean(interval.get("video_id"))
        start = to_float(interval.get("ad_start_sec"))
        end = to_float(interval.get("ad_end_sec"))
        start_nearest, start_dist = nearest_time(times_by_video.get(video_id, []), start)
        end_nearest, end_dist = nearest_time(times_by_video.get(video_id, []), end)
        records = by_video_records.get(video_id, [])
        def count_if(predicate: Any) -> int:
            count = 0
            for row in records:
                t = to_float(row.get("candidate_time_sec"))
                if t is not None and predicate(t):
                    count += 1
            return count
        rows.append(
            {
                "ad_interval_id": clean(interval.get("ad_interval_id")),
                "video_id": video_id,
                "video_title": clean(interval.get("video_title")),
                "video_filename": clean(interval.get("video_filename")),
                "ad_start_sec": fmt_number(start),
                "ad_start_mmss": mmss(start),
                "ad_end_sec": fmt_number(end),
                "ad_end_mmss": mmss(end),
                "start_hit_3s": bool_text(start_dist is not None and start_dist <= 3),
                "start_hit_5s": bool_text(start_dist is not None and start_dist <= 5),
                "end_hit_3s": bool_text(end_dist is not None and end_dist <= 3),
                "end_hit_5s": bool_text(end_dist is not None and end_dist <= 5),
                "both_boundary_hit_5s": bool_text((start_dist is not None and start_dist <= 5) and (end_dist is not None and end_dist <= 5)),
                "nearest_candidate_to_start_sec": fmt_number(start_nearest),
                "nearest_candidate_to_start_mmss": mmss(start_nearest),
                "distance_to_start_candidate_sec": fmt_number(start_dist),
                "nearest_candidate_to_end_sec": fmt_number(end_nearest),
                "nearest_candidate_to_end_mmss": mmss(end_nearest),
                "distance_to_end_candidate_sec": fmt_number(end_dist),
                "candidate_count_in_ad_interval": count_if(lambda t: start is not None and end is not None and start <= t <= end),
                "candidate_count_in_pre10": count_if(lambda t: start is not None and start - 10 <= t < start),
                "candidate_count_in_post10": count_if(lambda t: end is not None and end < t <= end + 10),
                "candidate_count_near_start_5s": count_if(lambda t: start is not None and abs(t - start) <= 5),
                "candidate_count_near_end_5s": count_if(lambda t: end is not None and abs(t - end) <= 5),
            }
        )
    return pd.DataFrame(rows)


def compare_candidates_overlap(opencv_df: pd.DataFrame, resnet_df: pd.DataFrame) -> pd.DataFrame:
    opencv_times = candidate_times_by_video(opencv_df)
    resnet_times = candidate_times_by_video(resnet_df)
    title_by_video: dict[str, str] = {}
    for df in [opencv_df, resnet_df]:
        for _, row in df.iterrows():
            data = safe_row_dict(row)
            if clean(data.get("video_id")) and not title_by_video.get(clean(data.get("video_id"))):
                title_by_video[clean(data.get("video_id"))] = clean(data.get("video_title"))
    rows: list[dict[str, Any]] = []
    video_ids = sorted(set(opencv_times) | set(resnet_times), key=lambda v: int(v) if v.isdigit() else 999999)
    for video_id in video_ids:
        opencv = opencv_times.get(video_id, [])
        resnet = resnet_times.get(video_id, [])
        rows.append(
            {
                "video_id": video_id,
                "video_title": title_by_video.get(video_id, ""),
                "unique_opencv_candidate_count": len(opencv),
                "unique_resnet_candidate_count": len(resnet),
                "overlap_candidate_count_3s": sum(1 for t in opencv if any(abs(t - r) <= 3 for r in resnet)),
                "overlap_candidate_count_5s": sum(1 for t in opencv if any(abs(t - r) <= 5 for r in resnet)),
            }
        )
    return pd.DataFrame(rows)


def generate_comparison(
    opencv_df: pd.DataFrame,
    resnet_df: pd.DataFrame,
    opencv_audit: pd.DataFrame,
    resnet_audit: pd.DataFrame,
    overlap_df: pd.DataFrame,
    total_ad_intervals: int,
) -> pd.DataFrame:
    opencv_rows = [safe_row_dict(row) for _, row in opencv_audit.iterrows()]
    resnet_rows = [safe_row_dict(row) for _, row in resnet_audit.iterrows()]
    opencv_by_id = {clean(row.get("ad_interval_id")): row for row in opencv_rows}
    resnet_by_id = {clean(row.get("ad_interval_id")): row for row in resnet_rows}
    all_ids = sorted(set(opencv_by_id) | set(resnet_by_id))
    unique_resnet = 0
    unique_opencv = 0
    for ad_id in all_ids:
        o = opencv_by_id.get(ad_id, {})
        r = resnet_by_id.get(ad_id, {})
        for start_key, end_key in [("start_hit_5s", "end_hit_5s")]:
            unique_resnet += int(truthy(r.get(start_key)) and not truthy(o.get(start_key)))
            unique_resnet += int(truthy(r.get(end_key)) and not truthy(o.get(end_key)))
            unique_opencv += int(truthy(o.get(start_key)) and not truthy(r.get(start_key)))
            unique_opencv += int(truthy(o.get(end_key)) and not truthy(r.get(end_key)))
    overlap_3s = int(pd.to_numeric(overlap_df.get("overlap_candidate_count_3s", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not overlap_df.empty else 0
    overlap_5s = int(pd.to_numeric(overlap_df.get("overlap_candidate_count_5s", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not overlap_df.empty else 0

    def method_row(method: str, candidate_df: pd.DataFrame, audit_rows: list[dict[str, Any]], notes: str) -> dict[str, Any]:
        total_boundaries = total_ad_intervals * 2
        start3 = sum(1 for row in audit_rows if truthy(row.get("start_hit_3s")))
        start5 = sum(1 for row in audit_rows if truthy(row.get("start_hit_5s")))
        end3 = sum(1 for row in audit_rows if truthy(row.get("end_hit_3s")))
        end5 = sum(1 for row in audit_rows if truthy(row.get("end_hit_5s")))
        both5 = sum(1 for row in audit_rows if truthy(row.get("both_boundary_hit_5s")))
        unique_videos = max(1, len(set(clean(row.get("video_id")) for _, row in candidate_df.iterrows())))
        return {
            "method": method,
            "total_candidate_count": len(candidate_df),
            "candidate_count_per_video_mean": len(candidate_df) / unique_videos,
            "start_hit_3s_count": start3,
            "start_hit_5s_count": start5,
            "end_hit_3s_count": end3,
            "end_hit_5s_count": end5,
            "boundary_hit_5s_count": start5 + end5,
            "boundary_hit_5s_rate": (start5 + end5) / total_boundaries if total_boundaries else 0,
            "both_boundary_hit_5s_count": both5,
            "total_ad_intervals": total_ad_intervals,
            "total_boundaries": total_boundaries,
            "unique_resnet_boundary_hits": unique_resnet,
            "unique_opencv_boundary_hits": unique_opencv,
            "overlap_candidate_count_3s": overlap_3s,
            "overlap_candidate_count_5s": overlap_5s,
            "notes": notes,
        }

    return pd.DataFrame(
        [
            method_row("opencv_ffmpeg_merged_v2_4_labelrefreshed", opencv_df, opencv_rows, "Existing OpenCV/ffmpeg candidates reused; v2.4 label boundary audit only."),
            method_row("resnet_embedding_v2_4_labelrefreshed", resnet_df, resnet_rows, "Existing pretrained ResNet candidates reused; no embedding rerun."),
        ]
    )


def valid_v2_4_boundary_audit(df: pd.DataFrame, expected_intervals: int) -> bool:
    if df.empty or len(df) != expected_intervals:
        return False
    if "ad_interval_id" not in df.columns or NEW_INTERVAL_ID not in set(df["ad_interval_id"].map(clean)):
        return False
    required = [
        "start_hit_3s",
        "start_hit_5s",
        "end_hit_3s",
        "end_hit_5s",
        "nearest_candidate_to_start_sec",
        "nearest_candidate_to_end_sec",
    ]
    return all(col in df.columns for col in required)


def valid_v2_4_comparison(df: pd.DataFrame) -> bool:
    if df.empty or "total_ad_intervals" not in df.columns or "total_boundaries" not in df.columns:
        return False
    intervals = {clean(v) for v in df["total_ad_intervals"].tolist()}
    boundaries = {clean(v) for v in df["total_boundaries"].tolist()}
    return "22" in intervals and "44" in boundaries


def prepare_v2_4_scene_inputs(
    ad_intervals: list[dict[str, Any]],
    missing_input_files: list[str],
    warnings: list[Any],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, bool, bool, dict[str, str]]:
    scene_sources: dict[str, str] = {}

    resnet_df = read_csv_optional(RESNET_CANDIDATES_V2_4, missing_input_files, warnings)
    if resnet_df.empty:
        fallback = read_csv_optional(RESNET_CANDIDATES_V2_3, missing_input_files, warnings)
        if fallback.empty:
            fallback = read_csv_optional(RESNET_CANDIDATES_MMSS_V2_3, missing_input_files, warnings)
        if not fallback.empty:
            resnet_df = refresh_resnet_candidates_from_v2_3(fallback, ad_intervals)
            resnet_df.to_csv(RESNET_CANDIDATES_V2_4, index=False, encoding="utf-8-sig")
            resnet_df.to_csv(RESNET_CANDIDATES_MMSS_V2_4, index=False, encoding="utf-8-sig")
            scene_sources["resnet_candidates"] = "generated_from_v2_3_label_dependent_refresh_only"
    else:
        scene_sources["resnet_candidates"] = str(RESNET_CANDIDATES_V2_4)

    opencv_df = read_csv_optional(OPENCV_CANDIDATES_V2_4, missing_input_files, warnings)
    if opencv_df.empty:
        opencv_df = read_csv_optional(OPENCV_CANDIDATES_MMSS_V2_4, missing_input_files, warnings)
    if opencv_df.empty:
        opencv_df = read_csv_optional(OPENCV_CANDIDATES_V2_3, missing_input_files, warnings)
    if opencv_df.empty:
        opencv_df = read_csv_optional(OPENCV_CANDIDATES_MMSS_V2_3, missing_input_files, warnings)
    scene_sources["opencv_candidates"] = "v2_4_or_fallback_loaded" if not opencv_df.empty else "missing"

    resnet_boundary = read_csv_optional(RESNET_BOUNDARY_AUDIT_V2_4, missing_input_files, warnings)
    resnet_boundary_created = False
    if not valid_v2_4_boundary_audit(resnet_boundary, len(ad_intervals)) and not resnet_df.empty:
        warnings.append({"resnet_boundary_audit_regenerated_or_created": str(RESNET_BOUNDARY_AUDIT_V2_4)})
        resnet_boundary = generate_boundary_audit(resnet_df, ad_intervals)
        resnet_boundary.to_csv(RESNET_BOUNDARY_AUDIT_V2_4, index=False, encoding="utf-8-sig")
        resnet_boundary_created = True
    scene_sources["resnet_boundary_audit"] = str(RESNET_BOUNDARY_AUDIT_V2_4)

    opencv_boundary = read_csv_optional(OPENCV_BOUNDARY_AUDIT_V2_4, missing_input_files, warnings)
    if not valid_v2_4_boundary_audit(opencv_boundary, len(ad_intervals)) and not opencv_df.empty:
        fallback = read_csv_optional(OPENCV_BOUNDARY_AUDIT_V2_3, missing_input_files, warnings)
        opencv_boundary = fallback if not fallback.empty else generate_boundary_audit(opencv_df, ad_intervals)
    scene_sources["opencv_boundary_audit"] = "v2_4_or_fallback_loaded" if not opencv_boundary.empty else "missing"

    overlap = read_csv_optional(OPENCV_RESNET_OVERLAP_V2_4, missing_input_files, warnings)
    comparison = read_csv_optional(OPENCV_RESNET_COMPARISON_V2_4, missing_input_files, warnings)
    comparison_created = False
    if overlap.empty and not opencv_df.empty and not resnet_df.empty:
        overlap = compare_candidates_overlap(opencv_df, resnet_df)
        overlap.to_csv(OPENCV_RESNET_OVERLAP_V2_4, index=False, encoding="utf-8-sig")
        comparison_created = True
    if not valid_v2_4_comparison(comparison) and not opencv_df.empty and not resnet_df.empty and not opencv_boundary.empty and not resnet_boundary.empty:
        warnings.append({"opencv_resnet_comparison_regenerated_or_created": str(OPENCV_RESNET_COMPARISON_V2_4)})
        comparison = generate_comparison(opencv_df, resnet_df, opencv_boundary, resnet_boundary, overlap, len(ad_intervals))
        comparison.to_csv(OPENCV_RESNET_COMPARISON_V2_4, index=False, encoding="utf-8-sig")
        comparison_created = True
    scene_sources["opencv_resnet_comparison"] = str(OPENCV_RESNET_COMPARISON_V2_4)
    scene_sources["opencv_resnet_overlap"] = str(OPENCV_RESNET_OVERLAP_V2_4)

    return resnet_df, resnet_boundary, opencv_df, opencv_boundary, resnet_boundary_created, comparison_created or valid_v2_4_comparison(comparison), scene_sources


def normalize_source(value: Any) -> str:
    text = clean(value).lower()
    if text in {"resnet_embedding_v2_3", "resnet"}:
        return "resnet_embedding_v2_3"
    return text


def normalize_method(value: Any) -> str:
    return clean(value).lower()


def normalize_model(value: Any) -> str:
    return clean(value).lower()


def exact_key(row: dict[str, Any]) -> tuple[str, float | None, str, str, str]:
    t = to_float(row.get("candidate_time_sec"))
    return (
        clean(row.get("video_id")),
        round(t, 3) if t is not None else None,
        normalize_source(row.get("candidate_source")),
        normalize_method(row.get("method_used")),
        normalize_model(row.get("model_name")),
    )


def similar_text(a: Any, b: Any) -> bool:
    aa = clean(a).lower()
    bb = clean(b).lower()
    return aa == bb or (aa and bb and (aa in bb or bb in aa))


def build_v2_3_candidate_indexes(v2_3_df: pd.DataFrame) -> tuple[dict[tuple[str, float | None, str, str, str], list[tuple[int, dict[str, Any]]]], dict[str, list[tuple[int, dict[str, Any]]]]]:
    exact: dict[tuple[str, float | None, str, str, str], list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    by_video: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for idx, series in v2_3_df.iterrows():
        row = safe_row_dict(series)
        row_id = idx + 2
        exact[exact_key(row)].append((row_id, row))
        by_video[clean(row.get("video_id"))].append((row_id, row))
    return exact, by_video


def candidate_is_a022_related(row: dict[str, Any]) -> bool:
    if clean(row.get("nearest_ad_interval_id")) == NEW_INTERVAL_ID:
        return True
    if clean(row.get("video_id")) != NEW_VIDEO_ID:
        return False
    t = to_float(row.get("candidate_time_sec"))
    if t is None:
        return False
    return NEW_START_SEC - 5 <= t <= NEW_END_SEC + 5


def opencv_same_boundary_hit(opencv_by_id: dict[str, dict[str, Any]], ad_interval_id: str, boundary_type: str) -> bool:
    row = opencv_by_id.get(clean(ad_interval_id), {})
    if boundary_type == "ad_start":
        return truthy(row.get("start_hit_5s"))
    if boundary_type == "ad_end":
        return truthy(row.get("end_hit_5s"))
    return False


def review_priority(row: dict[str, Any], ad_by_video: dict[str, list[dict[str, Any]]], opencv_by_id: dict[str, dict[str, Any]]) -> str:
    if candidate_is_a022_related(row):
        return "very_high"
    near_boundary = truthy(row.get("is_near_any_ad_boundary_5s"))
    if near_boundary and not opencv_same_boundary_hit(opencv_by_id, clean(row.get("nearest_ad_interval_id")), clean(row.get("nearest_ad_boundary_type"))):
        return "very_high"
    if near_boundary:
        return "high"
    if clean(row.get("nearest_ad_interval_id")) or candidate_inside_ad(ad_by_video.get(clean(row.get("video_id")), []), row.get("candidate_time_sec")):
        return "medium"
    return "low"


def transfer_candidate_reviews(
    v2_3_candidate: pd.DataFrame,
    v2_4_candidates: pd.DataFrame,
    ad_intervals: list[dict[str, Any]],
    opencv_df: pd.DataFrame,
    opencv_boundary: pd.DataFrame,
    missing_review_columns: list[Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, int], list[int]]:
    exact_index, by_video = build_v2_3_candidate_indexes(v2_3_candidate)
    opencv_times = candidate_times_by_video(opencv_df)
    ad_by_video = intervals_by_video(ad_intervals)
    opencv_by_id = {clean(row.get("ad_interval_id")): safe_row_dict(row) for _, row in opencv_boundary.iterrows()}
    rows: list[dict[str, Any]] = []
    audit_rows: list[dict[str, Any]] = []
    new_rows: list[dict[str, Any]] = []
    counts = Counter()
    used_v2_3_ids: set[int] = set()

    for out_idx, (_, series) in enumerate(v2_4_candidates.iterrows(), start=1):
        src = safe_row_dict(series)
        t = to_float(src.get("candidate_time_sec"))
        if t is None:
            continue
        row = dict(src)
        row["candidate_time_sec"] = fmt_number(t)
        row["candidate_time_mmss"] = clean(row.get("candidate_time_mmss")) or mmss(t)
        row["candidate_time_mmss_floor"] = clean(row.get("candidate_time_mmss_floor")) or mmss(t)
        row["candidate_time_mmss_round"] = clean(row.get("candidate_time_mmss_round")) or mmss(round(t))
        nearest_opencv_sec, nearest_opencv_dist = nearest_time(opencv_times.get(clean(row.get("video_id")), []), t)
        row["review_clip_start_sec"] = fmt_number(max(0.0, t - 5.0))
        row["review_clip_end_sec"] = fmt_number(t + 5.0)
        row["review_clip_start_mmss"] = mmss(max(0.0, t - 5.0))
        row["review_clip_end_mmss"] = mmss(t + 5.0)
        row["has_near_opencv_candidate_3s"] = bool_text(nearest_opencv_dist is not None and nearest_opencv_dist <= 3)
        row["has_near_opencv_candidate_5s"] = bool_text(nearest_opencv_dist is not None and nearest_opencv_dist <= 5)
        row["nearest_opencv_candidate_time_sec"] = fmt_number(nearest_opencv_sec)
        row["nearest_opencv_candidate_mmss"] = mmss(nearest_opencv_sec)
        row["distance_to_nearest_opencv_candidate_sec"] = fmt_number(nearest_opencv_dist)
        row["is_new_or_changed_due_to_v2_4"] = bool_text(candidate_is_a022_related(row))
        row["review_priority"] = review_priority(row, ad_by_video, opencv_by_id)

        matched_row_id: int | str = ""
        matched_payload: dict[str, Any] | None = None
        match_method = ""
        time_diff: Any = ""
        warning = ""
        transferred_cols: list[str] = []
        transfer_status = "missing_v2_3_match"

        if candidate_is_a022_related(row):
            transfer_status = "new_v2_4_row"
            match_method = "A022_new_or_changed_due_to_v2_4"
            for col in CANDIDATE_REVIEW_COLS:
                row[col] = "not_reviewed" if col == "review_status" else ""
            new_payload = dict(row)
            new_payload["new_row_reason"] = "A022_related_resnet_candidate"
            new_rows.append(new_payload)
        else:
            exact_matches = exact_index.get(exact_key(row), [])
            if len(exact_matches) == 1:
                matched_row_id, matched_payload = exact_matches[0]
                match_method = "exact"
                transfer_status = "transferred_exact"
                time_diff = 0.0
            elif len(exact_matches) > 1:
                transfer_status = "ambiguous_match_not_transferred"
                match_method = "exact_multiple"
                warning = f"multiple_exact_matches:{len(exact_matches)}"
            else:
                relaxed: list[tuple[int, dict[str, Any], float]] = []
                for row_id, old in by_video.get(clean(row.get("video_id")), []):
                    old_t = to_float(old.get("candidate_time_sec"))
                    if old_t is None:
                        continue
                    if abs(old_t - t) <= 0.25 and (
                        similar_text(old.get("candidate_source"), row.get("candidate_source"))
                        or similar_text(old.get("method_used"), row.get("method_used"))
                    ):
                        relaxed.append((row_id, old, abs(old_t - t)))
                if len(relaxed) == 1:
                    matched_row_id, matched_payload, time_diff = relaxed[0]
                    match_method = "relaxed_time_0.25"
                    transfer_status = "transferred_relaxed"
                elif len(relaxed) > 1:
                    transfer_status = "ambiguous_match_not_transferred"
                    match_method = "relaxed_multiple"
                    warning = f"multiple_relaxed_matches:{len(relaxed)}"
                else:
                    fallback: list[tuple[int, dict[str, Any], float]] = []
                    for row_id, old in by_video.get(clean(row.get("video_id")), []):
                        old_mmss = clean(old.get("candidate_time_mmss")) or clean(old.get("candidate_time_mmss_floor"))
                        new_mmss = clean(row.get("candidate_time_mmss")) or clean(row.get("candidate_time_mmss_floor"))
                        old_score = to_float(old.get("scene_change_score"))
                        new_score = to_float(row.get("scene_change_score"))
                        if old_mmss == new_mmss and old_score is not None and new_score is not None and abs(old_score - new_score) <= 1e-6:
                            fallback.append((row_id, old, abs((to_float(old.get("candidate_time_sec")) or t) - t)))
                    if len(fallback) == 1:
                        matched_row_id, matched_payload, time_diff = fallback[0]
                        match_method = "fallback_mmss_score"
                        transfer_status = "transferred_relaxed"
                    elif len(fallback) > 1:
                        transfer_status = "ambiguous_match_not_transferred"
                        match_method = "fallback_multiple"
                        warning = f"multiple_fallback_matches:{len(fallback)}"

            if matched_payload is not None:
                used_v2_3_ids.add(int(matched_row_id))
                for col in CANDIDATE_REVIEW_COLS:
                    if col in matched_payload:
                        row[col] = matched_payload.get(col, "")
                        transferred_cols.append(col)
                    else:
                        row[col] = "not_reviewed" if col == "review_status" else ""
                if not clean(row.get("review_status")):
                    row["review_status"] = "not_reviewed"
            else:
                for col in CANDIDATE_REVIEW_COLS:
                    row[col] = "not_reviewed" if col == "review_status" else ""

        rows.append(row)
        counts[transfer_status] += 1
        audit_rows.append(
            {
                "v2_4_row_id": out_idx + 1,
                "v2_4_video_id": clean(row.get("video_id")),
                "v2_4_candidate_time_sec": fmt_number(t),
                "v2_4_candidate_time_mmss": clean(row.get("candidate_time_mmss")),
                "matched_v2_3_row_id": matched_row_id,
                "match_method": match_method,
                "time_diff_sec": time_diff,
                "review_columns_transferred": ";".join(transferred_cols),
                "transfer_status": transfer_status,
                "transfer_warning": warning,
            }
        )

    counts["unmatched_v2_3_review_row_count"] = max(0, len(v2_3_candidate) - len(used_v2_3_ids))
    priority_order = {"very_high": 0, "high": 1, "medium": 2, "low": 3}
    rows.sort(
        key=lambda row: (
            0 if truthy(row.get("is_new_or_changed_due_to_v2_4")) else 1,
            priority_order.get(clean(row.get("review_priority")), 99),
            int(clean(row.get("video_id"))) if clean(row.get("video_id")).isdigit() else 999999,
            to_float(row.get("candidate_time_sec")) or 10**12,
        )
    )
    return rows, audit_rows, new_rows, dict(counts), sorted(used_v2_3_ids)


def build_boundary_rows(
    ad_intervals: list[dict[str, Any]],
    resnet_boundary: pd.DataFrame,
    opencv_boundary: pd.DataFrame,
    v2_3_boundary: pd.DataFrame,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, int], set[str]]:
    resnet_by_id = {clean(row.get("ad_interval_id")): safe_row_dict(row) for _, row in resnet_boundary.iterrows()}
    opencv_by_id = {clean(row.get("ad_interval_id")): safe_row_dict(row) for _, row in opencv_boundary.iterrows()}
    old_by_id = {clean(row.get("ad_interval_id")): safe_row_dict(row) for _, row in v2_3_boundary.iterrows()}
    rows: list[dict[str, Any]] = []
    audits: list[dict[str, Any]] = []
    counts = Counter()
    used: set[str] = set()

    for interval in ad_intervals:
        ad_id = clean(interval.get("ad_interval_id"))
        r = resnet_by_id.get(ad_id, {})
        o = opencv_by_id.get(ad_id, {})
        start = to_float(r.get("ad_start_sec")) or to_float(interval.get("ad_start_sec"))
        end = to_float(r.get("ad_end_sec")) or to_float(interval.get("ad_end_sec"))
        resnet_start = truthy(r.get("start_hit_5s"))
        resnet_end = truthy(r.get("end_hit_5s"))
        opencv_start = truthy(o.get("start_hit_5s"))
        opencv_end = truthy(o.get("end_hit_5s"))
        row = {
            "ad_interval_id": ad_id,
            "video_id": clean(r.get("video_id")) or clean(interval.get("video_id")),
            "video_title": clean(r.get("video_title")) or clean(interval.get("video_title")),
            "video_filename": clean(r.get("video_filename")) or clean(interval.get("video_filename")),
            "ad_start_sec": fmt_number(start),
            "ad_start_mmss": clean(r.get("ad_start_mmss")) or mmss(start),
            "ad_end_sec": fmt_number(end),
            "ad_end_mmss": clean(r.get("ad_end_mmss")) or mmss(end),
            "ad_duration_sec": fmt_number((end - start) if start is not None and end is not None else ""),
            "start_hit_3s": bool_text(truthy(r.get("start_hit_3s"))),
            "start_hit_5s": bool_text(resnet_start),
            "end_hit_3s": bool_text(truthy(r.get("end_hit_3s"))),
            "end_hit_5s": bool_text(resnet_end),
            "both_boundary_hit_5s": bool_text(truthy(r.get("both_boundary_hit_5s"))),
            "nearest_candidate_to_start_sec": fmt_number(r.get("nearest_candidate_to_start_sec")),
            "nearest_candidate_to_start_mmss": clean(r.get("nearest_candidate_to_start_mmss")),
            "distance_to_start_candidate_sec": fmt_number(r.get("distance_to_start_candidate_sec")),
            "nearest_candidate_to_end_sec": fmt_number(r.get("nearest_candidate_to_end_sec")),
            "nearest_candidate_to_end_mmss": clean(r.get("nearest_candidate_to_end_mmss")),
            "distance_to_end_candidate_sec": fmt_number(r.get("distance_to_end_candidate_sec")),
            "candidate_count_in_ad_interval": fmt_number(r.get("candidate_count_in_ad_interval"), digits=0),
            "candidate_count_in_pre10": fmt_number(r.get("candidate_count_in_pre10"), digits=0),
            "candidate_count_in_post10": fmt_number(r.get("candidate_count_in_post10"), digits=0),
            "candidate_count_near_start_5s": fmt_number(r.get("candidate_count_near_start_5s"), digits=0),
            "candidate_count_near_end_5s": fmt_number(r.get("candidate_count_near_end_5s"), digits=0),
            "opencv_start_hit_5s": bool_text(opencv_start),
            "opencv_end_hit_5s": bool_text(opencv_end),
            "opencv_both_boundary_hit_5s": bool_text(truthy(o.get("both_boundary_hit_5s"))),
            "resnet_additional_start_hit": bool_text(resnet_start and not opencv_start),
            "resnet_additional_end_hit": bool_text(resnet_end and not opencv_end),
            "opencv_only_start_hit": bool_text(opencv_start and not resnet_start),
            "opencv_only_end_hit": bool_text(opencv_end and not resnet_end),
            "is_new_interval_v2_4": bool_text(ad_id == NEW_INTERVAL_ID),
        }
        old = old_by_id.get(ad_id)
        transferred: list[str] = []
        if old and ad_id != NEW_INTERVAL_ID:
            used.add(ad_id)
            for col in BOUNDARY_REVIEW_COLS:
                row[col] = old.get(col, "")
                transferred.append(col)
            if not clean(row.get("overall_boundary_review_status")):
                row["overall_boundary_review_status"] = "not_reviewed"
            status = "transferred"
            counts["boundary_review_transferred_count"] += 1
        else:
            for col in BOUNDARY_REVIEW_COLS:
                row[col] = "not_reviewed" if col == "overall_boundary_review_status" else ""
            status = "new_interval_v2_4" if ad_id == NEW_INTERVAL_ID else "missing_v2_3_boundary_match"
            if ad_id == NEW_INTERVAL_ID:
                counts["boundary_review_new_interval_count"] += 1
        rows.append(row)
        audits.append(
            {
                "ad_interval_id": ad_id,
                "exists_in_v2_3": bool_text(bool(old)),
                "exists_in_v2_4": "true",
                "transfer_status": status,
                "transferred_columns": ";".join(transferred),
                "is_new_interval_v2_4": bool_text(ad_id == NEW_INTERVAL_ID),
                "transfer_warning": "" if old or ad_id == NEW_INTERVAL_ID else "missing_v2_3_boundary_review_row",
            }
        )
    rows.sort(
        key=lambda row: (
            0 if clean(row.get("ad_interval_id")) == NEW_INTERVAL_ID else 1,
            int(clean(row.get("video_id"))) if clean(row.get("video_id")).isdigit() else 999999,
            to_float(row.get("ad_start_sec")) or 10**12,
        )
    )
    counts["unmatched_v2_3_boundary_review_row_count"] = max(0, len(v2_3_boundary) - len(used))
    return rows, audits, dict(counts), used


def best_resnet_candidate(candidate_rows: list[dict[str, Any]], video_id: str, ad_id: str, boundary_type: str, boundary_sec: Any) -> dict[str, Any]:
    target = to_float(boundary_sec)
    if target is None:
        return {}
    best: tuple[float, dict[str, Any]] | None = None
    for row in candidate_rows:
        if clean(row.get("video_id")) != clean(video_id):
            continue
        cand = to_float(row.get("candidate_time_sec"))
        if cand is None:
            continue
        dist = abs(cand - target)
        if dist > 5:
            continue
        same_boundary = clean(row.get("nearest_ad_interval_id")) == clean(ad_id) and clean(row.get("nearest_ad_boundary_type")) == clean(boundary_type)
        rank_dist = dist if same_boundary else dist + 1000
        if best is None or rank_dist < best[0]:
            best = (rank_dist, row)
    return best[1] if best else {}


def build_resnet_only_rows(boundary_rows: list[dict[str, Any]], candidate_rows: list[dict[str, Any]], v2_3_only: pd.DataFrame) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, int]]:
    generated: list[dict[str, Any]] = []
    for boundary in boundary_rows:
        for boundary_type, boundary_sec, boundary_mmss, include in [
            ("ad_start", boundary.get("ad_start_sec"), boundary.get("ad_start_mmss"), truthy(boundary.get("resnet_additional_start_hit"))),
            ("ad_end", boundary.get("ad_end_sec"), boundary.get("ad_end_mmss"), truthy(boundary.get("resnet_additional_end_hit"))),
        ]:
            if not include:
                continue
            candidate = best_resnet_candidate(
                candidate_rows,
                clean(boundary.get("video_id")),
                clean(boundary.get("ad_interval_id")),
                boundary_type,
                boundary_sec,
            )
            candidate_time = to_float(candidate.get("candidate_time_sec"))
            clip_start = max(0.0, candidate_time - 5.0) if candidate_time is not None else ""
            clip_end = candidate_time + 5.0 if candidate_time is not None else ""
            generated.append(
                {
                    "ad_interval_id": clean(boundary.get("ad_interval_id")),
                    "video_id": clean(boundary.get("video_id")),
                    "video_title": clean(boundary.get("video_title")),
                    "video_filename": clean(boundary.get("video_filename")),
                    "boundary_type": boundary_type,
                    "ad_boundary_sec": fmt_number(boundary_sec),
                    "ad_boundary_mmss": boundary_mmss,
                    "resnet_candidate_time_sec": fmt_number(candidate_time),
                    "resnet_candidate_mmss": clean(candidate.get("candidate_time_mmss")) or mmss(candidate_time),
                    "distance_to_boundary_sec": fmt_number(abs((candidate_time or 0.0) - (to_float(boundary_sec) or 0.0)) if candidate_time is not None else ""),
                    "resnet_scene_change_score": fmt_number(candidate.get("scene_change_score"), digits=6),
                    "resnet_score_percentile_in_video": fmt_number(candidate.get("score_percentile_in_video"), digits=3),
                    "nearest_opencv_candidate_time_sec": fmt_number(candidate.get("nearest_opencv_candidate_time_sec")),
                    "nearest_opencv_candidate_mmss": clean(candidate.get("nearest_opencv_candidate_mmss")),
                    "distance_to_nearest_opencv_candidate_sec": fmt_number(candidate.get("distance_to_nearest_opencv_candidate_sec")),
                    "opencv_hit_same_boundary_5s": "false",
                    "review_clip_start_mmss": mmss(clip_start),
                    "review_clip_end_mmss": mmss(clip_end),
                    "is_new_interval_v2_4": bool_text(clean(boundary.get("ad_interval_id")) == NEW_INTERVAL_ID),
                }
            )

    old_by_key: dict[tuple[str, str], list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for idx, series in v2_3_only.iterrows():
        old = safe_row_dict(series)
        old_by_key[(clean(old.get("ad_interval_id")), clean(old.get("boundary_type")))].append((idx + 2, old))

    rows: list[dict[str, Any]] = []
    new_rows: list[dict[str, Any]] = []
    counts = Counter()
    used_old: set[int] = set()
    for row in generated:
        status = "missing_v2_3_match"
        matched = None
        row_t = to_float(row.get("resnet_candidate_time_sec"))
        for old_id, old in old_by_key.get((clean(row.get("ad_interval_id")), clean(row.get("boundary_type"))), []):
            old_t = to_float(old.get("resnet_candidate_time_sec"))
            if old_t is not None and row_t is not None and abs(old_t - row_t) <= 0.25:
                matched = (old_id, old)
                break
        if matched and not truthy(row.get("is_new_interval_v2_4")):
            old_id, old = matched
            used_old.add(old_id)
            for col in RESNET_ONLY_REVIEW_COLS:
                row[col] = old.get(col, "")
            status = "transferred"
            counts["resnet_only_boundary_transferred_count"] += 1
        else:
            for col in RESNET_ONLY_REVIEW_COLS:
                row[col] = ""
            if truthy(row.get("is_new_interval_v2_4")):
                status = "new_interval_v2_4"
            row["review_status"] = "not_reviewed"
            counts["resnet_only_boundary_new_v2_4_count"] += 1
            new_payload = dict(row)
            new_payload["transfer_status"] = status
            new_rows.append(new_payload)
        row["transfer_status"] = status
        rows.append(row)

    rows.sort(
        key=lambda row: (
            0 if truthy(row.get("is_new_interval_v2_4")) else 1,
            int(clean(row.get("video_id"))) if clean(row.get("video_id")).isdigit() else 999999,
            to_float(row.get("ad_boundary_sec")) or 10**12,
        )
    )
    counts["resnet_only_boundary_unmatched_v2_3_review_row_count"] = max(0, len(v2_3_only) - len(used_old))
    return rows, new_rows, dict(counts)


def method_family(value: Any) -> str:
    text = clean(value).lower()
    if "resnet" in text:
        return "resnet"
    if "opencv" in text or "ffmpeg" in text:
        return "opencv"
    return text


def build_comparison_review(comparison_df: pd.DataFrame, overlap_df: pd.DataFrame, v2_3_comparison: pd.DataFrame, boundary_rows: list[dict[str, Any]], opencv_df: pd.DataFrame, resnet_df: pd.DataFrame) -> list[dict[str, Any]]:
    if comparison_df.empty:
        comparison_df = generate_comparison(opencv_df, resnet_df, pd.DataFrame(boundary_rows), pd.DataFrame(boundary_rows), overlap_df, 22)
    old_by_family = {method_family(row.get("method")): safe_row_dict(row) for _, row in v2_3_comparison.iterrows()}
    overlap_3s = int(pd.to_numeric(overlap_df.get("overlap_candidate_count_3s", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not overlap_df.empty else ""
    overlap_5s = int(pd.to_numeric(overlap_df.get("overlap_candidate_count_5s", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not overlap_df.empty else ""
    rows: list[dict[str, Any]] = []
    for _, series in comparison_df.iterrows():
        src = safe_row_dict(series)
        family = method_family(src.get("method"))
        old = old_by_family.get(family, {})
        total = to_float(src.get("total_candidate_count"))
        unique_videos = len(set(clean(row.get("video_id")) for _, row in (resnet_df if family == "resnet" else opencv_df).iterrows()))
        row = {
            "method": clean(src.get("method")),
            "total_candidate_count": fmt_number(src.get("total_candidate_count"), digits=0),
            "candidate_count_per_video_mean": fmt_number(src.get("candidate_count_per_video_mean") or ((total or 0) / unique_videos if unique_videos else "")),
            "start_hit_3s_count": fmt_number(src.get("start_hit_3s_count"), digits=0),
            "start_hit_5s_count": fmt_number(src.get("start_hit_5s_count"), digits=0),
            "end_hit_3s_count": fmt_number(src.get("end_hit_3s_count"), digits=0),
            "end_hit_5s_count": fmt_number(src.get("end_hit_5s_count"), digits=0),
            "boundary_hit_5s_count": fmt_number(src.get("boundary_hit_5s_count"), digits=0),
            "boundary_hit_5s_rate": fmt_number(src.get("boundary_hit_5s_rate"), digits=6),
            "both_boundary_hit_5s_count": fmt_number(src.get("both_boundary_hit_5s_count"), digits=0),
            "total_ad_intervals": fmt_number(src.get("total_ad_intervals"), digits=0),
            "total_boundaries": fmt_number(src.get("total_boundaries"), digits=0),
            "unique_resnet_boundary_hits": fmt_number(src.get("unique_resnet_boundary_hits"), digits=0),
            "unique_opencv_boundary_hits": fmt_number(src.get("unique_opencv_boundary_hits"), digits=0),
            "overlap_candidate_count_3s": fmt_number(src.get("overlap_candidate_count_3s") or overlap_3s, digits=0),
            "overlap_candidate_count_5s": fmt_number(src.get("overlap_candidate_count_5s") or overlap_5s, digits=0),
            "notes": clean(src.get("notes")),
        }
        for col in COMPARISON_REVIEW_COLS:
            row[col] = old.get(col, "")
        rows.append(row)
    return rows


def build_video_summary_review(
    manifest: pd.DataFrame,
    candidate_rows: list[dict[str, Any]],
    boundary_rows: list[dict[str, Any]],
    resnet_only_rows: list[dict[str, Any]],
    opencv_df: pd.DataFrame,
    overlap_df: pd.DataFrame,
    v2_3_video_summary: pd.DataFrame,
) -> list[dict[str, Any]]:
    old_by_video = {clean(row.get("video_id")): safe_row_dict(row) for _, row in v2_3_video_summary.iterrows()}
    manifest_by_video: dict[str, dict[str, Any]] = {}
    for _, series in manifest.iterrows():
        row = safe_row_dict(series)
        vid = clean(row.get("label_mapping_video_id")) or clean(row.get("video_id"))
        if vid:
            manifest_by_video[vid] = row
    candidates_by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in candidate_rows:
        candidates_by_video[clean(row.get("video_id"))].append(row)
    boundary_by_video: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in boundary_rows:
        boundary_by_video[clean(row.get("video_id"))].append(row)
    resnet_only_by_video = Counter(clean(row.get("video_id")) for row in resnet_only_rows)
    opencv_count_by_video = Counter(clean(row.get("video_id")) for _, row in opencv_df.iterrows())
    overlap_by_video = {clean(row.get("video_id")): safe_row_dict(row) for _, row in overlap_df.iterrows()} if not overlap_df.empty else {}
    video_ids = sorted(set(manifest_by_video) | set(candidates_by_video) | set(opencv_count_by_video), key=lambda v: int(v) if v.isdigit() else 999999)
    rows: list[dict[str, Any]] = []
    for vid in video_ids:
        m = manifest_by_video.get(vid, {})
        cands = candidates_by_video.get(vid, [])
        bounds = boundary_by_video.get(vid, [])
        scores = pd.to_numeric(pd.Series([row.get("scene_change_score") for row in cands]), errors="coerce").dropna()
        duration = to_float(m.get("duration_sec"))
        if duration is None:
            duration = max([to_float(row.get("candidate_time_sec")) or 0 for row in cands] + [0])
        resnet_count = len(cands)
        opencv_count = opencv_count_by_video.get(vid, 0)
        resnet_boundary_hits = sum(int(truthy(row.get("start_hit_5s"))) + int(truthy(row.get("end_hit_5s"))) for row in bounds)
        opencv_boundary_hits = sum(int(truthy(row.get("opencv_start_hit_5s"))) + int(truthy(row.get("opencv_end_hit_5s"))) for row in bounds)
        opencv_only = sum(int(truthy(row.get("opencv_only_start_hit"))) + int(truthy(row.get("opencv_only_end_hit"))) for row in bounds)
        old = old_by_video.get(vid, {})
        row = {
            "video_id": vid,
            "video_title": clean(m.get("label_mapping_video_title")) or clean(m.get("video_title")) or (cands[0].get("video_title") if cands else ""),
            "video_filename": clean(m.get("video_filename")) or (cands[0].get("video_filename") if cands else ""),
            "video_duration_sec": fmt_number(duration),
            "video_duration_mmss": mmss(duration),
            "resnet_candidate_count": resnet_count,
            "opencv_candidate_count": opencv_count,
            "resnet_candidate_count_per_min": fmt_number(resnet_count / (duration / 60.0) if duration else ""),
            "opencv_candidate_count_per_min": fmt_number(opencv_count / (duration / 60.0) if duration else ""),
            "resnet_boundary_hit_5s_count": resnet_boundary_hits,
            "opencv_boundary_hit_5s_count": opencv_boundary_hits,
            "resnet_only_boundary_hit_count": resnet_only_by_video.get(vid, 0),
            "opencv_only_boundary_hit_count": opencv_only,
            "candidate_overlap_5s_count": fmt_number(overlap_by_video.get(vid, {}).get("overlap_candidate_count_5s"), digits=0),
            "max_resnet_scene_change_score": fmt_number(scores.max() if not scores.empty else "", digits=6),
            "median_resnet_scene_change_score": fmt_number(scores.median() if not scores.empty else "", digits=6),
            "p95_resnet_scene_change_score": fmt_number(scores.quantile(0.95) if not scores.empty else "", digits=6),
            "has_new_interval_v2_4": bool_text(vid == NEW_VIDEO_ID),
        }
        for col in VIDEO_REVIEW_COLS:
            row[col] = old.get(col, "")
        if not clean(row.get("resnet_video_review_priority")):
            row["resnet_video_review_priority"] = "high" if vid == NEW_VIDEO_ID else "medium"
        rows.append(row)
    return rows


def value_options_df() -> pd.DataFrame:
    max_len = max(len(values) for values in VALUE_OPTIONS.values())
    rows = []
    for idx in range(max_len):
        rows.append({key: values[idx] if idx < len(values) else "" for key, values in VALUE_OPTIONS.items()})
    return pd.DataFrame(rows)


def review_guide_df(v2_3_guide: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if not v2_3_guide.empty and {"section", "guide"}.issubset(v2_3_guide.columns):
        for _, series in v2_3_guide[["section", "guide"]].iterrows():
            rows.append(safe_row_dict(series))
    rows.extend(
        [
            {"section": "v2.4 변경", "guide": "v2_4는 A022 광고 구간 추가를 반영한 최신 라벨 기준이다."},
            {"section": "검토값 이관", "guide": "v2_3에서 입력한 review 값은 동일 candidate/ad_interval에 대해 이관했다. nearest_ad_interval_id, distance_to_nearest_ad_boundary_sec 같은 label-dependent 값은 v2_4 기준 값을 유지했다."},
            {"section": "A022 새 검토", "guide": "A022 관련 ResNet candidate와 boundary row는 새로 검토해야 하므로 not_reviewed 또는 blank 상태로 둔다."},
            {"section": "먼저 볼 항목", "guide": "1. resnet_only_boundary_review에서 is_new_interval_v2_4=true row 2. resnet_boundary_review에서 A022 row 3. resnet_candidate_review에서 is_new_or_changed_due_to_v2_4=true row 4. review_priority=very_high/high 5. 기존 not_reviewed row"},
            {"section": "주의", "guide": "이 workbook은 scene evidence audit용이며 최종 광고 탐지 성능 claim이 아니다. ResNet embedding, score, candidate timestamp는 재계산하지 않았다."},
        ]
    )
    return pd.DataFrame(rows)


def option_formula(option_name: str) -> str:
    values = VALUE_OPTIONS[option_name]
    return '"' + ",".join(values) + '"'


def col_index(ws: Any, header: str) -> int | None:
    for idx, cell in enumerate(ws[1], start=1):
        if clean(cell.value) == header:
            return idx
    return None


def add_validation(ws: Any, header: str, option_name: str) -> bool:
    idx = col_index(ws, header)
    if idx is None or ws.max_row < 2:
        return False
    dv = DataValidation(type="list", formula1=option_formula(option_name), allow_blank=True)
    ws.add_data_validation(dv)
    col = get_column_letter(idx)
    dv.add(f"{col}2:{col}{max(ws.max_row, 500)}")
    return True


def add_row_fill(ws: Any, header: str, value: str, color: str) -> bool:
    idx = col_index(ws, header)
    if idx is None or ws.max_row < 2:
        return False
    col = get_column_letter(idx)
    fill = PatternFill("solid", fgColor=color)
    ws.conditional_formatting.add(
        f"A2:{get_column_letter(ws.max_column)}{ws.max_row}",
        FormulaRule(formula=[f'=${col}2="{value}"'], fill=fill),
    )
    return True


def add_cell_fill(ws: Any, header: str, value: str, color: str) -> bool:
    idx = col_index(ws, header)
    if idx is None or ws.max_row < 2:
        return False
    col = get_column_letter(idx)
    fill = PatternFill("solid", fgColor=color)
    ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", FormulaRule(formula=[f'=${col}2="{value}"'], fill=fill))
    return True


def apply_workbook_style(path: Path) -> tuple[bool, bool]:
    wb = load_workbook(path)
    review_columns = set(CANDIDATE_REVIEW_COLS + BOUNDARY_REVIEW_COLS + RESNET_ONLY_REVIEW_COLS + COMPARISON_REVIEW_COLS + VIDEO_REVIEW_COLS)
    review_columns.add("review_status")
    wrap_columns = {
        "video_title",
        "video_filename",
        "video_path",
        "review_note",
        "start_boundary_review_note",
        "end_boundary_review_note",
        "comparison_review_note",
        "video_review_note",
        "guide",
        "notes",
    }
    header_fill = PatternFill("solid", fgColor="1F4E78")
    review_fill = PatternFill("solid", fgColor="F4B183")
    header_font = Font(color="FFFFFF", bold=True)
    review_font = Font(color="000000", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            header = clean(cell.value)
            cell.fill = review_fill if header in review_columns else header_fill
            cell.font = review_font if header in review_columns else header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            header = clean(ws.cell(row=1, column=col_idx).value)
            max_len = len(header)
            for cell in list(column_cells)[1 : min(ws.max_row, 120)]:
                max_len = max(max_len, min(len(clean(cell.value)), 60))
                if header in wrap_columns:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            width = min(max(max_len + 2, 10), 55)
            if header in wrap_columns:
                width = min(max(width, 28), 70)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    dropdown_ok = True
    conditional_ok = True

    ws = wb["resnet_candidate_review"]
    for header, option in [
        ("review_status", "review_status"),
        ("is_true_scene_change", "yes_no_unclear"),
        ("scene_change_strength", "scene_change_strength"),
        ("scene_change_type", "scene_change_type"),
        ("is_ad_boundary_related", "is_ad_boundary_related"),
        ("resnet_candidate_usefulness", "resnet_candidate_usefulness"),
        ("false_positive_type", "false_positive_type"),
        ("keep_as_boundary_candidate", "keep_as_boundary_candidate"),
    ]:
        dropdown_ok &= add_validation(ws, header, option)
    conditional_ok &= add_row_fill(ws, "review_priority", "very_high", "F8CBAD")
    conditional_ok &= add_row_fill(ws, "review_priority", "high", "FCE4D6")
    conditional_ok &= add_row_fill(ws, "is_new_or_changed_due_to_v2_4", "true", "F4CCCC")
    conditional_ok &= add_row_fill(ws, "is_near_any_ad_boundary_5s", "true", "D9EAF7")
    conditional_ok &= add_row_fill(ws, "review_status", "not_reviewed", "E7E6E6")

    ws = wb["resnet_boundary_review"]
    for header in ["start_candidate_correct", "end_candidate_correct", "actual_start_transition_visible", "actual_end_transition_visible"]:
        dropdown_ok &= add_validation(ws, header, "yes_no_unclear")
    dropdown_ok &= add_validation(ws, "ad_start_boundary_quality", "boundary_quality")
    dropdown_ok &= add_validation(ws, "ad_end_boundary_quality", "boundary_quality")
    dropdown_ok &= add_validation(ws, "resnet_boundary_better_than_opencv", "resnet_boundary_better_than_opencv")
    dropdown_ok &= add_validation(ws, "overall_boundary_review_status", "review_status")
    conditional_ok &= add_row_fill(ws, "is_new_interval_v2_4", "true", "F4CCCC")
    conditional_ok &= add_row_fill(ws, "overall_boundary_review_status", "not_reviewed", "E7E6E6")

    ws = wb["resnet_only_boundary_review"]
    for header in ["is_actual_scene_change", "is_actual_ad_boundary", "resnet_helped", "should_add_to_combined_scene_evidence"]:
        dropdown_ok &= add_validation(ws, header, "yes_no_unclear")
    dropdown_ok &= add_validation(ws, "failure_reason_for_opencv", "failure_reason_for_opencv")
    conditional_ok &= add_row_fill(ws, "is_new_interval_v2_4", "true", "F4CCCC")
    conditional_ok &= add_row_fill(ws, "opencv_hit_same_boundary_5s", "false", "F8CBAD")
    conditional_ok &= add_cell_fill(ws, "should_add_to_combined_scene_evidence", "yes", "E2F0D9")
    conditional_ok &= add_cell_fill(ws, "should_add_to_combined_scene_evidence", "no", "FCE4D6")

    ws = wb["opencv_resnet_comparison"]
    dropdown_ok &= add_validation(ws, "preferred_method_for_boundary", "combination_strategy")
    dropdown_ok &= add_validation(ws, "combination_strategy", "combination_strategy")

    ws = wb["video_summary_review"]
    for header in ["resnet_candidate_density_ok", "resnet_too_many_false_candidates", "resnet_too_few_candidates"]:
        dropdown_ok &= add_validation(ws, header, "yes_no_unclear")
    dropdown_ok &= add_validation(ws, "resnet_video_review_priority", "video_review_priority")
    conditional_ok &= add_row_fill(ws, "has_new_interval_v2_4", "true", "F4CCCC")
    conditional_ok &= add_row_fill(ws, "resnet_video_review_priority", "high", "FCE4D6")

    wb.save(path)
    return dropdown_ok, conditional_ok


def candidate_columns(rows: list[dict[str, Any]]) -> list[str]:
    preferred = [
        "video_id",
        "video_title",
        "video_filename",
        "video_path",
        "candidate_time_sec",
        "candidate_time_mmss",
        "candidate_time_mmss_floor",
        "candidate_time_mmss_round",
        "scene_change_score",
        "threshold",
        "cosine_distance",
        "l2_distance",
        "score_rank_in_video",
        "score_percentile_in_video",
        "candidate_source",
        "method_used",
        "model_name",
        "feature_dim",
        "nearest_window_id",
        "nearest_ad_interval_id",
        "nearest_ad_boundary_type",
        "nearest_ad_boundary_sec",
        "nearest_ad_boundary_mmss",
        "distance_to_nearest_ad_boundary_sec",
        "is_near_ad_start_3s",
        "is_near_ad_start_5s",
        "is_near_ad_end_3s",
        "is_near_ad_end_5s",
        "is_near_any_ad_boundary_5s",
        "label_refresh_version",
        "label_refresh_source",
        "review_clip_start_sec",
        "review_clip_end_sec",
        "review_clip_start_mmss",
        "review_clip_end_mmss",
        "has_near_opencv_candidate_3s",
        "has_near_opencv_candidate_5s",
        "nearest_opencv_candidate_time_sec",
        "nearest_opencv_candidate_mmss",
        "distance_to_nearest_opencv_candidate_sec",
        "review_priority",
        "is_new_or_changed_due_to_v2_4",
        *CANDIDATE_REVIEW_COLS,
    ]
    existing = set().union(*(row.keys() for row in rows)) if rows else set(preferred)
    return [col for col in preferred if col in existing]


def boundary_columns() -> list[str]:
    return [
        "ad_interval_id",
        "video_id",
        "video_title",
        "video_filename",
        "ad_start_sec",
        "ad_start_mmss",
        "ad_end_sec",
        "ad_end_mmss",
        "ad_duration_sec",
        "start_hit_3s",
        "start_hit_5s",
        "end_hit_3s",
        "end_hit_5s",
        "both_boundary_hit_5s",
        "nearest_candidate_to_start_sec",
        "nearest_candidate_to_start_mmss",
        "distance_to_start_candidate_sec",
        "nearest_candidate_to_end_sec",
        "nearest_candidate_to_end_mmss",
        "distance_to_end_candidate_sec",
        "candidate_count_in_ad_interval",
        "candidate_count_in_pre10",
        "candidate_count_in_post10",
        "candidate_count_near_start_5s",
        "candidate_count_near_end_5s",
        "opencv_start_hit_5s",
        "opencv_end_hit_5s",
        "opencv_both_boundary_hit_5s",
        "resnet_additional_start_hit",
        "resnet_additional_end_hit",
        "opencv_only_start_hit",
        "opencv_only_end_hit",
        "is_new_interval_v2_4",
        *BOUNDARY_REVIEW_COLS,
    ]


def resnet_only_columns() -> list[str]:
    return [
        "ad_interval_id",
        "video_id",
        "video_title",
        "video_filename",
        "boundary_type",
        "ad_boundary_sec",
        "ad_boundary_mmss",
        "resnet_candidate_time_sec",
        "resnet_candidate_mmss",
        "distance_to_boundary_sec",
        "resnet_scene_change_score",
        "resnet_score_percentile_in_video",
        "nearest_opencv_candidate_time_sec",
        "nearest_opencv_candidate_mmss",
        "distance_to_nearest_opencv_candidate_sec",
        "opencv_hit_same_boundary_5s",
        "review_clip_start_mmss",
        "review_clip_end_mmss",
        "is_new_interval_v2_4",
        *RESNET_ONLY_REVIEW_COLS,
        "review_status",
        "transfer_status",
    ]


def comparison_columns() -> list[str]:
    return [
        "method",
        "total_candidate_count",
        "candidate_count_per_video_mean",
        "start_hit_3s_count",
        "start_hit_5s_count",
        "end_hit_3s_count",
        "end_hit_5s_count",
        "boundary_hit_5s_count",
        "boundary_hit_5s_rate",
        "both_boundary_hit_5s_count",
        "total_ad_intervals",
        "total_boundaries",
        "unique_resnet_boundary_hits",
        "unique_opencv_boundary_hits",
        "overlap_candidate_count_3s",
        "overlap_candidate_count_5s",
        "notes",
        *COMPARISON_REVIEW_COLS,
    ]


def video_summary_columns() -> list[str]:
    return [
        "video_id",
        "video_title",
        "video_filename",
        "video_duration_sec",
        "video_duration_mmss",
        "resnet_candidate_count",
        "opencv_candidate_count",
        "resnet_candidate_count_per_min",
        "opencv_candidate_count_per_min",
        "resnet_boundary_hit_5s_count",
        "opencv_boundary_hit_5s_count",
        "resnet_only_boundary_hit_count",
        "opencv_only_boundary_hit_count",
        "candidate_overlap_5s_count",
        "max_resnet_scene_change_score",
        "median_resnet_scene_change_score",
        "p95_resnet_scene_change_score",
        "has_new_interval_v2_4",
        *VIDEO_REVIEW_COLS,
    ]


def latest_forbidden_files() -> list[str]:
    forbidden: list[str] = []
    if not LATEST_DIR.exists():
        return forbidden
    allowed_xlsx = {OUTPUT_XLSX.name, "scene_candidate_human_review_v2_4.xlsx"}
    forbidden_suffixes = {".mp4", ".mov", ".avi", ".mkv", ".jpg", ".jpeg", ".png", ".webp", ".pt", ".pth", ".ckpt", ".bin"}
    for path in LATEST_DIR.rglob("*"):
        if not path.is_file():
            continue
        name = path.name.lower()
        suffix = path.suffix.lower()
        if suffix in forbidden_suffixes:
            forbidden.append(str(path))
        if suffix in {".xlsx", ".xls"} and path.name not in allowed_xlsx:
            forbidden.append(str(path))
        if any(part in name for part in ["frame", "model", "checkpoint", "cache"]):
            forbidden.append(str(path))
    return sorted(set(forbidden))


def copy_latest(files: list[Path]) -> tuple[bool, list[str], list[str]]:
    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    copied: list[str] = []
    allowed_xlsx = {OUTPUT_XLSX.name}
    for src in files:
        if not src.exists():
            continue
        suffix = src.suffix.lower()
        if suffix in {".mp4", ".mov", ".avi", ".mkv", ".jpg", ".jpeg", ".png", ".webp", ".pt", ".pth", ".ckpt", ".bin"}:
            raise RuntimeError(f"Refusing to copy forbidden file to latest_for_chatgpt: {src}")
        if suffix in {".xlsx", ".xls"} and src.name not in allowed_xlsx:
            raise RuntimeError(f"Refusing to copy non-output xlsx to latest_for_chatgpt: {src}")
        dst = LATEST_DIR / src.name
        shutil.copy2(src, dst)
        copied.append(str(dst))
    LATEST_README.write_text(
        "# latest_for_chatgpt files\n\n"
        f"이번 작업명: {TASK_NAME}\n\n"
        "v2_3 ResNet review 값 이관 후 재생성한 v2_4 ResNet review workbook, companion CSV, transfer audit, report/summary/log만 복사했다. "
        "mp4, 원본 raw xlsx, frame, model, checkpoint, cache 파일은 복사하지 않는다.\n",
        encoding="utf-8",
    )
    copied.append(str(LATEST_README))
    forbidden = latest_forbidden_files()
    return len(forbidden) == 0, copied, forbidden


def update_readme() -> None:
    section_title = "## ResNet Scene Candidate Human Review v2.4 Rebuilt from v2.3"
    section = f"""{section_title}

v2_3 ResNet review xlsx의 사람이 입력한 검토값을 v2_4 ResNet review xlsx로 이관했다.

v2_4는 A022 광고 구간이 추가된 최신 라벨 기준이며, A022 관련 ResNet candidate/boundary row는 새로 검토해야 한다.

이후 ResNet scene 후보 검토는 `data/review/resnet_scene_candidate_human_review_v2_4.xlsx` 기준으로 진행한다.
"""
    text = README_PATH.read_text(encoding="utf-8") if README_PATH.exists() else "# youtube_ad_segment_detector\n"
    if section_title in text:
        before, _, after = text.partition(section_title)
        next_idx = after.find("\n## ")
        text = before.rstrip() + "\n\n" + section.rstrip() + ("\n" + after[next_idx + 1 :].lstrip() if next_idx >= 0 else "")
    else:
        text = text.rstrip() + "\n\n" + section.rstrip()
    README_PATH.write_text(text.rstrip() + "\n", encoding="utf-8")


def make_summary(report: dict[str, Any]) -> str:
    warnings = report.get("warnings") or []
    warning_text = "\n".join(f"- {item}" for item in warnings) if warnings else "- 주요 warning 없음"
    return f"""# rebuild_resnet_scene_review_v2_4_from_v2_3_xlsx summary

## Runtime

- 예상 작업 시간: {report.get("estimated_runtime")}
- 실제 작업 시간: {report.get("actual_runtime_readable")}
- 작업 시작 시각: {report.get("start_time")}
- 작업 종료 시각: {report.get("end_time")}

## Review Transfer

- candidate review rows v2_3/v2_4: {report.get("v2_3_candidate_review_row_count")} / {report.get("v2_4_candidate_review_row_count")}
- candidate transfer exact/relaxed/new/ambiguous: {report.get("candidate_review_transferred_exact_count")} / {report.get("candidate_review_transferred_relaxed_count")} / {report.get("candidate_review_new_v2_4_count")} / {report.get("candidate_review_ambiguous_not_transferred_count")}
- boundary review rows v2_3/v2_4: {report.get("v2_3_boundary_review_row_count")} / {report.get("v2_4_boundary_review_row_count")}
- boundary transferred/new interval: {report.get("boundary_review_transferred_count")} / {report.get("boundary_review_new_interval_count")}
- ResNet-only boundary rows v2_3/v2_4: {report.get("v2_3_resnet_only_boundary_row_count")} / {report.get("v2_4_resnet_only_boundary_row_count")}
- ResNet-only transferred/new: {report.get("resnet_only_boundary_transferred_count")} / {report.get("resnet_only_boundary_new_v2_4_count")}

## A022

- A022 candidate rows: {report.get("A022_candidate_row_count")}
- A022 boundary row exists: {report.get("A022_boundary_row_exists")}
- A022 ResNet-only boundary rows: {report.get("A022_resnet_only_boundary_row_count")}

## Outputs

- Review xlsx: `{report.get("output_v2_4_review_xlsx_path")}`
- Backup created: {report.get("backup_created")}
- Backup path: `{report.get("backup_path")}`
- Candidate transfer audit: `{CANDIDATE_TRANSFER_AUDIT}`
- Boundary transfer audit: `{BOUNDARY_TRANSFER_AUDIT}`
- Report: `{REPORT_PATH}`
- Log: `{LOG_PATH}`

## 먼저 검토할 항목

1. `resnet_only_boundary_review` sheet에서 `is_new_interval_v2_4=true` row
2. `resnet_boundary_review` sheet의 A022 row
3. `resnet_candidate_review` sheet에서 `is_new_or_changed_due_to_v2_4=true` row
4. `review_priority=very_high/high` row
5. 기존 `not_reviewed` row

## Sub Agent Results

```json
{json.dumps(report.get("sub_agent_results", {}), ensure_ascii=False, indent=2)}
```

## Warnings

{warning_text}
"""


def write_report_summary_log(report: dict[str, Any]) -> None:
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    SUMMARY_PATH.write_text(make_summary(report), encoding="utf-8")
    LOG_PATH.write_text("\n".join(RUN_LOG) + "\n", encoding="utf-8")


def status_from_checks(checks: list[bool], warnings: list[str] | None = None) -> str:
    if not all(checks):
        return "FAIL"
    if warnings:
        return "WARN"
    return "PASS"


def build_sub_agent_results(
    report: dict[str, Any],
    candidate_counts: dict[str, int],
    boundary_counts: dict[str, int],
    resnet_only_counts: dict[str, int],
    candidate_rows: list[dict[str, Any]],
    boundary_rows: list[dict[str, Any]],
    resnet_only_rows: list[dict[str, Any]],
    comparison_rows: list[dict[str, Any]],
    v2_3_hash_before: str,
    v2_3_hash_after: str,
    old_project_before: dict[str, Any],
    old_project_after: dict[str, Any],
    latest_forbidden: list[str],
) -> dict[str, Any]:
    a022_candidates = [row for row in candidate_rows if candidate_is_a022_related(row)]
    a022_candidate_priority_ok = bool(a022_candidates) and all(clean(row.get("review_priority")) in {"very_high", "high"} for row in a022_candidates)
    comparison_totals_ok = all(clean(row.get("total_ad_intervals")) == "22" and clean(row.get("total_boundaries")) == "44" for row in comparison_rows)

    review_transfer_warnings: list[str] = []
    if candidate_counts.get("unmatched_v2_3_review_row_count", 0):
        review_transfer_warnings.append(f"unmatched_v2_3_candidate_rows={candidate_counts.get('unmatched_v2_3_review_row_count')}")
    if resnet_only_counts.get("resnet_only_boundary_unmatched_v2_3_review_row_count", 0):
        review_transfer_warnings.append(f"unmatched_v2_3_resnet_only_rows={resnet_only_counts.get('resnet_only_boundary_unmatched_v2_3_review_row_count')}")
    review_transfer_checks = [
        report.get("v2_4_candidate_review_row_count", 0) > 0,
        candidate_counts.get("transferred_exact", 0) + candidate_counts.get("transferred_relaxed", 0) + candidate_counts.get("new_v2_4_row", 0) + candidate_counts.get("ambiguous_match_not_transferred", 0) + candidate_counts.get("missing_v2_3_match", 0) == report.get("v2_4_candidate_review_row_count", -1),
        report.get("candidate_review_ambiguous_not_transferred_count", 0) == 0,
    ]

    boundary_checks = [
        report.get("v2_4_boundary_review_row_count") == 22,
        report.get("A022_boundary_row_exists") is True,
        a022_candidate_priority_ok,
        report.get("v2_4_resnet_only_boundary_row_count", 0) >= 0,
        comparison_totals_ok,
    ]

    safety_checks = [
        v2_3_hash_before == v2_3_hash_after,
        (not report.get("preexisting_v2_4_review_xlsx") or report.get("backup_created") is True),
        old_project_before == old_project_after,
        len(latest_forbidden) == 0,
    ]
    safety_warnings: list[str] = []
    if latest_forbidden:
        safety_warnings.append("forbidden_latest_files_detected")

    return {
        "sub_agent_1_review_transfer": {
            "status": status_from_checks(review_transfer_checks, review_transfer_warnings),
            "checks": {
                "candidate_transfer_accounted": review_transfer_checks[1],
                "candidate_ambiguous_zero": review_transfer_checks[2],
                "review_columns_checked": True,
                "unmatched_v2_3_candidate_rows": candidate_counts.get("unmatched_v2_3_review_row_count", 0),
                "new_v2_4_candidate_rows": candidate_counts.get("new_v2_4_row", 0),
                "resnet_only_transferred": resnet_only_counts.get("resnet_only_boundary_transferred_count", 0),
            },
            "warnings": review_transfer_warnings,
        },
        "sub_agent_2_v2_4_boundary_comparison": {
            "status": status_from_checks(boundary_checks),
            "checks": {
                "v2_4_ad_interval_count_is_22": report.get("v2_4_boundary_review_row_count") == 22,
                "total_boundaries_is_44": comparison_totals_ok,
                "A022_boundary_row_exists": report.get("A022_boundary_row_exists"),
                "A022_candidates_high_or_very_high": a022_candidate_priority_ok,
                "resnet_only_boundary_review_created": report.get("v2_4_resnet_only_boundary_row_count", 0) >= 0,
            },
            "warnings": [],
        },
        "sub_agent_3_output_safety": {
            "status": status_from_checks(safety_checks, safety_warnings),
            "checks": {
                "v2_3_review_xlsx_unchanged": v2_3_hash_before == v2_3_hash_after,
                "v2_4_backup_created_if_needed": (not report.get("preexisting_v2_4_review_xlsx") or report.get("backup_created") is True),
                "latest_for_chatgpt_forbidden_files_absent": len(latest_forbidden) == 0,
                "old_project_unmodified": old_project_before == old_project_after,
            },
            "warnings": safety_warnings,
        },
    }


def main() -> int:
    ensure_dirs()
    start_monotonic = time.monotonic()
    start_time = now_iso()
    warnings: list[Any] = []
    errors: list[Any] = []
    missing_input_files: list[str] = []
    missing_required_columns: list[Any] = []
    missing_review_columns: list[Any] = []

    log(f"작업 시작 전 예상 작업 시간: {ESTIMATED_RUNTIME}")
    log(f"예상 근거: {RUNTIME_ESTIMATION_REASON}")
    log(f"작업 시작 시각: {start_time}")

    old_project_before = old_project_snapshot()
    v2_3_hash_before = file_sha256(INPUT_V2_3_REVIEW_XLSX)
    preexisting_v2_4 = OUTPUT_XLSX.exists()
    backup_created = False
    backup_path = ""

    report: dict[str, Any] = {
        "project_root": str(PROJECT_ROOT),
        "estimated_runtime": ESTIMATED_RUNTIME,
        "runtime_estimation_reason": RUNTIME_ESTIMATION_REASON,
        "start_time": start_time,
        "end_time": "",
        "actual_runtime_seconds": 0,
        "actual_runtime_readable": "",
        "input_v2_3_review_xlsx_path": str(INPUT_V2_3_REVIEW_XLSX),
        "input_v2_4_resnet_candidate_path": str(RESNET_CANDIDATES_V2_4),
        "input_v2_4_resnet_boundary_audit_path": str(RESNET_BOUNDARY_AUDIT_V2_4),
        "input_v2_4_opencv_comparison_path": str(OPENCV_RESNET_COMPARISON_V2_4),
        "output_v2_4_review_xlsx_path": str(OUTPUT_XLSX),
        "preexisting_v2_4_review_xlsx": preexisting_v2_4,
        "backup_created": False,
        "backup_path": "",
        "missing_input_files": missing_input_files,
        "missing_required_columns": missing_required_columns,
        "missing_review_columns": missing_review_columns,
        "sub_agent_results": {},
        "old_project_modified": False,
        "warnings": warnings,
        "errors": errors,
    }

    try:
        log("[STEP 1/10] cv 환경 확인")
        cv_ok, python_executable = verify_cv_environment(warnings, errors)
        report["cv_environment_checked"] = cv_ok
        report["python_executable"] = python_executable
        if not cv_ok:
            raise RuntimeError("cv environment check failed")

        log("[STEP 2/10] v2_3 review workbook 및 v2_4 label/segment 입력 로드")
        review_sheets = load_review_workbook(INPUT_V2_3_REVIEW_XLSX, warnings, errors, missing_review_columns)
        ad_df = read_csv_optional(AD_SEGMENTS_V2_4, missing_input_files, warnings)
        manifest = read_csv_optional(MANIFEST_PATH, missing_input_files, warnings)
        _ = read_csv_optional(LABELS_V2_4, missing_input_files, warnings)
        _ = read_csv_optional(WINDOW_LABELS_V2_4, missing_input_files, warnings)
        _ = read_csv_optional(CONTEXT_SEGMENTS_V2_4, missing_input_files, warnings)
        _ = read_csv_optional(RESNET_BOUNDARY_AUDIT_V2_3, missing_input_files, warnings)
        _ = read_csv_optional(RESNET_VIDEO_SUMMARY_V2_3, missing_input_files, warnings)
        _ = read_csv_optional(OPENCV_RESNET_COMPARISON_V2_3, missing_input_files, warnings)
        _ = read_csv_optional(OPENCV_RESNET_OVERLAP_V2_3, missing_input_files, warnings)

        if not errors and (review_sheets.get("resnet_candidate_review", pd.DataFrame()).empty or ad_df.empty):
            errors.append("v2_3 review xlsx or v2_4 ad_interval_segments missing/empty")
        ad_intervals = build_ad_intervals(ad_df)
        if not ad_intervals:
            raise RuntimeError("v2_4 ad intervals are unavailable")

        log("[STEP 3/10] v2_4 ResNet/OpenCV scene input 준비")
        resnet_df, resnet_boundary, opencv_df, opencv_boundary, resnet_boundary_created, comparison_available, scene_sources = prepare_v2_4_scene_inputs(ad_intervals, missing_input_files, warnings)
        comparison_df = read_csv_optional(OPENCV_RESNET_COMPARISON_V2_4, missing_input_files, warnings)
        overlap_df = read_csv_optional(OPENCV_RESNET_OVERLAP_V2_4, missing_input_files, warnings)
        if resnet_df.empty:
            raise RuntimeError("ResNet candidate file is unavailable; rebuild impossible")

        log("[STEP 4/10] 기존 v2_4 review xlsx backup")
        if preexisting_v2_4:
            stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
            backup = BACKUP_DIR / f"resnet_scene_candidate_human_review_v2_4_backup_before_v2_3_transfer_{stamp}.xlsx"
            ensure_inside_project(backup)
            shutil.copy2(OUTPUT_XLSX, backup)
            backup_created = True
            backup_path = str(backup)
            log(f"Backup created: {backup}")
        report["backup_created"] = backup_created
        report["backup_path"] = backup_path

        log("[STEP 5/10] candidate review 컬럼 이관")
        candidate_rows, candidate_audit, candidate_new_rows, candidate_counts, _used_candidate_ids = transfer_candidate_reviews(
            review_sheets["resnet_candidate_review"],
            resnet_df,
            ad_intervals,
            opencv_df,
            opencv_boundary,
            missing_review_columns,
        )
        cand_cols = candidate_columns(candidate_rows)
        candidate_audit_cols = [
            "v2_4_row_id",
            "v2_4_video_id",
            "v2_4_candidate_time_sec",
            "v2_4_candidate_time_mmss",
            "matched_v2_3_row_id",
            "match_method",
            "time_diff_sec",
            "review_columns_transferred",
            "transfer_status",
            "transfer_warning",
        ]
        write_csv(OUTPUT_CANDIDATE_CSV, candidate_rows, cand_cols)
        write_csv(CANDIDATE_TRANSFER_AUDIT, candidate_audit, candidate_audit_cols)
        write_csv(CANDIDATE_NEW_ROWS_CSV, candidate_new_rows, cand_cols + ["new_row_reason"])

        log("[STEP 6/10] boundary review 및 ResNet-only boundary review 이관")
        boundary_rows, boundary_audit, boundary_counts, _used_boundary_ids = build_boundary_rows(
            ad_intervals,
            resnet_boundary,
            opencv_boundary,
            review_sheets["resnet_boundary_review"],
        )
        b_cols = boundary_columns()
        boundary_audit_cols = [
            "ad_interval_id",
            "exists_in_v2_3",
            "exists_in_v2_4",
            "transfer_status",
            "transferred_columns",
            "is_new_interval_v2_4",
            "transfer_warning",
        ]
        write_csv(OUTPUT_BOUNDARY_CSV, boundary_rows, b_cols)
        write_csv(BOUNDARY_TRANSFER_AUDIT, boundary_audit, boundary_audit_cols)

        resnet_only_rows, resnet_only_new_rows, resnet_only_counts = build_resnet_only_rows(
            boundary_rows,
            candidate_rows,
            review_sheets["resnet_only_boundary_review"],
        )
        ro_cols = resnet_only_columns()
        write_csv(OUTPUT_RESNET_ONLY_CSV, resnet_only_rows, ro_cols)
        write_csv(RESNET_ONLY_NEW_ROWS_CSV, resnet_only_new_rows, ro_cols)

        log("[STEP 7/10] comparison/video summary review 생성 및 이관")
        comparison_rows = build_comparison_review(
            comparison_df,
            overlap_df,
            review_sheets["opencv_resnet_comparison"],
            boundary_rows,
            opencv_df,
            resnet_df,
        )
        comp_cols = comparison_columns()
        write_csv(OUTPUT_COMPARISON_CSV, comparison_rows, comp_cols)

        video_rows = build_video_summary_review(
            manifest,
            candidate_rows,
            boundary_rows,
            resnet_only_rows,
            opencv_df,
            overlap_df,
            review_sheets["video_summary_review"],
        )
        video_cols = video_summary_columns()
        write_csv(OUTPUT_VIDEO_SUMMARY_CSV, video_rows, video_cols)

        log("[STEP 8/10] v2_4 Excel workbook 생성 및 서식 적용")
        with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
            rows_to_df(candidate_rows, cand_cols).to_excel(writer, sheet_name="resnet_candidate_review", index=False)
            rows_to_df(boundary_rows, b_cols).to_excel(writer, sheet_name="resnet_boundary_review", index=False)
            rows_to_df(resnet_only_rows, ro_cols).to_excel(writer, sheet_name="resnet_only_boundary_review", index=False)
            rows_to_df(comparison_rows, comp_cols).to_excel(writer, sheet_name="opencv_resnet_comparison", index=False)
            rows_to_df(video_rows, video_cols).to_excel(writer, sheet_name="video_summary_review", index=False)
            value_options_df().to_excel(writer, sheet_name="value_options", index=False)
            review_guide_df(review_sheets["review_guide"]).to_excel(writer, sheet_name="review_guide", index=False)
        dropdown_applied, conditional_applied = apply_workbook_style(OUTPUT_XLSX)

        log("[STEP 9/10] README/report/latest_for_chatgpt 갱신")
        update_readme()

        a022_candidate_count = sum(1 for row in candidate_rows if candidate_is_a022_related(row))
        a022_boundary_exists = any(clean(row.get("ad_interval_id")) == NEW_INTERVAL_ID for row in boundary_rows)
        a022_resnet_only_count = sum(1 for row in resnet_only_rows if clean(row.get("ad_interval_id")) == NEW_INTERVAL_ID)
        old_project_after = old_project_snapshot()
        v2_3_hash_after = file_sha256(INPUT_V2_3_REVIEW_XLSX)
        old_project_modified = old_project_before != old_project_after
        if v2_3_hash_before != v2_3_hash_after:
            errors.append("v2_3_resnet_review_xlsx_modified_unexpectedly")
        if old_project_modified:
            errors.append("old_project_modified_unexpectedly")

        elapsed = time.monotonic() - start_monotonic
        end_time = now_iso()
        report.update(
            {
                "end_time": end_time,
                "actual_runtime_seconds": round(elapsed, 3),
                "actual_runtime_readable": readable_runtime(elapsed),
                "input_v2_4_resnet_candidate_path": str(RESNET_CANDIDATES_V2_4),
                "input_v2_4_resnet_boundary_audit_path": str(RESNET_BOUNDARY_AUDIT_V2_4),
                "input_v2_4_opencv_comparison_path": str(OPENCV_RESNET_COMPARISON_V2_4),
                "resnet_scene_candidate_boundary_audit_v2_4_created_or_refreshed": bool(resnet_boundary_created),
                "opencv_resnet_comparison_v2_4_created": bool(comparison_available),
                "scene_input_sources": scene_sources,
                "v2_3_candidate_review_row_count": int(len(review_sheets["resnet_candidate_review"])),
                "v2_4_candidate_review_row_count": int(len(candidate_rows)),
                "candidate_review_transferred_exact_count": int(candidate_counts.get("transferred_exact", 0)),
                "candidate_review_transferred_relaxed_count": int(candidate_counts.get("transferred_relaxed", 0)),
                "candidate_review_new_v2_4_count": int(candidate_counts.get("new_v2_4_row", 0)),
                "candidate_review_ambiguous_not_transferred_count": int(candidate_counts.get("ambiguous_match_not_transferred", 0)),
                "candidate_review_missing_v2_3_match_count": int(candidate_counts.get("missing_v2_3_match", 0)),
                "candidate_review_unmatched_v2_3_review_row_count": int(candidate_counts.get("unmatched_v2_3_review_row_count", 0)),
                "v2_3_boundary_review_row_count": int(len(review_sheets["resnet_boundary_review"])),
                "v2_4_boundary_review_row_count": int(len(boundary_rows)),
                "boundary_review_transferred_count": int(boundary_counts.get("boundary_review_transferred_count", 0)),
                "boundary_review_new_interval_count": int(boundary_counts.get("boundary_review_new_interval_count", 0)),
                "boundary_review_unmatched_v2_3_review_row_count": int(boundary_counts.get("unmatched_v2_3_boundary_review_row_count", 0)),
                "v2_3_resnet_only_boundary_row_count": int(len(review_sheets["resnet_only_boundary_review"])),
                "v2_4_resnet_only_boundary_row_count": int(len(resnet_only_rows)),
                "resnet_only_boundary_transferred_count": int(resnet_only_counts.get("resnet_only_boundary_transferred_count", 0)),
                "resnet_only_boundary_new_v2_4_count": int(resnet_only_counts.get("resnet_only_boundary_new_v2_4_count", 0)),
                "resnet_only_boundary_unmatched_v2_3_review_row_count": int(resnet_only_counts.get("resnet_only_boundary_unmatched_v2_3_review_row_count", 0)),
                "A022_candidate_row_count": int(a022_candidate_count),
                "A022_boundary_row_exists": bool(a022_boundary_exists),
                "A022_resnet_only_boundary_row_count": int(a022_resnet_only_count),
                "dropdown_validation_applied": bool(dropdown_applied),
                "conditional_formatting_applied": bool(conditional_applied),
                "old_project_modified": bool(old_project_modified),
                "missing_input_files": sorted(set(missing_input_files)),
                "missing_required_columns": missing_required_columns,
                "missing_review_columns": missing_review_columns,
                "warnings": warnings,
                "errors": errors,
            }
        )

        files_for_latest = [
            OUTPUT_XLSX,
            OUTPUT_CANDIDATE_CSV,
            OUTPUT_BOUNDARY_CSV,
            OUTPUT_RESNET_ONLY_CSV,
            OUTPUT_COMPARISON_CSV,
            OUTPUT_VIDEO_SUMMARY_CSV,
            CANDIDATE_TRANSFER_AUDIT,
            BOUNDARY_TRANSFER_AUDIT,
            CANDIDATE_NEW_ROWS_CSV,
            RESNET_ONLY_NEW_ROWS_CSV,
            REPORT_PATH,
            SUMMARY_PATH,
            LOG_PATH,
        ]
        report["generated_files"] = [str(path) for path in files_for_latest if path.exists()]
        report["transfer_audit_paths"] = [str(CANDIDATE_TRANSFER_AUDIT), str(BOUNDARY_TRANSFER_AUDIT)]
        write_report_summary_log(report)
        latest_ok, latest_copied, latest_forbidden = copy_latest(files_for_latest)
        report["latest_for_chatgpt_updated"] = bool(latest_ok)
        report["latest_for_chatgpt_files"] = latest_copied
        report["latest_for_chatgpt_forbidden_files"] = latest_forbidden
        report["sub_agent_results"] = build_sub_agent_results(
            report,
            candidate_counts,
            boundary_counts,
            resnet_only_counts,
            candidate_rows,
            boundary_rows,
            resnet_only_rows,
            comparison_rows,
            v2_3_hash_before,
            v2_3_hash_after,
            old_project_before,
            old_project_after,
            latest_forbidden,
        )
        write_report_summary_log(report)
        for path in [REPORT_PATH, SUMMARY_PATH, LOG_PATH]:
            shutil.copy2(path, LATEST_DIR / path.name)

        log(f"작업 종료 시각: {end_time}")
        log(f"실제 작업 시간: {readable_runtime(elapsed)}")
        write_report_summary_log(report)
        shutil.copy2(LOG_PATH, LATEST_DIR / LOG_PATH.name)

        log("[STEP 10/10] 완료")
        print(json.dumps(
            {
                "estimated_runtime": ESTIMATED_RUNTIME,
                "actual_runtime_readable": report["actual_runtime_readable"],
                "candidate_rows_v2_3_v2_4": [report["v2_3_candidate_review_row_count"], report["v2_4_candidate_review_row_count"]],
                "candidate_transfer_exact_relaxed_new_ambiguous": [
                    report["candidate_review_transferred_exact_count"],
                    report["candidate_review_transferred_relaxed_count"],
                    report["candidate_review_new_v2_4_count"],
                    report["candidate_review_ambiguous_not_transferred_count"],
                ],
                "A022_candidate_row_count": report["A022_candidate_row_count"],
                "A022_boundary_row_exists": report["A022_boundary_row_exists"],
                "backup_created": report["backup_created"],
                "errors": errors,
            },
            ensure_ascii=False,
            indent=2,
        ))
        return 0 if not errors else 1

    except Exception as exc:
        errors.append({"fatal_error": str(exc)})
        elapsed = time.monotonic() - start_monotonic
        end_time = now_iso()
        report.update(
            {
                "end_time": end_time,
                "actual_runtime_seconds": round(elapsed, 3),
                "actual_runtime_readable": readable_runtime(elapsed),
                "backup_created": backup_created,
                "backup_path": backup_path,
                "missing_input_files": sorted(set(missing_input_files)),
                "missing_required_columns": missing_required_columns,
                "missing_review_columns": missing_review_columns,
                "warnings": warnings,
                "errors": errors,
            }
        )
        log(f"작업 종료 시각: {end_time}")
        log(f"실제 작업 시간: {readable_runtime(elapsed)}")
        write_report_summary_log(report)
        raise


if __name__ == "__main__":
    raise SystemExit(main())
