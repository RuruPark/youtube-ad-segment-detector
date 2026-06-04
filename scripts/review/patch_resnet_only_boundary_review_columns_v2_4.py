#!/usr/bin/env python3
"""Patch review-input columns on the v2.4 ResNet-only boundary review sheet.

This script only edits the review workbook surface and its companion outputs.
It does not recalculate candidates, scores, embeddings, frames, OCR, audio, or models.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(".")
OLD_PROJECT_ROOT = Path("./_old_project_not_included")
CHECK_ENV_SCRIPT = PROJECT_ROOT / "scripts/utils/check_cv_environment.py"

TASK_NAME = "patch_resnet_only_boundary_review_columns_v2_4"
ESTIMATED_RUNTIME = "약 10분"
RUNTIME_ESTIMATION_REASON = (
    "기존 v2_4 ResNet review workbook을 backup하고, `resnet_only_boundary_review` sheet에 "
    "review 입력 컬럼/dropdown/formatting을 추가하는 작업 기준"
)

INPUT_WORKBOOK = PROJECT_ROOT / "data/review/resnet_scene_candidate_human_review_v2_4.xlsx"
COMPANION_CSV = PROJECT_ROOT / "data/review/resnet_only_boundary_review_v2_4.csv"
PATCH_AUDIT_CSV = PROJECT_ROOT / "data/review/resnet_only_boundary_review_columns_patch_audit_v2_4.csv"
BACKUP_DIR = PROJECT_ROOT / "data/review/backups"

REPORT_PATH = PROJECT_ROOT / "reports/patch_resnet_only_boundary_review_columns_v2_4_report.json"
SUMMARY_PATH = PROJECT_ROOT / "reports/patch_resnet_only_boundary_review_columns_v2_4_summary.md"
LOG_PATH = PROJECT_ROOT / "logs/patch_resnet_only_boundary_review_columns_v2_4_run_log.txt"
README_PATH = PROJECT_ROOT / "README.md"
LATEST_DIR = PROJECT_ROOT / "outputs/latest_for_chatgpt"
LATEST_README = LATEST_DIR / "README_latest_files.md"

SHEET_NAME = "resnet_only_boundary_review"
NEW_INTERVAL_ID = "A022"

EXPECTED_ANALYSIS_COLUMNS = [
    "ad_interval_id",
    "video_id",
    "video_title",
    "video_filename",
    "boundary_type",
    "ad_boundary_sec",
    "ad_boundary_mmss",
    "resnet_candidate_time_sec",
    "resnet_candidate_mmss",
    "distance_to_boundary_sec",
    "resnet_scene_change_score",
    "resnet_score_percentile_in_video",
    "nearest_opencv_candidate_time_sec",
    "nearest_opencv_candidate_mmss",
    "distance_to_nearest_opencv_candidate_sec",
    "opencv_hit_same_boundary_5s",
    "review_clip_start_mmss",
    "review_clip_end_mmss",
    "is_new_interval_v2_4",
]

REVIEW_COLUMNS = [
    "review_status",
    "is_actual_scene_change",
    "is_actual_ad_boundary",
    "resnet_helped",
    "should_add_to_combined_scene_evidence",
    "failure_reason_for_opencv",
    "review_note",
    "reviewer",
    "reviewed_at",
]

VALUE_OPTIONS = {
    "review_status": ["not_reviewed", "reviewed"],
    "yes_no_unclear": ["yes", "no", "unclear"],
    "failure_reason_for_opencv": [
        "subtle_semantic_change",
        "low_pixel_change",
        "threshold_missed",
        "no_actual_boundary",
        "unclear",
    ],
}

RUN_LOG: list[str] = []


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def log(message: str) -> None:
    line = f"[{now_iso()}] {message}"
    RUN_LOG.append(line)
    print(message, flush=True)


def readable_runtime(seconds: float) -> str:
    total = int(round(seconds))
    minutes, sec = divmod(total, 60)
    return f"{minutes}분 {sec}초"


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def ensure_inside_project(path: Path) -> None:
    resolved = path.resolve()
    root = PROJECT_ROOT.resolve()
    old = OLD_PROJECT_ROOT.resolve()
    if resolved != root and not str(resolved).startswith(str(root) + os.sep):
        raise RuntimeError(f"Refusing to write outside project root: {resolved}")
    if resolved == old or str(resolved).startswith(str(old) + os.sep):
        raise RuntimeError(f"Refusing to write inside old project root: {resolved}")


def ensure_dirs() -> None:
    for path in [BACKUP_DIR, REPORT_PATH.parent, LOG_PATH.parent, LATEST_DIR]:
        ensure_inside_project(path)
        path.mkdir(parents=True, exist_ok=True)


def old_project_snapshot() -> dict[str, Any]:
    if not OLD_PROJECT_ROOT.exists():
        return {"exists": False, "file_count": 0, "digest": ""}
    digest = hashlib.sha256()
    file_count = 0
    for path in sorted(p for p in OLD_PROJECT_ROOT.rglob("*") if p.is_file()):
        try:
            stat = path.stat()
        except OSError:
            continue
        rel = path.relative_to(OLD_PROJECT_ROOT).as_posix()
        digest.update(f"{rel}\t{stat.st_size}\t{stat.st_mtime_ns}\n".encode("utf-8", errors="replace"))
        file_count += 1
    return {"exists": True, "file_count": file_count, "digest": digest.hexdigest()}


def verify_cv_environment(warnings: list[Any], errors: list[Any]) -> tuple[bool, str]:
    executable = sys.executable
    cmd = ["conda", "run", "-n", "cv", "python"]
    if CHECK_ENV_SCRIPT.exists():
        cmd.append(str(CHECK_ENV_SCRIPT))
    else:
        cmd.extend(
            [
                "-c",
                "import sys; print(sys.executable); import pandas as pd; import openpyxl; "
                "print('pandas', pd.__version__); print('openpyxl', openpyxl.__version__)",
            ]
        )
    result = subprocess.run(cmd, cwd=PROJECT_ROOT, capture_output=True, text=True, check=False)
    log("cv environment command: " + " ".join(cmd))
    log("cv environment stdout: " + result.stdout.strip().replace("\n", " | "))
    if result.stderr.strip():
        log("cv environment stderr: " + result.stderr.strip().replace("\n", " | "))
    if result.returncode != 0:
        errors.append({"cv_environment_check_failed": result.returncode, "stderr": result.stderr.strip()})
        return False, executable
    if "/envs/cv/" not in executable:
        errors.append({"current_python_executable_not_in_cv": executable})
        return False, executable
    return True, executable


def headers(ws: Any) -> list[str]:
    return [clean(cell.value) for cell in ws[1]]


def header_index(ws: Any, name: str) -> int | None:
    for idx, value in enumerate(headers(ws), start=1):
        if value == name:
            return idx
    return None


def ensure_value_options_sheet(wb: Any) -> None:
    if "value_options" in wb.sheetnames:
        ws = wb["value_options"]
    else:
        ws = wb.create_sheet("value_options")
    existing_headers = headers(ws) if ws.max_row >= 1 else []
    for option_name, values in VALUE_OPTIONS.items():
        if option_name in existing_headers:
            col_idx = existing_headers.index(option_name) + 1
            existing = [clean(ws.cell(row=row_idx, column=col_idx).value) for row_idx in range(2, ws.max_row + 1)]
            row_idx = ws.max_row + 1
            for value in values:
                if value not in existing:
                    ws.cell(row=row_idx, column=col_idx).value = value
                    row_idx += 1
        else:
            col_idx = ws.max_column + 1 if ws.max_row >= 1 and any(clean(c.value) for c in ws[1]) else 1
            ws.cell(row=1, column=col_idx).value = option_name
            for row_offset, value in enumerate(values, start=2):
                ws.cell(row=row_offset, column=col_idx).value = value
            existing_headers.append(option_name)


def add_review_columns(ws: Any) -> tuple[list[str], list[str], int, int, int, int]:
    before_headers = headers(ws)
    row_count_before = max(ws.max_row - 1, 0)
    column_count_before = ws.max_column
    added: list[str] = []
    preserved = [col for col in REVIEW_COLUMNS if col in before_headers]
    for col_name in REVIEW_COLUMNS:
        if col_name in headers(ws):
            continue
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col).value = col_name
        if col_name == "review_status":
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=new_col).value = "not_reviewed"
        added.append(col_name)
    row_count_after = max(ws.max_row - 1, 0)
    column_count_after = ws.max_column
    return added, preserved, row_count_before, row_count_after, column_count_before, column_count_after


def fill_blank_review_status(ws: Any) -> int:
    idx = header_index(ws, "review_status")
    if idx is None:
        return 0
    filled = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=idx)
        if clean(cell.value) == "":
            cell.value = "not_reviewed"
            filled += 1
    return filled


def add_validation(ws: Any, column_name: str, values: list[str]) -> bool:
    idx = header_index(ws, column_name)
    if idx is None or ws.max_row < 2:
        return False
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    col = get_column_letter(idx)
    dv.add(f"{col}2:{col}{max(ws.max_row, 500)}")
    return True


def add_row_fill(ws: Any, column_name: str, expected_value: str, color: str) -> bool:
    idx = header_index(ws, column_name)
    if idx is None or ws.max_row < 2:
        return False
    col = get_column_letter(idx)
    fill = PatternFill("solid", fgColor=color)
    ws.conditional_formatting.add(
        f"A2:{get_column_letter(ws.max_column)}{ws.max_row}",
        FormulaRule(formula=[f'=${col}2="{expected_value}"'], fill=fill),
    )
    return True


def add_cell_fill(ws: Any, column_name: str, expected_value: str, color: str) -> bool:
    idx = header_index(ws, column_name)
    if idx is None or ws.max_row < 2:
        return False
    col = get_column_letter(idx)
    fill = PatternFill("solid", fgColor=color)
    ws.conditional_formatting.add(
        f"{col}2:{col}{ws.max_row}",
        FormulaRule(formula=[f'=${col}2="{expected_value}"'], fill=fill),
    )
    return True


def style_sheet(ws: Any) -> tuple[bool, bool]:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    review_header_fill = PatternFill("solid", fgColor="F4B183")
    analysis_header_fill = PatternFill("solid", fgColor="1F4E78")
    analysis_header_font = Font(color="FFFFFF", bold=True)
    review_header_font = Font(color="000000", bold=True)
    wrap_columns = {
        "video_title",
        "video_filename",
        "review_note",
        "reviewer",
        "reviewed_at",
    }

    for cell in ws[1]:
        name = clean(cell.value)
        if name in REVIEW_COLUMNS:
            cell.fill = review_header_fill
            cell.font = review_header_font
        else:
            cell.fill = analysis_header_fill
            cell.font = analysis_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        name = clean(ws.cell(row=1, column=col_idx).value)
        max_len = len(name)
        for cell in list(column_cells)[1 : min(ws.max_row, 100)]:
            max_len = max(max_len, min(len(clean(cell.value)), 60))
            if name in wrap_columns:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        width = min(max(max_len + 2, 10), 55)
        if name in wrap_columns:
            width = min(max(width, 24), 70)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.data_validations.dataValidation = []
    dropdown_ok = True
    dropdown_ok &= add_validation(ws, "review_status", VALUE_OPTIONS["review_status"])
    for column_name in [
        "is_actual_scene_change",
        "is_actual_ad_boundary",
        "resnet_helped",
        "should_add_to_combined_scene_evidence",
    ]:
        dropdown_ok &= add_validation(ws, column_name, VALUE_OPTIONS["yes_no_unclear"])
    dropdown_ok &= add_validation(ws, "failure_reason_for_opencv", VALUE_OPTIONS["failure_reason_for_opencv"])

    try:
        ws.conditional_formatting._cf_rules.clear()
    except Exception:
        pass
    conditional_ok = True
    conditional_ok &= add_row_fill(ws, "review_status", "not_reviewed", "E7E6E6")
    conditional_ok &= add_cell_fill(ws, "should_add_to_combined_scene_evidence", "yes", "E2F0D9")
    conditional_ok &= add_cell_fill(ws, "should_add_to_combined_scene_evidence", "no", "FCE4D6")
    conditional_ok &= add_row_fill(ws, "is_new_interval_v2_4", "true", "F4CCCC")
    conditional_ok &= add_row_fill(ws, "ad_interval_id", NEW_INTERVAL_ID, "F4CCCC")
    return dropdown_ok, conditional_ok


def write_sheet_to_csv(ws: Any, path: Path) -> None:
    ensure_inside_project(path)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    columns = [clean(value) for value in rows[0]]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(columns)
        for row in rows[1:]:
            writer.writerow(["" if value is None else value for value in row[: len(columns)]])


def worksheet_values(ws: Any) -> list[dict[str, Any]]:
    columns = headers(ws)
    values = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values.append({columns[idx]: row[idx] if idx < len(row) else "" for idx in range(len(columns))})
    return values


def update_readme() -> None:
    section_title = "## ResNet-only Boundary Review Columns Patch v2.4"
    section = f"""{section_title}

`resnet_only_boundary_review` sheet에 사람이 입력할 review 컬럼을 추가했다.

기존 v2_4 workbook은 backup 후 수정했으며, 사용자는 해당 sheet에서 ResNet-only boundary 후보가 실제 광고 boundary인지 검토하면 된다.

핵심 입력 컬럼:

- `is_actual_scene_change`
- `is_actual_ad_boundary`
- `resnet_helped`
- `should_add_to_combined_scene_evidence`
- `failure_reason_for_opencv`
- `review_note`
"""
    text = README_PATH.read_text(encoding="utf-8") if README_PATH.exists() else "# youtube_ad_segment_detector\n"
    if section_title in text:
        before, _, after = text.partition(section_title)
        next_idx = after.find("\n## ")
        text = before.rstrip() + "\n\n" + section.rstrip() + ("\n" + after[next_idx + 1 :].lstrip() if next_idx >= 0 else "")
    else:
        text = text.rstrip() + "\n\n" + section.rstrip()
    README_PATH.write_text(text.rstrip() + "\n", encoding="utf-8")


def latest_forbidden_files() -> list[str]:
    forbidden: list[str] = []
    raw_xlsx_names = {"new_ad_labeling.xlsx", "clean_ad_labels_v0_review.xlsx", "raw.xlsx"}
    forbidden_suffixes = {".mp4", ".mov", ".avi", ".mkv", ".jpg", ".jpeg", ".png", ".webp", ".pt", ".pth", ".ckpt", ".bin"}
    if not LATEST_DIR.exists():
        return forbidden
    for path in LATEST_DIR.rglob("*"):
        if not path.is_file():
            continue
        name = path.name.lower()
        if path.suffix.lower() in forbidden_suffixes:
            forbidden.append(str(path))
        if path.name in raw_xlsx_names:
            forbidden.append(str(path))
        if any(part in name for part in ["frame", "model", "checkpoint", "cache"]):
            forbidden.append(str(path))
    return sorted(set(forbidden))


def copy_latest(files: list[Path]) -> tuple[bool, list[str], list[str]]:
    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    copied: list[str] = []
    for src in files:
        if not src.exists():
            continue
        if src.suffix.lower() in {".mp4", ".mov", ".avi", ".mkv", ".jpg", ".jpeg", ".png", ".webp", ".pt", ".pth", ".ckpt", ".bin"}:
            raise RuntimeError(f"Refusing to copy forbidden file to latest_for_chatgpt: {src}")
        if src.name in {"new_ad_labeling.xlsx", "clean_ad_labels_v0_review.xlsx"}:
            raise RuntimeError(f"Refusing to copy raw xlsx to latest_for_chatgpt: {src}")
        dst = LATEST_DIR / src.name
        shutil.copy2(src, dst)
        copied.append(str(dst))
    LATEST_README.write_text(
        "# latest_for_chatgpt files\n\n"
        f"이번 작업명: {TASK_NAME}\n\n"
        "ResNet-only boundary review 컬럼 패치 workbook, companion CSV, audit, report/summary/log만 복사했다. "
        "mp4, 원본 raw xlsx, frame, model, checkpoint, cache 파일은 복사하지 않는다.\n",
        encoding="utf-8",
    )
    copied.append(str(LATEST_README))
    forbidden = latest_forbidden_files()
    return len(forbidden) == 0, copied, forbidden


def make_summary(report: dict[str, Any]) -> str:
    return f"""# patch_resnet_only_boundary_review_columns_v2_4 summary

## Output

- 수정 workbook: `{report.get("input_workbook_path")}`
- backup: `{report.get("backup_path")}`
- companion CSV: `{report.get("companion_csv_path")}`
- audit: `{report.get("audit_path")}`

## Patch

- row 수 before/after: {report.get("row_count_before")} / {report.get("row_count_after")}
- column 수 before/after: {report.get("column_count_before")} / {report.get("column_count_after")}
- 추가된 review 컬럼: {", ".join(report.get("added_review_columns", [])) or "없음, 이미 존재"}
- dropdown validation 적용: {report.get("dropdown_validation_applied")}
- conditional formatting 적용: {report.get("conditional_formatting_applied")}

## 사용자가 입력할 핵심 컬럼

- `is_actual_scene_change`
- `is_actual_ad_boundary`
- `resnet_helped`
- `should_add_to_combined_scene_evidence`
- `failure_reason_for_opencv`
- `review_note`

## Sub Agent Results

```json
{json.dumps(report.get("sub_agent_results", {}), ensure_ascii=False, indent=2)}
```
"""


def write_report_summary_log(report: dict[str, Any]) -> None:
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    SUMMARY_PATH.write_text(make_summary(report), encoding="utf-8")
    LOG_PATH.write_text("\n".join(RUN_LOG) + "\n", encoding="utf-8")


def data_validations_for_columns(ws: Any, column_names: list[str]) -> dict[str, bool]:
    result = {name: False for name in column_names}
    target_ranges: dict[str, str] = {}
    for name in column_names:
        idx = header_index(ws, name)
        if idx is not None:
            col = get_column_letter(idx)
            target_ranges[name] = col
    for dv in ws.data_validations.dataValidation:
        ranges_text = str(dv.sqref)
        for name, col in target_ranges.items():
            if col in ranges_text:
                result[name] = True
    return result


def build_sub_agent_results(
    sheet_found: bool,
    row_count_before: int,
    row_count_after: int,
    before_headers: list[str],
    after_headers: list[str],
    added_columns: list[str],
    preserved_review_values: bool,
    validation_status: dict[str, bool],
    backup_created: bool,
    latest_forbidden: list[str],
    old_project_modified: bool,
) -> dict[str, Any]:
    analysis_preserved = all(col in after_headers for col in before_headers)
    review_complete = all(col in after_headers for col in REVIEW_COLUMNS)
    row_preserved = row_count_before == row_count_after
    validation_ok = all(validation_status.values())
    latest_ok = len(latest_forbidden) == 0

    return {
        "sub_agent_1_sheet_structure": {
            "status": "PASS" if sheet_found and row_preserved and analysis_preserved and review_complete else "FAIL",
            "checks": {
                "sheet_found": sheet_found,
                "row_count_preserved": row_preserved,
                "existing_columns_preserved": analysis_preserved,
                "all_review_columns_present": review_complete,
                "added_columns": added_columns,
            },
            "warnings": [],
        },
        "sub_agent_2_review_validation": {
            "status": "PASS" if validation_ok and preserved_review_values else "FAIL",
            "checks": {
                "dropdown_validation_by_column": validation_status,
                "existing_review_values_preserved": preserved_review_values,
                "review_status_default_applied_when_added": "review_status" in after_headers,
            },
            "warnings": [],
        },
        "sub_agent_3_output_safety": {
            "status": "PASS" if backup_created and latest_ok and not old_project_modified else "FAIL",
            "checks": {
                "backup_created": backup_created,
                "latest_for_chatgpt_forbidden_files_absent": latest_ok,
                "old_project_unmodified": not old_project_modified,
            },
            "warnings": latest_forbidden,
        },
    }


def main() -> int:
    ensure_dirs()
    start_monotonic = time.monotonic()
    start_time = now_iso()
    warnings: list[Any] = []
    errors: list[Any] = []
    missing_input_files: list[str] = []

    log(f"작업 시작 전 예상 작업 시간: {ESTIMATED_RUNTIME}")
    log(f"예상 근거: {RUNTIME_ESTIMATION_REASON}")
    log(f"작업 시작 시각: {start_time}")

    old_project_before = old_project_snapshot()
    report: dict[str, Any] = {
        "project_root": str(PROJECT_ROOT),
        "estimated_runtime": ESTIMATED_RUNTIME,
        "runtime_estimation_reason": RUNTIME_ESTIMATION_REASON,
        "start_time": start_time,
        "end_time": "",
        "actual_runtime_seconds": 0,
        "actual_runtime_readable": "",
        "input_workbook_path": str(INPUT_WORKBOOK),
        "backup_created": False,
        "backup_path": "",
        "sheet_found": False,
        "row_count_before": 0,
        "row_count_after": 0,
        "column_count_before": 0,
        "column_count_after": 0,
        "added_review_columns": [],
        "filled_blank_review_status_count": 0,
        "existing_review_columns_preserved": [],
        "missing_expected_analysis_columns": [],
        "dropdown_validation_applied": False,
        "conditional_formatting_applied": False,
        "companion_csv_updated": False,
        "companion_csv_path": str(COMPANION_CSV),
        "audit_path": str(PATCH_AUDIT_CSV),
        "sub_agent_results": {},
        "old_project_modified": False,
        "warnings": warnings,
        "errors": errors,
    }

    try:
        log("[STEP 1/7] cv 환경 확인")
        cv_ok, executable = verify_cv_environment(warnings, errors)
        report["cv_environment_checked"] = cv_ok
        report["python_executable"] = executable
        if not cv_ok:
            raise RuntimeError("cv environment check failed")

        log("[STEP 2/7] 입력 workbook 확인 및 backup 생성")
        if not INPUT_WORKBOOK.exists():
            errors.append({"missing_input_workbook": str(INPUT_WORKBOOK)})
            raise RuntimeError(f"Input workbook not found: {INPUT_WORKBOOK}")
        stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
        backup_path = BACKUP_DIR / f"resnet_scene_candidate_human_review_v2_4_backup_before_resnet_only_columns_patch_{stamp}.xlsx"
        ensure_inside_project(backup_path)
        shutil.copy2(INPUT_WORKBOOK, backup_path)
        report["backup_created"] = True
        report["backup_path"] = str(backup_path)
        log(f"Backup created: {backup_path}")

        log("[STEP 3/7] resnet_only_boundary_review sheet 컬럼 보강")
        wb = load_workbook(INPUT_WORKBOOK)
        if SHEET_NAME not in wb.sheetnames:
            errors.append({"missing_sheet": SHEET_NAME})
            raise RuntimeError(f"Workbook does not contain sheet: {SHEET_NAME}")
        ws = wb[SHEET_NAME]
        report["sheet_found"] = True
        before_headers = headers(ws)
        before_rows = worksheet_values(ws)
        existing_review_before = [
            {col: row.get(col, "") for col in REVIEW_COLUMNS if col in before_headers}
            for row in before_rows
        ]
        added, preserved, row_before, row_after, col_before, col_after = add_review_columns(ws)
        filled_review_status = fill_blank_review_status(ws)
        after_headers = headers(ws)
        report["row_count_before"] = row_before
        report["row_count_after"] = row_after
        report["column_count_before"] = col_before
        report["column_count_after"] = col_after
        report["added_review_columns"] = added
        report["filled_blank_review_status_count"] = filled_review_status
        report["existing_review_columns_preserved"] = preserved
        missing_analysis = [col for col in EXPECTED_ANALYSIS_COLUMNS if col not in after_headers]
        report["missing_expected_analysis_columns"] = missing_analysis
        if missing_analysis:
            warnings.append({"missing_expected_analysis_columns": missing_analysis})

        log("[STEP 4/7] value_options/dropdown/formatting 적용")
        ensure_value_options_sheet(wb)
        dropdown_ok, conditional_ok = style_sheet(ws)
        report["dropdown_validation_applied"] = bool(dropdown_ok)
        report["conditional_formatting_applied"] = bool(conditional_ok)
        wb.save(INPUT_WORKBOOK)

        log("[STEP 5/7] companion CSV 및 patch audit 생성")
        wb_check = load_workbook(INPUT_WORKBOOK)
        ws_check = wb_check[SHEET_NAME]
        write_sheet_to_csv(ws_check, COMPANION_CSV)
        report["companion_csv_updated"] = True
        validation_status = data_validations_for_columns(
            ws_check,
            [
                "review_status",
                "is_actual_scene_change",
                "is_actual_ad_boundary",
                "resnet_helped",
                "should_add_to_combined_scene_evidence",
                "failure_reason_for_opencv",
            ],
        )
        after_rows = worksheet_values(ws_check)
        preserved_review_values = True
        for idx, before in enumerate(existing_review_before):
            if idx >= len(after_rows):
                preserved_review_values = False
                break
            for col, value in before.items():
                if col == "review_status" and clean(value) == "" and clean(after_rows[idx].get(col)) == "not_reviewed":
                    continue
                if clean(after_rows[idx].get(col)) != clean(value):
                    preserved_review_values = False
                    break
            if not preserved_review_values:
                break

        audit_row = {
            "sheet_name": SHEET_NAME,
            "row_count_before": row_before,
            "row_count_after": max(ws_check.max_row - 1, 0),
            "column_count_before": col_before,
            "column_count_after": ws_check.max_column,
            "added_columns": ";".join(added),
            "existing_review_columns_preserved": ";".join(preserved),
            "missing_expected_analysis_columns": ";".join(missing_analysis),
            "dropdown_validation_applied": str(dropdown_ok).lower(),
            "conditional_formatting_applied": str(conditional_ok).lower(),
            "backup_created": str(report["backup_created"]).lower(),
            "backup_path": report["backup_path"],
            "patch_status": "PASS" if row_before == max(ws_check.max_row - 1, 0) and all(col in headers(ws_check) for col in REVIEW_COLUMNS) else "FAIL",
            "patch_warning": ";".join([str(w) for w in warnings]),
        }
        with PATCH_AUDIT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(audit_row.keys()))
            writer.writeheader()
            writer.writerow(audit_row)

        log("[STEP 6/7] README/report/latest_for_chatgpt 갱신")
        update_readme()
        old_project_after = old_project_snapshot()
        old_project_modified = old_project_before != old_project_after
        if old_project_modified:
            errors.append("old_project_modified_unexpectedly")
        latest_files = [INPUT_WORKBOOK, COMPANION_CSV, PATCH_AUDIT_CSV, REPORT_PATH, SUMMARY_PATH, LOG_PATH]

        elapsed = time.monotonic() - start_monotonic
        end_time = now_iso()
        report.update(
            {
                "end_time": end_time,
                "actual_runtime_seconds": round(elapsed, 3),
                "actual_runtime_readable": readable_runtime(elapsed),
                "row_count_after": max(ws_check.max_row - 1, 0),
                "column_count_after": ws_check.max_column,
                "old_project_modified": old_project_modified,
                "warnings": warnings,
                "errors": errors,
            }
        )
        write_report_summary_log(report)
        latest_ok, latest_copied, latest_forbidden = copy_latest(latest_files)
        report["latest_for_chatgpt_updated"] = latest_ok
        report["latest_for_chatgpt_files"] = latest_copied
        report["latest_for_chatgpt_forbidden_files"] = latest_forbidden
        report["sub_agent_results"] = build_sub_agent_results(
            True,
            row_before,
            report["row_count_after"],
            before_headers,
            headers(ws_check),
            added,
            preserved_review_values,
            validation_status,
            bool(report["backup_created"]),
            latest_forbidden,
            old_project_modified,
        )
        write_report_summary_log(report)
        for path in [REPORT_PATH, SUMMARY_PATH, LOG_PATH]:
            shutil.copy2(path, LATEST_DIR / path.name)

        log(f"작업 종료 시각: {end_time}")
        log(f"실제 작업 시간: {readable_runtime(elapsed)}")
        write_report_summary_log(report)
        shutil.copy2(LOG_PATH, LATEST_DIR / LOG_PATH.name)
        log("[STEP 7/7] 완료")
        print(json.dumps(
            {
                "estimated_runtime": ESTIMATED_RUNTIME,
                "actual_runtime_readable": report["actual_runtime_readable"],
                "row_count_before_after": [report["row_count_before"], report["row_count_after"]],
                "added_review_columns": added,
                "dropdown_validation_applied": report["dropdown_validation_applied"],
                "conditional_formatting_applied": report["conditional_formatting_applied"],
                "backup_created": report["backup_created"],
                "errors": errors,
            },
            ensure_ascii=False,
            indent=2,
        ))
        return 0 if not errors else 1

    except Exception as exc:
        errors.append({"fatal_error": str(exc)})
        elapsed = time.monotonic() - start_monotonic
        end_time = now_iso()
        report.update(
            {
                "end_time": end_time,
                "actual_runtime_seconds": round(elapsed, 3),
                "actual_runtime_readable": readable_runtime(elapsed),
                "warnings": warnings,
                "errors": errors,
            }
        )
        log(f"작업 종료 시각: {end_time}")
        log(f"실제 작업 시간: {readable_runtime(elapsed)}")
        write_report_summary_log(report)
        raise


if __name__ == "__main__":
    raise SystemExit(main())
