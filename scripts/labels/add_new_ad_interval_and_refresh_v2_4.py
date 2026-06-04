#!/usr/bin/env python3
"""수동 확인 광고 구간을 추가하고 v2.4 label/audit을 갱신한다."""

from __future__ import annotations

import csv
import json
import math
import os
import re
import shutil
import subprocess
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path(".")
OLD_PROJECT_ROOT = Path("./_old_project_not_included")
CHECK_ENV_SCRIPT = PROJECT_ROOT / "scripts/utils/check_cv_environment.py"

INPUT_LABEL = PROJECT_ROOT / "data/labels/clean_ad_labels_v0_scope_refreshed_v2_2.csv"
MANIFEST_PATH = PROJECT_ROOT / "data/video_metadata/video_manifest_v2_2.csv"
WINDOW_GRID_PATH = PROJECT_ROOT / "data/windows/window_grid_5s_v2.csv"
WINDOW_LABELS_V2_3 = PROJECT_ROOT / "data/windows/window_labels_5s_v2_3.csv"
AD_SEGMENTS_V2_3 = PROJECT_ROOT / "data/segments/ad_interval_segments_v2_3.csv"
CONTEXT_SEGMENTS_V2_3 = PROJECT_ROOT / "data/segments/ad_context_segments_v2_3.csv"
COMBINED_SEGMENTS_V2_3 = PROJECT_ROOT / "data/segments/label_interval_context_segments_v2_3.csv"
REPORT_V2_3 = PROJECT_ROOT / "reports/refresh_window_segments_v2_3_after_A003_video_id_fix_report.json"

OPENCV_CANDIDATES_V2_3 = PROJECT_ROOT / "data/scene/scene_candidates_v2_3_merged_ffmpeg_fallback.csv"
OPENCV_CANDIDATES_MMSS_V2_3 = PROJECT_ROOT / "data/scene/scene_candidates_v2_3_merged_ffmpeg_fallback_mmss.csv"
OPENCV_AUDIT_V2_3 = PROJECT_ROOT / "data/scene/scene_candidate_boundary_audit_v2_3_merged_ffmpeg_fallback.csv"
RESNET_CANDIDATES_V2_3 = PROJECT_ROOT / "data/scene/resnet_scene_candidates_v2_3.csv"
RESNET_CANDIDATES_MMSS_V2_3 = PROJECT_ROOT / "data/scene/resnet_scene_candidates_v2_3_mmss.csv"
RESNET_AUDIT_V2_3 = PROJECT_ROOT / "data/scene/resnet_scene_candidate_boundary_audit_v2_3.csv"

OUTPUT_LABEL = PROJECT_ROOT / "data/labels/clean_ad_labels_v0_scope_refreshed_v2_4.csv"
NEW_LABEL_AUDIT = PROJECT_ROOT / "data/labels/new_ad_interval_addition_audit_v2_4.csv"
MAPPING_AUDIT = PROJECT_ROOT / "data/video_metadata/label_video_mapping_audit_v2_4.csv"
WINDOW_LABELS_V2_4 = PROJECT_ROOT / "data/windows/window_labels_5s_v2_4.csv"
AD_SEGMENTS_V2_4 = PROJECT_ROOT / "data/segments/ad_interval_segments_v2_4.csv"
CONTEXT_SEGMENTS_V2_4 = PROJECT_ROOT / "data/segments/ad_context_segments_v2_4.csv"
COMBINED_SEGMENTS_V2_4 = PROJECT_ROOT / "data/segments/label_interval_context_segments_v2_4.csv"
NEW_INTERVAL_INCLUSION_AUDIT = PROJECT_ROOT / "data/segments/new_ad_interval_inclusion_audit_v2_4.csv"
CLIPPED_AUDIT = PROJECT_ROOT / "data/segments/clipped_segment_audit_v2_4.csv"
OVERLAP_AUDIT = PROJECT_ROOT / "data/segments/ad_interval_overlap_audit_v2_4.csv"

OPENCV_CANDIDATES_V2_4 = PROJECT_ROOT / "data/scene/scene_candidates_v2_4_merged_ffmpeg_fallback_labelrefreshed.csv"
OPENCV_CANDIDATES_MMSS_V2_4 = PROJECT_ROOT / "data/scene/scene_candidates_v2_4_merged_ffmpeg_fallback_labelrefreshed_mmss.csv"
OPENCV_AUDIT_V2_4 = PROJECT_ROOT / "data/scene/scene_candidate_boundary_audit_v2_4_merged_ffmpeg_fallback.csv"
RESNET_CANDIDATES_V2_4 = PROJECT_ROOT / "data/scene/resnet_scene_candidates_v2_4_labelrefreshed.csv"
RESNET_CANDIDATES_MMSS_V2_4 = PROJECT_ROOT / "data/scene/resnet_scene_candidates_v2_4_labelrefreshed_mmss.csv"
RESNET_AUDIT_V2_4 = PROJECT_ROOT / "data/scene/resnet_scene_candidate_boundary_audit_v2_4.csv"
COMPARISON_V2_4 = PROJECT_ROOT / "data/scene/opencv_vs_resnet_scene_comparison_v2_4.csv"
OVERLAP_SCENE_V2_4 = PROJECT_ROOT / "data/scene/opencv_resnet_candidate_overlap_v2_4.csv"

REVIEW_DIR = PROJECT_ROOT / "data/review"
OPENCV_REVIEW_XLSX = REVIEW_DIR / "scene_candidate_human_review_v2_4.xlsx"
OPENCV_REVIEW_CANDIDATE_CSV = REVIEW_DIR / "scene_candidate_human_review_candidate_sheet_v2_4.csv"
OPENCV_REVIEW_BOUNDARY_CSV = REVIEW_DIR / "scene_candidate_human_review_boundary_sheet_v2_4.csv"
RESNET_REVIEW_XLSX = REVIEW_DIR / "resnet_scene_candidate_human_review_v2_4.xlsx"
RESNET_REVIEW_CANDIDATE_CSV = REVIEW_DIR / "resnet_scene_candidate_human_review_candidate_sheet_v2_4.csv"
RESNET_REVIEW_BOUNDARY_CSV = REVIEW_DIR / "resnet_scene_candidate_human_review_boundary_sheet_v2_4.csv"
RESNET_ONLY_BOUNDARY_REVIEW = REVIEW_DIR / "resnet_only_boundary_review_v2_4.csv"
COMPARISON_REVIEW = REVIEW_DIR / "opencv_resnet_scene_comparison_review_v2_4.csv"

REPORT_PATH = PROJECT_ROOT / "reports/add_new_ad_interval_and_refresh_v2_4_report.json"
SUMMARY_PATH = PROJECT_ROOT / "reports/add_new_ad_interval_and_refresh_v2_4_summary.md"
LOG_PATH = PROJECT_ROOT / "logs/add_new_ad_interval_and_refresh_v2_4_run_log.txt"
SCRIPT_PATH = PROJECT_ROOT / "scripts/labels/add_new_ad_interval_and_refresh_v2_4.py"
README_PATH = PROJECT_ROOT / "README.md"
LATEST_DIR = PROJECT_ROOT / "outputs/latest_for_chatgpt"
LATEST_README = LATEST_DIR / "README_latest_files.md"

ESTIMATED_RUNTIME = "약 25분"
RUNTIME_ESTIMATION_REASON = "라벨 1건 추가, window overlap 재계산, segment 재생성, 기존 scene candidate 기준 boundary audit 재계산, review 파일 갱신 범위 기준"
CONTEXT_WINDOW_SEC = 10.0
NEW_SOURCE = "user_manual_addition"
NEW_REFRESH_SOURCE = "add_new_interval_video12_19m20s_20m06s"
NEW_VIDEO_ID = "12"
NEW_VIDEO_TITLE = "eng) 신입사원의 일주일 🥸 | 친구 없는 사람의 일상, 상경 한달차 | 출근룩 찍기 | 5:30 퇴근 후 열심히 샌드위치 공장돌리기"
NEW_START_SEC = 1160.0
NEW_END_SEC = 1206.0
NEW_DURATION_SEC = 46.0
RUN_LOG: list[str] = []


def log(message: str) -> None:
    stamp = datetime.now().astimezone().isoformat(timespec="seconds")
    RUN_LOG.append(f"[{stamp}] {message}")
    print(message, flush=True)


def step(number: int, message: str) -> None:
    log(f"[STEP {number}/12] {message}")


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_float(value: Any) -> float | None:
    try:
        if value is None or pd.isna(value):
            return None
        text = str(value).strip()
        if not text:
            return None
        return float(text)
    except Exception:
        return None


def fmt(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        rounded = round(value, 6)
        return int(round(rounded)) if abs(rounded - round(rounded)) < 1e-9 else rounded
    return value


def bool_text(value: bool) -> str:
    return "true" if bool(value) else "false"


def is_true(value: Any) -> bool:
    return clean(value).lower() in {"true", "1", "yes", "y"}


def mmss_floor(seconds: Any) -> str:
    sec = to_float(seconds)
    if sec is None:
        return ""
    total = max(0, int(math.floor(sec)))
    return f"{total // 60:02d}분 {total % 60:02d}초"


def mmss_round(seconds: Any) -> str:
    sec = to_float(seconds)
    if sec is None:
        return ""
    total = max(0, int(round(sec)))
    return f"{total // 60:02d}분 {total % 60:02d}초"


def readable(seconds: float) -> str:
    total = int(round(seconds))
    return f"{total // 60}분 {total % 60}초"


def ensure_inside_project(path: Path) -> None:
    resolved = path.resolve()
    root = PROJECT_ROOT.resolve()
    old = OLD_PROJECT_ROOT.resolve()
    if resolved != root and not str(resolved).startswith(str(root) + os.sep):
        raise RuntimeError(f"Refusing to write outside project root: {resolved}")
    if resolved == old or str(resolved).startswith(str(old) + os.sep):
        raise RuntimeError(f"Refusing to write inside old project root: {resolved}")


def ensure_dirs() -> None:
    for path in [
        PROJECT_ROOT / "data/labels",
        PROJECT_ROOT / "data/video_metadata",
        PROJECT_ROOT / "data/windows",
        PROJECT_ROOT / "data/segments",
        PROJECT_ROOT / "data/scene",
        REVIEW_DIR,
        PROJECT_ROOT / "reports",
        PROJECT_ROOT / "logs",
        PROJECT_ROOT / "scripts/labels",
        LATEST_DIR,
    ]:
        ensure_inside_project(path)
        path.mkdir(parents=True, exist_ok=True)


def read_csv(path: Path, required: bool, warnings: list[Any], errors: list[Any]) -> pd.DataFrame:
    if not path.exists():
        payload = {"missing_file": str(path)}
        if required:
            errors.append(payload)
        else:
            warnings.append(payload)
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str, keep_default_na=False)


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ensure_inside_project(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: fmt(row.get(col, "")) for col in columns})


def write_df(path: Path, df: pd.DataFrame) -> None:
    ensure_inside_project(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")


def distribution(values: Any) -> dict[str, int]:
    if values is None:
        return {}
    return dict(sorted(Counter(clean(v) for v in list(values)).items()))


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def old_project_snapshot() -> dict[str, Any]:
    if not OLD_PROJECT_ROOT.exists():
        return {"exists": False}
    try:
        stat = OLD_PROJECT_ROOT.stat()
        return {"exists": True, "mtime_ns": stat.st_mtime_ns, "size": stat.st_size}
    except Exception as exc:
        return {"exists": True, "error": str(exc)}


def verify_cv_environment(warnings: list[Any]) -> tuple[bool, str]:
    executable = sys.executable
    in_cv = "/envs/cv/" in executable or executable.endswith("/envs/cv/bin/python") or executable.endswith("/envs/cv/bin/python3.10")
    if CHECK_ENV_SCRIPT.exists():
        cmd = ["conda", "run", "-n", "cv", "python", str(CHECK_ENV_SCRIPT)]
        log("Command: " + " ".join(cmd))
        result = subprocess.run(cmd, cwd=PROJECT_ROOT, capture_output=True, text=True, check=False)
        log("cv check stdout: " + result.stdout.strip().replace("\n", " | "))
        if result.stderr.strip():
            log("cv check stderr: " + result.stderr.strip().replace("\n", " | "))
        if result.returncode != 0:
            warnings.append("cv_environment_check_failed")
            return False, executable
    else:
        cmd = ["conda", "run", "-n", "cv", "python", "-c", "import sys; print(sys.executable); import pandas as pd; import openpyxl; print('pandas', pd.__version__); print('openpyxl', openpyxl.__version__)"]
        log("Command: " + " ".join(cmd))
        result = subprocess.run(cmd, cwd=PROJECT_ROOT, capture_output=True, text=True, check=False)
        log("cv fallback stdout: " + result.stdout.strip().replace("\n", " | "))
        if result.stderr.strip():
            log("cv fallback stderr: " + result.stderr.strip().replace("\n", " | "))
        if result.returncode != 0:
            warnings.append("cv_environment_fallback_check_failed")
            return False, executable
    if not in_cv:
        warnings.append({"current_python_executable_not_in_cv": executable})
    return in_cv, executable


def label_valid_mask(df: pd.DataFrame) -> pd.Series:
    if "label_valid" in df.columns:
        return df["label_valid"].astype(str).str.strip().str.lower().isin(["true", "1", "yes", "y"])
    starts = df["ad_start_sec"].map(to_float)
    ends = df["ad_end_sec"].map(to_float)
    return starts.notna() & ends.notna() & (starts < ends)


def valid_label_rows(df: pd.DataFrame) -> pd.DataFrame:
    labels = df.copy()
    labels["_source_label_row_index"] = labels.index + 2
    labels["_video_id_text"] = labels["video_id"].map(clean)
    labels["_video_title_text"] = labels["video_title"].map(clean)
    labels["_ad_start_sec_num"] = labels["ad_start_sec"].map(to_float)
    labels["_ad_end_sec_num"] = labels["ad_end_sec"].map(to_float)
    labels["_label_valid_bool"] = label_valid_mask(labels)
    return labels[
        labels["_label_valid_bool"]
        & labels["_video_id_text"].ne("")
        & labels["_video_title_text"].ne("")
        & labels["_ad_start_sec_num"].notna()
        & labels["_ad_end_sec_num"].notna()
        & (labels["_ad_start_sec_num"] < labels["_ad_end_sec_num"])
    ].copy()


def next_ad_interval_id(labels: pd.DataFrame) -> str:
    max_num = 0
    if "ad_interval_id" in labels.columns:
        for value in labels["ad_interval_id"]:
            match = re.match(r"A(\d+)$", clean(value))
            if match:
                max_num = max(max_num, int(match.group(1)))
    return f"A{max_num + 1:03d}"


def manifest_lookup(manifest: pd.DataFrame) -> tuple[dict[tuple[str, str], dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    by_id: dict[str, dict[str, Any]] = {}
    by_filename: dict[str, dict[str, Any]] = {}
    for _, row in manifest.iterrows():
        mapped_id = clean(row.get("label_mapping_video_id")) or clean(row.get("video_id"))
        mapped_title = clean(row.get("label_mapping_video_title")) or clean(row.get("matched_label_video_title")) or clean(row.get("video_title"))
        payload = {
            "mapped_video_id": mapped_id,
            "mapped_video_title": mapped_title,
            "matched_manifest_video_id": clean(row.get("video_id")),
            "matched_video_filename": clean(row.get("video_filename")),
            "matched_video_path": clean(row.get("video_path")),
            "file_stem": clean(row.get("file_stem")),
            "title_match_status": clean(row.get("title_match_status")),
            "label_mapping_status": clean(row.get("label_mapping_status")) or clean(row.get("title_match_status")),
            "label_mapping_warning": clean(row.get("label_mapping_warning")),
            "video_duration_sec": to_float(row.get("duration_sec")),
            "fps": to_float(row.get("fps")),
            "frame_count": to_float(row.get("frame_count")),
            "width": to_float(row.get("width")),
            "height": to_float(row.get("height")),
        }
        if mapped_id and mapped_title:
            by_key[(mapped_id, mapped_title)] = payload
            by_id.setdefault(mapped_id, payload)
        if payload["matched_video_filename"]:
            by_filename[payload["matched_video_filename"]] = payload
    return by_key, by_id, by_filename


def add_new_label_row(labels: pd.DataFrame, manifest_by_id: dict[str, dict[str, Any]]) -> tuple[pd.DataFrame, dict[str, Any]]:
    out = labels.copy()
    for col in ["ad_start_time_label", "ad_end_time_label", "source"]:
        if col not in out.columns:
            out[col] = ""
    new_id = next_ad_interval_id(out)
    mapping = manifest_by_id.get(NEW_VIDEO_ID, {})
    existing_ids = set(out["ad_interval_id"].map(clean))
    overlaps: list[str] = []
    for _, row in valid_label_rows(out).iterrows():
        if clean(row.get("video_id")) != NEW_VIDEO_ID:
            continue
        start = float(row["_ad_start_sec_num"])
        end = float(row["_ad_end_sec_num"])
        overlap = max(0.0, min(end, NEW_END_SEC) - max(start, NEW_START_SEC))
        if overlap > 0:
            overlaps.append(clean(row.get("ad_interval_id")))
    row = {col: "" for col in out.columns}
    row.update(
        {
            "ad_interval_id": new_id,
            "video_id": NEW_VIDEO_ID,
            "video_title": NEW_VIDEO_TITLE,
            "ad_start_min": 19,
            "ad_start_sec_part": 20,
            "ad_end_min": 20,
            "ad_end_sec_part": 6,
            "ad_start_time_label": "19분 20초",
            "ad_end_time_label": "20분 06초",
            "ad_start_sec": int(NEW_START_SEC),
            "ad_end_sec": int(NEW_END_SEC),
            "ad_duration_sec": int(NEW_DURATION_SEC),
            "is_abrupt_transition_ad": "yes",
            "label_valid": "true",
            "invalid_reason": "",
            "source_file": NEW_SOURCE,
            "source_sheet": "",
            "source_row_number": "",
            "source": NEW_SOURCE,
            "is_abrupt_transition_ad_original": "yes",
            "is_abrupt_transition_ad_refreshed": "yes",
            "scope_value_source": NEW_SOURCE,
            "scope_value_valid": "true",
            "scope_warning": "",
            "label_refresh_version": "v2_4",
        }
    )
    out = pd.concat([out, pd.DataFrame([row], columns=out.columns)], ignore_index=True)
    duration = mapping.get("video_duration_sec")
    warnings: list[str] = []
    if new_id in existing_ids:
        warnings.append("new_ad_interval_id_duplicate")
    if NEW_END_SEC <= NEW_START_SEC:
        warnings.append("end_not_after_start")
    if NEW_DURATION_SEC != NEW_END_SEC - NEW_START_SEC:
        warnings.append("duration_mismatch")
    if not mapping:
        warnings.append("video_id_12_not_found_in_manifest")
    if duration is not None and NEW_END_SEC > float(duration):
        warnings.append("ad_end_sec_exceeds_video_duration")
    if overlaps:
        warnings.append("overlaps_existing_interval")
    audit = {
        "new_ad_interval_id": new_id,
        "video_id": NEW_VIDEO_ID,
        "video_title": NEW_VIDEO_TITLE,
        "ad_start_sec": NEW_START_SEC,
        "ad_end_sec": NEW_END_SEC,
        "ad_duration_sec": NEW_DURATION_SEC,
        "manifest_video_found": bool_text(bool(mapping)),
        "video_duration_sec": duration if duration is not None else "",
        "end_within_video_duration": bool_text(duration is not None and NEW_END_SEC <= float(duration)),
        "overlaps_existing_interval": bool_text(bool(overlaps)),
        "overlap_with_ad_interval_ids": ";".join(overlaps),
        "add_status": "added" if not warnings else "added_with_warning",
        "add_warning": ";".join(warnings),
    }
    return out, audit


def create_mapping_audit(labels: pd.DataFrame, mapping_key: dict[tuple[str, str], dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for _, row in valid_label_rows(labels).iterrows():
        key = (clean(row.get("video_id")), clean(row.get("video_title")))
        mapping = mapping_key.get(key)
        mapped = mapping is not None
        rows.append(
            {
                "ad_interval_id": clean(row.get("ad_interval_id")),
                "label_video_id": key[0],
                "label_video_title": key[1],
                "matched_manifest_video_id": mapping.get("matched_manifest_video_id", "") if mapping else "",
                "matched_video_filename": mapping.get("matched_video_filename", "") if mapping else "",
                "matched_video_path": mapping.get("matched_video_path", "") if mapping else "",
                "video_duration_sec": mapping.get("video_duration_sec", "") if mapping else "",
                "mapping_status": "mapped" if mapped else "unmapped",
                "mapping_warning": "" if mapped else "no manifest row with same video_id and video_title",
            }
        )
    return rows


def labels_grouped(labels: pd.DataFrame, mapping_key: dict[tuple[str, str], dict[str, Any]]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for _, row in valid_label_rows(labels).iterrows():
        key = (clean(row.get("video_id")), clean(row.get("video_title")))
        if key not in mapping_key:
            continue
        grouped[key].append(
            {
                "ad_interval_id": clean(row.get("ad_interval_id")),
                "ad_start_sec": float(row["_ad_start_sec_num"]),
                "ad_end_sec": float(row["_ad_end_sec_num"]),
                "scope": clean(row.get("is_abrupt_transition_ad_refreshed")),
                "scope_value_source": clean(row.get("scope_value_source")),
                "source_label_row_index": int(row["_source_label_row_index"]),
            }
        )
    return grouped


def window_scope(overlaps: list[dict[str, Any]]) -> str:
    if not overlaps:
        return "outside_ad"
    qualifying = sorted({item["scope"] for item in overlaps if item["overlap_ratio"] >= 0.5})
    if len(qualifying) != 1:
        return "mixed_or_boundary"
    value = qualifying[0]
    if value == "yes":
        return "abrupt_ad"
    if value == "no":
        return "out_of_scope_ad"
    if value == "unclear":
        return "uncertain_ad"
    if value == "not_reviewed":
        return "not_reviewed_ad"
    return "mixed_or_boundary"


def create_window_labels(window_grid: pd.DataFrame, labels: pd.DataFrame, mapping_by_filename: dict[str, dict[str, Any]], mapping_key: dict[tuple[str, str], dict[str, Any]]) -> pd.DataFrame:
    grouped = labels_grouped(labels, mapping_key)
    rows: list[dict[str, Any]] = []
    for _, window in window_grid.iterrows():
        filename = clean(window.get("video_filename"))
        mapping = mapping_by_filename.get(filename, {})
        mapped_id = mapping.get("mapped_video_id", clean(window.get("video_id")))
        mapped_title = mapping.get("mapped_video_title", clean(window.get("video_title")))
        win_start = float(window.get("window_start_sec"))
        win_end = float(window.get("window_end_sec"))
        win_duration = max(float(window.get("window_duration_sec")), 1e-9)
        overlaps: list[dict[str, Any]] = []
        for label in grouped.get((mapped_id, mapped_title), []):
            overlap_sec = max(0.0, min(win_end, label["ad_end_sec"]) - max(win_start, label["ad_start_sec"]))
            if overlap_sec > 0:
                overlaps.append({**label, "overlap_sec": overlap_sec, "overlap_ratio": overlap_sec / win_duration})
        overlap_total = min(sum(item["overlap_sec"] for item in overlaps), win_duration)
        max_overlap = max(overlaps, key=lambda item: item["overlap_sec"], default=None)
        scopes = sorted({item["scope"] for item in overlaps})
        base = window.to_dict()
        source_window_id = clean(base.get("window_id"))
        if mapped_id:
            base["video_id"] = mapped_id
            base["video_title"] = mapped_title
            base["window_id"] = f"{mapped_id}_W{int(float(base['window_index'])):06d}"
        base.update(
            {
                "source_window_id_v2": source_window_id,
                "label_mapping_status": mapping.get("label_mapping_status", "mapping_missing"),
                "label_mapping_warning": mapping.get("label_mapping_warning", ""),
                "label_refresh_version": "v2_4",
                "window_label_refresh_source": NEW_REFRESH_SOURCE,
                "overlap_any_ad": bool_text(bool(overlaps)),
                "overlap_ad_sec": overlap_total,
                "overlap_ad_ratio": overlap_total / win_duration,
                "matched_ad_interval_ids": ";".join(item["ad_interval_id"] for item in overlaps),
                "max_overlap_ad_interval_id": max_overlap["ad_interval_id"] if max_overlap else "",
                "max_overlap_ad_sec": max_overlap["overlap_sec"] if max_overlap else 0,
                "max_overlap_ad_ratio": max_overlap["overlap_ratio"] if max_overlap else 0,
                "matched_abrupt_scope_values": ";".join(scopes),
                "window_label_scope": window_scope(overlaps),
            }
        )
        rows.append(base)
    return pd.DataFrame(rows)


def create_segment_row(label: pd.Series, mapping: dict[str, Any], segment_id: str, segment_type: str, boundary_role: str, original_start: float, original_end: float, context_window_sec: Any) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    start = original_start
    end = original_end
    duration = mapping.get("video_duration_sec")
    segment_clipped = False
    clipping_reason = ""
    clipped_audit: dict[str, Any] | None = None
    if duration is not None:
        clipped_start = max(0.0, start)
        clipped_end = min(float(duration), end)
        segment_clipped = clipped_start != start or clipped_end != end
        start, end = clipped_start, clipped_end
        if segment_clipped:
            clipping_reason = "segment exceeded video duration or started before zero"
            clipped_audit = {
                "source_version": "v2_4",
                "segment_id": segment_id,
                "segment_type": segment_type,
                "video_id": clean(label.get("video_id")),
                "video_title": clean(label.get("video_title")),
                "video_filename": mapping.get("matched_video_filename", ""),
                "ad_interval_id": clean(label.get("ad_interval_id")),
                "original_segment_start_sec": original_start,
                "original_segment_end_sec": original_end,
                "clipped_segment_start_sec": start,
                "clipped_segment_end_sec": end,
                "video_duration_sec": duration,
                "clipping_applied": "true",
                "clipping_reason": clipping_reason,
                "needs_manual_review": "true",
            }
    if start >= end:
        return None, clipped_audit
    row = {
        "segment_id": segment_id,
        "segment_type": segment_type,
        "boundary_role": boundary_role,
        "video_id": clean(label.get("video_id")),
        "video_title": clean(label.get("video_title")),
        "video_filename": mapping.get("matched_video_filename", ""),
        "video_path": mapping.get("matched_video_path", ""),
        "file_stem": mapping.get("file_stem", ""),
        "ad_interval_id": clean(label.get("ad_interval_id")),
        "source_label_row_index": int(label.get("_source_label_row_index")),
        "segment_start_sec": start,
        "segment_end_sec": end,
        "segment_duration_sec": end - start,
        "context_window_sec": context_window_sec,
        "ad_start_sec": float(label.get("_ad_start_sec_num")),
        "ad_end_sec": float(label.get("_ad_end_sec_num")),
        "is_abrupt_transition_ad_original": clean(label.get("is_abrupt_transition_ad_original")),
        "is_abrupt_transition_ad_refreshed": clean(label.get("is_abrupt_transition_ad_refreshed")),
        "scope_value_source": clean(label.get("scope_value_source")),
        "label_valid": "true",
        "video_duration_sec": duration,
        "label_mapping_status": mapping.get("label_mapping_status", ""),
        "segment_refresh_version": "v2_4",
        "segment_clipped": bool_text(segment_clipped),
        "clipping_reason": clipping_reason,
        "segment_valid": "true",
    }
    return row, clipped_audit


def create_segments(labels: pd.DataFrame, mapping_key: dict[tuple[str, str], dict[str, Any]], warnings: list[Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], int]:
    ad_rows: list[dict[str, Any]] = []
    context_rows: list[dict[str, Any]] = []
    clipped_rows: list[dict[str, Any]] = []
    empty_count = 0
    for _, label in valid_label_rows(labels).iterrows():
        key = (clean(label.get("video_id")), clean(label.get("video_title")))
        mapping = mapping_key.get(key)
        if not mapping:
            continue
        video_id = clean(label.get("video_id"))
        interval_id = clean(label.get("ad_interval_id"))
        ad_start = float(label["_ad_start_sec_num"])
        ad_end = float(label["_ad_end_sec_num"])
        specs = [
            (f"{video_id}_{interval_id}_AD", "ad_interval", "", ad_start, ad_end, ""),
            (f"{video_id}_{interval_id}_PRE10", "pre_ad_start_10s", "before_ad_start", max(0.0, ad_start - CONTEXT_WINDOW_SEC), ad_start, CONTEXT_WINDOW_SEC),
            (f"{video_id}_{interval_id}_POST10", "post_ad_end_10s", "after_ad_end", ad_end, ad_end + CONTEXT_WINDOW_SEC, CONTEXT_WINDOW_SEC),
        ]
        for segment_id, segment_type, boundary_role, start, end, context_sec in specs:
            row, audit = create_segment_row(label, mapping, segment_id, segment_type, boundary_role, start, end, context_sec)
            if audit:
                clipped_rows.append(audit)
            if row is None:
                empty_count += 1
                warnings.append({"empty_segment_skipped": segment_id, "reason": "segment_start_sec_gte_segment_end_sec"})
                continue
            if segment_type == "ad_interval":
                ad_rows.append(row)
            else:
                context_rows.append(row)
    return ad_rows, context_rows, ad_rows + context_rows, clipped_rows, empty_count


def create_overlap_audit(labels: pd.DataFrame, mapping_key: dict[tuple[str, str], dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    rows: list[dict[str, Any]] = []
    grouped: dict[tuple[str, str], list[pd.Series]] = defaultdict(list)
    for _, row in valid_label_rows(labels).iterrows():
        key = (clean(row.get("video_id")), clean(row.get("video_title")))
        if key in mapping_key:
            grouped[key].append(row)
    overlap_count = 0
    for (video_id, video_title), items in grouped.items():
        if len(items) < 2:
            continue
        items = sorted(items, key=lambda item: (float(item["_ad_start_sec_num"]), clean(item.get("ad_interval_id"))))
        for idx, first in enumerate(items):
            for second in items[idx + 1 :]:
                start1 = float(first["_ad_start_sec_num"])
                end1 = float(first["_ad_end_sec_num"])
                start2 = float(second["_ad_start_sec_num"])
                end2 = float(second["_ad_end_sec_num"])
                overlap = max(0.0, min(end1, end2) - max(start1, start2))
                min_duration = max(min(end1 - start1, end2 - start2), 1e-9)
                if overlap > 0:
                    overlap_count += 1
                rows.append(
                    {
                        "video_id": video_id,
                        "video_title": video_title,
                        "ad_interval_id_1": clean(first.get("ad_interval_id")),
                        "ad_interval_id_2": clean(second.get("ad_interval_id")),
                        "interval_1_start_sec": start1,
                        "interval_1_end_sec": end1,
                        "interval_2_start_sec": start2,
                        "interval_2_end_sec": end2,
                        "overlap_sec": overlap,
                        "overlap_ratio_min_duration": overlap / min_duration,
                        "overlap_warning": "overlap_detected" if overlap > 0 else "no_overlap",
                    }
                )
    return rows, overlap_count


def nearest_window_id(windows: pd.DataFrame, video_id: str, timestamp: float) -> str:
    if windows.empty:
        return ""
    subset = windows[windows["video_id"].map(clean).eq(video_id)]
    if subset.empty:
        return ""
    starts = subset["window_start_sec"].astype(float)
    ends = subset["window_end_sec"].astype(float)
    hit = subset[(starts <= timestamp) & (timestamp < ends)]
    if hit.empty:
        hit = subset[(starts <= timestamp) & (timestamp <= ends)]
    return clean(hit.iloc[0].get("window_id")) if not hit.empty else ""


def boundary_context(ad_segments: pd.DataFrame, video_id: str, timestamp: float) -> dict[str, Any]:
    subset = ad_segments[ad_segments["video_id"].map(clean).eq(video_id)] if not ad_segments.empty else pd.DataFrame()
    best: tuple[float, str, float, pd.Series] | None = None
    start_dists: list[float] = []
    end_dists: list[float] = []
    for _, row in subset.iterrows():
        start = to_float(row.get("ad_start_sec")) or to_float(row.get("segment_start_sec"))
        end = to_float(row.get("ad_end_sec")) or to_float(row.get("segment_end_sec"))
        for btype, sec in [("ad_start", start), ("ad_end", end)]:
            if sec is None:
                continue
            dist = abs(timestamp - sec)
            if btype == "ad_start":
                start_dists.append(dist)
            else:
                end_dists.append(dist)
            if best is None or dist < best[0]:
                best = (dist, btype, sec, row)
    if best is None:
        return {
            "nearest_ad_interval_id": "",
            "nearest_ad_boundary_type": "",
            "nearest_ad_boundary_sec": "",
            "nearest_ad_boundary_mmss": "",
            "distance_to_nearest_ad_boundary_sec": "",
            "is_near_ad_start_3s": "false",
            "is_near_ad_start_5s": "false",
            "is_near_ad_end_3s": "false",
            "is_near_ad_end_5s": "false",
            "is_near_any_ad_boundary_5s": "false",
        }
    min_start = min(start_dists) if start_dists else math.inf
    min_end = min(end_dists) if end_dists else math.inf
    return {
        "nearest_ad_interval_id": clean(best[3].get("ad_interval_id")),
        "nearest_ad_boundary_type": best[1],
        "nearest_ad_boundary_sec": best[2],
        "nearest_ad_boundary_mmss": mmss_floor(best[2]),
        "distance_to_nearest_ad_boundary_sec": best[0],
        "is_near_ad_start_3s": bool_text(min_start <= 3),
        "is_near_ad_start_5s": bool_text(min_start <= 5),
        "is_near_ad_end_3s": bool_text(min_end <= 3),
        "is_near_ad_end_5s": bool_text(min_end <= 5),
        "is_near_any_ad_boundary_5s": bool_text(min(min_start, min_end) <= 5),
    }


def refresh_candidates(candidate_df: pd.DataFrame, ad_segments: pd.DataFrame, windows: pd.DataFrame, source_name: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if candidate_df.empty:
        return rows
    df = candidate_df.copy()
    for _, row in df.iterrows():
        t = to_float(row.get("candidate_time_sec"))
        if t is None:
            continue
        video_id = clean(row.get("video_id"))
        payload = row.to_dict()
        payload["candidate_time_sec"] = t
        payload["candidate_time_mmss"] = clean(payload.get("candidate_time_mmss")) or mmss_floor(t)
        payload["candidate_time_mmss_floor"] = mmss_floor(t)
        payload["candidate_time_mmss_round"] = mmss_round(t)
        payload["nearest_window_id"] = nearest_window_id(windows, video_id, t) or clean(payload.get("nearest_window_id"))
        payload.update(boundary_context(ad_segments, video_id, t))
        payload["label_refresh_version"] = "v2_4"
        payload["label_refresh_source"] = NEW_REFRESH_SOURCE
        payload["candidate_source_for_audit"] = source_name
        rows.append(payload)
    return rows


def count_between(candidates: pd.DataFrame, video_id: str, start: float, end: float) -> int:
    if candidates.empty or "candidate_time_sec" not in candidates.columns:
        return 0
    subset = candidates[candidates["video_id"].map(clean).eq(video_id)]
    if subset.empty:
        return 0
    times = subset["candidate_time_sec"].astype(float)
    return int(((times >= start) & (times <= end)).sum())


def nearest_candidate(candidates: pd.DataFrame, video_id: str, boundary_sec: float) -> tuple[str, str, str]:
    if candidates.empty or "candidate_time_sec" not in candidates.columns:
        return "", "", ""
    subset = candidates[candidates["video_id"].map(clean).eq(video_id)].copy()
    if subset.empty:
        return "", "", ""
    subset["_distance"] = (subset["candidate_time_sec"].astype(float) - boundary_sec).abs()
    score_col = "scene_change_score" if "scene_change_score" in subset.columns else "_distance"
    ascending = [True, False] if score_col == "scene_change_score" else [True, True]
    row = subset.sort_values(["_distance", score_col], ascending=ascending).iloc[0]
    sec = float(row["candidate_time_sec"])
    return str(fmt(sec)), mmss_floor(sec), str(fmt(float(row["_distance"])))


def boundary_audit(candidates: list[dict[str, Any]], ad_segments: pd.DataFrame, all_segments: pd.DataFrame) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    cands = pd.DataFrame(candidates)
    rows: list[dict[str, Any]] = []
    sh3 = sh5 = eh3 = eh5 = both5 = 0
    for _, ad in ad_segments.iterrows():
        video_id = clean(ad.get("video_id"))
        ad_id = clean(ad.get("ad_interval_id"))
        start = to_float(ad.get("ad_start_sec")) or to_float(ad.get("segment_start_sec")) or 0.0
        end = to_float(ad.get("ad_end_sec")) or to_float(ad.get("segment_end_sec")) or 0.0
        subset = cands[cands["video_id"].map(clean).eq(video_id)].copy() if not cands.empty else pd.DataFrame()
        times = subset["candidate_time_sec"].astype(float) if not subset.empty else pd.Series(dtype=float)
        start3 = int(((times - start).abs() <= 3).sum())
        start5 = int(((times - start).abs() <= 5).sum())
        end3 = int(((times - end).abs() <= 3).sum())
        end5 = int(((times - end).abs() <= 5).sum())
        sh3 += int(start3 > 0)
        sh5 += int(start5 > 0)
        eh3 += int(end3 > 0)
        eh5 += int(end5 > 0)
        both5 += int(start5 > 0 and end5 > 0)
        ns_sec, ns_mmss, ns_dist = nearest_candidate(cands, video_id, start)
        ne_sec, ne_mmss, ne_dist = nearest_candidate(cands, video_id, end)
        context = all_segments[(all_segments["video_id"].map(clean).eq(video_id)) & (all_segments["ad_interval_id"].map(clean).eq(ad_id))] if not all_segments.empty else pd.DataFrame()
        pre = context[context["segment_type"].map(clean).eq("pre_ad_start_10s")] if not context.empty else pd.DataFrame()
        post = context[context["segment_type"].map(clean).eq("post_ad_end_10s")] if not context.empty else pd.DataFrame()
        rows.append(
            {
                "ad_interval_id": ad_id,
                "video_id": video_id,
                "video_title": clean(ad.get("video_title")),
                "video_filename": clean(ad.get("video_filename")),
                "ad_start_sec": start,
                "ad_start_mmss": mmss_floor(start),
                "ad_end_sec": end,
                "ad_end_mmss": mmss_floor(end),
                "start_hit_3s": bool_text(start3 > 0),
                "start_hit_5s": bool_text(start5 > 0),
                "end_hit_3s": bool_text(end3 > 0),
                "end_hit_5s": bool_text(end5 > 0),
                "both_boundary_hit_5s": bool_text(start5 > 0 and end5 > 0),
                "nearest_candidate_to_start_sec": ns_sec,
                "nearest_candidate_to_start_mmss": ns_mmss,
                "distance_to_start_candidate_sec": ns_dist,
                "nearest_candidate_to_end_sec": ne_sec,
                "nearest_candidate_to_end_mmss": ne_mmss,
                "distance_to_end_candidate_sec": ne_dist,
                "candidate_count_in_ad_interval": count_between(cands, video_id, start, end),
                "candidate_count_in_pre10": count_between(cands, video_id, float(pre.iloc[0]["segment_start_sec"]), float(pre.iloc[0]["segment_end_sec"])) if not pre.empty else 0,
                "candidate_count_in_post10": count_between(cands, video_id, float(post.iloc[0]["segment_start_sec"]), float(post.iloc[0]["segment_end_sec"])) if not post.empty else 0,
                "candidate_count_near_start_5s": start5,
                "candidate_count_near_end_5s": end5,
            }
        )
    total = len(ad_segments) * 2
    return rows, {
        "total_ad_intervals": int(len(ad_segments)),
        "total_boundaries": int(total),
        "start_hit_3s_count": int(sh3),
        "start_hit_5s_count": int(sh5),
        "end_hit_3s_count": int(eh3),
        "end_hit_5s_count": int(eh5),
        "boundary_hit_5s_count": int(sh5 + eh5),
        "boundary_hit_5s_rate": float((sh5 + eh5) / total) if total else 0.0,
        "both_boundary_hit_5s_count": int(both5),
    }


def audit_hit_set(audit_df: pd.DataFrame) -> set[str]:
    hits: set[str] = set()
    if audit_df.empty:
        return hits
    for _, row in audit_df.iterrows():
        ad_id = clean(row.get("ad_interval_id"))
        if is_true(row.get("start_hit_5s")):
            hits.add(f"{ad_id}:start")
        if is_true(row.get("end_hit_5s")):
            hits.add(f"{ad_id}:end")
    return hits


def candidate_times(df: pd.DataFrame, video_id: str) -> list[float]:
    if df.empty or "candidate_time_sec" not in df.columns:
        return []
    subset = df[df["video_id"].map(clean).eq(video_id)]
    return sorted(float(v) for v in subset["candidate_time_sec"].astype(float).tolist())


def count_overlap_times(source: list[float], target: list[float], tolerance: float) -> int:
    return sum(1 for t in source if any(abs(t - other) <= tolerance for other in target))


def create_scene_overlap_rows(videos: list[dict[str, Any]], opencv_df: pd.DataFrame, resnet_df: pd.DataFrame) -> tuple[list[dict[str, Any]], dict[str, int]]:
    rows: list[dict[str, Any]] = []
    total3 = total5 = 0
    for video in videos:
        vid = clean(video.get("video_id"))
        ocv = candidate_times(opencv_df, vid)
        res = candidate_times(resnet_df, vid)
        overlap3 = count_overlap_times(res, ocv, 3.0)
        overlap5 = count_overlap_times(res, ocv, 5.0)
        total3 += overlap3
        total5 += overlap5
        rows.append(
            {
                "video_id": vid,
                "video_title": clean(video.get("video_title")),
                "unique_opencv_candidate_count": len(ocv),
                "unique_resnet_candidate_count": len(res),
                "overlap_candidate_count_3s": overlap3,
                "overlap_candidate_count_5s": overlap5,
            }
        )
    return rows, {"overlap_candidate_count_3s": total3, "overlap_candidate_count_5s": total5}


def comparison_rows(opencv_df: pd.DataFrame, opencv_audit: pd.DataFrame, resnet_df: pd.DataFrame, resnet_audit: pd.DataFrame, overlap_totals: dict[str, int]) -> tuple[list[dict[str, Any]], int, int]:
    opencv_hits = audit_hit_set(opencv_audit)
    resnet_hits = audit_hit_set(resnet_audit)
    unique_resnet = len(resnet_hits - opencv_hits)
    unique_opencv = len(opencv_hits - resnet_hits)

    def row(method: str, cands: pd.DataFrame, audit: pd.DataFrame, notes: str) -> dict[str, Any]:
        start3 = int(audit["start_hit_3s"].map(is_true).sum()) if not audit.empty else 0
        start5 = int(audit["start_hit_5s"].map(is_true).sum()) if not audit.empty else 0
        end3 = int(audit["end_hit_3s"].map(is_true).sum()) if not audit.empty else 0
        end5 = int(audit["end_hit_5s"].map(is_true).sum()) if not audit.empty else 0
        both = int(audit["both_boundary_hit_5s"].map(is_true).sum()) if not audit.empty else 0
        total_intervals = len(audit)
        total_boundaries = total_intervals * 2
        hit = start5 + end5
        return {
            "method": method,
            "total_candidate_count": int(len(cands)),
            "start_hit_3s_count": start3,
            "start_hit_5s_count": start5,
            "end_hit_3s_count": end3,
            "end_hit_5s_count": end5,
            "boundary_hit_5s_count": hit,
            "boundary_hit_5s_rate": hit / total_boundaries if total_boundaries else 0.0,
            "both_boundary_hit_5s_count": both,
            "total_ad_intervals": total_intervals,
            "total_boundaries": total_boundaries,
            "unique_resnet_boundary_hits": unique_resnet,
            "unique_opencv_boundary_hits": unique_opencv,
            "overlap_candidate_count_3s": overlap_totals.get("overlap_candidate_count_3s", 0),
            "overlap_candidate_count_5s": overlap_totals.get("overlap_candidate_count_5s", 0),
            "notes": notes,
        }

    rows = [
        row("opencv_ffmpeg_merged_v2_4_labelrefreshed", opencv_df, opencv_audit, "Existing OpenCV/ffmpeg candidates reused; v2.4 label boundary audit only."),
        row("resnet_embedding_v2_4_labelrefreshed", resnet_df, resnet_audit, "Existing pretrained ResNet candidates reused; no embedding rerun."),
    ]
    return rows, unique_resnet, unique_opencv


def review_priority(row: dict[str, Any]) -> str:
    interval_id = clean(row.get("nearest_ad_interval_id"))
    dist = to_float(row.get("distance_to_nearest_ad_boundary_sec"))
    near5 = is_true(row.get("is_near_any_ad_boundary_5s"))
    if interval_id == "A022" and (near5 or (dist is not None and dist <= 10)):
        return "very_high"
    if near5:
        return "high"
    if interval_id == "A022":
        return "high"
    if dist is not None and dist <= 10:
        return "medium"
    return "low"


def make_candidate_review_rows(candidates: list[dict[str, Any]], method: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in candidates:
        t = to_float(row.get("candidate_time_sec"))
        if t is None:
            continue
        out = dict(row)
        out["candidate_time_sec"] = t
        out["candidate_time_mmss"] = clean(out.get("candidate_time_mmss")) or mmss_floor(t)
        out["review_clip_start_sec"] = max(0.0, t - 5.0)
        out["review_clip_end_sec"] = t + 5.0
        out["review_clip_start_mmss"] = mmss_floor(out["review_clip_start_sec"])
        out["review_clip_end_mmss"] = mmss_floor(out["review_clip_end_sec"])
        out["review_priority"] = review_priority(out)
        out["review_status"] = "not_reviewed"
        out["is_true_scene_change"] = ""
        out["scene_change_strength"] = ""
        out["scene_change_type"] = ""
        out["is_ad_boundary_related"] = ""
        out["false_positive_type"] = ""
        out["keep_as_boundary_candidate"] = ""
        out["review_note"] = clean(out.get("review_note"))
        out["reviewer"] = ""
        out["reviewed_at"] = ""
        out["review_method"] = method
        rows.append(out)
    order = {"very_high": 0, "high": 1, "medium": 2, "low": 3}
    rows.sort(key=lambda r: (order.get(clean(r.get("review_priority")), 9), clean(r.get("video_id")), to_float(r.get("candidate_time_sec")) or 0))
    return rows


def make_boundary_review_rows(audit_rows: list[dict[str, Any]], method: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in audit_rows:
        priority = "very_high" if clean(row.get("ad_interval_id")) == "A022" else ("high" if is_true(row.get("start_hit_5s")) or is_true(row.get("end_hit_5s")) else "medium")
        out = dict(row)
        out["review_method"] = method
        out["review_priority"] = priority
        out["boundary_review_status"] = "not_reviewed"
        out["start_boundary_useful"] = ""
        out["end_boundary_useful"] = ""
        out["review_note"] = ""
        out["reviewer"] = ""
        out["reviewed_at"] = ""
        rows.append(out)
    order = {"very_high": 0, "high": 1, "medium": 2, "low": 3}
    rows.sort(key=lambda r: (order.get(clean(r.get("review_priority")), 9), clean(r.get("video_id")), clean(r.get("ad_interval_id"))))
    return rows


def rows_to_df(rows: list[dict[str, Any]], columns: list[str] | None = None) -> pd.DataFrame:
    if columns is None:
        cols: list[str] = []
        for row in rows:
            for key in row.keys():
                if key not in cols:
                    cols.append(key)
        columns = cols
    return pd.DataFrame([{col: fmt(row.get(col, "")) for col in columns} for row in rows], columns=columns)


def add_validation(ws: Any, column_name: str, options: str) -> None:
    header = [cell.value for cell in ws[1]]
    if column_name not in header:
        return
    col_idx = header.index(column_name) + 1
    col = get_column_letter(col_idx)
    dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{max(ws.max_row, 2)}")


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fills = {
        "very_high": PatternFill("solid", fgColor="F4B183"),
        "high": PatternFill("solid", fgColor="FCE4D6"),
        "medium": PatternFill("solid", fgColor="FFF2CC"),
    }
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 10
            for cell in column_cells[:200]:
                max_len = max(max_len, min(len(str(cell.value or "")), 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
        add_validation(ws, "review_status", "not_reviewed,reviewed,skip")
        add_validation(ws, "boundary_review_status", "not_reviewed,reviewed,skip")
        add_validation(ws, "is_true_scene_change", "yes,no,unclear")
        add_validation(ws, "scene_change_strength", "strong,medium,weak,none")
        add_validation(ws, "is_ad_boundary_related", "yes,no,unclear")
        add_validation(ws, "keep_as_boundary_candidate", "yes,no,unclear")
        header = [cell.value for cell in ws[1]]
        if "review_priority" in header and ws.max_row >= 2:
            col = get_column_letter(header.index("review_priority") + 1)
            for priority, fill in fills.items():
                rule = FormulaRule(formula=[f'${col}2="{priority}"'], fill=fill)
                ws.conditional_formatting.add(f"A2:{get_column_letter(ws.max_column)}{ws.max_row}", rule)
    wb.save(path)


def create_review_files(candidates: list[dict[str, Any]], audit_rows: list[dict[str, Any]], xlsx_path: Path, candidate_csv: Path, boundary_csv: Path, method: str, extra_sheets: dict[str, pd.DataFrame] | None = None) -> tuple[list[str], dict[str, Any]]:
    candidate_rows = make_candidate_review_rows(candidates, method)
    boundary_rows = make_boundary_review_rows(audit_rows, method)
    candidate_cols = list(candidate_rows[0].keys()) if candidate_rows else []
    boundary_cols = list(boundary_rows[0].keys()) if boundary_rows else []
    write_csv(candidate_csv, candidate_rows, candidate_cols)
    write_csv(boundary_csv, boundary_rows, boundary_cols)
    value_rows = [
        {"field": "review_priority", "value": "very_high/high/medium/low"},
        {"field": "review_status", "value": "not_reviewed/reviewed/skip"},
        {"field": "is_true_scene_change", "value": "yes/no/unclear"},
        {"field": "keep_as_boundary_candidate", "value": "yes/no/unclear"},
    ]
    guide_rows = [
        {"section": "purpose", "guide": "v2.4 label boundary 기준으로 기존 scene candidate를 사람이 검토하기 위한 파일"},
        {"section": "note", "guide": "scene evidence audit이며 final 광고 탐지 성능 claim이 아님"},
        {"section": "new_interval", "guide": "A022 video_id=12 19분20초~20분06초 row는 high 이상 review_priority로 표시"},
    ]
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        rows_to_df(candidate_rows, candidate_cols).to_excel(writer, sheet_name=("resnet_candidate_review" if method == "resnet" else "scene_candidate_review"), index=False)
        rows_to_df(boundary_rows, boundary_cols).to_excel(writer, sheet_name=("resnet_boundary_review" if method == "resnet" else "ad_boundary_review"), index=False)
        if extra_sheets:
            for sheet, df in extra_sheets.items():
                df.to_excel(writer, sheet_name=sheet[:31], index=False)
        rows_to_df(value_rows, ["field", "value"]).to_excel(writer, sheet_name="value_options", index=False)
        rows_to_df(guide_rows, ["section", "guide"]).to_excel(writer, sheet_name="review_guide", index=False)
    style_workbook(xlsx_path)
    return [str(xlsx_path), str(candidate_csv), str(boundary_csv)], {
        "candidate_review_rows": len(candidate_rows),
        "boundary_review_rows": len(boundary_rows),
        "very_high_review_rows": sum(1 for row in candidate_rows if clean(row.get("review_priority")) == "very_high"),
        "high_review_rows": sum(1 for row in candidate_rows if clean(row.get("review_priority")) == "high"),
    }


def update_readme() -> None:
    section = """## New Ad Interval Addition and Refresh v2.4

video_id=12에 새 광고 구간 19분20초~20분06초를 추가했다. 새 구간은 전환형 광고로 보고 `is_abrupt_transition_ad=yes`로 처리했다.

v2_4 기준 label/window/segment/scene boundary audit/review 파일을 새로 생성했다. 이후 검토 및 rule 설계는 v2_4 기준 파일을 사용해야 한다.
"""
    text = README_PATH.read_text(encoding="utf-8") if README_PATH.exists() else "# youtube_ad_segment_detector\n"
    marker = "## New Ad Interval Addition and Refresh v2.4"
    if marker in text:
        before, _, after = text.partition(marker)
        next_idx = after.find("\n## ")
        text = before.rstrip() + "\n\n" + section + (after[next_idx:] if next_idx >= 0 else "")
    else:
        text = text.rstrip() + "\n\n" + section
    README_PATH.write_text(text.rstrip() + "\n", encoding="utf-8")


def latest_copy(files: list[Path]) -> tuple[bool, list[str]]:
    ensure_inside_project(LATEST_DIR)
    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    for child in LATEST_DIR.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()
    copied: list[str] = []
    forbidden_suffixes = {".mp4", ".jpg", ".jpeg", ".png", ".webp", ".pt", ".pth", ".ckpt", ".bin"}
    allowed_xlsx = {OPENCV_REVIEW_XLSX.name, RESNET_REVIEW_XLSX.name}
    for src in files:
        if not src.exists():
            continue
        if src.suffix.lower() in forbidden_suffixes:
            raise RuntimeError(f"Refusing to copy forbidden latest file: {src}")
        if src.suffix.lower() in {".xlsx", ".xls"} and src.name not in allowed_xlsx:
            raise RuntimeError(f"Refusing to copy non-review xlsx to latest: {src}")
        dst = LATEST_DIR / src.name
        shutil.copy2(src, dst)
        copied.append(str(dst))
    LATEST_README.write_text(
        "# latest_for_chatgpt files\n\n"
        "이번 작업명: add_new_ad_interval_and_refresh_v2_4\n\n"
        "v2_4 label/window/segment/audit/review/report/log/script 핵심 파일만 복사했다. mp4, 원본 xlsx, frame, model, checkpoint, cache 파일은 복사하지 않는다. 생성된 review xlsx 2개만 포함한다.\n",
        encoding="utf-8",
    )
    copied.append(str(LATEST_README))
    return True, copied


def summary_text(report: dict[str, Any]) -> str:
    warnings = report.get("warnings", [])
    warning_text = "\n".join(f"- {w}" for w in warnings) if warnings else "- 주요 warning 없음"
    return f"""# add_new_ad_interval_and_refresh_v2_4 summary

## 새 광고 구간

- added: {report.get('new_ad_interval_segment_included')}
- ad_interval_id: {report.get('new_ad_interval_id')}
- video_id: {report.get('new_ad_interval_video_id')}
- start/end: {report.get('new_ad_interval_start_sec')} / {report.get('new_ad_interval_end_sec')}
- mapping_status: {report.get('new_ad_interval_mapping_status')}
- overlap window count: {report.get('new_ad_interval_overlap_window_count')}

## Window / Segment

- label rows before/after: {report.get('label_row_count_before_v2_3')} / {report.get('label_row_count_after_v2_4')}
- ad interval segments before/after: {report.get('ad_interval_segment_count_before_v2_3')} / {report.get('ad_interval_segment_count_after_v2_4')}
- combined segments before/after: {report.get('combined_segment_count_before_v2_3')} / {report.get('combined_segment_count_after_v2_4')}
- window scope before: `{report.get('window_label_scope_distribution_before_v2_3')}`
- window scope after: `{report.get('window_label_scope_distribution_after_v2_4')}`
- clipped segments v2_4: {report.get('clipped_segment_count_v2_4')}
- ad interval overlaps v2_4: {report.get('ad_interval_overlap_count_v2_4')}

## Scene Boundary Audit

- total ad intervals / boundaries: {report.get('total_ad_intervals_v2_4')} / {report.get('total_boundaries_v2_4')}
- OpenCV/ffmpeg ±5s hit before/after: {report.get('opencv_boundary_hit_5s_before_v2_3')} / {report.get('opencv_boundary_hit_5s_after_v2_4')}
- ResNet ±5s hit before/after: {report.get('resnet_boundary_hit_5s_before_v2_3')} / {report.get('resnet_boundary_hit_5s_after_v2_4')}
- New interval OpenCV start/end hit@5s: {report.get('new_interval_opencv_start_hit_5s')} / {report.get('new_interval_opencv_end_hit_5s')}
- New interval ResNet start/end hit@5s: {report.get('new_interval_resnet_start_hit_5s')} / {report.get('new_interval_resnet_end_hit_5s')}

이 값은 scene evidence audit이며 final 광고 탐지 성능 claim이 아니다.

## Review

- OpenCV/ffmpeg review xlsx: `{OPENCV_REVIEW_XLSX}`
- ResNet review xlsx: `{RESNET_REVIEW_XLSX}`

## Warning

{warning_text}

## 다음 작업 추천

A022 주변 OpenCV/ResNet candidate review를 먼저 확인한 뒤, v2_4 기준 visual evidence 결합 rule을 갱신하는 것을 권장한다.
"""


def write_report_summary_log(report: dict[str, Any]) -> None:
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    SUMMARY_PATH.write_text(summary_text(report), encoding="utf-8")
    LOG_PATH.write_text("\n".join(RUN_LOG) + "\n", encoding="utf-8")


def base_candidate_columns(rows: list[dict[str, Any]]) -> list[str]:
    preferred = [
        "merged_candidate_id", "video_id", "video_title", "video_filename", "video_path", "candidate_time_sec", "candidate_time_mmss", "candidate_time_mmss_floor", "candidate_time_mmss_round", "scene_change_score", "threshold", "cosine_distance", "l2_distance", "score_rank_in_video", "score_percentile_in_video", "candidate_source", "method_used", "model_name", "feature_dim", "nearest_window_id", "nearest_ad_interval_id", "nearest_ad_boundary_type", "nearest_ad_boundary_sec", "nearest_ad_boundary_mmss", "distance_to_nearest_ad_boundary_sec", "is_near_ad_start_3s", "is_near_ad_start_5s", "is_near_ad_end_3s", "is_near_ad_end_5s", "is_near_any_ad_boundary_5s", "label_refresh_version", "label_refresh_source", "review_note"
    ]
    cols: list[str] = []
    for col in preferred:
        if any(col in row for row in rows):
            cols.append(col)
    for row in rows:
        for key in row.keys():
            if key not in cols and not key.startswith("Unnamed"):
                cols.append(key)
    return cols


def main() -> None:
    ensure_dirs()
    start_dt = datetime.now().astimezone()
    start_time = start_dt.isoformat(timespec="seconds")
    started = time.time()
    warnings: list[Any] = []
    errors: list[Any] = []
    before_old = old_project_snapshot()
    log(f"작업 시작 전 예상 작업 시간: {ESTIMATED_RUNTIME}")
    log(f"예상 근거: {RUNTIME_ESTIMATION_REASON}")
    log(f"작업 시작 시각: {start_time}")

    report: dict[str, Any] = {
        "project_root": str(PROJECT_ROOT),
        "estimated_runtime": ESTIMATED_RUNTIME,
        "runtime_estimation_reason": RUNTIME_ESTIMATION_REASON,
        "start_time": start_time,
        "end_time": "",
        "actual_runtime_seconds": 0,
        "actual_runtime_readable": "",
        "cv_environment_checked": False,
        "python_executable": sys.executable,
        "old_project_modified": False,
        "warnings": warnings,
        "errors": errors,
    }

    try:
        step(1, "작업 시간 기록 및 cv 환경 확인")
        cv_ok, python_executable = verify_cv_environment(warnings)
        report["cv_environment_checked"] = cv_ok
        report["python_executable"] = python_executable
        if not cv_ok:
            errors.append("cv_environment_check_failed")
            raise RuntimeError("cv environment check failed")

        step(2, "입력 파일 탐색 및 현재 v2_3 상태 확인")
        labels = read_csv(INPUT_LABEL, True, warnings, errors)
        manifest = read_csv(MANIFEST_PATH, True, warnings, errors)
        window_grid = read_csv(WINDOW_GRID_PATH, True, warnings, errors)
        window_labels_v2_3 = read_csv(WINDOW_LABELS_V2_3, False, warnings, errors)
        ad_segments_v2_3 = read_csv(AD_SEGMENTS_V2_3, False, warnings, errors)
        combined_v2_3 = read_csv(COMBINED_SEGMENTS_V2_3, False, warnings, errors)
        opencv_candidates_v2_3 = read_csv(OPENCV_CANDIDATES_V2_3, False, warnings, errors)
        opencv_audit_v2_3 = read_csv(OPENCV_AUDIT_V2_3, False, warnings, errors)
        resnet_candidates_v2_3 = read_csv(RESNET_CANDIDATES_V2_3, False, warnings, errors)
        resnet_audit_v2_3 = read_csv(RESNET_AUDIT_V2_3, False, warnings, errors)
        if labels.empty or manifest.empty or window_grid.empty:
            raise RuntimeError("required label/manifest/window_grid input missing")
        mapping_key, mapping_by_id, mapping_by_filename = manifest_lookup(manifest)
        report_v2_3 = load_json(REPORT_V2_3)

        step(3, "video_id=12 manifest/title 매칭 확인")
        video12_mapping = mapping_by_id.get(NEW_VIDEO_ID, {})
        if not video12_mapping:
            errors.append("video_id_12_not_found_in_manifest")
            raise RuntimeError("video_id=12 not found in manifest")
        if clean(video12_mapping.get("mapped_video_title")) != NEW_VIDEO_TITLE:
            warnings.append({"video12_title_mismatch": {"manifest": video12_mapping.get("mapped_video_title"), "new": NEW_VIDEO_TITLE}})
        if float(video12_mapping.get("video_duration_sec") or 0) < NEW_END_SEC:
            errors.append("new_interval_end_exceeds_video_duration")
            raise RuntimeError("new interval end exceeds video duration")

        step(4, "새 광고 interval row 생성 및 v2_4 label 생성")
        labels_v2_4, addition_audit = add_new_label_row(labels, mapping_by_id)
        write_df(OUTPUT_LABEL, labels_v2_4)
        write_csv(NEW_LABEL_AUDIT, [addition_audit], ["new_ad_interval_id", "video_id", "video_title", "ad_start_sec", "ad_end_sec", "ad_duration_sec", "manifest_video_found", "video_duration_sec", "end_within_video_duration", "overlaps_existing_interval", "overlap_with_ad_interval_ids", "add_status", "add_warning"])
        new_ad_interval_id = clean(addition_audit["new_ad_interval_id"])

        step(5, "label-to-video mapping audit 생성")
        mapping_audit_rows = create_mapping_audit(labels_v2_4, mapping_key)
        write_csv(MAPPING_AUDIT, mapping_audit_rows, ["ad_interval_id", "label_video_id", "label_video_title", "matched_manifest_video_id", "matched_video_filename", "matched_video_path", "video_duration_sec", "mapping_status", "mapping_warning"])
        new_mapping = next((row for row in mapping_audit_rows if row.get("ad_interval_id") == new_ad_interval_id), {})

        step(6, "window_labels_5s_v2_4 재계산")
        window_labels_v2_4 = create_window_labels(window_grid, labels_v2_4, mapping_by_filename, mapping_key)
        write_df(WINDOW_LABELS_V2_4, window_labels_v2_4)

        step(7, "ad/context/combined segment v2_4 재생성")
        ad_rows, context_rows, combined_rows, clipped_rows, empty_segment_count = create_segments(labels_v2_4, mapping_key, warnings)
        segment_cols = ["segment_id", "segment_type", "boundary_role", "video_id", "video_title", "video_filename", "video_path", "file_stem", "ad_interval_id", "source_label_row_index", "segment_start_sec", "segment_end_sec", "segment_duration_sec", "context_window_sec", "ad_start_sec", "ad_end_sec", "is_abrupt_transition_ad_original", "is_abrupt_transition_ad_refreshed", "scope_value_source", "label_valid", "video_duration_sec", "label_mapping_status", "segment_refresh_version", "segment_clipped", "clipping_reason", "segment_valid"]
        write_csv(AD_SEGMENTS_V2_4, ad_rows, segment_cols)
        write_csv(CONTEXT_SEGMENTS_V2_4, context_rows, segment_cols)
        write_csv(COMBINED_SEGMENTS_V2_4, combined_rows, segment_cols)
        ad_segments_v2_4 = pd.DataFrame(ad_rows)
        combined_v2_4 = pd.DataFrame(combined_rows)

        step(8, "clipped/overlap/new interval audit 생성")
        overlap_rows, actual_overlap_count = create_overlap_audit(labels_v2_4, mapping_key)
        write_csv(CLIPPED_AUDIT, clipped_rows, ["source_version", "segment_id", "segment_type", "video_id", "video_title", "video_filename", "ad_interval_id", "original_segment_start_sec", "original_segment_end_sec", "clipped_segment_start_sec", "clipped_segment_end_sec", "video_duration_sec", "clipping_applied", "clipping_reason", "needs_manual_review"])
        write_csv(OVERLAP_AUDIT, overlap_rows, ["video_id", "video_title", "ad_interval_id_1", "ad_interval_id_2", "interval_1_start_sec", "interval_1_end_sec", "interval_2_start_sec", "interval_2_end_sec", "overlap_sec", "overlap_ratio_min_duration", "overlap_warning"])
        new_windows = window_labels_v2_4[window_labels_v2_4["matched_ad_interval_ids"].astype(str).str.split(";").map(lambda xs: new_ad_interval_id in xs)]
        new_ad_exists = any(row.get("ad_interval_id") == new_ad_interval_id and row.get("segment_type") == "ad_interval" for row in ad_rows)
        new_pre_exists = any(row.get("ad_interval_id") == new_ad_interval_id and row.get("segment_type") == "pre_ad_start_10s" for row in context_rows)
        new_post_exists = any(row.get("ad_interval_id") == new_ad_interval_id and row.get("segment_type") == "post_ad_end_10s" for row in context_rows)
        new_segment_clipped = any(row.get("ad_interval_id") == new_ad_interval_id for row in clipped_rows)
        inclusion_warning: list[str] = []
        if new_windows.empty:
            inclusion_warning.append("no_overlap_windows")
        if not new_ad_exists:
            inclusion_warning.append("ad_interval_segment_missing")
        if not new_pre_exists:
            inclusion_warning.append("pre_segment_missing")
        if not new_post_exists:
            inclusion_warning.append("post_segment_missing")
        inclusion_row = {
            "new_ad_interval_id": new_ad_interval_id,
            "video_id": NEW_VIDEO_ID,
            "video_title": NEW_VIDEO_TITLE,
            "ad_start_sec": NEW_START_SEC,
            "ad_end_sec": NEW_END_SEC,
            "ad_duration_sec": NEW_DURATION_SEC,
            "matched_manifest_video_filename": video12_mapping.get("matched_video_filename", ""),
            "video_duration_sec": video12_mapping.get("video_duration_sec", ""),
            "overlap_window_count": int(len(new_windows)),
            "abrupt_ad_window_count": int((new_windows["window_label_scope"] == "abrupt_ad").sum()) if not new_windows.empty else 0,
            "mixed_or_boundary_window_count": int((new_windows["window_label_scope"] == "mixed_or_boundary").sum()) if not new_windows.empty else 0,
            "ad_interval_segment_exists": bool_text(new_ad_exists),
            "pre_segment_exists": bool_text(new_pre_exists),
            "post_segment_exists": bool_text(new_post_exists),
            "segment_clipped": bool_text(new_segment_clipped),
            "inclusion_status": "included" if not inclusion_warning else "warning",
            "inclusion_warning": ";".join(inclusion_warning),
        }
        write_csv(NEW_INTERVAL_INCLUSION_AUDIT, [inclusion_row], ["new_ad_interval_id", "video_id", "video_title", "ad_start_sec", "ad_end_sec", "ad_duration_sec", "matched_manifest_video_filename", "video_duration_sec", "overlap_window_count", "abrupt_ad_window_count", "mixed_or_boundary_window_count", "ad_interval_segment_exists", "pre_segment_exists", "post_segment_exists", "segment_clipped", "inclusion_status", "inclusion_warning"])

        step(9, "OpenCV/ffmpeg scene candidate v2_4 boundary audit 재계산")
        opencv_rows = refresh_candidates(opencv_candidates_v2_3, ad_segments_v2_4, window_labels_v2_4, "opencv_ffmpeg")
        opencv_cols = base_candidate_columns(opencv_rows)
        write_csv(OPENCV_CANDIDATES_V2_4, opencv_rows, opencv_cols)
        write_csv(OPENCV_CANDIDATES_MMSS_V2_4, opencv_rows, opencv_cols)
        opencv_audit_rows, opencv_summary = boundary_audit(opencv_rows, ad_segments_v2_4, combined_v2_4)
        audit_cols = ["ad_interval_id", "video_id", "video_title", "video_filename", "ad_start_sec", "ad_start_mmss", "ad_end_sec", "ad_end_mmss", "start_hit_3s", "start_hit_5s", "end_hit_3s", "end_hit_5s", "both_boundary_hit_5s", "nearest_candidate_to_start_sec", "nearest_candidate_to_start_mmss", "distance_to_start_candidate_sec", "nearest_candidate_to_end_sec", "nearest_candidate_to_end_mmss", "distance_to_end_candidate_sec", "candidate_count_in_ad_interval", "candidate_count_in_pre10", "candidate_count_in_post10", "candidate_count_near_start_5s", "candidate_count_near_end_5s"]
        write_csv(OPENCV_AUDIT_V2_4, opencv_audit_rows, audit_cols)

        step(10, "ResNet scene candidate v2_4 boundary audit 재계산")
        resnet_rows = refresh_candidates(resnet_candidates_v2_3, ad_segments_v2_4, window_labels_v2_4, "resnet")
        resnet_cols = base_candidate_columns(resnet_rows)
        write_csv(RESNET_CANDIDATES_V2_4, resnet_rows, resnet_cols)
        write_csv(RESNET_CANDIDATES_MMSS_V2_4, resnet_rows, resnet_cols)
        resnet_audit_rows, resnet_summary = boundary_audit(resnet_rows, ad_segments_v2_4, combined_v2_4)
        write_csv(RESNET_AUDIT_V2_4, resnet_audit_rows, audit_cols)
        videos = [{"video_id": row.get("video_id"), "video_title": row.get("video_title")} for _, row in manifest.iterrows()]
        overlap_scene_rows, overlap_totals = create_scene_overlap_rows(videos, pd.DataFrame(opencv_rows), pd.DataFrame(resnet_rows))
        write_csv(OVERLAP_SCENE_V2_4, overlap_scene_rows, ["video_id", "video_title", "unique_opencv_candidate_count", "unique_resnet_candidate_count", "overlap_candidate_count_3s", "overlap_candidate_count_5s"])
        comparison, unique_resnet_hits, unique_opencv_hits = comparison_rows(pd.DataFrame(opencv_rows), pd.DataFrame(opencv_audit_rows), pd.DataFrame(resnet_rows), pd.DataFrame(resnet_audit_rows), overlap_totals)
        comparison_cols = ["method", "total_candidate_count", "start_hit_3s_count", "start_hit_5s_count", "end_hit_3s_count", "end_hit_5s_count", "boundary_hit_5s_count", "boundary_hit_5s_rate", "both_boundary_hit_5s_count", "total_ad_intervals", "total_boundaries", "unique_resnet_boundary_hits", "unique_opencv_boundary_hits", "overlap_candidate_count_3s", "overlap_candidate_count_5s", "notes"]
        write_csv(COMPARISON_V2_4, comparison, comparison_cols)

        step(11, "review용 CSV/Excel v2_4 갱신")
        opencv_review_files, opencv_review_stats = create_review_files(opencv_rows, opencv_audit_rows, OPENCV_REVIEW_XLSX, OPENCV_REVIEW_CANDIDATE_CSV, OPENCV_REVIEW_BOUNDARY_CSV, "opencv")
        resnet_only_rows = []
        opencv_hit_set = audit_hit_set(pd.DataFrame(opencv_audit_rows))
        resnet_hit_set = audit_hit_set(pd.DataFrame(resnet_audit_rows))
        for hit in sorted(resnet_hit_set - opencv_hit_set):
            ad_id, boundary_type = hit.split(":")
            row = next((r for r in resnet_audit_rows if r.get("ad_interval_id") == ad_id), {})
            resnet_only_rows.append({"ad_interval_id": ad_id, "boundary_type": boundary_type, **row})
        write_csv(RESNET_ONLY_BOUNDARY_REVIEW, resnet_only_rows, list(resnet_only_rows[0].keys()) if resnet_only_rows else ["ad_interval_id", "boundary_type"])
        comparison_review_rows = [dict(row) for row in comparison]
        write_csv(COMPARISON_REVIEW, comparison_review_rows, comparison_cols)
        extra = {"resnet_only_boundary_review": rows_to_df(resnet_only_rows), "opencv_resnet_comparison": rows_to_df(comparison_review_rows, comparison_cols)}
        resnet_review_files, resnet_review_stats = create_review_files(resnet_rows, resnet_audit_rows, RESNET_REVIEW_XLSX, RESNET_REVIEW_CANDIDATE_CSV, RESNET_REVIEW_BOUNDARY_CSV, "resnet", extra)
        review_files_created = opencv_review_files + resnet_review_files + [str(RESNET_ONLY_BOUNDARY_REVIEW), str(COMPARISON_REVIEW)]

        step(12, "report/summary/log/README/latest_for_chatgpt 갱신")
        new_opencv_audit = next((row for row in opencv_audit_rows if row.get("ad_interval_id") == new_ad_interval_id), {})
        new_resnet_audit = next((row for row in resnet_audit_rows if row.get("ad_interval_id") == new_ad_interval_id), {})
        window_scope_before = distribution(window_labels_v2_3["window_label_scope"]) if not window_labels_v2_3.empty and "window_label_scope" in window_labels_v2_3.columns else {}
        window_scope_after = distribution(window_labels_v2_4["window_label_scope"])
        before_old_after = old_project_snapshot()
        old_modified = before_old != before_old_after
        if old_modified:
            warnings.append({"old_project_modified": True})
        generated_files = [
            OUTPUT_LABEL, NEW_LABEL_AUDIT, MAPPING_AUDIT, WINDOW_LABELS_V2_4, AD_SEGMENTS_V2_4, CONTEXT_SEGMENTS_V2_4, COMBINED_SEGMENTS_V2_4, NEW_INTERVAL_INCLUSION_AUDIT, CLIPPED_AUDIT, OVERLAP_AUDIT, OPENCV_CANDIDATES_V2_4, OPENCV_CANDIDATES_MMSS_V2_4, OPENCV_AUDIT_V2_4, RESNET_CANDIDATES_V2_4, RESNET_CANDIDATES_MMSS_V2_4, RESNET_AUDIT_V2_4, COMPARISON_V2_4, OVERLAP_SCENE_V2_4, OPENCV_REVIEW_XLSX, OPENCV_REVIEW_CANDIDATE_CSV, OPENCV_REVIEW_BOUNDARY_CSV, RESNET_REVIEW_XLSX, RESNET_REVIEW_CANDIDATE_CSV, RESNET_REVIEW_BOUNDARY_CSV, RESNET_ONLY_BOUNDARY_REVIEW, COMPARISON_REVIEW, REPORT_PATH, SUMMARY_PATH, LOG_PATH, SCRIPT_PATH]
        latest_files = [OUTPUT_LABEL, NEW_LABEL_AUDIT, WINDOW_LABELS_V2_4, AD_SEGMENTS_V2_4, CONTEXT_SEGMENTS_V2_4, COMBINED_SEGMENTS_V2_4, NEW_INTERVAL_INCLUSION_AUDIT, CLIPPED_AUDIT, OVERLAP_AUDIT, OPENCV_AUDIT_V2_4, RESNET_AUDIT_V2_4, COMPARISON_V2_4, OPENCV_REVIEW_XLSX, RESNET_REVIEW_XLSX, REPORT_PATH, SUMMARY_PATH, LOG_PATH, SCRIPT_PATH]
        update_readme()
        end_time = datetime.now().astimezone().isoformat(timespec="seconds")
        elapsed = time.time() - started
        report.update(
            {
                "end_time": end_time,
                "actual_runtime_seconds": elapsed,
                "actual_runtime_readable": readable(elapsed),
                "old_project_modified": old_modified,
                "input_label_path": str(INPUT_LABEL),
                "output_label_v2_4_path": str(OUTPUT_LABEL),
                "new_ad_interval_id": new_ad_interval_id,
                "new_ad_interval_video_id": NEW_VIDEO_ID,
                "new_ad_interval_start_sec": NEW_START_SEC,
                "new_ad_interval_end_sec": NEW_END_SEC,
                "new_ad_interval_duration_sec": NEW_DURATION_SEC,
                "new_ad_interval_mapping_status": new_mapping.get("mapping_status", ""),
                "new_ad_interval_overlap_window_count": int(inclusion_row["overlap_window_count"]),
                "new_ad_interval_segment_included": bool(new_ad_exists and new_pre_exists and new_post_exists),
                "new_ad_interval_ad_segment_exists": bool(new_ad_exists),
                "new_ad_interval_pre_segment_exists": bool(new_pre_exists),
                "new_ad_interval_post_segment_exists": bool(new_post_exists),
                "label_row_count_before_v2_3": int(len(labels)),
                "label_row_count_after_v2_4": int(len(labels_v2_4)),
                "ad_interval_segment_count_before_v2_3": int(len(ad_segments_v2_3)) if not ad_segments_v2_3.empty else 0,
                "ad_interval_segment_count_after_v2_4": int(len(ad_rows)),
                "combined_segment_count_before_v2_3": int(len(combined_v2_3)) if not combined_v2_3.empty else 0,
                "combined_segment_count_after_v2_4": int(len(combined_rows)),
                "window_count": int(len(window_labels_v2_4)),
                "window_label_scope_distribution_before_v2_3": window_scope_before,
                "window_label_scope_distribution_after_v2_4": window_scope_after,
                "clipped_segment_count_v2_4": int(len(clipped_rows)),
                "ad_interval_overlap_count_v2_4": int(actual_overlap_count),
                "opencv_boundary_hit_5s_before_v2_3": int(opencv_audit_v2_3["start_hit_5s"].map(is_true).sum() + opencv_audit_v2_3["end_hit_5s"].map(is_true).sum()) if not opencv_audit_v2_3.empty else 0,
                "opencv_boundary_hit_5s_after_v2_4": int(opencv_summary["boundary_hit_5s_count"]),
                "resnet_boundary_hit_5s_before_v2_3": int(resnet_audit_v2_3["start_hit_5s"].map(is_true).sum() + resnet_audit_v2_3["end_hit_5s"].map(is_true).sum()) if not resnet_audit_v2_3.empty else 0,
                "resnet_boundary_hit_5s_after_v2_4": int(resnet_summary["boundary_hit_5s_count"]),
                "total_ad_intervals_v2_4": int(opencv_summary["total_ad_intervals"]),
                "total_boundaries_v2_4": int(opencv_summary["total_boundaries"]),
                "new_interval_opencv_start_hit_5s": new_opencv_audit.get("start_hit_5s", ""),
                "new_interval_opencv_end_hit_5s": new_opencv_audit.get("end_hit_5s", ""),
                "new_interval_resnet_start_hit_5s": new_resnet_audit.get("start_hit_5s", ""),
                "new_interval_resnet_end_hit_5s": new_resnet_audit.get("end_hit_5s", ""),
                "new_interval_opencv_audit": new_opencv_audit,
                "new_interval_resnet_audit": new_resnet_audit,
                "unique_resnet_boundary_hits": unique_resnet_hits,
                "unique_opencv_boundary_hits": unique_opencv_hits,
                "review_files_created": review_files_created,
                "opencv_review_stats": opencv_review_stats,
                "resnet_review_stats": resnet_review_stats,
                "generated_files": [str(path) for path in generated_files],
                "latest_for_chatgpt_updated": False,
                "latest_for_chatgpt_files": [],
                "sub_agent_results": {
                    "sub_agent_1_label": "PENDING_EXTERNAL_REVIEW",
                    "sub_agent_2_window_segment": "PENDING_EXTERNAL_REVIEW",
                    "sub_agent_3_scene_audit_review": "PENDING_EXTERNAL_REVIEW",
                    "sub_agent_4_safety_output": "PENDING_EXTERNAL_REVIEW",
                },
                "warnings": warnings,
                "errors": errors,
            }
        )
        write_report_summary_log(report)
        latest_ok, latest_copied = latest_copy(latest_files)
        report["latest_for_chatgpt_updated"] = latest_ok
        report["latest_for_chatgpt_files"] = latest_copied
        write_report_summary_log(report)
        for src in [REPORT_PATH, SUMMARY_PATH, LOG_PATH]:
            shutil.copy2(src, LATEST_DIR / src.name)
        log(f"작업 종료 시각: {end_time}")
        log(f"실제 작업 시간: {readable(elapsed)}")
        write_report_summary_log(report)
        if latest_ok:
            shutil.copy2(LOG_PATH, LATEST_DIR / LOG_PATH.name)
        print(json.dumps({"new_ad_interval_id": new_ad_interval_id, "label_rows": [len(labels), len(labels_v2_4)], "ad_segments": [len(ad_segments_v2_3), len(ad_rows)], "combined_segments": [len(combined_v2_3), len(combined_rows)], "opencv_hit_5s": opencv_summary["boundary_hit_5s_count"], "resnet_hit_5s": resnet_summary["boundary_hit_5s_count"], "actual_runtime_readable": readable(elapsed), "errors": errors}, ensure_ascii=False, indent=2))
    except Exception as exc:
        errors.append({"fatal_error": str(exc)})
        end_time = datetime.now().astimezone().isoformat(timespec="seconds")
        elapsed = time.time() - started
        report.update({"end_time": end_time, "actual_runtime_seconds": elapsed, "actual_runtime_readable": readable(elapsed), "warnings": warnings, "errors": errors})
        write_report_summary_log(report)
        log(f"작업 종료 시각: {end_time}")
        log(f"실제 작업 시간: {readable(elapsed)}")
        raise


if __name__ == "__main__":
    main()
